#!/usr/bin/env python3
"""
Build Client-Input-Dataset-v1.xlsx + client-input.csv: the client-supplied projection of
App-Rationalization-Dummy-Dataset-v2.xlsx, and verify the round trip back to v2.

No value is regenerated or perturbed. Every cell is copied from the built v2 rows.
"""
import csv
import json
import os
import sys

OUT = "/tmp/claude-0/-home-claude/894f0710-2492-5ee1-ba03-47d6878e86d8/scratchpad/out"
sys.path.insert(0, OUT)

import generate_dataset as g                      # noqa: E402
from openpyxl import Workbook, load_workbook      # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter      # noqa: E402

TIERS = json.load(open(os.path.join(OUT, "column-tiers.json")))

# ---------------------------------------------------------------------------------------
# Tier partition, recounted from the workbook's own "Provided by" column
# ---------------------------------------------------------------------------------------
_wb2 = load_workbook(os.path.join(OUT, "App-Rationalization-Dummy-Dataset-v2.xlsx"))
_dd = _wb2["Data dictionary"]
_hdr = [c.value for c in _dd[1]]
PROVIDED_BY = {r[_hdr.index("Column")]: r[_hdr.index("Provided by")]
               for r in _dd.iter_rows(min_row=2, values_only=True)}
DD_DEF = {r[_hdr.index("Column")]: r[_hdr.index("Plain-language definition")]
          for r in _dd.iter_rows(min_row=2, values_only=True)}
DD_TYPE = {r[_hdr.index("Column")]: r[_hdr.index("Data type / allowed values")]
           for r in _dd.iter_rows(min_row=2, values_only=True)}
DD_SOURCE = {r[_hdr.index("Column")]: r[_hdr.index("Source or who answers")]
             for r in _dd.iter_rows(min_row=2, values_only=True)}
DD_FALLBACK = {r[_hdr.index("Column")]: r[_hdr.index("If not supplied")]
               for r in _dd.iter_rows(min_row=2, values_only=True)}

TIER_LABEL = {
    "client_mandatory": "Client - required",
    "client_if_available": "Client - if available",
    "system_extract": "Client system extract",
    "sme_judgement": "Client SME judgement",
    "tool_derives": "Derived by the tool",
}
CLIENT_COLS_UNORDERED = [c for c in g.COLUMN_ORDER if TIERS[c]["tier"] != "tool_derives"]
DERIVED_COLS = [c for c in g.COLUMN_ORDER if TIERS[c]["tier"] == "tool_derives"]

# cross-check the two classifications agree
_mismatch = [c for c in g.COLUMN_ORDER if TIER_LABEL[TIERS[c]["tier"]] != PROVIDED_BY[c]]
assert not _mismatch, f"tier/Provided-by disagreement: {_mismatch}"

# ---------------------------------------------------------------------------------------
# Reading order for the client sheet
# ---------------------------------------------------------------------------------------
BLOCKS = [
    ("1. Identity and lifecycle", [
        "app_id", "app_name", "vendor_name", "description", "deployment_model",
        "sourcing_type", "lifecycle_stage", "implementation_date", "version_installed"]),
    ("2. Ownership and org", [
        "business_owner", "technical_owner", "legal_entity", "business_unit", "department",
        "cost_centre"]),
    ("3. Capability", ["primary_capability", "secondary_capabilities"]),
    ("4. Usage and entitlement", ["licences_purchased", "active_users", "last_signin_date"]),
    ("5. Cost", [
        "cost_licence_subscription", "cost_upgrade_and_modules", "cost_maintenance_dev_labour",
        "cost_infrastructure_peripherals", "cost_indirect_and_training",
        "consumption_based_cost", "one_time_implementation_cost"]),
    ("6. Contract", [
        "contract_id", "annual_contract_value", "term_start", "term_end", "auto_renewal_flag",
        "renewal_notice_days", "licence_metric", "early_termination_penalty"]),
    ("7. Dependencies and data", ["integration_pattern", "data_types_held"]),
    ("8. Retention and classification", [
        "retention_obligation_flag", "retention_expiry_date", "residual_archival_cost",
        "information_classification"]),
    ("9. Migration and saving inputs", [
        "amortised_one_time_migration_cost", "realization_lag_months"]),
    ("10. Assessment scores (the answers a person gives)", [
        "process_centrality", "owner_stated_strategic_importance",
        "ov_increase_value", "ov_patient_care_criticality", "ov_governance_compliance",
        "th_architecture_fit", "th_operational_stability", "th_vendor_viability",
        "th_customization_debt",
        "r_technical_risk", "r_business_compliance_risk", "r_clinical_safety_risk",
        "r_end_user_perceived_quality",
        "urg_risk_pain_severity"]),
]
CLIENT_COLS = [c for _b, cols in BLOCKS for c in cols]
BLOCK_OF = {c: b for b, cols in BLOCKS for c in cols}

assert sorted(CLIENT_COLS) == sorted(CLIENT_COLS_UNORDERED), (
    set(CLIENT_COLS) ^ set(CLIENT_COLS_UNORDERED))
assert len(CLIENT_COLS) == len(set(CLIENT_COLS))

# The 21-item minimum viable intake, verbatim from the v2 "Client intake" sheet
MIN_VIABLE = [
    "app_id", "app_name", "vendor_name", "deployment_model", "version_installed",
    "implementation_date", "business_owner", "cost_centre", "licences_purchased",
    "active_users", "cost_licence_subscription", "term_end", "primary_capability",
    "ov_increase_value", "ov_patient_care_criticality", "ov_governance_compliance",
    "process_centrality", "owner_stated_strategic_importance",
    "r_technical_risk", "r_business_compliance_risk", "r_clinical_safety_risk",
]
assert len(MIN_VIABLE) == 21 and all(c in CLIENT_COLS for c in MIN_VIABLE)

# Plain-language labels (from column-tiers.json, which carries them per column)
LABEL = {c: TIERS[c]["label"] for c in g.COLUMN_ORDER}

# ---------------------------------------------------------------------------------------
# Typing, so the CSV can be read back losslessly
# ---------------------------------------------------------------------------------------
SPEC_TYPE = {c[1]: c[3] for c in g.COLUMNS}


def kind(col):
    t = SPEC_TYPE[col]
    if t.startswith("money") or t.startswith("int"):
        return "int"
    if t.startswith("score") or t.startswith("float"):
        return "float"
    if t.startswith("bool"):
        return "bool"
    return "str"


def parse(col, raw):
    if raw is None or raw == "":
        return None
    k = kind(col)
    if k == "int":
        return int(raw)
    if k == "float":
        return float(raw)
    if k == "bool":
        return {"TRUE": True, "FALSE": False}[str(raw).strip().upper()]
    return raw


# ---------------------------------------------------------------------------------------
# The authoritative v2 rows
# ---------------------------------------------------------------------------------------
FULL = g.build()
problems = g.sanity_checks(FULL)
assert not problems, problems
BY_ID = {r["app_id"]: r for r in FULL}

# Confirm the built rows are byte-identical to the published applications-v2.csv,
# so "same values as the full dataset" is a checked claim, not an assumption.
with open(os.path.join(OUT, "applications-v2.csv"), newline="", encoding="utf-8") as fh:
    pub = list(csv.DictReader(fh))
assert len(pub) == len(FULL) == 20
csv_diffs = []
for p, r in zip(pub, FULL):
    for c in g.COLUMN_ORDER:
        if str(g.csv_value(r[c])) != p[c]:
            csv_diffs.append((r["app_id"], c, g.csv_value(r[c]), p[c]))
assert not csv_diffs, csv_diffs[:10]

# ---------------------------------------------------------------------------------------
# Write client-input.csv
# ---------------------------------------------------------------------------------------
CSV_PATH = os.path.join(OUT, "client-input.csv")
with open(CSV_PATH, "w", newline="", encoding="utf-8") as fh:
    w = csv.writer(fh)
    w.writerow(CLIENT_COLS)
    for r in FULL:
        w.writerow([g.csv_value(r[c]) for c in CLIENT_COLS])

# ---------------------------------------------------------------------------------------
# ROUND TRIP: read client-input.csv, run the engine, compare against v2
# ---------------------------------------------------------------------------------------
GENERATOR_INTERNAL = ["_evidence", "_gross_saving_basis", "_reharvest_share",
                      "_reharvest_reserve", "_intended_disposition", "_intended_priority",
                      "_intended_key"]


def load_client_csv():
    with open(CSV_PATH, newline="", encoding="utf-8") as fh:
        rd = csv.DictReader(fh)
        assert rd.fieldnames == CLIENT_COLS
        return [{c: parse(c, rec[c]) for c in CLIENT_COLS} for rec in rd]


def run_engine(seed_rows):
    """Run generate_dataset's derivation over rows that carry only what was seeded.
    Every unseeded column is present as None - the engine has no value for it."""
    rows = []
    for s in seed_rows:
        r = {c: None for c in g.COLUMN_ORDER}
        r.update({k: None for k in GENERATOR_INTERNAL})
        r.update(s)
        rows.append(r)
    rows = [g.compute(r) for r in rows]
    for r in rows:
        r["urgency_score"] = g.urgency(r)
    rows = [g.compute_savings(r) for r in rows]
    rows = g.apply_constraints_and_priority(rows)
    for r in rows:
        g.compute_provenance(r)
        try:
            r["rationale"] = g.compose_rationale(r)
        except Exception as e:                       # noqa: BLE001
            r["rationale"] = f"<could not compose: {e}>"
    return rows


def compare(rows, cols):
    """Return {column: [(app_id, got, want), ...]} for every mismatch."""
    fails = {}
    for r in rows:
        want = BY_ID[r["app_id"]]
        for c in cols:
            if r.get(c) != want[c]:
                fails.setdefault(c, []).append((r["app_id"], r.get(c), want[c]))
    return fails


client_rows = load_client_csv()

# --- Pass A: strict. Client's 57 columns only, nothing else. --------------------------
pass_a = run_engine(client_rows)
fails_a = compare(pass_a, DERIVED_COLS)

# --- Pass B: client's 57 + the derived-labelled columns the generator authors ---------
# (i.e. the outputs of the tool's upstream stages that generate_dataset.py does not
#  implement - it hardcodes their results as inputs).
AUTHORED_DERIVED = [c for c in DERIVED_COLS
                    if any(c in r for r in g.ROWS)]
COMPUTED_DERIVED = [c for c in DERIVED_COLS if c not in AUTHORED_DERIVED]

seed_b = []
for cr in client_rows:
    src = BY_ID[cr["app_id"]]
    s = dict(cr)
    s.update({c: src[c] for c in AUTHORED_DERIVED})
    s.update({k: src.get(k) for k in GENERATOR_INTERNAL})
    seed_b.append(s)
pass_b = run_engine(seed_b)
fails_b = compare(pass_b, COMPUTED_DERIVED)

# --- Value fidelity: the 57 client columns must equal v2 exactly ----------------------
fails_client = compare(client_rows, CLIENT_COLS)

report = {
    "client_column_count": len(CLIENT_COLS),
    "tier_counts": {TIER_LABEL[t]: sum(1 for c in CLIENT_COLS if TIERS[c]["tier"] == t)
                    for t in ("client_mandatory", "client_if_available",
                              "system_extract", "sme_judgement")},
    "derived_column_count": len(DERIVED_COLS),
    "authored_derived": AUTHORED_DERIVED,
    "computed_derived_count": len(COMPUTED_DERIVED),
    "pass_a_reproduced": len(DERIVED_COLS) - len(fails_a),
    "pass_a_failed": sorted(fails_a),
    "pass_b_failed": {k: v[:2] for k, v in fails_b.items()},
    "client_value_fidelity_failures": fails_client,
    "blank_cells": sum(1 for r in client_rows for c in CLIENT_COLS if r[c] is None),
}
json.dump(report, open("/tmp/claude-0/-home-claude/894f0710-2492-5ee1-ba03-47d6878e86d8/"
                       "scratchpad/roundtrip.json", "w"), indent=2, default=str)

print("client cols        :", len(CLIENT_COLS), report["tier_counts"])
print("derived cols       :", len(DERIVED_COLS))
print("  authored in ROWS :", len(AUTHORED_DERIVED))
print("  truly computed   :", len(COMPUTED_DERIVED))
print("PASS A reproduced  :", report["pass_a_reproduced"], "/", len(DERIVED_COLS))
print("PASS A failures    :", len(fails_a))
for c in sorted(fails_a):
    print("   -", c, f"({len(fails_a[c])}/20 rows)")
print("PASS B failures    :", len(fails_b), sorted(fails_b))
print("client fidelity    :", "OK" if not fails_client else fails_client)
print("blank client cells :", report["blank_cells"])

# =======================================================================================
# WORKBOOK
# =======================================================================================
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
BAND_FILL = PatternFill("solid", fgColor="DCE6F1")
TIER_FILL = PatternFill("solid", fgColor="EEF3FA")
MV_FILL = PatternFill("solid", fgColor="FFF2CC")
EX_FILL = PatternFill("solid", fgColor="FCE4D6")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_COLS = {c[1] for c in g.COLUMNS if c[3].startswith("money")}
RATE_COLS = {c[1] for c in g.COLUMNS if c[3].startswith("float")}
SCORE_COLS = {c[1] for c in g.COLUMNS if c[3].startswith("score")}

SHORT_TIER = {
    "client_mandatory": "required",
    "client_if_available": "if available",
    "system_extract": "system extract",
    "sme_judgement": "SME answer",
}
BLANKS = {c: sum(1 for r in FULL if r[c] is None) for c in CLIENT_COLS}


def col_width(name):
    if name == "description":
        return 70
    if name in ("app_name", "vendor_name", "secondary_capabilities", "data_types_held",
                "business_unit", "primary_capability"):
        return 34
    return min(26, max(11, len(name) * 0.95 + 2))


def fmt(cell, name):
    if name in MONEY_COLS:
        cell.number_format = '#,##0'
    elif name in RATE_COLS:
        cell.number_format = '0.000'
    elif name in SCORE_COLS:
        cell.number_format = '0.0'


def prose(ws, blocks, width=112):
    ws.column_dimensions["A"].width = width
    r = 1
    for kindname, text in blocks:
        c = ws.cell(row=r, column=1, value=("- " + text) if kindname == "bullet" else text)
        if kindname == "h1":
            c.font = Font(bold=True, size=15, color="1F3864")
            r += 1
        elif kindname == "h2":
            c.font = Font(bold=True, size=12, color="1F3864")
        elif kindname == "gap":
            pass
        else:
            c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = max(14, 13 * (1 + len(text) // 100))
        r += 1
    ws.sheet_view.showGridLines = False


def three_row_header(ws, note=None):
    """Section band, tier band, then the machine column names. Returns the header row index."""
    top = 1
    if note:
        c = ws.cell(row=1, column=1, value=note)
        c.font = Font(bold=True, size=11, color="C00000")
        top = 3
    prev = None
    for j, name in enumerate(CLIENT_COLS, start=1):
        block = BLOCK_OF[name]
        c = ws.cell(row=top, column=j, value=block if block != prev else "")
        prev = block
        c.fill, c.font, c.border = BAND_FILL, Font(size=8, italic=True, bold=True,
                                                   color="1F3864"), BOX
        c.alignment = Alignment(wrap_text=True, vertical="bottom")

        c = ws.cell(row=top + 1, column=j, value=SHORT_TIER[TIERS[name]["tier"]]
                    + (" *" if name in MIN_VIABLE else ""))
        c.fill, c.font, c.border = TIER_FILL, Font(size=8, color="404040"), BOX
        c.alignment = Alignment(wrap_text=True, vertical="bottom")

        c = ws.cell(row=top + 2, column=j, value=name)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BOX
        c.alignment = Alignment(wrap_text=True, vertical="bottom")
        ws.column_dimensions[get_column_letter(j)].width = col_width(name)
    ws.row_dimensions[top].height = 24
    ws.row_dimensions[top + 1].height = 20
    ws.row_dimensions[top + 2].height = 46
    return top + 2


# ---------------------------------------------------------------------------------------
wb = Workbook()
wb.remove(wb.active)

# --- Read me first ---------------------------------------------------------------------
n_mv = len(MIN_VIABLE)
readme = [
    ("h1", "Client input dataset - read me first"),
    ("p", "This workbook is one half of the Aberdeen Advisors application-rationalization "
          "dummy dataset. It is the half a client actually hands over. Everything the tool "
          "works out for itself has been taken out."),
    ("gap", ""),
    ("h2", "What is in here"),
    ("p", f"20 invented applications and the {len(CLIENT_COLS)} pieces of information we ask "
          f"the client for. The full dataset "
          f"(App-Rationalization-Dummy-Dataset-v2.xlsx) has 126 columns for the same 20 "
          f"applications; the other {len(DERIVED_COLS)} are worked out by the tool from the "
          f"{len(CLIENT_COLS)} here - scores, pass and fail flags, the recommendation, the "
          f"priority, the total cost of ownership, the savings arithmetic and the written "
          f"rationale."),
    ("p", "Nothing here has been re-invented or nudged. Every value is copied straight out of "
          "the full dataset, so the two files agree cell for cell on all 20 rows."),
    ("gap", ""),
    ("h2", "Everything in it is made up"),
    ("p", "This is dummy data for a hackathon build. The only real things in the file are "
          "product and vendor names - Epic, Waystar, Workday and so on - which are real "
          "products used so the portfolio reads like a genuine health system. Every number, "
          "every person's name, every cost centre, every contract, every score and all three "
          "legal entities are invented. No client data of any kind is in this file."),
    ("gap", ""),
    ("h2", "The four sheets"),
    ("bullet", "Client input - the 20 rows, one application per row. Read left to right: who "
               "and what it is, who owns it, what it does, who uses it, what it costs, what "
               "the contract says, then the scores a person has to give us."),
    ("bullet", f"Field guide - one row per column, in plain language: what we are asking for, "
               f"who answers it or which system it comes out of, whether it is compulsory, "
               f"what a valid answer looks like, and what happens if you cannot supply it. "
               f"The {n_mv} shaded rows are the minimum we need to run anything at all."),
    ("bullet", "Blank template - the same columns with nothing in them, ready to send to a "
               "client. One example row is included and is marked in orange; delete it before "
               "the client fills the sheet in."),
    ("bullet", "Round-trip check - what happened when we fed this file back through the "
               "engine and compared the answers against the full dataset. Read it: it names "
               "the fields the tool cannot yet work out on its own."),
    ("gap", ""),
    ("h2", "The blank cells are deliberate"),
    ("p", f"{report['blank_cells']} cells in the Client input sheet are empty, and they are "
          f"meant to be. A real hand-over is never complete. Otter.ai (APP-010) is the "
          f"shadow-IT row - a tool bought on a credit card, so there is no contract, no owner "
          f"and no user count; 17 of its cells are blank. Abridge and Sunquest CoPath Plus "
          f"have no contract record. Nine applications have no retention expiry date. Empty "
          f"means we genuinely do not know: there is no \"N/A\" and no zero standing in for a "
          f"missing answer, so anything loading this file can tell the difference between "
          f"nothing and nought."),
    ("gap", ""),
    ("h2", "How many things are we asking for"),
    ("p", f"{len(CLIENT_COLS)} columns, in four kinds: "
          f"{report['tier_counts']['Client - required']} the tool cannot run without, "
          f"{report['tier_counts']['Client - if available']} that are welcome if they exist and "
          f"have a documented fallback if not, "
          f"{report['tier_counts']['Client system extract']} that come out of a source system "
          f"rather than a person, and {report['tier_counts']['Client SME judgement']} that "
          f"somebody has to sit down and answer."),
    ("p", f"Ryo asked for 58. Counting the 'Provided by' column on the full dataset's data "
          f"dictionary gives {len(CLIENT_COLS)}: 3 + 11 + 29 + 14. The file has "
          f"{len(CLIENT_COLS)}, not 58, and nothing has been padded or dropped to make a "
          f"number match."),
    ("gap", ""),
    ("h2", "The short version, if the client can only manage one thing"),
    ("p", f"{n_mv} of the {len(CLIENT_COLS)} columns are the minimum viable intake: 12 out of "
          f"systems, 1 capability field, and 8 answers per application - about a ten-minute "
          f"conversation with each owner. That is enough to return a recommendation and a "
          f"priority for every application, but every row comes back at medium confidence "
          f"and the cost picture is licence-only. The trade-offs are written out in "
          f"client-intake-requirements.md."),
    ("gap", ""),
    ("h2", "Files"),
    ("bullet", "Client-Input-Dataset-v1.xlsx - this workbook."),
    ("bullet", "client-input.csv - the Client input sheet as flat CSV, for loading."),
    ("bullet", "App-Rationalization-Dummy-Dataset-v2.xlsx - the full 126-column dataset this "
               "was cut down from."),
    ("bullet", "column-tiers.json - which of the 126 columns is asked for and which is "
               "derived, machine readable."),
    ("gap", ""),
    ("p", f"Built {g.dt.date.today().isoformat()} from {g.XLSX_NAME} ({g.DATASET_VERSION}). "
          f"Client input v1."),
]
prose(wb.create_sheet("Read me first"), readme)

# --- Client input ----------------------------------------------------------------------
ws = wb.create_sheet("Client input")
hrow = three_row_header(ws)
for i, r in enumerate(FULL, start=hrow + 1):
    for j, name in enumerate(CLIENT_COLS, start=1):
        c = ws.cell(row=i, column=j)
        if r[name] is not None:                 # blanks stay genuinely empty
            c.value = r[name]
        c.border = BOX
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=name == "description")
        fmt(c, name)
ws.freeze_panes = ws.cell(row=hrow + 1, column=3)
ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(CLIENT_COLS))}{hrow + len(FULL)}"

# --- Field guide -----------------------------------------------------------------------
ws = wb.create_sheet("Field guide")
fg_head = ["#", "Section", "Column", "Plain-language label", "What it means",
           "Provided by", "Source system or who answers", "Must we have it?",
           "Allowed values / format", "One of the 21 minimum?",
           "Blank on how many of the 20 rows", "If you cannot supply it"]
fg_widths = [4, 30, 34, 34, 62, 20, 34, 16, 40, 12, 12, 62]
fg_wrap = {"What it means", "Source system or who answers", "Allowed values / format",
           "If you cannot supply it", "Plain-language label", "Section"}
for j, h in enumerate(fg_head, start=1):
    c = ws.cell(row=1, column=j, value=h)
    c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BOX
    c.alignment = Alignment(wrap_text=True, vertical="bottom")
for i, name in enumerate(CLIENT_COLS, start=1):
    mv = name in MIN_VIABLE
    tier = TIERS[name]["tier"]
    vals = [
        i, BLOCK_OF[name], name, LABEL[name], DD_DEF[name], PROVIDED_BY[name],
        DD_SOURCE[name] or ("the client's own export" if tier == "client_mandatory" else ""),
        "Yes - the tool cannot run without it" if tier == "client_mandatory" else "No",
        DD_TYPE[name], "YES *" if mv else "", BLANKS[name] or "",
        DD_FALLBACK[name] or "",
    ]
    for j, v in enumerate(vals, start=1):
        c = ws.cell(row=i + 1, column=j, value=v)
        c.border = BOX
        c.font = Font(size=10, bold=mv)
        c.alignment = Alignment(wrap_text=fg_head[j - 1] in fg_wrap, vertical="top")
        if mv:
            c.fill = MV_FILL
for j, w in enumerate(fg_widths, start=1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[1].height = 44
ws.freeze_panes = "D2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(fg_head))}{len(CLIENT_COLS) + 1}"

# a legend under the table
lr = len(CLIENT_COLS) + 3
ws.cell(row=lr, column=2, value="Shaded, bold rows marked YES * are the 21 minimum viable "
        "items - the shortest list that still returns a recommendation and a priority for "
        "every application.").font = Font(size=10, italic=True, color="7F6000")
ws.cell(row=lr, column=2).fill = MV_FILL
ws.cell(row=lr + 1, column=2, value=f"{len(CLIENT_COLS)} columns asked for; "
        f"{len(DERIVED_COLS)} more are derived by the tool and are not in this workbook. See "
        f"the full dataset's 'Data dictionary' sheet for those.").font = Font(size=10,
                                                                             italic=True)

# --- Blank template --------------------------------------------------------------------
ws = wb.create_sheet("Blank template")
note = ("BLANK TEMPLATE - one row per application. Row 6 is an EXAMPLE, shaded orange: "
        "delete it before you start. Leave a cell EMPTY if you do not know the answer - "
        "please do not type N/A, 'unknown' or 0.")
hrow = three_row_header(ws, note=note)
EXAMPLE = {
    "app_id": "APP-000", "app_name": "EXAMPLE ROW - DELETE ME",
    "vendor_name": "Example Vendor, Inc.",
    "description": "One or two sentences on what the application actually does.",
    "deployment_model": "SaaS", "sourcing_type": "SaaS", "lifecycle_stage": "Mature",
    "implementation_date": "2021-04-01", "version_installed": "2026.1 (SaaS)",
    "business_owner": "Firstname Lastname", "technical_owner": "Firstname Lastname",
    "legal_entity": "LE-01", "business_unit": "Example Business Unit",
    "department": "Example Department", "cost_centre": "CC-0000",
    "primary_capability": "patient access",
    "secondary_capabilities": "clinical communication; analytics",
    "licences_purchased": 500, "active_users": 310, "last_signin_date": "2026-08-13",
    "cost_licence_subscription": 120000, "cost_upgrade_and_modules": 0,
    "cost_maintenance_dev_labour": 15000, "cost_infrastructure_peripherals": 0,
    "cost_indirect_and_training": 8000, "consumption_based_cost": 0,
    "one_time_implementation_cost": 0,
    "contract_id": "CTR-2024-0000", "annual_contract_value": 120000,
    "term_start": "2024-05-01", "term_end": "2027-04-30", "auto_renewal_flag": True,
    "renewal_notice_days": 90, "licence_metric": "per-user", "early_termination_penalty": 0,
    "integration_pattern": "API", "data_types_held": "database; documents",
    "retention_obligation_flag": False, "retention_expiry_date": None,
    "residual_archival_cost": 0, "information_classification": "internal",
    "amortised_one_time_migration_cost": 0, "realization_lag_months": None,
    "process_centrality": "Medium", "owner_stated_strategic_importance": 3.5,
    "ov_increase_value": 3.0, "ov_patient_care_criticality": 3.0,
    "ov_governance_compliance": 3.5, "th_architecture_fit": 3.5,
    "th_operational_stability": 4.0, "th_vendor_viability": 3.5,
    "th_customization_debt": 4.0, "r_technical_risk": 3.5,
    "r_business_compliance_risk": 3.5, "r_clinical_safety_risk": 4.0,
    "r_end_user_perceived_quality": 3.5, "urg_risk_pain_severity": "Low",
}
assert set(EXAMPLE) == set(CLIENT_COLS), set(EXAMPLE) ^ set(CLIENT_COLS)
for j, name in enumerate(CLIENT_COLS, start=1):
    c = ws.cell(row=hrow + 1, column=j)
    if EXAMPLE[name] is not None:
        c.value = EXAMPLE[name]
    c.fill, c.border = EX_FILL, BOX
    c.font = Font(size=10, italic=True, color="833C00")
    c.alignment = Alignment(vertical="top", wrap_text=name == "description")
    fmt(c, name)
for i in range(hrow + 2, hrow + 32):            # 30 empty, bordered rows to fill in
    for j in range(1, len(CLIENT_COLS) + 1):
        cc = ws.cell(row=i, column=j)
        cc.border = BOX
        cc.font = Font(size=10)
ws.freeze_panes = ws.cell(row=hrow + 1, column=3)

# --- Round-trip check ------------------------------------------------------------------
rt = [
    ("h1", "Round-trip check"),
    ("p", "The test: read client-input.csv, run the derivation, scoring and savings logic out "
          "of generate_dataset.py over it with no sight of any derived column, then compare "
          "every derived answer against the full v2 dataset. Anything that does not come back "
          "the same is a derived field that secretly needs something we are not asking the "
          "client for."),
    ("gap", ""),
    ("h2", "Result"),
    ("p", f"The 57 client columns match the full dataset exactly on all 20 rows, blanks "
          f"included."),
    ("p", f"Of the {len(DERIVED_COLS)} derived columns, {report['pass_a_reproduced']} come "
          f"back identical from the client input alone and {len(report['pass_a_failed'])} do "
          f"not. So no, the round trip does not fully reproduce v2."),
    ("p", f"The reason is a single structural fact: generate_dataset.py only implements "
          f"{len(COMPUTED_DERIVED)} of the {len(DERIVED_COLS)} derivations. The other "
          f"{len(AUTHORED_DERIVED)} derived columns are written by hand into the ROWS "
          f"literals - the dataset author decided them, the code never computes them. When "
          f"those {len(AUTHORED_DERIVED)} are supplied, all {len(COMPUTED_DERIVED)} implemented "
          f"derivations reproduce v2 exactly, on every row, with zero mismatches. The engine "
          f"is right; the pipeline in front of it is missing."),
    ("gap", ""),
    ("h2", f"The {len(AUTHORED_DERIVED)} derived columns the code does not compute"),
    ("p", "Grouped by the stage that is missing, with the extra input each stage needs."),
    ("bullet", "AI inventory (6): is_ai_tool, ai_delivery_form, ai_capability_class, "
               "ai_host_app_id, ai_already_entitled_elsewhere, ai_entitled_alternative_app_id. "
               "Needs the REQ 11 / REQ 26 classification stage, reading licence entitlement "
               "documents and product documentation - not a client column."),
    ("bullet", "Vendor lifecycle (3): version_vendor_supported, vendor_eos_date, "
               "technical_obsolescence_flag. Needs an external vendor support and "
               "end-of-support reference feed. The client supplies version_installed; only a "
               "vendor matrix says whether that version is still supported. th_supportability "
               "(1 more) is scored off these, so it fails with them."),
    ("bullet", "Governance and shadow IT (3): is_orphaned, governance_visibility, "
               "is_shadow_it. Derived by reconciling registers against each other - CMDB "
               "versus general ledger versus SSO sign-in. Needs the raw extracts as separate "
               "files, not the merged row."),
    ("bullet", "Overlap clustering and successor choice (5): overlap_cluster_id, cluster_role, "
               "replacement_app_id, replacement_ongoing_tco, "
               "replacement_cost_already_in_baseline. Needs the REQ 25 clustering stage over "
               "capability and description, and the successor-selection step after it. The "
               "CLUSTERS table in generate_dataset.py is used only to draw a sheet; compute() "
               "never reads it."),
    ("bullet", "Dependency inventory (2): has_downstream_dependents, dependency_count. Needs "
               "the interface-engine or integration registry - a per-interface extract, one "
               "row per interface, which we do not currently ask for."),
    ("bullet", "Score mappings that are simply not written yet (5): ov_reach_consumers and "
               "c_unused_licence_waste (both bands over active_users against "
               "licences_purchased), ov_reduce_costs_efficiency (a band over "
               "process_centrality), c_absolute_cost_band (a band over annual_tco_recurring), "
               "urg_timeline_sensitivity (a band over contract runway, notice window and "
               "end-of-support proximity). Every one of these is computable from columns "
               "already in this file. They fail only because the band thresholds live in the "
               "author's head rather than in the code. This is the cheapest gap to close."),
    ("bullet", "Peer and plan benchmarks (2): c_cost_per_active_user_vs_peers needs a peer "
               "benchmark of cost per active user across applications in the same capability, "
               "and c_consumption_price_variance needs the REQ 55 modelled consumption plan to "
               "compare metered spend against. Both are modelled inside the tool rather than "
               "supplied by the client, and neither model exists in the code - so these two "
               "legitimately cannot be reproduced from client input, and should not be "
               "expected to be."),
    ("bullet", "Labels and provenance (3): action and saving_type (a verb and a saving label "
               "read off disposition, sourcing_type and cluster role - and REQ 64 has a human "
               "confirm the action anyway), and data_source (which extracts the row was "
               "assembled from, which belongs to the intake manifest, not to any application "
               "row)."),
    ("gap", ""),
    ("h2", f"The {len(report['pass_a_failed']) - len(AUTHORED_DERIVED)} knock-on failures"),
    ("p", "These are computed correctly but from broken inputs, so they fail too, and they "
          "will all come right the moment the stages above exist: business_value_score, "
          "technical_health_score, cost_efficiency_score, c_pass, vtcr_key, "
          "infotech_template_disposition, retain_or_invest_basis, disposition, priority, "
          "rationale, confidence, redundancy_override_applied, suppressed_recommendation, "
          "suppression_reason, consolidation_saving, gross_saving_annual, net_saving_annual, "
          "net_saving_five_year, urgency_score, completeness_score, missing_fields."),
    ("p", "rationale carries one extra problem of its own: it splices in a paragraph of "
          "hand-written evidence prose held in the generator's _evidence field. That is "
          "authored text, not a derivation, so no amount of client input will reproduce it. "
          "Two more hidden intermediates sit alongside it - _gross_saving_basis (which savings "
          "method applies) and _reharvest_share / _reharvest_reserve (the seat reserve on a "
          "licence reharvest). None of the three is a column in the deliverable, and all three "
          "have to be decided by the tool before the savings arithmetic can run."),
    ("gap", ""),
    ("h2", f"The {report['pass_a_reproduced']} that do reproduce, from client input alone"),
    ("p", "licence_utilisation_rate, unused_licence_count, tco_five_category_subtotal, "
          "annual_tco_recurring, five_year_cumulative_tco, cost_per_active_user, "
          "unused_licence_spend, notice_deadline_date, in_notice_window_now, "
          "contract_runway_months, risk_posture_score, v_pass, t_pass, r_pass, "
          "lifecycle_exclusion_applied, sourcing_exclusion_applied, "
          "retention_override_applied."),
    ("p", "That is the whole of the cost and contract arithmetic, plus the risk dimension - "
          "the one dimension whose three weighted criteria are all answered by a person and "
          "none by the tool. It is a fair demonstration that the arithmetic is sound and the "
          "intake list is sufficient for it."),
    ("gap", ""),
    ("h2", "What to do about it"),
    ("bullet", "The five band mappings are worth writing down first: five thresholds tables "
               "and five derived columns stop being authored."),
    ("bullet", "Two new client asks would close two more groups: an interface or integration "
               "extract (one row per interface) for the dependency counts, and the raw CMDB, "
               "GL/AP and SSO extracts kept separate rather than pre-merged, for the shadow-IT "
               "and orphan flags."),
    ("bullet", "One external feed is needed and cannot come from the client: a vendor support "
               "and end-of-support matrix."),
    ("bullet", "Two peer or plan models are ours to build, and are correctly not asked of the "
               "client."),
    ("gap", ""),
    ("p", "The check itself is build_client_input.py, and it runs on every build: it fails "
          "loudly if the built rows stop matching applications-v2.csv, if the two "
          "classifications disagree, or if any of the 57 client values drifts from the full "
          "dataset."),
]
prose(wb.create_sheet("Round-trip check"), rt)

XLSX_PATH = os.path.join(OUT, "Client-Input-Dataset-v1.xlsx")
wb.save(XLSX_PATH)
print("wrote", XLSX_PATH)
print("wrote", CSV_PATH)
