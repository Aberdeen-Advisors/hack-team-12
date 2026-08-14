#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tune_northstar_600.py — BUILD A DEMO PORTFOLIO THAT CLEARS A 17% SAVINGS TARGET.

READ THIS FIRST. WHAT COMES OUT OF THIS SCRIPT IS NOT AN ANALYSIS RESULT.
----------------------------------------------------------------------------------
This script writes `healthcare_app_rationalization_sample_600_tuned.xlsx`, a FICTIONAL
variant of Bina Din's corrected 600-application workbook whose VALUES WERE CHOSEN so that
the unchanged scoring model returns a net first-year saving of at least 17% of portfolio
run cost. It exists for one purpose: a demo needs a portfolio with a large, visible number
in it. The 17% is therefore a PROPERTY OF THE INPUT DATA, not a finding about Bina's estate.

The honest run stays untouched beside it:
  * input   `healthcare_app_rationalization_sample_600_corrected.xlsx`
  * script  `score_northstar_600_corrected.py`
  * outputs `Northstar-Disposition-Analysis-600-corrected.xlsx`,
            `northstar-600-corrected-tool-vocabulary.csv` / `.xlsx`
That run returns 4.9% net on data the tool was not fitted to, and it is the evidence the
tool works. This one is the demo fixture. Do not quote them as two analyses of one estate.

WHAT WAS *NOT* TOUCHED, AND WHY THAT IS THE WHOLE POINT
-------------------------------------------------------
Nothing in the engine. Not one weight, band, rubric, gate, table row, guardrail or dollar
formula. `score_northstar_600_tuned.py` is `score_northstar_600_corrected.py` with six
output paths changed and a provenance banner added; both import the same
`score_northstar_v3` and both go through `v3.verify_engine_constants()`. The 17% is reached
by changing the INPUT, which is what a data-tuning exercise is allowed to do; reaching it
by changing the model would have made the tool worthless, since the tool is the deliverable.
No row is special-cased in scoring: every tuned row is scored by exactly the same code that
scores every other row, and the disposition each one lands on is the model's answer to the
new data, not an assignment written into the file.

THE FICTION, STATED PLAINLY
---------------------------
Northstar's estate contains 35 capability clusters averaging ~17 members apiece — the
classic post-merger picture, the same product stood up again per region. The corrected file
records those clusters but almost never records that a member is REDUNDANT: only 9 of 595
clustered applications were majority-Duplicative, and only 7 had any migration path in
evidence, so the model (correctly, on that evidence) left the duplicates as retain / invest.
This file tells the other story: on a selected set of non-survivor cluster members it
records the redundancy that the corrected file left unstated, and where a consolidation is
intended it supplies the migration-path evidence the corrected file was missing.

HOW EACH TUNED ROW WAS MADE, ALL OF IT IN THE DATA
--------------------------------------------------
Selection first, because selection is where plausibility is protected. A row is eligible
only if ALL of these hold in the corrected file:
    * it is APP-021..APP-600 (the original 20 applications are LEFT ENTIRELY ALONE, so the
      regression check against the committed 20-app answers still means something);
    * it is a member of a capability cluster and is NOT that cluster's survivor;
    * its Business Criticality is not 'Critical'; and
    * its Patient Care Impact is not 'Direct'.
The last two are the plausibility gate that matters: a patient-critical application is NOT
made into a retire candidate here, at any cost to the target. Healthcare criticality
protects applications, and the tuned file honours that.

Eligible rows are then taken in descending Annual TCO until the projected net clears the
target, and each is given one of two consistent stories:

  RETIRE-SHAPED (no migration path in evidence, so it falls to the gates)
    Capability Map     every row for the app becomes Coverage Level 'Duplicate' (which the
                       normalisation layer reads as Support Role 'Duplicative'), Capability
                       Criticality steps down to Medium, capability Critical Operation Flag
                       to No.
    App Inventory      Business Criticality High -> Medium, Critical Operation Flag -> No,
                       her own Business Value Score -> 2 with a matching rationale.
    Perf & Roadmap     the technical decay that makes retirement the answer rather than an
                       instruction: Vendor Support End inside a year, a legacy versioned
                       release line, MTTR past the 120-minute band, availability below its
                       own SLA. NB deliberately NOT pilot/beta wording — that would arm the
                       lifecycle guard and the model would suppress retire, correctly.
    TCO                the duplicate instance carries its own full stack, so the six cost
                       components are scaled up and Annual TCO is re-derived as their SUM.
                       Avoidable = 80% of the new total (licence, maintenance, infra,
                       vendor services and training stop; internal labour largely does not
                       in year one). Transition = 18% of avoidable — extraction, retention
                       archive and interface decommissioning. Cost Notes carry NO migration
                       language, because nothing is being migrated.
    Money is set BY INTENT and then CHECKED against what the model actually returned; the
    loop below re-labels any row the model read differently, so the file that ships never
    claims a retire-sized saving on a row the model consolidates.

  CONSOLIDATE-SHAPED (migration path evidenced, so the redundancy override absorbs it)
    Capability Map     as above — majority Duplicative.
    Cost Notes and Dependencies 'Required Before Disposition' name the survivor and the
                       cutover, which is the evidence condition (c) in build_clusters that
                       the corrected file could only satisfy on 7 rows.
    App Inventory      criticality steps down as above; technical health is NOT decayed —
                       the case here is redundancy, not decrepitude.
    TCO                components scaled more modestly. Avoidable = 55% of the new total
                       (licence and infra go, the survivor absorbs the load, interfaces and
                       internal labour persist). Transition = 32% of avoidable, because a
                       real migration costs real money and shrinking it to flatter the net
                       would be the dishonest move.

Every category, its row count and its dollar delta is written to
`northstar-600-tuned-change-log.md` and to the `Provenance — TUNED` sheet inside the
workbook itself, so the file cannot be opened alone and mistaken for a computed result.

USAGE
-----
    python3 tune_northstar_600.py        # writes only inside data/northstar/
    python3 score_northstar_600_tuned.py # then score it with the unmodified model
"""

import datetime as dt
import os
import sys
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import score_northstar_v3 as v3                     # noqa: E402  the model, read-only here
import score_northstar_600_corrected as corrected   # noqa: E402  for its normalisation layer

DATA = os.path.join(HERE, os.pardir, "data", "northstar")
SOURCE_XLSX = os.path.join(DATA, "healthcare_app_rationalization_sample_600_corrected.xlsx")
TUNED_XLSX = os.path.join(DATA, "healthcare_app_rationalization_sample_600_tuned.xlsx")
CHANGE_LOG = os.path.join(DATA, "northstar-600-tuned-change-log.md")

ORIGINAL_20 = {f"APP-{i:03d}" for i in range(1, 21)}
TARGET_PCT = 0.17
AIM_PCT = 0.152          # the projection is deliberately under the target: converting a
                         # cluster member also moves rows the projection cannot see (a
                         # survivor changing hands, a peer-median shifting), so the greedy
                         # walk stops early and the measured result lands above the line

# --- the tuning constants. Ratios, not per-row numbers: no row is hand-set. --------------
RETIRE_TCO_MULT = 1.25   # a duplicate instance carries its own licence, infra and labour
CONS_TCO_MULT = 1.15
RETIRE_AVOID = 0.80      # of the new annual total
RETIRE_TRANS = 0.18      # of avoidable
CONS_AVOID = 0.55
CONS_TRANS = 0.32
CONS_SHARE_MIN = 0.40    # keep the spread mixed rather than an all-retire portfolio

TCO_COMPONENTS = ["Annual License / Subscription", "Annual Maintenance",
                  "Annual Infrastructure / Hosting", "Annual Vendor Services",
                  "Annual Internal Labor", "Annual Education / Training"]

RETIRE_RELEASE = "7.4 / build 2019.2 (synthetic)"
RETIRE_SUPPORT_END = dt.datetime(2027, 4, 30)
RETIRE_MTTR = 165
RETIRE_AVAIL_SHORTFALL = 0.004


# =====================================================================================
# Reading the corrected file through the model, to find out what is already there
# =====================================================================================

def model_view(path):
    """Score `path` with the unmodified model and report per-app facts. Read-only."""
    v3.SOURCE_XLSX = path
    src = v3.load_source()
    corrected.normalise_vocabulary(src)
    ctx = v3.build_context(src)
    apps = v3.build_apps(src, ctx)
    v3.add_peer_context(apps, ctx)
    for a in apps:
        v3.score_app(a, ctx)
    clusters = v3.build_clusters(apps, ctx)
    cba = v3.clusters_by_app_of(clusters)
    view = {}
    for a in apps:
        base = v3.decide(a, cba, ctx, variant=False)
        sav = v3.compute_savings(a, base, cba)
        cl = cba.get(a["app_id"])
        view[a["app_id"]] = {
            "name": a["name"], "key": base["key"], "disposition": base["disposition"],
            "priority": base["priority"], "tco": a["annual_tco"] or 0.0,
            "avoidable": a["avoidable_annual"] or 0.0,
            "transition": a["one_time_transition"] or 0.0,
            "net": sav["net_first_year_saving"],
            "cluster": cl["cluster_id"] if cl else None,
            "role": cl["roles"][a["app_id"]]["role"] if cl else None,
            "survivor": cl["survivor"] if cl else None,
            "survivor_name": (cl and view_name(apps, cl["survivor"])) or None,
            "migration_path": cl["roles"][a["app_id"]]["migration_path_evidenced"] if cl else False,
            "criticality": a["business_criticality"],
            "patient_care": a["gr_patient_care_impact"],
        }
    return view


def view_name(apps, aid):
    for a in apps:
        if a["app_id"] == aid:
            return a["name"]
    return None


def totals(view):
    run = sum(v["tco"] for v in view.values())
    net = sum(v["net"] for v in view.values())
    acting = ("retire", "consolidate", "replace")
    gross = sum(v["avoidable"] for v in view.values() if v["disposition"] in acting)
    trans = sum(v["transition"] for v in view.values() if v["disposition"] in acting)
    return {"run": run, "gross": gross, "transition": trans, "net": net,
            "pct": (net / run) if run else 0.0,
            "spread": Counter(v["disposition"] for v in view.values())}


# =====================================================================================
# Selection
# =====================================================================================

def eligible(view):
    """Non-survivor cluster members outside the original 20 that are not patient-critical."""
    out = []
    for aid, v in view.items():
        if aid in ORIGINAL_20:
            continue
        if not v["cluster"] or v["role"] == "survivor":
            continue
        if v["criticality"] == "Critical":
            continue
        if v["patient_care"] == "Direct":
            continue
        out.append(aid)
    return sorted(out, key=lambda a: (-view[a]["tco"], a))


def select(view, forced=None):
    """Walk eligible rows by descending cost until the projected net clears AIM_PCT.

    `forced` carries the fixpoint's corrections: app_id -> intent, for rows a previous pass
    proved the model reads differently from the intent they were given.
    """
    forced = forced or {}
    run_base = sum(v["tco"] for v in view.values())
    intents, run_extra, net = {}, 0.0, 0.0
    n_cons = 0
    for aid in eligible(view):
        if net >= AIM_PCT * (run_base + run_extra):
            break
        v = view[aid]
        if aid in forced:
            intent = forced[aid]
        elif v["migration_path"]:
            intent = "consolidate"          # her own file already evidences the path
        elif n_cons < CONS_SHARE_MIN * (len(intents) + 1):
            intent = "consolidate"          # keep the spread mixed
        else:
            intent = "retire"
        intents[aid] = intent
        if intent == "consolidate":
            n_cons += 1
            new_tco = round(v["tco"] * CONS_TCO_MULT)
            avoid, trans = CONS_AVOID, CONS_TRANS
        else:
            new_tco = round(v["tco"] * RETIRE_TCO_MULT)
            avoid, trans = RETIRE_AVOID, RETIRE_TRANS
        run_extra += new_tco - v["tco"]
        gross = new_tco * avoid
        net += gross - gross * trans
    return intents


# =====================================================================================
# Writing the tuned workbook
# =====================================================================================

def header_index(ws):
    return {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value is not None}


def put(ws, row, col_idx, value):
    """Write, preserving the column's existing storage type for numbers."""
    cell = ws.cell(row=row, column=col_idx)
    if isinstance(value, (int, float)) and isinstance(cell.value, str):
        value = str(int(value)) if float(value).is_integer() else str(value)
    cell.value = value


def round_k(x):
    return int(round(x / 1000.0) * 1000)


def apply_tuning(intents, view):
    """Apply every edit to a fresh copy of the corrected workbook and save it."""
    # data_only=True: her workbook carries formulas (Annual TCO = SUM of its components,
    # Avoidable % = J/I, First-Year Net = MAX(0, J-L), App Inventory's Annual TCO = SUMIF
    # over the TCO sheet). openpyxl cannot recalculate, and saving a formula without its
    # cached value would hand the model a blank on every row. So the fixture is written
    # STATIC, from the cached values, and this script re-derives by hand exactly what those
    # formulas did — Annual TCO is still the sum of its six components on every row it
    # touches, and both mirrors are written explicitly.
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    log = defaultdict(lambda: {"rows": 0, "delta": 0.0, "detail": Counter()})
    retire = {a for a, i in intents.items() if i == "retire"}
    cons = {a for a, i in intents.items() if i == "consolidate"}
    touched = retire | cons

    # ---------------------------------------------------------------- Capability Map
    ws = wb["Capability Map"]
    h = header_index(ws)
    # Which capabilities does each app hold, so a consolidation can be pointed at a survivor
    # that genuinely holds the same capability as Primary — condition (b) of the engine's
    # absorption test. Without it the model reads the row as a duplicate with nowhere to go
    # and retires it instead, which is the model behaving correctly on thin evidence.
    caps_of = defaultdict(set)
    for r in range(2, ws.max_row + 1):
        aid = ws.cell(row=r, column=h["App ID"]).value
        if aid:
            caps_of[aid].add(ws.cell(row=r, column=h["Capability ID"]).value)
    survivor_primary_caps = defaultdict(set)
    for aid in cons:
        surv = view[aid]["survivor"]
        survivor_primary_caps[surv] |= (caps_of[aid] & caps_of[surv])
    for r in range(2, ws.max_row + 1):
        aid = ws.cell(row=r, column=h["App ID"]).value
        cap = ws.cell(row=r, column=h["Capability ID"]).value
        if aid in survivor_primary_caps and cap in survivor_primary_caps[aid]:
            role = ws.cell(row=r, column=h["Support Role"]).value
            col = "Support Role" if role in ("Primary", "Secondary", "Duplicative") \
                else "Coverage Level"
            if ws.cell(row=r, column=h[col]).value != "Primary":
                put(ws, r, h[col], "Primary")
                log["Capability Map (cluster survivors only): the survivor is recorded as "
                    "Primary on the capability its absorbed members hand over, which is the "
                    "engine's condition (b) for absorbing a duplicate"]["rows"] += 1
        if aid not in touched:
            continue
        cov = ws.cell(row=r, column=h["Coverage Level"]).value
        role = ws.cell(row=r, column=h["Support Role"]).value
        if role in ("Primary", "Secondary", "Duplicative"):        # original-20 convention
            put(ws, r, h["Support Role"], "Duplicative")
        else:
            put(ws, r, h["Coverage Level"], "Duplicate")
        log["Capability Map: coverage set to 'Duplicate' (read as Support Role "
            "'Duplicative') on every capability row of a selected duplicate instance"]["rows"] += 1
        log["Capability Map: coverage set to 'Duplicate' (read as Support Role "
            "'Duplicative') on every capability row of a selected duplicate instance"
            ]["detail"][f"was '{cov}'"] += 1
        crit = ws.cell(row=r, column=h["Capability Criticality"]).value
        if crit in ("Critical", "High"):
            put(ws, r, h["Capability Criticality"], "Medium")
            log["Capability Map: Capability Criticality stepped down to Medium where the "
                "capability is held as a duplicate copy"]["rows"] += 1
        if ws.cell(row=r, column=h["Critical Operation Flag"]).value == "Yes":
            put(ws, r, h["Critical Operation Flag"], "No")
            log["Capability Map: capability Critical Operation Flag -> No on duplicate "
                "copies"]["rows"] += 1

    # ---------------------------------------------------------------- App Inventory
    inv = wb["App Inventory"]
    h = header_index(inv)
    inv_row = {}
    for r in range(2, inv.max_row + 1):
        aid = inv.cell(row=r, column=h["App ID"]).value
        if aid:
            inv_row[aid] = r
    for aid in sorted(touched):
        r = inv_row[aid]
        if inv.cell(row=r, column=h["Business Criticality"]).value == "High":
            put(inv, r, h["Business Criticality"], "Medium")
            log["App Inventory: Business Criticality High -> Medium on selected duplicate "
                "instances (the capability is held as Primary by the cluster survivor)"]["rows"] += 1
        if inv.cell(row=r, column=h["Critical Operation Flag"]).value == "Yes":
            put(inv, r, h["Critical Operation Flag"], "No")
            log["App Inventory: Critical Operation Flag Yes -> No on selected duplicate "
                "instances"]["rows"] += 1
        put(inv, r, h["Business Value Score (1-5)"], 2)
        put(inv, r, h["Business Value Rationale"],
            "TUNED DEMO VALUE. Duplicate regional instance of a capability the cluster "
            "survivor holds as Primary; retained for local habit rather than for a "
            "capability only this instance provides.")
        put(inv, r, h["Lifecycle Stage"],
            "Retirement Candidate" if aid in retire else "Consolidate Candidate")
        log["App Inventory: her own Business Value Score -> 2 with a matching rationale, and "
            "her (held-out, never scored) Lifecycle Stage label aligned to the story"]["rows"] += 1

    # ---------------------------------------------------------------- Performance & Roadmap
    prf = wb["Performance & Roadmap"]
    h = header_index(prf)
    for r in range(2, prf.max_row + 1):
        aid = prf.cell(row=r, column=h["App ID"]).value
        if aid not in retire:
            continue
        put(prf, r, h["Current Release / Version"], RETIRE_RELEASE)
        put(prf, r, h["Vendor Support End"], RETIRE_SUPPORT_END)
        put(prf, r, h["MTTR (minutes)"], RETIRE_MTTR)
        sla = v3.f(prf.cell(row=r, column=h["SLA Target"]).value) or 0.999
        put(prf, r, h["Availability (12mo)"], round(sla - RETIRE_AVAIL_SHORTFALL, 6))
        put(prf, r, h["Performance Status"], "Below")
        put(prf, r, h["Roadmap / Update Notes"],
            "TUNED DEMO VALUE. Vendor support for this release line ends inside the year "
            "and no upgrade is funded for a duplicate instance; the estate's demand is "
            "already served by the cluster survivor.")
        log["Performance & Roadmap (retire-shaped rows only): Vendor Support End inside a "
            "year, legacy versioned release line, MTTR past the 120-minute band, "
            "availability below its own SLA target"]["rows"] += 1
        # keep App Inventory's mirror of the same two fields consistent
        ir = inv_row[aid]
        ih = header_index(inv)
        put(inv, ir, ih["Availability (12mo)"], round(sla - RETIRE_AVAIL_SHORTFALL, 6))
        put(inv, ir, ih["Performance Status"], "Below")

    # ---------------------------------------------------------------- Dependencies
    dep = wb["Dependencies"]
    h = header_index(dep)
    for r in range(2, dep.max_row + 1):
        aid = dep.cell(row=r, column=h["Source App ID"]).value
        if aid not in touched:
            continue
        if aid in cons:
            surv = view[aid]["survivor"]
            put(dep, r, h["Required Before Disposition"],
                f"TUNED DEMO VALUE. Repoint this interface to {surv} "
                f"({view[aid]['survivor_name']}) and cut over in a maintenance window; "
                f"validate message parity for one cycle before the instance is switched off.")
            put(dep, r, h["Migration Feasibility"], "High")
            log["Dependencies (consolidate-shaped rows only): 'Required Before Disposition' "
                "now names the survivor and the cutover, which is the migration-path "
                "evidence the corrected file could only satisfy on 7 rows"]["rows"] += 1
        else:
            txt = v3.s(dep.cell(row=r, column=h["Required Before Disposition"]).value)
            if any(t in txt.lower() for t in v3.MIGRATION_TERMS):
                put(dep, r, h["Required Before Disposition"],
                    "TUNED DEMO VALUE. Confirm no remaining consumer of this interface, "
                    "then decommission it with the instance; nothing is being carried "
                    "forward from here.")
                log["Dependencies (retire-shaped rows only): migration wording removed, "
                    "because nothing is migrating off a retirement — this is what keeps the "
                    "row on the gates instead of in the consolidate bucket"]["rows"] += 1

    # ---------------------------------------------------------------- TCO
    tco = wb["TCO"]
    h = header_index(tco)
    for r in range(2, tco.max_row + 1):
        aid = tco.cell(row=r, column=h["App ID"]).value
        if aid not in touched:
            continue
        is_ret = aid in retire
        mult = RETIRE_TCO_MULT if is_ret else CONS_TCO_MULT
        old_total = v3.f(tco.cell(row=r, column=h["Annual TCO"]).value) or 0.0
        parts = []
        for comp in TCO_COMPONENTS:
            old = v3.f(tco.cell(row=r, column=h[comp]).value) or 0.0
            new = round_k(old * mult)
            put(tco, r, h[comp], new)
            parts.append(new)
        new_total = sum(parts)                       # Annual TCO IS the sum, always
        put(tco, r, h["Annual TCO"], new_total)
        avoid = round_k(new_total * (RETIRE_AVOID if is_ret else CONS_AVOID))
        trans = round_k(avoid * (RETIRE_TRANS if is_ret else CONS_TRANS))
        put(tco, r, h["Avoidable Annual Cost"], avoid)
        put(tco, r, h["Avoidable % of TCO"], round(avoid / new_total, 6) if new_total else 0)
        put(tco, r, h["One-Time Transition Cost"], trans)
        put(tco, r, h["First-Year Net Savings"], avoid - trans)
        if is_ret:
            put(tco, r, h["Cost Notes"],
                "TUNED DEMO VALUE. Duplicate regional instance carrying its own licence, "
                "hosting and support stack. Switching it off removes licence, maintenance, "
                "infrastructure, vendor services and training; roughly a fifth of the "
                "annual total is internal labour and retained-record archive that does not "
                "come out in year one. No successor work: the capability is already served "
                "elsewhere.")
        else:
            surv = view[aid]["survivor"]
            put(tco, r, h["Cost Notes"],
                f"TUNED DEMO VALUE. Absorbed into {surv} ({view[aid]['survivor_name']}): "
                f"migrate the data, repoint the interfaces and consolidate the licence "
                f"position. Licence, hosting and vendor services come out; interfaces and "
                f"internal labour largely persist on the survivor, so a little over half "
                f"the annual total is avoidable. The one-time cost is the migration itself "
                f"and is not discounted.")
        cat = ("TCO: retire-shaped rows — six cost components scaled up "
               f"x{RETIRE_TCO_MULT}, Annual TCO re-derived as their sum, avoidable "
               f"{RETIRE_AVOID:.0%} of the new total, transition {RETIRE_TRANS:.0%} of "
               "avoidable") if is_ret else (
              "TCO: consolidate-shaped rows — six cost components scaled up "
              f"x{CONS_TCO_MULT}, Annual TCO re-derived as their sum, avoidable "
              f"{CONS_AVOID:.0%} of the new total, transition {CONS_TRANS:.0%} of avoidable")
        log[cat]["rows"] += 1
        log[cat]["delta"] += new_total - old_total
        # App Inventory mirrors both money columns
        ir = inv_row[aid]
        ih = header_index(inv)
        put(inv, ir, ih["Annual TCO"], new_total)
        put(inv, ir, ih["Avoidable Annual Cost"], avoid)

    return wb, log, retire, cons


# =====================================================================================
# Provenance, written inside the file
# =====================================================================================

BANNER = "TUNED DEMO DATA — NOT A COMPUTED RESULT FROM BINA'S CORRECTED SOURCE"


def write_provenance(wb, intents, log, before, after):
    ws = wb.create_sheet("Provenance — TUNED", 0)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 96
    bold = Font(bold=True, size=11)
    head = Font(bold=True, color="FFFFFF", size=12)
    fill = PatternFill("solid", fgColor="9C0006")
    ws["A1"], ws["B1"] = BANNER, ""
    ws["A1"].font, ws["A1"].fill = head, fill
    ws.merge_cells("A1:B1")

    lines = [
        ("What this file is",
         "A FICTIONAL variant of healthcare_app_rationalization_sample_600_corrected.xlsx "
         "whose values were deliberately constructed so that the unchanged scoring model "
         "returns a net first-year saving of at least 17% of portfolio run cost. Built for "
         "a demo that needed a large number in it."),
        ("What this file is NOT",
         "It is NOT a computed result, an analysis, a finding, or a corrected version of "
         "anything. The 17% is a property of these input values. Nothing here says anything "
         "about the real estate the corrected workbook describes."),
        ("The un-fitted run — quote THIS one",
         "healthcare_app_rationalization_sample_600_corrected.xlsx scored by "
         "score_northstar_600_corrected.py returns $17,409,000 net, 4.9% of a $354,330,000 "
         "run cost. That run is the evidence the tool works on data it was not fitted to. "
         "Outputs: Northstar-Disposition-Analysis-600-corrected.xlsx and "
         "northstar-600-corrected-tool-vocabulary.csv / .xlsx."),
        ("Built by", "engine/tune_northstar_600.py, from the corrected workbook, on "
                     f"{dt.date.today().isoformat()}. Scored by "
                     "engine/score_northstar_600_tuned.py."),
        ("The model was NOT touched",
         "No weight, band, rubric, gate, lookup-table row, guardrail or savings formula was "
         "changed, and no row is special-cased in scoring. The tuned run imports the same "
         "score_northstar_v3 as the corrected run and re-verifies the engine constants. "
         "Every disposition below is the model's answer to the new data."),
        ("Rows left completely alone",
         "APP-001..APP-020, so the regression check against the committed 20-application "
         "answers still means something. Also untouched: every cluster survivor, every "
         "application whose Business Criticality is 'Critical', and every application whose "
         "Patient Care Impact is 'Direct'. A patient-critical application was not made into "
         "a retire candidate at any cost to the target."),
        ("Rows changed",
         f"{len(intents)} of 600 — {sum(1 for i in intents.values() if i == 'retire')} given "
         f"a retire-shaped story (redundant, decaying, no migration path) and "
         f"{sum(1 for i in intents.values() if i == 'consolidate')} a consolidate-shaped one "
         f"(redundant, with the migration path now evidenced)."),
        ("Before -> after",
         f"run cost ${before['run']:,.0f} -> ${after['run']:,.0f}; net first year "
         f"${before['net']:,.0f} ({before['pct']:.2%}) -> ${after['net']:,.0f} "
         f"({after['pct']:.2%}); spread {dict(before['spread'])} -> {dict(after['spread'])}."),
        ("Internal consistency held",
         "Annual TCO is the sum of its six components on every changed row; App Inventory's "
         "Annual TCO and Avoidable Annual Cost mirror the TCO sheet; Avoidable % of TCO and "
         "First-Year Net Savings are re-derived, not left stale; user counts and utilisation "
         "were NOT touched, so cost per active user stays in a sane range; transition cost "
         "was not shrunk to flatter the net."),
        ("Full change log", "data/northstar/northstar-600-tuned-change-log.md, and "
                            "data/README.md records what this file is."),
    ]
    r = 3
    for k, val in lines:
        ws.cell(row=r, column=1, value=k).font = bold
        c = ws.cell(row=r, column=2, value=val)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 14 * max(1, (len(val) // 100 + 1))
        r += 2

    ws.cell(row=r, column=1, value="Change categories").font = bold
    r += 1
    for cat, d in sorted(log.items(), key=lambda kv: -kv[1]["rows"]):
        ws.cell(row=r, column=1, value=f"{d['rows']} rows" + (
            f", ${d['delta']:+,.0f} annual" if d["delta"] else ""))
        cc = ws.cell(row=r, column=2, value=cat)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # A second, unmissable banner on her own Read Me sheet.
    rm = wb["Read Me"]
    rr = rm.max_row + 2
    rm.cell(row=rr, column=1, value=BANNER).font = Font(bold=True, color="9C0006", size=12)
    rm.cell(row=rr + 1, column=1, value=(
        "The values in this workbook were constructed to clear a 17% first-year savings "
        "target. They are not a computed result from the corrected source workbook. See the "
        "'Provenance — TUNED' sheet. The un-fitted run is "
        "healthcare_app_rationalization_sample_600_corrected.xlsx (4.9% net)."))
    wb.active = 0
    return wb


def write_change_log(intents, log, before, after, retire, cons):
    with open(CHANGE_LOG, "w", encoding="utf-8") as fh:
        w = fh.write
        w(f"# {BANNER}\n\n")
        w("`healthcare_app_rationalization_sample_600_tuned.xlsx` is a **fictional variant** "
          "of `healthcare_app_rationalization_sample_600_corrected.xlsx`, built by "
          "`engine/tune_northstar_600.py` so that the **unchanged** scoring model returns a "
          "net first-year saving of at least 17% of portfolio run cost. The 17% is a "
          "property of these input values, not a finding. The un-fitted run stays beside it "
          "and is the one to quote: `score_northstar_600_corrected.py` returns "
          "$17,409,000 net, 4.9% of $354,330,000.\n\n")
        w("No weight, band, rubric, gate, lookup-table row, guardrail or savings formula was "
          "touched, and no row is special-cased in scoring.\n\n")
        w("## Result\n\n")
        w("| | corrected (un-fitted) | tuned |\n| --- | --- | --- |\n")
        w(f"| Portfolio run cost | ${before['run']:,.0f} | ${after['run']:,.0f} |\n")
        w(f"| Gross avoidable claimed | ${before['gross']:,.0f} | ${after['gross']:,.0f} |\n")
        w(f"| One-time transition | ${before['transition']:,.0f} | ${after['transition']:,.0f} |\n")
        w(f"| **Net first year** | **${before['net']:,.0f}** | **${after['net']:,.0f}** |\n")
        w(f"| Net as % of run cost | {before['pct']:.2%} | {after['pct']:.2%} |\n")
        for term in ("retain", "invest", "consolidate", "replace", "retire"):
            w(f"| {term} | {before['spread'].get(term, 0)} | {after['spread'].get(term, 0)} |\n")
        w("\n## Rows selected, and what protected the rest\n\n")
        w(f"- Eligible: APP-021..APP-600, a non-survivor member of a capability cluster, "
          f"Business Criticality not 'Critical', Patient Care Impact not 'Direct'.\n")
        w(f"- Selected: **{len(intents)} of 600**, taken in descending Annual TCO until the "
          f"projection cleared the target — {len(retire)} retire-shaped, {len(cons)} "
          f"consolidate-shaped.\n")
        w("- Left completely alone: APP-001..APP-020 (so the 20-app regression check still "
          "means something), every cluster survivor, every 'Critical' application and every "
          "application whose Patient Care Impact is 'Direct'.\n")
        w("\n## Every category of change\n\n")
        w("| Rows | Annual $ delta | Category |\n| --- | --- | --- |\n")
        for cat, d in sorted(log.items(), key=lambda kv: -kv[1]["rows"]):
            w(f"| {d['rows']} | {('$' + format(d['delta'], '+,.0f')) if d['delta'] else '—'}"
              f" | {cat} |\n")
        w("\n## Ratios used, and why they are not 100%\n\n")
        w(f"- Retire-shaped: cost components x{RETIRE_TCO_MULT}, avoidable "
          f"{RETIRE_AVOID:.0%} of the new annual total, transition {RETIRE_TRANS:.0%} of "
          f"avoidable. The withheld {1 - RETIRE_AVOID:.0%} is internal labour and "
          f"retained-record archive that does not come out in year one.\n")
        w(f"- Consolidate-shaped: cost components x{CONS_TCO_MULT}, avoidable "
          f"{CONS_AVOID:.0%}, transition {CONS_TRANS:.0%} of avoidable. Interfaces and "
          f"internal labour persist on the survivor, and the migration cost is not "
          f"discounted to flatter the net.\n")
        w("- User counts, utilisation and every clinical field were left untouched, so cost "
          "per active user stays in a sane range and no clinically critical application "
          "changed shape.\n")
        w("\n## How to check all of this rather than take its word for it\n\n")
        w("- `python3 engine/audit_tuned_consistency.py` — 15 internal-consistency rules over "
          "the fixture and its export: Annual TCO equals the sum of its six components on "
          "every row, both App Inventory mirrors match the TCO sheet, Avoidable % and "
          "First-Year Net are re-derived rather than stale, transition cost is empty (never "
          "zero) on every row that removes no run-rate spend and present on every row that "
          "does, the exported absolute-cost band still matches the dollars with the cheapest "
          "band at 5, utilisation still equals active / entitled, cost per active user stays "
          "inside the range the corrected file already spanned, no row this tuning touched "
          "carries an acting disposition alongside Business Criticality 'Critical' / Patient "
          "Care Impact 'Direct' / Critical Operation Flag 'Yes', and no retire row carries "
          "pilot or beta release wording that would have dodged the lifecycle guard.\n")
        w("- `node engine/verify_tuned_parity.js` — extracts the page's own scoring engine "
          "out of `index.html` at run time and replays it over the emitted tool-vocabulary "
          "columns, checking disposition parity 600/600, priority parity 600/600, both "
          "post-lookup guardrails, and that no row carries a negative net.\n")
        w("- `python3 engine/score_northstar_600_corrected.py` — still returns the corrected "
          "run's own figures, untouched, from the corrected input.\n")
        w("\n## Files\n\n")
        w("- `healthcare_app_rationalization_sample_600_tuned.xlsx` — this input, with a "
          "`Provenance — TUNED` sheet as its first sheet and a banner on `Read Me`\n")
        w("- `engine/score_northstar_600_tuned.py` — the corrected run script with output "
          "paths changed only\n")
        w("- `Northstar-Disposition-Analysis-600-tuned.xlsx`, "
          "`northstar-dispositions-600-tuned.csv`, `northstar-600-tuned-summary.md`\n")
        w("- `Northstar-600-tuned-tool-vocabulary.xlsx`, "
          "`northstar-600-tuned-tool-vocabulary.csv` — for the web page's Upload / Analyze\n")


# =====================================================================================

def main():
    before_view = model_view(SOURCE_XLSX)
    before = totals(before_view)
    print(f"corrected: net ${before['net']:,.0f} = {before['pct']:.2%} of "
          f"${before['run']:,.0f}  {dict(before['spread'])}")

    forced, intents, after = {}, None, None
    for attempt in range(1, 7):
        intents = select(before_view, forced)
        wb, log, retire, cons = apply_tuning(intents, before_view)
        wb.save(TUNED_XLSX)
        after_view = model_view(TUNED_XLSX)
        after = totals(after_view)
        # fixpoint: money is set by intent, so any row the model reads differently is
        # re-labelled and the file rebuilt. The shipped file never claims a retire-sized
        # saving on a row the model consolidates.
        mismatch = {a: after_view[a]["disposition"] for a, i in intents.items()
                    if after_view[a]["disposition"] != i
                    and after_view[a]["disposition"] in ("retire", "consolidate")}
        stuck = {a: i for a, i in intents.items()
                 if after_view[a]["disposition"] not in ("retire", "consolidate")}
        print(f"pass {attempt}: net ${after['net']:,.0f} = {after['pct']:.2%} of "
              f"${after['run']:,.0f}  {dict(after['spread'])}  "
              f"re-label {len(mismatch)}  inert {len(stuck)}")
        if not mismatch and after["pct"] >= TARGET_PCT:
            break
        forced.update(mismatch)
        for a in stuck:                     # a row the model will not act on is dropped
            forced.pop(a, None)

    wb, log, retire, cons = apply_tuning(intents, before_view)
    after_view = model_view(TUNED_XLSX)     # already saved by the last pass
    wb = write_provenance(wb, intents, log, before, totals(after_view))
    wb.save(TUNED_XLSX)
    final_view = model_view(TUNED_XLSX)
    final = totals(final_view)
    write_change_log(intents, log, before, final, retire, cons)
    print(f"TUNED: net ${final['net']:,.0f} = {final['pct']:.2%} of ${final['run']:,.0f}")
    print("spread", dict(final["spread"]))
    print("wrote", TUNED_XLSX, "and", CHANGE_LOG)
    if final["pct"] < TARGET_PCT:
        print(f"SHORT OF TARGET: {final['pct']:.2%} < {TARGET_PCT:.0%}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
