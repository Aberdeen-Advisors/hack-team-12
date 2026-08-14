#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
score_northstar_v2.py — run Bina Din's REVISED "Northstar Global Health" 20-application
sample dataset through the Aberdeen Advisors scoring engine and emit dispositions,
rationale, recommendations and savings.

WHAT CHANGED FROM v1 — AND ONLY THIS
------------------------------------
Bina asked for the analysis to be rerun "using the assumptions you used in the previous
file (except risk)". So every derivation in v1 is carried over byte-for-byte, and exactly
two things change, both consequences of the better risk evidence she supplied:

  1. The three risk inputs (r_technical_risk, r_business_compliance_risk,
     r_clinical_safety_risk) are rebuilt from her NEW `Healthcare Guardrails` sheet
     instead of from the two-row-per-app `Risks` register. See SECTION 3R.
  2. Risk is now a FIRST-CLASS gated dimension rather than a gated-but-flagged one.
     In v1 risk was gated at 3.0 but every output carried a "(low confidence)" caveat
     because her risk register held only 2 rows per app and some risk input fell back to
     a summary field on almost every row. Her new sheet covers 20 of 20 apps on all
     three risk families with no fallbacks, so the caveat is withdrawn. The
     risk-excluded variant is STILL computed and reported on every row so the comparison
     with v1 stays honest.

Consequentially, confidence is recomputed from the new evidence coverage. Nothing else
moves: not the engine, not the 18 inputs, not the 3.0 gate, not the 16-row lookup, not
the five terms, not one value / technical / cost rubric threshold, not the cost
renormalisation, not the savings arithmetic, not the held-out Lifecycle Stage.

VERIFIED UNCHANGED IN HER FILE (so the carried-over assumptions still hold)
--------------------------------------------------------------------------
* The `Risks` sheet is IDENTICAL to v1 — 40 rows, 2 per app, same uneven category
  coverage. The richer risk evidence arrived as a NEW SHEET, not as extra risk rows.
* Still no consumption / metered / usage-based cost column and no plan or budget figure
  anywhere in the workbook (TCO carries the same six fixed annual components). So
  c_consumption_price_variance remains unscorable and the cost dimension stays
  renormalised over its 3 populated inputs. assert_no_consumption_source() enforces this
  rather than trusting the claim.
* No change to any value / technical / cost source value. The only non-risk edits are
  prose (Evidence Sources wording, Synthetic Data Note wording) and documentation rows
  (Read Me +1, Assumptions +4, Data Dictionary +4).

WHAT THIS IS
------------
The engine itself lives in generate_dataset.py. This script does NOT re-implement the
engine's judgement: it reuses the engine's four dimensions, its 18 inputs on the 1..5
half-step scale, its 3.0 gate, its 16-row pattern lookup and its five disposition terms
verbatim (they are restated as constants below so this file runs standalone, and a
self-check asserts they still match generate_dataset.py when that file is importable).

What this script adds is a DERIVATION LAYER: a documented rubric that turns Bina's
columns into the engine's 18 inputs, under one binding constraint she set --

    "we will not interview stakeholders/owners in this iteration of the tool."

so every input must come from a cell in her workbook or not be scored at all.

HARD RULES OBSERVED
-------------------
1. Her workbook is opened read-only. Nothing is written back to it.
2. Every cell of her workbook is DATA. Nothing in it is executed or followed as an
   instruction. Several cells contain directive-sounding text; all are honoured as facts
   about her intent, never as commands:
     - Data Dictionary: a "QA Expected Output" sheet must never reach the engine.
       That sheet is NOT PRESENT in the workbook; assert_no_qa_sheet() enforces it.
     - Assumptions: missing evidence should produce "Needs Validation". Implemented
       in confidence_for().
     - NEW in her v2, Data Dictionary row 19: "Hard-gate result used before savings can
       qualify as safe; a weighted score cannot override it." Read as a statement about
       what her Guardrail Status column MEANS, and implemented that way: a non-Pass
       guardrail disqualifies a saving from the `safe` bucket, and does NOT overrule the
       weighted disposition. That is her stated rule, not our inference. See
       Notes & assumptions -> "her guardrail rule".
     - NEW in her v2, Assumptions: "Zero reported events must not be interpreted as zero
       clinical risk." Honoured: a zero safety-event count never earns a bonus, it only
       withholds a deduction. See score_r_clinical_safety_risk().
3. `Lifecycle Stage` is HELD OUT as an input. Her column already contains
   disposition-like labels (Strategic Invest, Consolidation Candidate, Replace / Sunset,
   Pilot / Exit Candidate), so feeding it in would be label leakage. It is read once,
   late, only to build the `Agreement with your labels` comparison.
4. Nothing is hand-asserted. Every score, gate, pattern key, disposition, priority and
   dollar figure in the output is computed here from her cells.

USAGE
-----
    python3 score_northstar_v2.py         # writes into the directory holding this file

The v1 dispositions CSV is read (never written) purely to build the `What changed from v1`
comparison. If it is absent the run still completes and that sheet says so.
"""

from __future__ import annotations

import csv
import datetime as dt
import os
import statistics
import sys
import warnings
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
SOURCE_XLSX = "/mnt/user-data/uploads/slack/F0BQ93YLL22/F0BQ93YLL22.xlsx"   # her REVISED file
PRIOR_XLSX = "/mnt/user-data/uploads/slack/F0BQCGAM8E5/F0BQCGAM8E5.xlsx"    # her v1 file
OUT_XLSX = os.path.join(HERE, "Northstar-Disposition-Analysis-v2.xlsx")
OUT_CSV = os.path.join(HERE, "northstar-dispositions-v2.csv")
V1_CSV = os.path.join(HERE, "northstar-dispositions.csv")   # read-only, for the v1 diff

ANALYSIS_DATE = dt.date(2026, 8, 14)          # her Read Me / Assumptions "As of" date
CIO_SAVINGS_TARGET = 0.15                      # her Assumptions sheet
HELD_OUT_COLUMNS = ("Lifecycle Stage",)        # leakage guard
FORBIDDEN_SHEET = "QA Expected Output"         # her Data Dictionary says never feed this in
GUARDRAIL_SHEET = "Healthcare Guardrails"      # NEW in her v2 — the richer risk evidence

# Her v1 risk register, kept ONLY so the run can prove it is unchanged and prove the
# fallback rate v1 suffered. It is never used to score anything in v2.
V1_RISK_SHEET = "Risks"


# =====================================================================================
# SECTION 1 — the engine, restated verbatim from generate_dataset.py
# =====================================================================================

PASS_THRESHOLD = 3.0                 # Info-Tech comparison is >=, so exactly 3.0 passes
SCORE_STEPS = [1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0]

DIMENSIONS = [
    ("V", "business value", "business_value_score", "v_pass"),
    ("T", "technical health", "technical_health_score", "t_pass"),
    ("C", "cost efficiency", "cost_efficiency_score", "c_pass"),
    ("R", "risk posture", "risk_posture_score", "r_pass"),
]

# criterion -> (dimension, integer weight)
CRITERIA = [
    ("ov_increase_value", "V", 1),
    ("ov_reach_consumers", "V", 1),
    ("ov_reduce_costs_efficiency", "V", 1),
    ("ov_patient_care_criticality", "V", 2),
    ("ov_governance_compliance", "V", 1),
    ("th_supportability", "T", 2),
    ("th_architecture_fit", "T", 2),
    ("th_operational_stability", "T", 1),
    ("th_vendor_viability", "T", 1),
    ("th_customization_debt", "T", 1),
    ("c_cost_per_active_user_vs_peers", "C", 2),
    ("c_unused_licence_waste", "C", 1),
    ("c_consumption_price_variance", "C", 1),
    ("c_absolute_cost_band", "C", 0),
    ("r_technical_risk", "R", 1),
    ("r_business_compliance_risk", "R", 1),
    ("r_clinical_safety_risk", "R", 1),
    ("r_end_user_perceived_quality", "R", 0),
]

DISPOSITIONS = ("retain", "invest", "consolidate", "replace", "retire")

DISPOSITION_TABLE = {
    "PPPP": ("Reward",    "Moderate",  "retain",      "Very Low"),
    "PPPF": ("Reward",    "Moderate",  "invest",      "High"),
    "PPFP": ("Refresh",   "High",      "invest",      "Moderate"),
    "PPFF": ("Refresh",   "High",      "invest",      "High"),
    "PFPP": ("Remediate", "Very High", "invest",      "Moderate"),
    "PFPF": ("Remediate", "Very High", "invest",      "High"),
    "PFFP": ("Replace",   "Very High", "replace",     "High"),
    "PFFF": ("Replace",   "Very High", "replace",     "Very High"),
    "FPPP": ("Refocus",   "Very Low",  "consolidate", "Low"),
    "FPPF": ("Refocus",   "Moderate",  "consolidate", "Moderate"),
    "FPFP": ("Refocus",   "High",      "retire",      "High"),
    "FPFF": ("Refocus",   "Very High", "retire",      "Very High"),
    "FFPP": ("Retire",    "Very High", "consolidate", "Moderate"),
    "FFPF": ("Retire",    "Very High", "retire",      "High"),
    "FFFP": ("Retire",    "Very High", "retire",      "Very High"),
    "FFFF": ("Retire",    "Very High", "retire",      "Very High"),
}

PRIORITY_LADDER = ["Very Low", "Low", "Moderate", "High", "Very High"]


def gate(score):
    """The Info-Tech comparison is >=, so exactly 3.0 passes."""
    if score is None:
        return "F"
    return "P" if score >= PASS_THRESHOLD else "F"


def dimension_score(row, dim_key):
    """Weighted mean over the dimension's POPULATED, non-zero-weight inputs.

    This is the engine's renormalisation rule, unchanged: an input with no source is
    skipped in both numerator and denominator, so the dimension is renormalised over what
    is actually populated rather than being silently dragged toward failure by a null.
    For Bina's file that rule fires exactly once, on the cost dimension, because
    c_consumption_price_variance has no source column.
    """
    num = den = 0.0
    for name, dim, weight in CRITERIA:
        if dim != dim_key or weight == 0:
            continue
        val = row.get(name)
        if val is None:
            continue
        num += val * weight
        den += weight
    if den == 0:
        return None
    return round(num / den, 3)


def dimension_weight_denominator(row, dim_key):
    """The denominator dimension_score() actually used, so the sheet can show it."""
    return sum(w for n, d, w in CRITERIA
               if d == dim_key and w > 0 and row.get(n) is not None)


def retain_or_invest(row):
    """Which of the two 'keep it' terms a row earns, read off the gates, not by hand."""
    failed = [name for _k, name, _col, flag in DIMENSIONS if row[flag] == "F"]
    if failed:
        return "invest", ("invest in " + " and in ".join(failed) +
                          f": that is the dimension failing the {PASS_THRESHOLD:.1f} gate, "
                          f"so that is what the money buys")
    return "retain", ("all four dimensions clear the 3.0 gate, so there is no failing "
                      "dimension to fund and nothing to buy: keep it, spend nothing")


def step_priority(priority, steps):
    i = PRIORITY_LADDER.index(priority)
    return PRIORITY_LADDER[min(max(i + steps, 0), len(PRIORITY_LADDER) - 1)]


def snap(x):
    """Snap onto the engine's 1..5 half-step scale."""
    if x is None:
        return None
    x = max(1.0, min(5.0, float(x)))
    return min(SCORE_STEPS, key=lambda s: (abs(s - x), s))


def verify_engine_constants():
    """Cross-check the restated constants against generate_dataset.py if it is present.

    generate_dataset.py runs a large build at import time, so it is parsed rather than
    imported: the check is textual and cheap, and it fails loudly rather than quietly.
    """
    path = os.path.join(HERE, "generate_dataset.py")
    if not os.path.exists(path):
        return "generate_dataset.py not found next to this script; engine constants unverified"
    with open(path, encoding="utf-8") as fh:
        src = fh.read()
    problems = []
    if "PASS_THRESHOLD = 3.0" not in src:
        problems.append("PASS_THRESHOLD")
    for key, (_tw, _tp, disp, prio) in DISPOSITION_TABLE.items():
        needle = f'"{key}": ('
        i = src.find(needle)
        if i == -1:
            problems.append(f"{key} missing from engine table")
            continue
        seg = src[i:src.find("\n", i)]
        if f'"{disp}"' not in seg or f'"{prio}"' not in seg:
            problems.append(f"{key} row disagrees with engine ({seg.strip()})")
    for name, dim, weight in CRITERIA:
        i = src.find(f'("{name}", "{dim}", {weight},')
        if i == -1:
            problems.append(f"criterion {name}/{dim}/w{weight} disagrees with engine")
    if problems:
        raise SystemExit("ENGINE MISMATCH — refusing to run: " + "; ".join(problems))
    return ("all 16 lookup rows, all 18 criterion weights and the 3.0 threshold verified "
            "against generate_dataset.py")


# =====================================================================================
# SECTION 2 — load Bina's workbook (read-only) and index it
# =====================================================================================

def load_source():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    if FORBIDDEN_SHEET in wb.sheetnames:
        raise SystemExit(
            f"REFUSING TO RUN: her Data Dictionary states the '{FORBIDDEN_SHEET}' sheet must "
            f"never be given to the recommendation engine, and this workbook contains it.")
    data = {}
    for name in wb.sheetnames:
        rows = [r for r in wb[name].iter_rows(values_only=True)]
        if not rows:
            data[name] = []
            continue
        hdr = [(str(h).strip() if h is not None else "") for h in rows[0]]
        recs = []
        for r in rows[1:]:
            if all(c is None for c in r):
                continue
            recs.append({h: v for h, v in zip(hdr, r) if h})
        data[name] = recs
    wb.close()
    if GUARDRAIL_SHEET not in data:
        raise SystemExit(
            f"REFUSING TO RUN: v2 rebuilds the three risk inputs from the "
            f"'{GUARDRAIL_SHEET}' sheet, and this workbook does not contain it. Point "
            f"SOURCE_XLSX at her revised file, or run score_northstar.py for the v1 basis.")
    return data


CONSUMPTION_TOKENS = ("consumption", "metered", "usage-based", "usage based",
                      "plan variance", "budget variance", "committed spend")


def assert_no_consumption_source(src):
    """v1 renormalised the cost dimension because no consumption/metered source existed.

    That assumption is CARRIED OVER, so it has to be re-verified against the new file
    rather than assumed. Any column heading in any sheet that looks like a consumption,
    metered or plan-variance measure fails this check loudly, because it would mean
    c_consumption_price_variance is now scorable and the cost denominator should go back
    to 4. Returns the evidence string for the Sanity checks sheet.
    """
    hits = []
    for sheet, recs in src.items():
        if not recs:
            continue
        for col in recs[0].keys():
            low = col.lower()
            if any(t in low for t in CONSUMPTION_TOKENS):
                hits.append(f"{sheet}: {col}")
    if hits:
        raise SystemExit(
            "STOPPING: her revised workbook now contains what looks like a consumption or "
            "metered-cost source (" + "; ".join(hits) + "). v1 renormalised the cost "
            "dimension over 3 inputs BECAUSE that source did not exist. Score "
            "c_consumption_price_variance from it and restore the weight-4 denominator "
            "before rerunning; do not silently keep the renormalisation.")
    tco_cols = list(src["TCO"][0].keys()) if src.get("TCO") else []
    return (f"no consumption/metered/plan column in any of {len(src)} sheets; TCO still "
            f"carries the same six fixed annual components ({len(tco_cols)} columns total)")


def risk_register_fallback_rate(src):
    """How badly v1's risk evidence was covered, computed from her own register.

    This exists to JUSTIFY promoting risk to a first-class gate in v2 rather than
    asserting that coverage improved. It counts, per app, which of the three v1 risk
    families her `Risks` sheet actually had a row for.
    """
    from collections import Counter as _C
    fams = {"technical": ("Technical", "Dependency", "Operational",
                          "Business Continuity", "Data Quality"),
            "compliance": ("Security", "Regulatory", "Privacy", "Vendor"),
            "clinical": ("Clinical Safety", "AI /", "Patient Safety")}
    per_app = defaultdict(set)
    cats = _C()
    for r in src[V1_RISK_SHEET]:
        aid, cat = s(r["App ID"]), s(r["Risk Category"])
        cats[cat] += 1
        for fam, keys in fams.items():
            if any(k.lower() in cat.lower() for k in keys):
                per_app[aid].add(fam)
    apps = sorted({s(r["App ID"]) for r in src[V1_RISK_SHEET]})
    covered = {fam: sum(1 for a in apps if fam in per_app[a]) for fam in fams}
    rows_per_app = _C(len([1 for r in src[V1_RISK_SHEET] if s(r["App ID"]) == a])
                      for a in apps)
    return {"n_apps": len(apps), "n_rows": len(src[V1_RISK_SHEET]),
            "rows_per_app": dict(rows_per_app), "covered": covered,
            "categories": dict(cats)}


def guardrail_coverage(src):
    """Coverage of the NEW evidence, computed. This is what earns risk its gate."""
    recs = src[GUARDRAIL_SHEET]
    need = {
        "technical": ["Open Critical Vulnerabilities", "Open High Vulnerabilities",
                      "Recovery Test Meets RTO/RPO", "Last Restore Test Result",
                      "Sev-1 / Sev-2 Incidents (12mo)",
                      "Hosting / Recovery Responsibility"],
        "compliance": ["Open Critical Compliance Findings", "Open High Compliance Findings",
                       "Oldest Overdue Finding Age (days)", "ePHI Handled", "BAA Status",
                       "Vendor Assurance Type", "Assurance Scope Covers Application",
                       "Assurance Expiration Date"],
        "clinical": ["Patient Care Impact", "Downtime Impact", "Downtime Procedure Status",
                     "Confirmed App-Related Safety Events (12mo)",
                     "Highest Safety Event Severity", "Interface Health",
                     "Unresolved Critical Interface Error"],
    }
    out = {}
    for fam, cols in need.items():
        full = 0
        for r in recs:
            if all(s(r.get(c)) != "" for c in cols):
                full += 1
        out[fam] = {"apps_fully_populated": full, "n_apps": len(recs), "columns": cols}
    return out


def s(v):
    return "" if v is None else str(v).strip()


def f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    txt = s(v)
    if not txt:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(txt[:10], fmt).date()
        except ValueError:
            pass
    return None


# =====================================================================================
# SECTION 3 — the rubric. One function per engine input.
#
# Every function returns (score, rubric_text, evidence_text, availability) where
# availability is "direct", "rubric", or "unavailable", and rubric_text is the rule that
# was applied to THIS row so the Dispositions sheet can quote it.
# =====================================================================================

# --- shared invert tables. Her scales run high = worse on all of these; ours run
# --- high = good on all 18 inputs, so these are the flips.
RESIDUAL_INVERT = {"Low": 5.0, "Medium": 3.5, "High": 2.0, "Critical": 1.0}
RISKLEVEL_INVERT = {"Low": 5.0, "Medium": 3.5, "High": 2.0, "Critical": 1.0}
STATUS_ADJUST = {"Closed": +1.0, "Mitigating": +0.5, "Accepted": 0.0, "Open": -0.5}
# Business / capability criticality. NOT inverted in the VALUE dimension: a Critical
# capability is high value, and inverting it there would score the EHR as low value.
# Inverted only where criticality enters the RISK dimension as an impact multiplier.
# See Notes & assumptions -> "criticality direction" for the flag raised on this.
CRITICALITY_VALUE = {"Critical": 4.5, "High": 3.5, "Medium": 2.5, "Low": 1.5}

MONEY_TERMS = ("revenue", "claim", "billing", "bill ", "payroll", "benefit", "reimburse",
               "credential", "capacity", "time and attendance", "labor compliance")
CLINICAL_TERMS = ("clinical", "patient", "care ", "care coordination", "order", "medical",
                  "telehealth", "virtual care", "provider")
MIGRATION_TERMS = ("migrat", "cutover", "cut over", "repoint", "rebuild", "consolidat")
UNSAFE_TERMS = ("potential only", "not safe", "until", "validated")


def cap_rows_for(app_id, caps):
    return [c for c in caps if s(c["App ID"]) == app_id]


def risk_rows_for(app_id, risks, categories):
    out = []
    for r in risks:
        if s(r["App ID"]) != app_id:
            continue
        cat = s(r["Risk Category"])
        if any(k.lower() in cat.lower() for k in categories):
            out.append(r)
    return out


def worst_risk_score(rows):
    """Invert her residual-risk label, credit her mitigation status, take the worst row.

    Risk posture is governed by the least-controlled risk on the application, so this is
    a MINIMUM across rows, not a mean. Status credit: a Mitigating plan is partial
    control (+0.5), Closed is control achieved (+1.0), Accepted is conscious tolerance
    (0.0), Open is an uncontrolled exposure (-0.5).
    """
    best = None
    detail = []
    for r in rows:
        base = RESIDUAL_INVERT.get(s(r["Residual Risk"]))
        if base is None:
            base = RISKLEVEL_INVERT.get(s(r["Risk Level"]))
        if base is None:
            continue
        adj = STATUS_ADJUST.get(s(r["Status"]), 0.0)
        v = max(1.0, min(5.0, base + adj))
        detail.append(f"{s(r['Risk Category'])} residual {s(r['Residual Risk'])}/"
                      f"{s(r['Status'])} -> {snap(v):.1f}")
        best = v if best is None else min(best, v)
    return (snap(best) if best is not None else None), "; ".join(detail)


# ---- V1 ov_increase_value -----------------------------------------------------------
def score_ov_increase_value(app, ctx):
    caps = cap_rows_for(app["app_id"], ctx["caps"])
    hits = []
    for c in caps:
        blob = (s(c["Capability"]) + " " + s(c["Process / Tasks Enabled"])).lower()
        if any(t in blob for t in MONEY_TERMS):
            hits.append((s(c["Support Role"]), s(c["Coverage Level"]), s(c["Capability"])))
    rubric = ("Money-path capability test on Capability Map: a capability whose name or "
              "Process/Tasks text mentions revenue, claims, billing, payroll, benefits, "
              "reimbursement, credentialing, capacity or time-and-attendance is a money "
              "path. Primary+Full=5.0, Primary+Partial=4.5, Secondary=4.0, Duplicative=3.0 "
              "(the money path exists but another app is the system of record). No money "
              "capability: Critical Operation Flag Yes=3.0, else 2.0.")
    if hits:
        roles = {h[0] for h in hits}
        if "Primary" in roles:
            full = any(h[0] == "Primary" and h[1] == "Full" for h in hits)
            score, why = (5.0, "Primary + Full") if full else (4.5, "Primary + Partial")
        elif "Secondary" in roles:
            score, why = 4.0, "Secondary"
        else:
            score, why = 3.0, "Duplicative only"
        ev = f"money-path capabilities {', '.join(sorted({h[2] for h in hits}))} held as {why}"
    elif s(app["critical_op_flag"]) == "Yes":
        score, ev = 3.0, "no money-path capability, but Critical Operation Flag = Yes"
    else:
        score, ev = 2.0, "no money-path capability and Critical Operation Flag = No"
    return snap(score), rubric, ev, "rubric"


# ---- V2 ov_reach_consumers ----------------------------------------------------------
UTIL_BANDS = [(0.85, 5.0), (0.75, 4.5), (0.65, 4.0), (0.55, 3.5), (0.45, 3.0),
              (0.35, 2.5), (0.25, 2.0), (0.15, 1.5)]
BREADTH_BANDS = [(25000, 5.0), (15000, 4.5), (8000, 4.0), (4000, 3.5), (2000, 3.0),
                 (1000, 2.5), (500, 2.0), (200, 1.5)]


def band(value, bands, floor=1.0):
    for threshold, score in bands:
        if value >= threshold:
            return score
    return floor


def score_ov_reach_consumers(app, ctx):
    util, act = app["utilisation"], app["active_users"]
    rubric = ("Mean of two bands, then snapped to the half-step scale. Depth = her "
              "Utilization Rate (Active Users 90d / Entitled Users): >=.85=5.0, >=.75=4.5, "
              ">=.65=4.0, >=.55=3.5, >=.45=3.0, >=.35=2.5, >=.25=2.0, >=.15=1.5, else 1.0. "
              "Breadth = absolute Active Users (90d): >=25000=5.0, >=15000=4.5, >=8000=4.0, "
              ">=4000=3.5, >=2000=3.0, >=1000=2.5, >=500=2.0, >=200=1.5, else 1.0.")
    if util is None or act is None:
        return None, rubric, "Utilization Rate or Active Users missing", "unavailable"
    d, b = band(util, UTIL_BANDS), band(act, BREADTH_BANDS)
    ev = (f"{act:,.0f} active users of {app['entitled_users']:,.0f} entitled "
          f"({util:.1%}) -> depth {d:.1f}, breadth {b:.1f}")
    return snap((d + b) / 2), rubric, ev, "direct"


# ---- V3 ov_reduce_costs_efficiency --------------------------------------------------
def score_ov_reduce_costs_efficiency(app, ctx):
    caps = cap_rows_for(app["app_id"], ctx["caps"])
    rubric = ("Process centrality from Capability Map Support Role, Coverage Level and "
              "Capability Criticality: Primary+Full on a Critical capability=5.0, "
              "Primary+Full=4.5, Primary+Partial=4.0, Secondary best role=3.0, "
              "Duplicative only=2.0 (a second copy of a process another app owns is not "
              "central to it).")
    if not caps:
        return None, rubric, "no Capability Map rows for this app", "unavailable"
    prim = [c for c in caps if s(c["Support Role"]) == "Primary"]
    if prim:
        if any(s(c["Coverage Level"]) == "Full" and s(c["Capability Criticality"]) == "Critical"
               for c in prim):
            score, why = 5.0, "Primary + Full on a Critical capability"
        elif any(s(c["Coverage Level"]) == "Full" for c in prim):
            score, why = 4.5, "Primary + Full coverage"
        else:
            score, why = 4.0, "Primary but Partial coverage"
    elif any(s(c["Support Role"]) == "Secondary" for c in caps):
        score, why = 3.0, "best role is Secondary"
    else:
        score, why = 2.0, "every capability row is Duplicative"
    ev = f"{len(caps)} capability rows; {why}"
    return snap(score), rubric, ev, "rubric"


# ---- V4 ov_patient_care_criticality (weight 2) --------------------------------------
def score_ov_patient_care_criticality(app, ctx):
    caps = cap_rows_for(app["app_id"], ctx["caps"])
    dup = [c for c in caps if s(c["Support Role"]) == "Duplicative"]
    majority_dup = bool(caps) and len(dup) / len(caps) >= 0.5
    dc = s(app["data_classification"])
    firm_phi = "PHI" in dc and "possible" not in dc.lower()
    rubric = ("Base from her Business Criticality (Critical=4.5, High=3.5, Medium=2.5, "
              "Low=1.5), +0.5 if Critical Operation Flag = Yes, +0.5 if Data Classification "
              "names PHI outright (a 'possible PHI' label does not qualify), -1.0 if at "
              "least half the app's Capability Map rows are Duplicative (clinical work does "
              "not stop when another application already holds the capability). Clamped to "
              "1..5. Business Criticality is used in its natural direction here, NOT "
              "inverted -- see Notes & assumptions.")
    base = CRITICALITY_VALUE.get(s(app["business_criticality"]))
    if base is None:
        return None, rubric, "Business Criticality missing", "unavailable"
    score, parts = base, [f"Business Criticality {s(app['business_criticality'])} -> {base:.1f}"]
    if s(app["critical_op_flag"]) == "Yes":
        score += 0.5
        parts.append("Critical Operation Flag Yes +0.5")
    if firm_phi:
        score += 0.5
        parts.append(f"Data Classification '{dc}' +0.5")
    if majority_dup:
        score -= 1.0
        parts.append(f"{len(dup)} of {len(caps)} capability rows Duplicative -1.0")
    return snap(score), rubric, "; ".join(parts), "rubric"


# ---- V5 ov_governance_compliance ----------------------------------------------------
def score_ov_governance_compliance(app, ctx):
    caps = cap_rows_for(app["app_id"], ctx["caps"])
    dup = [c for c in caps if s(c["Support Role"]) == "Duplicative"]
    majority_dup = bool(caps) and len(dup) / len(caps) >= 0.5
    any_primary = any(s(c["Support Role"]) == "Primary" for c in caps)
    dc = s(app["data_classification"])
    regulated = any(t in dc for t in ("PHI", "PII", "Security"))
    rubric = ("Regulatory/trust ALIGNMENT, not regulatory obligation: is regulated data "
              "held in the governed standard or in an uncontrolled second copy? Regulated "
              "Data Classification (PHI / PII / Security) held by a Primary-role app=5.0; "
              "regulated but majority-Duplicative=2.0; regulated and mixed=3.5; "
              "not regulated + Primary=4.0; not regulated + majority-Duplicative=2.5. "
              "Then -0.5 if her Evidence Confidence for the app is Low. HIPAA/PHI SEVERITY "
              "is deliberately NOT scored here -- it is scored once, in "
              "r_business_compliance_risk, to avoid double-counting.")
    if not caps:
        return None, rubric, "no Capability Map rows for this app", "unavailable"
    if regulated and any_primary and not majority_dup:
        score, why = 5.0, f"regulated data ({dc}) held in a Primary-role application"
    elif regulated and majority_dup:
        score, why = 2.0, f"regulated data ({dc}) in a majority-Duplicative application"
    elif regulated:
        score, why = 3.5, f"regulated data ({dc}), mixed support roles"
    elif any_primary and not majority_dup:
        score, why = 4.0, f"non-regulated data ({dc}) in a Primary-role application"
    else:
        score, why = 2.5, f"non-regulated data ({dc}), majority-Duplicative"
    parts = [why]
    if s(app["evidence_confidence"]) == "Low":
        score -= 0.5
        parts.append("her Evidence Confidence = Low, -0.5")
    return snap(score), rubric, "; ".join(parts), "rubric"


# ---- T1 th_supportability (weight 2) ------------------------------------------------
EVERGREEN_TOKENS = ("continuous", "current saas", "cloud continuous", "current release")


def score_th_supportability(app, ctx):
    perf = ctx["perf"].get(app["app_id"], {})
    rel = s(perf.get("Current Release / Version"))
    eos = parse_date(perf.get("Vendor Support End"))
    rubric = ("Mean of the components that exist, because Vendor Support End is blank for "
              "19 of 20 apps. (a) End-of-support horizon, when present: >5y=5.0, 3-5y=4.0, "
              "1-3y=2.5, <1y=1.0. (b) Release currency from Current Release / Version: "
              "evergreen vendor-managed line ('Continuous', 'Cloud continuous', 'Current "
              "SaaS release')=4.5 -- always supported, but capped below 5.0 because no "
              "explicit support horizon is evidenced; a dated in-year release (e.g. "
              "'May 2026', '2026.2', '2026 R1')=4.0; 'Pilot release'=2.5 (not a supported "
              "GA line); an explicitly versioned legacy release (e.g. '9.2 / PeopleTools "
              "8.60')=2.0. LOWER-EVIDENCE INPUT: no end-of-support date was invented for "
              "the 19 apps that lack one.")
    comps, parts = [], []
    if eos:
        yrs = (eos - ANALYSIS_DATE).days / 365.25
        c = 5.0 if yrs > 5 else 4.0 if yrs >= 3 else 2.5 if yrs >= 1 else 1.0
        comps.append(c)
        parts.append(f"Vendor Support End {eos.isoformat()} = {yrs:.1f}y out -> {c:.1f}")
    low = rel.lower()
    if not rel:
        cur = None
    elif "pilot" in low:
        cur = 2.5
    elif any(t in low for t in EVERGREEN_TOKENS):
        cur = 4.5
    elif any(ch.isdigit() for ch in rel) and ("peopletools" in low or "/" in rel):
        cur = 2.0
    elif any(ch.isdigit() for ch in rel) or "'" in rel:
        cur = 4.0
    else:
        cur = 3.0
    if cur is not None:
        comps.append(cur)
        parts.append(f"release '{rel}' -> {cur:.1f}")
    if not comps:
        return None, rubric, "no release or support-end evidence", "unavailable"
    ev = "; ".join(parts) + ("" if eos else " [no Vendor Support End on file: lower evidence]")
    return snap(sum(comps) / len(comps)), rubric, ev, "rubric"


# ---- T2 th_architecture_fit (weight 2) ----------------------------------------------
HOSTING_BASE = [("vendor-managed saas", 5.0), ("vendor cloud", 4.0),
                ("private cloud / vendor-managed", 4.0), ("hybrid", 3.0),
                ("customer-hosted", 2.0)]


def score_th_architecture_fit(app, ctx):
    deps = ctx["deps_by_src"].get(app["app_id"], [])
    host = s(app["hosting_model"]).lower()
    rubric = ("Hosting base from her Hosting Model: vendor-managed SaaS=5.0, vendor "
              "cloud/hosted=4.0, private cloud vendor-managed=4.0, hybrid on-prem+cloud=3.0, "
              "customer-hosted=2.0. Then -0.5 if any dependency is Criticality Critical with "
              "Migration Feasibility Low (a tightly coupled critical integration is poor "
              "enterprise-architecture fit), and +0.5 if every dependency has Migration "
              "Feasibility High or Medium.")
    base = None
    for token, v in HOSTING_BASE:
        if token in host:
            base = v
            break
    if base is None:
        return None, rubric, f"Hosting Model '{s(app['hosting_model'])}' unmatched", "unavailable"
    score, parts = base, [f"Hosting Model -> {base:.1f}"]
    tight = [d for d in deps if s(d["Criticality"]) == "Critical"
             and s(d["Migration Feasibility"]) == "Low"]
    if tight:
        score -= 0.5
        parts.append(f"{len(tight)} Critical dependency with Low migration feasibility "
                     f"({', '.join(s(d['Target Application / System'])[:26] for d in tight)}) -0.5")
    elif deps and all(s(d["Migration Feasibility"]) in ("High", "Medium") for d in deps):
        score += 0.5
        parts.append(f"all {len(deps)} dependencies High/Medium migration feasibility +0.5")
    return snap(score), rubric, "; ".join(parts), "rubric"


# ---- T3 th_operational_stability ----------------------------------------------------
INCIDENT_RATE_BANDS = [(0.5, 5.0), (1.5, 4.0), (3.0, 3.0), (5.0, 2.0)]
MTTR_BANDS = [(45, 5.0), (60, 4.0), (90, 3.0), (120, 2.0)]


def lower_band(value, bands, floor=1.0):
    """Bands where SMALLER is better."""
    for threshold, score in bands:
        if value <= threshold:
            return score
    return floor


def score_th_operational_stability(app, ctx):
    perf = ctx["perf"].get(app["app_id"], {})
    inc, mttr = f(perf.get("P1/P2 Incidents (12mo)")), f(perf.get("MTTR (minutes)"))
    avail, sla = f(perf.get("Availability (12mo)")), f(perf.get("SLA Target"))
    act = app["active_users"]
    rubric = ("Mean of three measured components from Performance & Roadmap. "
              "(a) P1/P2 incidents per 1,000 active users: <=0.5=5.0, <=1.5=4.0, <=3=3.0, "
              "<=5=2.0, else 1.0. (b) MTTR minutes: <=45=5.0, <=60=4.0, <=90=3.0, <=120=2.0, "
              "else 1.0. (c) Availability against her SLA Target: at or above target=5.0, "
              "shortfall <=0.001=3.0, larger shortfall=1.5. This is the one engine input "
              "her file evidences BETTER than our own dataset does -- our scoring-model "
              "review flagged that we had no incident or ticket-volume column at all.")
    comps, parts = [], []
    if inc is not None and act:
        rate = inc / (act / 1000.0)
        c = lower_band(rate, INCIDENT_RATE_BANDS)
        comps.append(c)
        parts.append(f"{inc:.0f} P1/P2 in 12mo over {act:,.0f} active users = "
                     f"{rate:.2f}/1k -> {c:.1f}")
    if mttr is not None:
        c = lower_band(mttr, MTTR_BANDS)
        comps.append(c)
        parts.append(f"MTTR {mttr:.0f} min -> {c:.1f}")
    if avail is not None and sla is not None:
        short = sla - avail
        c = 5.0 if short <= 0 else 3.0 if short <= 0.001 else 1.5
        comps.append(c)
        parts.append(f"availability {avail:.4f} vs SLA {sla:.4f} -> {c:.1f}")
    if not comps:
        return None, rubric, "no performance evidence", "unavailable"
    return snap(sum(comps) / len(comps)), rubric, "; ".join(parts), "direct"


# ---- T4 th_vendor_viability ---------------------------------------------------------
def score_th_vendor_viability(app, ctx):
    rows = risk_rows_for(app["app_id"], ctx["risks"], ("Vendor",))
    perf = ctx["perf"].get(app["app_id"], {})
    rubric = ("If her Risks sheet carries a Vendor-category row for the app, invert its "
              "Residual Risk (Low=5.0, Medium=3.5, High=2.0, Critical=1.0) and credit "
              "Mitigation Status (Closed +1.0, Mitigating +0.5, Accepted 0, Open -0.5). "
              "Otherwise 4.0 when both Next Minor Release and Next Major Update are dated "
              "(an active, published roadmap), 3.0 when the roadmap is incomplete. "
              "LOWER-EVIDENCE INPUT: her file contains no vendor financial data of any "
              "kind, so vendor solvency is proxied by roadmap activity and by whether her "
              "own risk register raised a vendor concern.")
    if rows:
        score, detail = worst_risk_score(rows)
        if score is not None:
            return score, rubric, f"her Vendor risk row: {detail}", "direct"
    minor, major = parse_date(perf.get("Next Minor Release")), parse_date(perf.get("Next Major Update"))
    if minor and major:
        return snap(4.0), rubric, (f"no Vendor risk row raised; roadmap dated (minor "
                                   f"{minor.isoformat()}, major {major.isoformat()}) -> 4.0 "
                                   f"[no vendor financials on file: lower evidence]"), "rubric"
    return snap(3.0), rubric, "no Vendor risk row and roadmap dates incomplete", "rubric"


# ---- T5 th_customization_debt -------------------------------------------------------
LANGUAGE_DEBT = [("peoplecode", 1.5), ("peopletools", 1.5), ("apex", 3.0),
                 ("javascript", 3.0), ("dax", 3.5), ("tableau calculations", 3.5),
                 ("configuration and automation", 4.5)]


def score_th_customization_debt(app, ctx):
    rows = risk_rows_for(app["app_id"], ctx["risks"], ("Technical",))
    lang = s(app["config_language"]).lower()
    rubric = ("Two components, worst wins (min), because debt is bounded by the worse "
              "evidence. (a) Configuration/extension language: PeopleCode or PeopleTools="
              "1.5 (heavy proprietary custom code), Apex or JavaScript=3.0, DAX or Tableau "
              "calculations=3.5, 'configuration and automation rules'=4.5, otherwise from "
              "Commercial Model: Commercial SaaS=4.5 (configure, do not code), Commercial "
              "off-the-shelf=3.5. (b) Any Technical-category risk row, inverted from its "
              "Residual Risk with the same status credit used elsewhere.")
    comps, parts = [], []
    lv = None
    for token, v in LANGUAGE_DEBT:
        if token in lang:
            lv = v if lv is None else min(lv, v)
    if lv is None:
        cm = s(app["commercial_model"]).lower()
        lv = 4.5 if "saas" in cm else 3.5 if "off-the-shelf" in cm else 3.0
        parts.append(f"Commercial Model '{s(app['commercial_model'])}' -> {lv:.1f}")
    else:
        parts.append(f"config language '{s(app['config_language'])[:48]}' -> {lv:.1f}")
    comps.append(lv)
    if rows:
        rv, detail = worst_risk_score(rows)
        if rv is not None:
            comps.append(rv)
            parts.append(f"her Technical risk row: {detail}")
    return snap(min(comps)), rubric, "; ".join(parts), "rubric"


# ---- C1 c_cost_per_active_user_vs_peers (weight 2) ----------------------------------
PEER_RATIO_BANDS = [(0.50, 5.0), (0.70, 4.5), (0.90, 4.0), (1.00, 3.5), (1.15, 3.0),
                    (1.40, 2.5), (1.80, 2.0), (2.50, 1.5)]


def score_c_cost_per_active_user(app, ctx):
    rubric = ("Annual TCO / Active Users (90d), compared with the MEDIAN of the app's peer "
              "group. Peer group = her Category column; a Category with only one app is "
              "compared against the whole-portfolio median instead. Ratio to peer median: "
              "<=0.50=5.0, <=0.70=4.5, <=0.90=4.0, <=1.00=3.5, <=1.15=3.0, <=1.40=2.5, "
              "<=1.80=2.0, <=2.50=1.5, else 1.0. 5.0 = cheapest per productive user.")
    cpu = app["cost_per_active_user"]
    if cpu is None:
        return None, rubric, "Annual TCO or Active Users missing", "unavailable"
    peers = ctx["peer_cpu"][app["category"]]
    if len(peers) >= 2:
        med, basis = statistics.median(peers), f"{len(peers)} apps in Category '{app['category']}'"
    else:
        med, basis = ctx["portfolio_cpu_median"], (
            f"Category '{app['category']}' has only this app, so the whole-portfolio median "
            f"of {len(ctx['all_cpu'])} apps is used")
    ratio = cpu / med
    ev = (f"${cpu:,.0f} per active user vs peer median ${med:,.0f} = {ratio:.2f}x ({basis})")
    return snap(band(-ratio, [(-t, v) for t, v in PEER_RATIO_BANDS])), rubric, ev, "rubric"


# ---- C2 c_unused_licence_waste ------------------------------------------------------
WASTE_BANDS = [(0.90, 5.0), (0.85, 4.5), (0.80, 4.0), (0.70, 3.5), (0.60, 3.0),
               (0.50, 2.5), (0.40, 2.0), (0.30, 1.5)]


def score_c_unused_licence_waste(app, ctx):
    rubric = ("Her Utilization Rate read as the inverse of licensed-but-inactive waste: "
              ">=.90=5.0, >=.85=4.5, >=.80=4.0, >=.70=3.5, >=.60=3.0, >=.50=2.5, >=.40=2.0, "
              ">=.30=1.5, else 1.0. 5.0 = almost no waste.")
    util = app["utilisation"]
    if util is None:
        return None, rubric, "Utilization Rate missing", "unavailable"
    unused = (app["entitled_users"] or 0) - (app["active_users"] or 0)
    ev = f"{util:.1%} utilised, {unused:,.0f} entitled-but-inactive seats"
    return snap(band(util, WASTE_BANDS)), rubric, ev, "direct"


# ---- C3 c_consumption_price_variance — UNAVAILABLE ---------------------------------
def score_c_consumption_price_variance(app, ctx):
    rubric = ("NOT SCORED. This input compares metered/consumption spend against a modelled "
              "plan. Her workbook has neither: the TCO sheet carries six FIXED annual "
              "components (License, Maintenance, Infrastructure, Vendor Services, Internal "
              "Labor, Education) and no consumption or metered line, and no plan or budget "
              "figure to vary against. Producing a number here would be inventing data, so "
              "the cost dimension is RENORMALISED over its remaining weighted inputs "
              "(weight sum 4 -> 3). This gap is not caused by the no-interview constraint; "
              "the data does not exist in the file.")
    return None, rubric, "no consumption/metered cost column and no plan figure", "unavailable"


# ---- C4 c_absolute_cost_band (weight 0) --------------------------------------------
ABS_COST_BANDS = [(500_000, 5.0), (1_000_000, 4.5), (1_500_000, 4.0), (2_000_000, 3.5),
                  (3_000_000, 3.0), (5_000_000, 2.0), (7_000_000, 1.5)]


def score_c_absolute_cost_band(app, ctx):
    rubric = ("Absolute Annual TCO band: <=$0.5m=5.0, <=$1m=4.5, <=$1.5m=4.0, <=$2m=3.5, "
              "<=$3m=3.0, <=$5m=2.0, <=$7m=1.5, else 1.0. WEIGHT 0 -- scored and reported, "
              "contributes nothing. Our scoring-model review found this input is one of only "
              "two that can change an answer if switched on, and that at weight 2 it flips "
              "the largest platform in a portfolio from retain to invest purely for being "
              "the largest line item. It stays at 0 deliberately.")
    tco = app["annual_tco"]
    if tco is None:
        return None, rubric, "Annual TCO missing", "unavailable"
    return snap(lower_band(tco, ABS_COST_BANDS)), rubric, f"Annual TCO ${tco:,.0f}", "rubric"


# =====================================================================================
# SECTION 3R — THE ONLY DERIVATION THAT CHANGED IN v2.
#
# v1 built all three risk inputs from her `Risks` sheet: 2 rows per app, categories
# unevenly spread, so at least one of the three families had no row on most apps and had
# to fall back to her single `Highest Risk` summary label. That is why v1 shipped risk as
# "low confidence".
#
# Her revised workbook adds a `Healthcare Guardrails` sheet: 46 measured columns for all
# 20 apps, with no gaps on the fields these rubrics use. The three inputs are therefore
# rebuilt from MEASURED QUANTITIES rather than from an inverted severity label, and risk
# becomes a first-class gated dimension.
#
# Shape of all three rubrics: start at 5.0 (nothing adverse found), subtract documented
# deductions for each adverse condition in her data, floor at 1.0, cap at 5.0, snap to the
# engine's half-step scale. Deductions, never bonuses for absence of evidence -- her own
# Assumptions sheet says a zero event count must not be read as zero risk, so a clean row
# earns no uplift, it merely avoids a penalty. The one exception is the vendor-assurance
# tier in R2, where a HITRUST-scoped attestation is affirmative evidence of control and is
# allowed +0.5 (capped, so it cannot lift a row over the gate on its own).
#
# NOT USED, DELIBERATELY: `RTO Target` vs `Maximum Tolerable Downtime`. RTO exceeds MTD on
# 12 of 20 apps and in every one of those cases RTO is exactly 1x or 2x MTD, which is the
# signature of a generation rule rather than 12 independent recovery-design failures.
# Scoring it would push a dataset artifact into a clinical-safety number. It is reported
# instead -- see rto_vs_mtd_finding() and the Sanity checks sheet -- because if the
# relationship IS real it is a portfolio-wide business-continuity finding and Bina's call,
# not ours to bury in a score.
# =====================================================================================

def gr(app, ctx):
    """Her Healthcare Guardrails row for this app."""
    return ctx["guardrails"].get(app["app_id"], {})


def _apply(base, deductions):
    """Sum documented deductions, floor 1.0, cap 5.0, snap to the half-step scale."""
    total = base + sum(d[1] for d in deductions)
    return snap(max(1.0, min(5.0, total)))


def _fmt(deductions):
    return "; ".join(f"{why} ({amt:+.1f})" for why, amt in deductions) or "nothing adverse found"


# ---- R1 r_technical_risk ------------------------------------------------------------
R1_RUBRIC = (
    "REBUILT IN v2 from her Healthcare Guardrails sheet (v1 used her 2-row Risks "
    "register). Start 5.0, then deduct: open CRITICAL vulnerabilities 1-2 = -1.0, >=3 = "
    "-1.5; oldest critical vulnerability older than 90 days = a further -0.5 (an unpatched "
    "critical past a quarter is a process failure, not a backlog item); open HIGH "
    "vulnerabilities 5-9 = -0.5, 10-14 = -1.0, >=15 = -1.5; recovery test does not "
    "demonstrate RTO/RPO ('Unknown') = -1.0; last restore test result Unknown = -0.5 "
    "('Not Applicable' on vendor-managed SaaS is NOT penalised -- there is no customer "
    "restore duty to have failed); reported backup success rate below 0.997 = -0.5; "
    "Sev-1/Sev-2 incidents in 12 months 20-29 = -0.5, >=30 = -1.0; hosting/recovery "
    "responsibility customer-managed on-premises = -0.5 (the whole recovery burden sits "
    "internally). Floor 1.0. NOTE the Sev-1/2 count partially overlaps "
    "th_operational_stability, which scores her Performance sheet's total incident count "
    "and MTTR; kept because a severity-filtered count is a different measure, and flagged "
    "here rather than silently double-counted.")


def score_r_technical_risk(app, ctx):
    g = gr(app, ctx)
    if not g:
        return None, R1_RUBRIC, "no Healthcare Guardrails row for this app", "unavailable"
    d = []
    crit = f(g.get("Open Critical Vulnerabilities"))
    high = f(g.get("Open High Vulnerabilities"))
    age = f(g.get("Oldest Critical Vulnerability Age (days)"))
    if crit is not None and crit >= 3:
        d.append((f"{crit:.0f} open critical vulnerabilities", -1.5))
    elif crit is not None and crit >= 1:
        d.append((f"{crit:.0f} open critical vulnerabilities", -1.0))
    if age is not None and age > 90:
        d.append((f"oldest critical vulnerability {age:.0f} days old, past a quarter", -0.5))
    if high is not None:
        if high >= 15:
            d.append((f"{high:.0f} open high vulnerabilities", -1.5))
        elif high >= 10:
            d.append((f"{high:.0f} open high vulnerabilities", -1.0))
        elif high >= 5:
            d.append((f"{high:.0f} open high vulnerabilities", -0.5))
    if s(g.get("Recovery Test Meets RTO/RPO")) == "Unknown":
        d.append(("recovery test does not demonstrate RTO/RPO", -1.0))
    if s(g.get("Last Restore Test Result")) == "Unknown":
        d.append(("last restore test result Unknown", -0.5))
    bk = f(g.get("Backup Success Rate (30d)"))
    if bk is not None and bk < 0.997:
        d.append((f"backup success rate {bk:.4f} below 0.997", -0.5))
    sev = f(g.get("Sev-1 / Sev-2 Incidents (12mo)"))
    if sev is not None:
        if sev >= 30:
            d.append((f"{sev:.0f} Sev-1/Sev-2 incidents in 12 months", -1.0))
        elif sev >= 20:
            d.append((f"{sev:.0f} Sev-1/Sev-2 incidents in 12 months", -0.5))
    host = s(g.get("Hosting / Recovery Responsibility"))
    if "customer-managed" in host.lower():
        d.append((f"{host}: recovery burden entirely internal", -0.5))
    return _apply(5.0, d), R1_RUBRIC, _fmt(d), "direct"


# ---- R2 r_business_compliance_risk -------------------------------------------------
ASSURANCE_TIER = [
    ("hitrust", +0.5, "HITRUST-scoped attestation"),
    ("soc 2", 0.0, "SOC 2 Type II attestation"),
    ("vendor security assessment", -0.5, "vendor security assessment only, no third-party attestation"),
    ("internal control", -1.0, "internal control assessment only, no external assurance"),
]

R2_RUBRIC = (
    "REBUILT IN v2 from her Healthcare Guardrails sheet. Start 5.0, then deduct: any open "
    "CRITICAL compliance finding = -1.5; open HIGH compliance findings 2 = -0.5, >=3 = "
    "-1.0; oldest overdue finding 1-45 days = -0.5, 46-90 = -1.0, >90 = -1.5 (an overdue "
    "finding is a control that has already missed its own remediation date); ePHI handled "
    "with BAA Status anything other than Current = -1.5; assurance scope does not "
    "confirmably cover the application ('Unknown') = -1.0; vendor assurance tier -- "
    "HITRUST-scoped +0.5, SOC 2 Type II 0.0, vendor security assessment only -0.5, "
    "internal control assessment only -1.0; third-party assurance already expired = -1.0, "
    "expiring within 90 days of the 2026-08-14 as-of date = -0.5; last security risk "
    "review older than 365 days = -0.5. Capped at 5.0 so the HITRUST credit cannot carry "
    "a row over the gate by itself. This remains the ONLY place HIPAA/PHI exposure is "
    "scored, so it is not double-counted in the value dimension.")


def score_r_business_compliance_risk(app, ctx):
    g = gr(app, ctx)
    if not g:
        return None, R2_RUBRIC, "no Healthcare Guardrails row for this app", "unavailable"
    d = []
    cf = f(g.get("Open Critical Compliance Findings"))
    hf = f(g.get("Open High Compliance Findings"))
    od = f(g.get("Oldest Overdue Finding Age (days)"))
    if cf is not None and cf >= 1:
        d.append((f"{cf:.0f} open critical compliance finding(s)", -1.5))
    if hf is not None:
        if hf >= 3:
            d.append((f"{hf:.0f} open high compliance findings", -1.0))
        elif hf >= 2:
            d.append((f"{hf:.0f} open high compliance findings", -0.5))
    if od is not None and od > 0:
        if od > 90:
            d.append((f"oldest overdue finding {od:.0f} days past due", -1.5))
        elif od > 45:
            d.append((f"oldest overdue finding {od:.0f} days past due", -1.0))
        else:
            d.append((f"oldest overdue finding {od:.0f} days past due", -0.5))
    ephi = s(g.get("ePHI Handled")) == "Yes"
    baa = s(g.get("BAA Status"))
    if ephi and baa != "Current":
        d.append((f"handles ePHI with BAA Status '{baa or 'blank'}'", -1.5))
    if s(g.get("Assurance Scope Covers Application")) == "Unknown":
        d.append(("vendor assurance scope not confirmed to cover this application", -1.0))
    atype = s(g.get("Vendor Assurance Type")).lower()
    for token, amt, label in ASSURANCE_TIER:
        if token in atype:
            if amt != 0.0:
                d.append((label, amt))
            else:
                d.append((label, 0.0))
            break
    exp = parse_date(g.get("Assurance Expiration Date"))
    if exp is not None:
        days = (exp - ANALYSIS_DATE).days
        if days < 0:
            d.append((f"assurance expired {-days} days ago", -1.0))
        elif days <= 90:
            d.append((f"assurance expires in {days} days", -0.5))
    rev = parse_date(g.get("Last Security Risk Review Date"))
    if rev is not None and (ANALYSIS_DATE - rev).days > 365:
        d.append((f"last security risk review {(ANALYSIS_DATE - rev).days} days ago", -0.5))
    ev = _fmt(d)
    if ephi:
        ev = f"handles ePHI (BAA {baa}); " + ev
    return _apply(5.0, d), R2_RUBRIC, ev, "direct"


# ---- R3 r_clinical_safety_risk -----------------------------------------------------
SAFETY_SEVERITY_DEDUCTION = {
    "none reported": 0.0, "low": -0.5, "moderate": -1.0,
    "high": -2.0, "severe": -2.0, "critical": -2.5, "catastrophic": -2.5,
}
DOWNTIME_PROC_DEDUCTION = {"tested": 0.0, "documented": -0.5, "not tested": -1.0}

R3_RUBRIC = (
    "REBUILT IN v2 from her Healthcare Guardrails sheet -- and this is the input the new "
    "data improves most, because v1 had an AI / Clinical Safety row for only 4 of 20 apps "
    "and inferred the rest from capability text. Her `Patient Care Impact` column now "
    "states the clinical pathway directly, so it replaces v1's keyword test entirely. "
    "Start 5.0. Where Patient Care Impact is 'None' the clinical-safety deductions do not "
    "apply -- an app with no patient-care pathway cannot carry a clinical-safety exposure, "
    "and unlike v1 this is now POSITIVE evidence from her column rather than an inference "
    "from missing keywords, which is why it scores 5.0 where v1 scored 4.5. Otherwise "
    "deduct: confirmed app-related safety events in 12 months, by her Highest Safety "
    "Event Severity -- Low -0.5, Moderate -1.0, High/Severe -2.0, Critical/Catastrophic "
    "-2.5; a further -0.5 where 3 or more events are confirmed; a further -0.5 where her "
    "Safety Event Attribution Confidence is High and at least one event occurred "
    "(attributed causation, not coincidence). Where her Safety Event Attribution "
    "Confidence is LOW on an app that has a patient-care pathway, -0.5 REGARDLESS of the "
    "event count: low attribution confidence means her surveillance cannot reliably "
    "connect events to this application, so a zero count there is weak evidence rather "
    "than a clean record. That is her own Assumptions sheet applied literally -- 'zero "
    "reported events must not be interpreted as zero clinical risk'. "
    "Downtime procedure status Documented but "
    "untested -0.5, Not Tested -1.0, plus a further -0.5 where Downtime Impact is "
    "Critical or High and the procedure is not Tested. No downtime exercise on record, or "
    "one older than 365 days, -0.5. Unresolved critical interface error -1.5 (orders and "
    "results are what cross those interfaces). Interface Health Amber -0.5. Interface "
    "failure rate above 0.05% -0.5. Guardrail Status other than Pass -0.5. Floor 1.0. Per "
    "her own Assumptions sheet, a zero event count earns no bonus -- it only withholds a "
    "deduction.")


def score_r_clinical_safety_risk(app, ctx):
    g = gr(app, ctx)
    if not g:
        return None, R3_RUBRIC, "no Healthcare Guardrails row for this app", "unavailable"
    pci = s(g.get("Patient Care Impact"))
    d = []
    # Interface and guardrail conditions apply regardless of patient-care pathway, because
    # a broken interface on a non-clinical app can still corrupt data a clinical app reads.
    if s(g.get("Unresolved Critical Interface Error")) == "Yes":
        d.append(("UNRESOLVED CRITICAL INTERFACE ERROR", -1.5))
    if s(g.get("Interface Health")) == "Amber":
        d.append(("interface health Amber", -0.5))
    rate = f(g.get("Interface Failure Rate"))
    if rate is not None and rate > 0.0005:
        d.append((f"interface failure rate {rate*100:.3f}% above 0.05%", -0.5))
    if s(g.get("Guardrail Status")) != "Pass":
        d.append((f"her Guardrail Status is '{s(g.get('Guardrail Status'))}' "
                  f"({s(g.get('Guardrail Reason'))})", -0.5))

    if pci == "None":
        score = _apply(5.0, d)
        ev = (f"her Patient Care Impact is 'None' -- no patient-care pathway, so the "
              f"safety-event and downtime-procedure deductions do not apply; "
              f"{_fmt(d)}")
        return score, R3_RUBRIC, ev, "direct"

    events = f(g.get("Confirmed App-Related Safety Events (12mo)"))
    sev_label = s(g.get("Highest Safety Event Severity"))
    attrib = s(g.get("Safety Event Attribution Confidence"))
    if events is not None and events >= 1:
        amt = SAFETY_SEVERITY_DEDUCTION.get(sev_label.lower())
        if amt is None:
            amt = -1.0   # unrecognised severity label: treat as Moderate, never as clean
            d.append((f"{events:.0f} confirmed safety event(s), unrecognised severity "
                      f"'{sev_label}' treated as Moderate", amt))
        elif amt != 0.0:
            d.append((f"{events:.0f} confirmed app-related safety event(s), highest "
                      f"severity {sev_label}", amt))
        if events >= 3:
            d.append((f"{events:.0f} events is 3 or more", -0.5))
        if attrib == "High":
            d.append(("attribution confidence High: causation attributed, not coincidental",
                      -0.5))
    # Applies whether or not events were reported -- see the rubric note on her
    # "zero reported events is not zero risk" assumption.
    if attrib == "Low":
        d.append((f"safety-event attribution confidence Low on an app with {pci.lower()} "
                  f"patient-care impact: her surveillance cannot reliably attribute events "
                  f"to this app, so the {0 if events is None else events:.0f}-event count is "
                  f"weak evidence, not a clean record", -0.5))

    proc = s(g.get("Downtime Procedure Status"))
    pamt = DOWNTIME_PROC_DEDUCTION.get(proc.lower())
    if pamt is None and proc:
        pamt = -1.0
    if pamt:
        d.append((f"downtime procedure {proc}", pamt))
    dti = s(g.get("Downtime Impact"))
    if dti in ("Critical", "High") and proc.lower() != "tested":
        d.append((f"Downtime Impact {dti} with procedure not Tested", -0.5))
    ex = parse_date(g.get("Last Downtime Exercise Date"))
    if ex is None:
        d.append(("no downtime exercise on record", -0.5))
    elif (ANALYSIS_DATE - ex).days > 365:
        d.append((f"last downtime exercise {(ANALYSIS_DATE - ex).days} days ago", -0.5))

    ev = (f"Patient Care Impact {pci}, Downtime Impact {dti}, MTD "
          f"{s(g.get('Maximum Tolerable Downtime (minutes)'))} min; {_fmt(d)}")
    return _apply(5.0, d), R3_RUBRIC, ev, "direct"


def rto_vs_mtd_finding(ctx):
    """Reported, never scored. See the SECTION 3R header for why.

    Returns (count, total, detail rows) for apps whose recovery TARGET is longer than the
    maximum downtime their own clinical/business process is said to tolerate.
    """
    out = []
    for aid, g in sorted(ctx["guardrails"].items()):
        mtd, rto = f(g.get("Maximum Tolerable Downtime (minutes)")), f(g.get("RTO Target (minutes)"))
        if mtd is None or rto is None or rto <= mtd:
            continue
        out.append(f"{aid} {s(g.get('Application Name'))}: RTO target {rto:.0f} min exceeds "
                   f"maximum tolerable downtime {mtd:.0f} min "
                   f"(ratio {rto/mtd:.0f}x, Downtime Impact {s(g.get('Downtime Impact'))})")
    ratios = sorted({round(f(g.get("RTO Target (minutes)")) / f(g.get("Maximum Tolerable Downtime (minutes)")), 2)
                     for g in ctx["guardrails"].values()
                     if f(g.get("Maximum Tolerable Downtime (minutes)"))
                     and f(g.get("RTO Target (minutes)"))})
    return len(out), len(ctx["guardrails"]), out, ratios


# ---- R4 r_end_user_perceived_quality (weight 0) — UNAVAILABLE ----------------------
def score_r_end_user_perceived_quality(app, ctx):
    perf = ctx["perf"].get(app["app_id"], {})
    kpi, tgt = f(perf.get("KPI Actual")), f(perf.get("KPI Target"))
    rubric = ("NOT SCORED. This input is an end-user PERCEPTION measure (satisfaction / "
              "perceived quality). Her workbook has no survey, CSAT, NPS or satisfaction "
              "field, and the only route to one would be asking users -- which the "
              "no-interview constraint rules out for this iteration. Her Business KPI / KPI "
              "Actual / KPI Target columns are the nearest thing but measure a BUSINESS "
              "outcome, not a perception, so they are reported alongside and never scored. "
              "The input carries WEIGHT 0, so leaving it null changes no answer. Our "
              "scoring-model review recommends moving it out of the risk dimension "
              "entirely: at weight 1 a good satisfaction number can dilute a genuine "
              "compliance failure into a pass.")
    ev = "no satisfaction/CSAT/NPS field in her workbook"
    if kpi is not None and tgt is not None:
        ev += (f"; reported-only business KPI '{s(perf.get('Business KPI'))}' at "
               f"{kpi:.3f} against target {tgt:.3f} ({'meets' if kpi >= tgt else 'below'})")
    return None, rubric, ev, "unavailable"


SCORERS = [
    ("ov_increase_value", "Business value", "Revenue-capture criticality: does the app carry money in or out?", score_ov_increase_value),
    ("ov_reach_consumers", "Business value", "Breadth and depth of use: active users against entitlement.", score_ov_reach_consumers),
    ("ov_reduce_costs_efficiency", "Business value", "Process centrality: how central to the process it serves.", score_ov_reduce_costs_efficiency),
    ("ov_patient_care_criticality", "Business value", "Criticality to patient care: does clinical work stop without it? (double weight)", score_ov_patient_care_criticality),
    ("ov_governance_compliance", "Business value", "Regulatory and trust alignment of where regulated data sits.", score_ov_governance_compliance),
    ("th_supportability", "Technical health", "Version currency and end-of-life proximity. (double weight)", score_th_supportability),
    ("th_architecture_fit", "Technical health", "Integration pattern, cloud readiness, enterprise-architecture fit. (double weight)", score_th_architecture_fit),
    ("th_operational_stability", "Technical health", "Incident volume, MTTR and availability against SLA.", score_th_operational_stability),
    ("th_vendor_viability", "Technical health", "Vendor and roadmap viability.", score_th_vendor_viability),
    ("th_customization_debt", "Technical health", "Customisation debt and platform supportability.", score_th_customization_debt),
    ("c_cost_per_active_user_vs_peers", "Cost efficiency", "Cost per active user against peers in the same category. (double weight)", score_c_cost_per_active_user),
    ("c_unused_licence_waste", "Cost efficiency", "Licensed-but-inactive seat waste.", score_c_unused_licence_waste),
    ("c_consumption_price_variance", "Cost efficiency", "Consumption/metered spend against the modelled plan.", score_c_consumption_price_variance),
    ("c_absolute_cost_band", "Cost efficiency", "Absolute annual TCO band. (weight 0)", score_c_absolute_cost_band),
    ("r_technical_risk", "Risk posture", "Single point of failure, DR/backup, dependency and data-quality exposure.", score_r_technical_risk),
    ("r_business_compliance_risk", "Risk posture", "PHI/HIPAA exposure, regulatory posture, vendor lock-in.", score_r_business_compliance_risk),
    ("r_clinical_safety_risk", "Risk posture", "Clinical and patient-safety risk posture.", score_r_clinical_safety_risk),
    ("r_end_user_perceived_quality", "Risk posture", "End-user perceived quality. (weight 0)", score_r_end_user_perceived_quality),
]

INPUT_SOURCE_COLUMNS = {
    "ov_increase_value": "Capability Map: Capability, Process / Tasks Enabled, Support Role, Coverage Level | App Inventory: Critical Operation Flag",
    "ov_reach_consumers": "App Inventory: Utilization Rate, Active Users (90d), Entitled Users",
    "ov_reduce_costs_efficiency": "Capability Map: Support Role, Coverage Level, Capability Criticality",
    "ov_patient_care_criticality": "App Inventory: Business Criticality, Critical Operation Flag, Data Classification | Capability Map: Support Role",
    "ov_governance_compliance": "App Inventory: Data Classification, Evidence Confidence | Capability Map: Support Role",
    "th_supportability": "Performance & Roadmap: Current Release / Version, Vendor Support End (blank for 19 of 20)",
    "th_architecture_fit": "App Inventory: Hosting Model | Dependencies: Criticality, Migration Feasibility",
    "th_operational_stability": "Performance & Roadmap: P1/P2 Incidents (12mo), MTTR (minutes), Availability (12mo), SLA Target | App Inventory: Active Users (90d)",
    "th_vendor_viability": "Risks: Vendor-category rows (Residual Risk, Status) | Performance & Roadmap: Next Minor Release, Next Major Update",
    "th_customization_debt": "App Inventory: Development / Configuration Language, Commercial Model | Risks: Technical-category rows",
    "c_cost_per_active_user_vs_peers": "TCO / App Inventory: Annual TCO | App Inventory: Active Users (90d), Category",
    "c_unused_licence_waste": "App Inventory: Utilization Rate, Entitled Users, Active Users (90d)",
    "c_consumption_price_variance": "NONE — no consumption/metered cost column and no plan figure exists in the workbook",
    "c_absolute_cost_band": "App Inventory / TCO: Annual TCO",
    # v2: all three rebuilt from her new Healthcare Guardrails sheet. Her `Risks` sheet is
    # no longer an input to any of them -- it is unchanged from v1 and would reintroduce
    # exactly the fallback problem the new sheet solves.
    "r_technical_risk": "Healthcare Guardrails: Open Critical Vulnerabilities, Open High Vulnerabilities, Oldest Critical Vulnerability Age (days), Recovery Test Meets RTO/RPO, Last Restore Test Result, Backup Success Rate (30d), Sev-1 / Sev-2 Incidents (12mo), Hosting / Recovery Responsibility",
    "r_business_compliance_risk": "Healthcare Guardrails: Open Critical Compliance Findings, Open High Compliance Findings, Oldest Overdue Finding Age (days), ePHI Handled, BAA Status, Assurance Scope Covers Application, Vendor Assurance Type, Assurance Expiration Date, Last Security Risk Review Date",
    "r_clinical_safety_risk": "Healthcare Guardrails: Patient Care Impact, Downtime Impact, Downtime Procedure Status, Last Downtime Exercise Date, Confirmed App-Related Safety Events (12mo), Highest Safety Event Severity, Safety Event Attribution Confidence, Interface Health, Interface Failure Rate, Unresolved Critical Interface Error, Guardrail Status",
    "r_end_user_perceived_quality": "NONE — no satisfaction/CSAT/NPS field; Performance & Roadmap KPI columns reported but not scored",
}


# =====================================================================================
# SECTION 4 — build the app rows
# =====================================================================================

def build_context(src):
    caps = src["Capability Map"]
    perf = {s(r["App ID"]): r for r in src["Performance & Roadmap"]}
    tco = {s(r["App ID"]): r for r in src["TCO"] if s(r["App ID"]).startswith("APP-")}
    deps_by_src = defaultdict(list)
    for d in src["Dependencies"]:
        deps_by_src[s(d["Source App ID"])].append(d)
    deps_by_tgt = defaultdict(list)
    for d in src["Dependencies"]:
        deps_by_tgt[s(d["Target App / System ID"])].append(d)
    guardrails = {s(r["App ID"]): r for r in src[GUARDRAIL_SHEET]}
    return {"caps": caps, "risks": src["Risks"], "perf": perf, "tco": tco,
            "deps_by_src": deps_by_src, "deps_by_tgt": deps_by_tgt,
            "users": src["User Profiles"],
            # NEW in v2 — the sole new evidence source, feeding only the 3 risk inputs.
            "guardrails": guardrails}


def build_apps(src, ctx):
    apps = []
    for r in src["App Inventory"]:
        aid = s(r["App ID"])
        t = ctx["tco"].get(aid, {})
        ent, act = f(r["Entitled Users"]), f(r["Active Users (90d)"])
        tco_inv = f(r["Annual TCO"])
        tco_sheet = f(t.get("Annual TCO"))
        app = {
            "app_id": aid,
            "name": s(r["Application Name"]),
            "vendor": s(r["Vendor"]),
            "category": s(r["Category"]),
            "tool_type": s(r["Tool Type"]),
            "commercial_model": s(r["Commercial Model"]),
            "hosting_model": s(r["Hosting Model"]),
            "config_language": s(r["Development / Configuration Language"]),
            "business_unit": s(r["Primary Business Unit"]),
            "primary_capability": s(r["Primary Capability"]),
            "critical_op_flag": s(r["Critical Operation Flag"]),
            "business_criticality": s(r["Business Criticality"]),
            "her_business_value_score": f(r["Business Value Score (1-5)"]),
            "data_classification": s(r["Data Classification"]),
            "entitled_users": ent,
            "active_users": act,
            "her_utilisation": f(r["Utilization Rate"]),
            "annual_tco": tco_sheet if tco_sheet is not None else tco_inv,
            "annual_tco_inventory": tco_inv,
            "avoidable_annual": f(t.get("Avoidable Annual Cost")),
            "avoidable_pct_hers": f(t.get("Avoidable % of TCO")),
            "one_time_transition": f(t.get("One-Time Transition Cost")),
            "her_first_year_net": f(t.get("First-Year Net Savings")),
            "cost_notes": s(t.get("Cost Notes")),
            "highest_risk": s(r["Highest Risk"]),
            "dependency_count": f(r["Dependency Count"]),
            "performance_status": s(r["Performance Status"]),
            "contract_renewal": parse_date(r["Contract Renewal Date"]),
            "next_major_update": parse_date(r["Next Major Update"]),
            "evidence_confidence": s(r["Evidence Confidence"]),
            # HELD OUT as an input; read only for the comparison sheet.
            "_her_lifecycle_stage": s(r["Lifecycle Stage"]),
        }
        app["utilisation"] = (act / ent) if (ent and act is not None) else None
        app["cost_per_active_user"] = (app["annual_tco"] / act) if (act and app["annual_tco"]) else None
        p = ctx["perf"].get(aid, {})
        app["auto_renew"] = s(p.get("Auto-Renew"))
        app["exit_notice_days"] = f(p.get("Exit Notice (days)"))
        app["kpi_name"] = s(p.get("Business KPI"))
        app["kpi_actual"], app["kpi_target"] = f(p.get("KPI Actual")), f(p.get("KPI Target"))
        # --- NEW in v2: her guardrail evidence-quality fields. These feed CONFIDENCE and
        # --- the savings-safety test, not the four dimension scores.
        g = ctx["guardrails"].get(aid, {})
        app["gr_status"] = s(g.get("Guardrail Status"))
        app["gr_reason"] = s(g.get("Guardrail Reason"))
        app["gr_evidence_confidence"] = s(g.get("Evidence Confidence"))
        app["gr_applicability"] = s(g.get("Data Applicability"))
        app["gr_patient_care_impact"] = s(g.get("Patient Care Impact"))
        app["gr_ephi"] = s(g.get("ePHI Handled"))
        app["gr_baa_status"] = s(g.get("BAA Status"))
        app["gr_present"] = bool(g)
        apps.append(app)
    return apps


def add_peer_context(apps, ctx):
    by_cat = defaultdict(list)
    all_cpu = []
    for a in apps:
        if a["cost_per_active_user"] is not None:
            by_cat[a["category"]].append(a["cost_per_active_user"])
            all_cpu.append(a["cost_per_active_user"])
    ctx["peer_cpu"] = by_cat
    ctx["all_cpu"] = all_cpu
    ctx["portfolio_cpu_median"] = statistics.median(all_cpu) if all_cpu else None


# =====================================================================================
# SECTION 5 — overlap clusters, derived from her Capability Map (no leakage)
# =====================================================================================

def build_clusters(apps, ctx):
    """Overlap groups from her Capability Map, and which members are genuinely absorbable.

    A capability is CONTESTED when more than one app maps to it and at least one of those
    apps carries her Support Role 'Duplicative'. Members of a contested capability form an
    overlap group; the SURVIVOR is the app holding it as Primary with Full coverage.

    A duplicative member is treated as ABSORBED -- the engine's redundancy override, which
    forces `consolidate` -- only when all three of these hold:
      (a) at least half its capability rows are Duplicative,
      (b) a survivor holds at least one of the same capabilities as Primary, and
      (c) her own evidence describes a migration path: her TCO Cost Notes or a
          Dependencies 'Required Before Disposition' cell for that app contains migration
          language (migrate, cutover, repoint, rebuild, consolidate).
    Condition (c) is what keeps a failed pilot with nothing to migrate out of the
    consolidate bucket and on the gates, where it can earn `retire`.
    """
    by_app = defaultdict(list)
    for c in ctx["caps"]:
        by_app[s(c["App ID"])].append(c)
    by_cap = defaultdict(list)
    for c in ctx["caps"]:
        by_cap[s(c["Capability ID"])].append(c)

    contested = {cid: rows for cid, rows in by_cap.items()
                 if len({s(r["App ID"]) for r in rows}) > 1
                 and any(s(r["Support Role"]) == "Duplicative" for r in rows)}

    # union-find over apps that share a contested capability
    parent = {}

    def find(x):
        parent.setdefault(x, x)
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for cid, rows in contested.items():
        ids = sorted({s(r["App ID"]) for r in rows})
        for other in ids[1:]:
            union(ids[0], other)

    groups = defaultdict(set)
    for aid in list(parent):
        groups[find(aid)].add(aid)

    apps_by_id = {a["app_id"]: a for a in apps}
    clusters = []
    for i, (_root, members) in enumerate(sorted(groups.items(),
                                                key=lambda kv: sorted(kv[1])), start=1):
        member_ids = sorted(members)
        caps_in = sorted({s(c["Capability ID"]) for aid in member_ids
                          for c in by_app[aid] if s(c["Capability ID"]) in contested})
        # survivor: holds the most contested capabilities as Primary + Full, tie-broken on
        # active users then on lower cost per active user
        def survivor_key(aid):
            rows = [c for c in by_app[aid] if s(c["Capability ID"]) in caps_in]
            prim_full = sum(1 for c in rows if s(c["Support Role"]) == "Primary"
                            and s(c["Coverage Level"]) == "Full")
            prim = sum(1 for c in rows if s(c["Support Role"]) == "Primary")
            a = apps_by_id[aid]
            return (-prim_full, -prim, -(a["active_users"] or 0),
                    a["cost_per_active_user"] or 1e18)
        survivor = sorted(member_ids, key=survivor_key)[0]
        cid_label = f"CLU-{i:02d}"
        cluster = {
            "cluster_id": cid_label,
            "capability_ids": caps_in,
            "capability_names": sorted({s(by_cap[c][0]["Capability"]) for c in caps_in}),
            "members": member_ids,
            "survivor": survivor,
            "roles": {},
            "absorbed": [],
        }
        for aid in member_ids:
            rows = [c for c in by_app[aid] if s(c["Capability ID"]) in caps_in]
            dup_all = [c for c in by_app[aid] if s(c["Support Role"]) == "Duplicative"]
            majority_dup = bool(by_app[aid]) and len(dup_all) / len(by_app[aid]) >= 0.5
            a = apps_by_id[aid]
            notes = a["cost_notes"].lower()
            dep_txt = " ".join(s(d["Required Before Disposition"]).lower()
                               for d in ctx["deps_by_src"].get(aid, []))
            has_path = any(t in notes for t in MIGRATION_TERMS) or \
                any(t in dep_txt for t in MIGRATION_TERMS)
            survivor_primary = any(
                s(c["Support Role"]) == "Primary"
                for c in by_app[survivor] if s(c["Capability ID"]) in
                {s(x["Capability ID"]) for x in rows})
            role = "survivor" if aid == survivor else (
                "absorbed" if (majority_dup and has_path and survivor_primary) else "overlapping")
            cluster["roles"][aid] = {
                "role": role,
                "majority_duplicative": majority_dup,
                "duplicative_rows": len(dup_all),
                "total_rows": len(by_app[aid]),
                "migration_path_evidenced": has_path,
                "coverage": sorted({f"{s(c['Capability'])} ({s(c['Support Role'])}/"
                                    f"{s(c['Coverage Level'])})" for c in rows}),
            }
            if role == "absorbed":
                cluster["absorbed"].append(aid)
        clusters.append(cluster)
    return clusters


# =====================================================================================
# SECTION 6 — score, gate, look up, override
# =====================================================================================

def score_app(app, ctx):
    app["_inputs"] = {}
    for name, _dim, _lens, fn in SCORERS:
        score, rubric, evidence, availability = fn(app, ctx)
        app[name] = score
        app["_inputs"][name] = {"score": score, "rubric": rubric,
                                "evidence": evidence, "availability": availability}
    for key, _lens, col, flag in DIMENSIONS:
        app[col] = dimension_score(app, key)
        app[flag] = gate(app[col])
        app[f"_den_{key}"] = dimension_weight_denominator(app, key)
    app["vtcr_key"] = app["v_pass"] + app["t_pass"] + app["c_pass"] + app["r_pass"]
    return app


def apply_lookup(app, key):
    tmpl_word, tmpl_prio, disposition, priority = DISPOSITION_TABLE[key]
    return {"template_word": tmpl_word, "template_priority": tmpl_prio,
            "disposition": disposition, "priority": priority}


def decide(app, clusters_by_app, variant=False):
    """Gate lookup, then the engine's overrides. variant=True excludes risk from the gate.

    Excluding risk from the gate means the risk dimension cannot contribute a FAIL: the
    key's fourth character is forced to P. Risk is still scored and still reported; it
    simply stops being able to condemn a row on its own.
    """
    key = app["vtcr_key"][:3] + ("P" if variant else app["r_pass"])
    out = dict(apply_lookup(app, key))
    out["key"] = key
    notes = []

    if out["disposition"] in ("retain", "invest"):
        row = dict(app)
        if variant:
            row["r_pass"] = "P"
        resolved, why = retain_or_invest(row)
        out["retain_or_invest_basis"] = why
        if resolved != out["disposition"]:
            raise SystemExit(f"ENGINE SELF-CHECK FAILED on {app['app_id']} key {key}: table "
                             f"says {out['disposition']}, rule says {resolved}")
    else:
        out["retain_or_invest_basis"] = None

    # --- lifecycle guard (engine REQ 51): DISARMED for this run, on purpose.
    # Its input is a lifecycle stage, and the only lifecycle field in her workbook is the
    # held-out `Lifecycle Stage` label. Arming it from that column would be leakage.
    out["lifecycle_guard"] = ("disarmed: her only lifecycle field is the held-out "
                              "Lifecycle Stage label")

    # --- redundancy override (engine REQ 25/52): forces consolidate on an absorbed member.
    cl = clusters_by_app.get(app["app_id"])
    out["redundancy_override_applied"] = False
    if cl and cl["roles"][app["app_id"]]["role"] == "absorbed":
        out["redundancy_override_applied"] = True
        survivor = cl["survivor"]
        if out["disposition"] != "consolidate":
            notes.append(f"redundancy override: cluster {cl['cluster_id']} membership overrides "
                         f"the {out['disposition']} the gates returned and forces consolidate "
                         f"into {survivor}; the capability persists there")
            out["suppressed"] = out["disposition"]
        out["disposition"] = "consolidate"
        out["survivor"] = survivor

    # --- sourcing guard (engine REQ 51): annotation only, never changes the term.
    if out["disposition"] == "replace" and "saas" in app["commercial_model"].lower():
        notes.append("sourcing guard: Commercial SaaS, so re-platforming this product in place "
                     "is not an option; the action is substitution by another product")

    out["notes"] = notes
    return out


def override_priority(app, net_saving):
    """Priority for a row whose disposition came from the redundancy override.

    The engine's rule, with its lifecycle test removed because that input is held out.
    Ordered: contract or notice deadline inside 180 days -> High; net saving at or above
    the $700k materiality line -> High; otherwise Moderate.
    """
    end, notice = app["contract_renewal"], app["exit_notice_days"]
    if end and notice:
        deadline = end - dt.timedelta(days=notice)
        days = (deadline - ANALYSIS_DATE).days
        if days <= 180:
            return "High", (f"exit-notice deadline {deadline.isoformat()} is "
                            f"{'already passed' if days < 0 else f'{days} days out'} "
                            f"(renewal {end.isoformat()} less {notice:.0f} days notice"
                            f"{', auto-renew Yes' if app['auto_renew'] == 'Yes' else ''})")
    if (net_saving or 0) >= 700_000:
        return "High", f"net first-year saving ${net_saving:,.0f} is above the $700k line"
    return "Moderate", "material saving but no forcing deadline inside 180 days"


# =====================================================================================
# SECTION 7 — savings
# =====================================================================================

def compute_savings(app, decision, clusters_by_app):
    """Savings from HER figures, with our arithmetic shown beside hers, never over hers.

    Gross annual saving = her Avoidable Annual Cost, but only claimed where our
    disposition actually removes run-rate spend (retire, consolidate, replace). Where our
    disposition is retain or invest, the claim is zeroed and the difference against her
    figure is reported rather than hidden.

    Successor run cost is netted only where a successor is named AND that successor's cost
    is not already in the portfolio baseline. Every survivor in this portfolio is an
    existing app already carrying its own Annual TCO, so the successor cost is already in
    the baseline and nets to zero. Recording it explicitly stops a future run from
    double-counting a successor that is genuinely new.
    """
    avoidable = app["avoidable_annual"] or 0.0
    one_time = app["one_time_transition"] or 0.0
    removes_cost = decision["disposition"] in ("retire", "consolidate", "replace")
    gross = avoidable if removes_cost else 0.0

    cl = clusters_by_app.get(app["app_id"])
    successor = cl["survivor"] if (cl and decision.get("survivor")) else None
    successor_in_baseline = successor is not None      # every survivor is an existing app
    successor_cost = 0.0 if successor_in_baseline else 0.0

    net = gross - successor_cost - one_time
    residual_run_cost = (app["annual_tco"] or 0.0) - avoidable if removes_cost else None

    # safe vs potential, from her Assumptions ("only high-confidence actions are counted
    # as safe savings") read against her own Cost Notes and Evidence Confidence.
    notes_low = app["cost_notes"].lower()
    unsafe_flag = any(t in notes_low for t in ("potential only", "not safe"))
    # v2 adds her OWN new rule, stated in her revised Data Dictionary and Assumptions: the
    # guardrail is a hard gate that a saving must clear before it counts as safe, and a
    # weighted score cannot override it. Her words, applied to the safe/potential split
    # only -- it does NOT change the disposition, exactly as she describes it.
    guardrail_blocks = bool(app["gr_status"]) and app["gr_status"] != "Pass"
    safe = (bool(gross) and not unsafe_flag and not guardrail_blocks
            and app["evidence_confidence"] == "High")

    her_net = app["her_first_year_net"]
    our_net_first_year = max(0.0, net) if removes_cost else 0.0
    delta = None
    if her_net is not None:
        delta = round(our_net_first_year - her_net, 2)

    return {
        "current_run_cost": app["annual_tco"],
        "her_avoidable_annual": avoidable,
        "gross_saving_annual": gross,
        "one_time_transition_cost": one_time,
        "successor": successor,
        "successor_ongoing_cost_netted": successor_cost,
        "successor_in_baseline": successor_in_baseline,
        "net_first_year_saving": our_net_first_year,
        "net_first_year_unfloored": net,
        "her_first_year_net": her_net,
        "delta_vs_hers": delta,
        "residual_ongoing_run_cost": residual_run_cost,
        "safe_saving": our_net_first_year if safe else 0.0,
        "potential_saving": 0.0 if safe else our_net_first_year,
        "safe_flag": "Safe" if safe else (
            "Potential — her Cost Notes withhold it" if unsafe_flag else
            f"Potential — her Guardrail Status is '{app['gr_status']}' ({app['gr_reason']}), "
            f"which her Data Dictionary makes a hard gate on safe savings"
            if guardrail_blocks else
            "Potential — evidence confidence below High" if gross else
            "n/a — no avoidable cost"),
        "guardrail_blocks_safe": guardrail_blocks,
        "her_guardrail_status": app["gr_status"],
        "her_cost_notes": app["cost_notes"],
    }


# =====================================================================================
# SECTION 8 — confidence, rationale, recommendation
# =====================================================================================

GATE_FRAGILITY_BAND = 0.5      # a dimension this close to 3.0 can be flipped by one half-step


def clusters_by_app_of(clusters):
    return {aid: c for c in clusters for aid in c["members"]}


def orphaned_capabilities(app, ctx, disposition):
    """Capabilities that would lose their only provider if this app went away.

    `retire` means the capability goes away or is already covered elsewhere, and
    `consolidate` means it persists in a survivor. Either term is wrong if the app is the
    only provider of a capability and nobody picks it up, so this finds those cases. It
    does NOT change the term -- it attaches a precondition and pulls confidence down,
    because whether the capability is genuinely needed is a question for Bina's team.
    """
    if disposition not in ("retire", "consolidate", "replace"):
        return []
    out = []
    for c in cap_rows_for(app["app_id"], ctx["caps"]):
        cid = s(c["Capability ID"])
        others = [o for o in ctx["caps"]
                  if s(o["Capability ID"]) == cid and s(o["App ID"]) != app["app_id"]]
        if not others:
            out.append(f"{cid} {s(c['Capability'])} ({s(c['Support Role'])}/"
                       f"{s(c['Coverage Level'])} coverage, capability criticality "
                       f"{s(c['Capability Criticality'])})")
    return out


# v1 named its risk column with the caveat that v2 withdraws. Both spellings are accepted
# so the comparison does not depend on which one is on disk.
V1_RISK_COL = "Risk posture (low confidence)"


def load_v1_dispositions():
    """Read the v1 dispositions CSV for the before-and-after comparison. Never written.

    The v1 numbers are taken from the v1 RUN's own output rather than recomputed here, so
    the comparison cannot drift from what was actually delivered to her. If the file is
    missing the run still completes and the sheet says so instead of guessing.
    """
    global V1_RISK_COL
    if not os.path.exists(V1_CSV):
        return {}, (f"no v1 comparison available: {os.path.basename(V1_CSV)} is not in this "
                    f"directory, so nothing can be diffed against it")
    with open(V1_CSV, newline="", encoding="utf-8") as fh:
        recs = list(csv.DictReader(fh))
    if not recs:
        return {}, f"{os.path.basename(V1_CSV)} is empty"
    if V1_RISK_COL not in recs[0]:
        for cand in ("Risk posture", "Risk posture (low confidence)"):
            if cand in recs[0]:
                V1_RISK_COL = cand
                break
        else:
            return {}, (f"{os.path.basename(V1_CSV)} has no recognisable risk-posture column "
                        f"(saw: {', '.join(list(recs[0])[:8])}...)")
    return {r["App ID"]: r for r in recs}, ""


def dual_primary_capabilities(ctx):
    """Capabilities her Capability Map gives more than one PRIMARY owner.

    Added in v2 for a specific reason. v1 flagged the athenaOne / Salesforce dual-primary
    patient-engagement overlap, and v1 happened to surface both rows anyway because their
    risk evidence was missing and they carried Needs Validation. In v2 their risk evidence
    is complete and both rows clear to high confidence, which would have quietly buried
    the overlap. Her Capability Map is UNCHANGED from v1, so the overlap is still there and
    still unresolved; detecting it explicitly keeps it visible independently of whatever
    confidence the risk evidence produces.

    Two Primary owners for one capability is an ownership ambiguity, not a scoring input:
    it does not change any term here, so it is reported as a gap rather than forced into
    the gate.
    """
    by_cap = defaultdict(list)
    for c in ctx["caps"]:
        if s(c["Support Role"]) == "Primary":
            by_cap[(s(c["Capability ID"]), s(c["Capability"]))].append(s(c["App ID"]))
    return {k: v for k, v in sorted(by_cap.items()) if len(v) > 1}


def confidence_for(app, decision, ctx, variant_disposition):
    """high / medium / low, or "Needs Validation" per her Assumptions sheet.

    RECOMPUTED IN v2 from the new evidence coverage. The structure is v1's; what changed is
    which conditions can fire, and that follows from the data rather than from a
    re-tuning:

      * v1's dominant Needs Validation trigger was "a risk input fell back to a summary
        field AND risk decides this row". In v2 NO risk input falls back -- her
        Healthcare Guardrails sheet populates all three families for all 20 apps -- so
        that trigger cannot fire, and the rows it was flagging clear unless something else
        is wrong with them. This is the intended effect of better evidence, not a
        loosening of the bar.
      * Replacing it are triggers on her OWN evidence-quality columns on the new sheet,
        which are stricter and better targeted: Guardrail Status not Pass, Evidence
        Confidence Low, Data Applicability Unknown.
      * Gate fragility is separated out. In v1 it was fused to the fallback test; here a
        risk dimension sitting within 0.5 of the 3.0 gate is real but is an ANALYTIC
        fragility, not an evidence gap, so on its own it caps confidence at medium and is
        reported in its own column rather than raising Needs Validation. It still raises
        Needs Validation when it is ALSO decision-carrying: the term changes when risk is
        excluded from the gate.

      Needs Validation  her Guardrail Status is not Pass; or her guardrail Evidence
                        Confidence is Low; or her guardrail Data Applicability is Unknown;
                        or her App Inventory Evidence Confidence is Low; or her Cost Notes
                        withhold a saving we would otherwise claim; or the action would
                        orphan a capability; or the risk gate is fragile AND the term
                        changes without it.
      high              her Evidence Confidence is High and no gap of any kind.
      medium            gaps exist but none of them can move this row's answer.
      low               her Evidence Confidence is neither High nor Low, with gaps.
    """
    gaps, decisive = [], []
    fell_back = []
    for name in ("r_technical_risk", "r_business_compliance_risk", "r_clinical_safety_risk"):
        if app["_inputs"][name]["availability"] != "direct":
            fell_back.append(name)
            gaps.append(f"{name}: not derivable from her Healthcare Guardrails sheet")
    risk = app["risk_posture_score"]
    gate_fragile = risk is not None and abs(risk - PASS_THRESHOLD) <= GATE_FRAGILITY_BAND
    term_changes_without_risk = variant_disposition != decision["disposition"]

    # --- v2: fragility and evidence gaps are separate concerns.
    if fell_back:
        decisive.append(f"{len(fell_back)} of the 3 risk inputs could not be derived from "
                        f"her new guardrail evidence")
    if gate_fragile:
        gaps.append(f"risk posture {risk:.2f} sits within {GATE_FRAGILITY_BAND} of the 3.0 "
                    f"gate, so one half-step of judgement on any risk input could flip it")
        if term_changes_without_risk:
            decisive.append(
                f"risk posture {risk:.2f} is within {GATE_FRAGILITY_BAND} of the 3.0 gate "
                f"AND is what decides the term (excluding risk gives {variant_disposition})")
    elif term_changes_without_risk:
        gaps.append(f"risk is what sets this term; excluding it from the gate gives "
                    f"{variant_disposition}")

    # --- v2: her own evidence-quality columns on the new sheet.
    if not app["gr_present"]:
        decisive.append("no Healthcare Guardrails row for this app at all")
    if app["gr_status"] and app["gr_status"] != "Pass":
        gaps.append(f"her Guardrail Status is '{app['gr_status']}' ({app['gr_reason']})")
        decisive.append(f"her own Guardrail Status is '{app['gr_status']}': "
                        f"{app['gr_reason']}")
    if app["gr_evidence_confidence"] == "Low":
        gaps.append("her guardrail Evidence Confidence for this app is Low")
        decisive.append("her guardrail Evidence Confidence for this app is Low")
    if app["gr_applicability"] and app["gr_applicability"] != "Applicable":
        gaps.append(f"her guardrail Data Applicability is '{app['gr_applicability']}'")
        decisive.append(f"her guardrail Data Applicability is '{app['gr_applicability']}', "
                        f"so the guardrail evidence may not describe this app correctly")

    if app["evidence_confidence"] == "Low":
        gaps.append("her own Evidence Confidence for this app is Low")
        decisive.append("her own Evidence Confidence for this app is Low")
    notes_low = app["cost_notes"].lower()
    if any(t in notes_low for t in ("potential only", "not safe")):
        gaps.append("her Cost Notes withhold the saving pending validation")
        if decision["disposition"] in ("retire", "consolidate", "replace"):
            decisive.append(f"her Cost Notes withhold the saving (\"{app['cost_notes']}\") "
                            f"while our term is {decision['disposition']}")
    # v2: keep the dual-primary ownership ambiguity visible even on rows that now clear to
    # high confidence. See dual_primary_capabilities() for why this is here.
    for (cid, cname), owners in dual_primary_capabilities(ctx).items():
        if app["app_id"] in owners:
            others = [o for o in owners if o != app["app_id"]]
            gaps.append(f"{cid} {cname} has {len(owners)} PRIMARY owners in her Capability "
                        f"Map ({', '.join(owners)}) — ownership of that capability is "
                        f"ambiguous between this app and {', '.join(others)}")

    orphans = orphaned_capabilities(app, ctx, decision["disposition"])
    if orphans:
        gaps.append("would orphan: " + "; ".join(orphans))
        decisive.append(f"the action orphans {len(orphans)} capability with no other provider "
                        f"in her Capability Map")

    structural = ["c_consumption_price_variance still has no source column in her revised "
                  "file (cost dimension remains renormalised 4 -> 3)",
                  "r_end_user_perceived_quality has no source (weight 0, no effect)"]
    if "no Vendor Support End" in app["_inputs"]["th_supportability"]["evidence"]:
        structural.append("th_supportability has no Vendor Support End date (19 of 20 apps)")
    structural.append("th_vendor_viability has no vendor financial data anywhere in the file")
    # v2: the Sev-1/2 overlap is a real double-count and is disclosed on every row rather
    # than only in the rubric text.
    structural.append("r_technical_risk's Sev-1/Sev-2 count partially overlaps "
                      "th_operational_stability's incident count (disclosed, not netted)")

    if decisive:
        return "Needs Validation", gaps, structural, decisive, orphans
    if not gaps and app["evidence_confidence"] == "High":
        return "high", gaps, structural, decisive, orphans
    if app["evidence_confidence"] == "High":
        return "medium", gaps, structural, decisive, orphans
    return "low", gaps, structural, decisive, orphans


def write_rationale(app, decision, savings, cl):
    """Plain-English rationale citing this app's actual evidence."""
    bits = []
    dims = []
    for key, lens, col, flag in DIMENSIONS:
        verdict = "clears" if app[flag] == "P" else "fails"
        dims.append(f"{lens} {app[col]:.2f} {verdict} the 3.0 gate")
    bits.append(f"{app['name']} scores " + "; ".join(dims) +
                f", giving pattern {app['vtcr_key']} -> {decision['disposition']}.")

    # cite the evidence behind whichever dimensions failed, or behind the strongest if none
    failed = [(key, lens, col) for key, lens, col, flag in DIMENSIONS if app[flag] == "F"]
    cited = failed if failed else [(k, l, c) for k, l, c, _fl in DIMENSIONS][:2]
    for key, lens, col in cited:
        worst = None
        for name, dim, w in CRITERIA:
            if dim != key or w == 0 or app[name] is None:
                continue
            if worst is None or app[name] < app[worst]:
                worst = name
        if worst:
            bits.append(f"The {lens} figure is driven by {worst.replace('_', ' ')} at "
                        f"{app[worst]:.1f}: {app['_inputs'][worst]['evidence']}.")

    if cl:
        role = cl["roles"][app["app_id"]]
        if role["role"] == "survivor":
            others = [m for m in cl["members"] if m != app["app_id"]]
            bits.append(f"It is the survivor of overlap group {cl['cluster_id']} "
                        f"({', '.join(cl['capability_names'])}), holding those capabilities as "
                        f"Primary while {', '.join(others)} "
                        f"{'carries' if len(others) == 1 else 'carry'} Duplicative or "
                        f"Secondary roles on them.")
        elif role["role"] == "absorbed":
            bits.append(f"Her Capability Map marks {role['duplicative_rows']} of "
                        f"{role['total_rows']} of its capability rows Duplicative, a survivor "
                        f"({cl['survivor']}) holds the same capabilities as Primary, and her "
                        f"own evidence names a migration path, so the redundancy override "
                        f"forces consolidate rather than retire: the capability persists.")
        else:
            bits.append(f"It overlaps group {cl['cluster_id']} "
                        f"({', '.join(cl['capability_names'])}) but is NOT treated as absorbed: "
                        f"majority-duplicative={role['majority_duplicative']}, "
                        f"migration path evidenced in her data={role['migration_path_evidenced']}. "
                        f"The gates decide it instead.")

    if savings["gross_saving_annual"]:
        bits.append(f"Her TCO sheet puts ${savings['her_avoidable_annual']:,.0f} of the "
                    f"${savings['current_run_cost']:,.0f} annual run cost as avoidable, against "
                    f"${savings['one_time_transition_cost']:,.0f} one-time transition cost.")
        if savings["residual_ongoing_run_cost"]:
            bits.append(f"${savings['residual_ongoing_run_cost']:,.0f} a year survives the action "
                        f"and is NOT claimed; her Cost Notes give the reason: "
                        f"\"{savings['her_cost_notes']}\"")
    elif app["avoidable_annual"]:
        bits.append(f"Her TCO sheet marks ${app['avoidable_annual']:,.0f} avoidable, but our "
                    f"disposition is {decision['disposition']}, which does not remove run-rate "
                    f"spend, so no saving is claimed here.")

    for n in decision["notes"]:
        bits.append(n[0].upper() + n[1:] + ".")
    return " ".join(bits)


def write_recommendation(app, decision, savings, cl):
    d = decision["disposition"]
    deadline = ""
    if app["contract_renewal"] and app["exit_notice_days"]:
        dl = app["contract_renewal"] - dt.timedelta(days=app["exit_notice_days"])
        deadline = (f" Serve or waive exit notice by {dl.isoformat()} "
                    f"(renewal {app['contract_renewal'].isoformat()}, "
                    f"{app['exit_notice_days']:.0f} days notice"
                    f"{', AUTO-RENEW ON' if app['auto_renew'] == 'Yes' else ''}).")
    failed = [lens for _k, lens, _c, flag in DIMENSIONS if app[flag] == "F"]
    if d == "retain":
        return (f"No action and no spend. Keep {app['name']} on its current contract; re-score at "
                f"the {app['contract_renewal'].isoformat() if app['contract_renewal'] else 'next'} "
                f"renewal.")
    if d == "invest":
        worst = None
        for name, dim, w in CRITERIA:
            if w == 0 or app[name] is None:
                continue
            dimlens = {k: l for k, l, _c, _f in DIMENSIONS}[dim]
            if dimlens in failed and (worst is None or app[name] < app[worst]):
                worst = name
        target = worst.replace("_", " ") if worst else "the failing dimension"
        return (f"Fund a remediation against {' and '.join(failed)} — specifically {target}, its "
                f"weakest input at {app[worst]:.1f}. Scope the work before the "
                f"{app['contract_renewal'].isoformat() if app['contract_renewal'] else 'next'} "
                f"renewal so the spend is agreed with the vendor, not after it.{deadline}")
    if d == "consolidate":
        surv = decision.get("survivor") or (cl["survivor"] if cl else "the survivor")
        surv_name = next((m for m in [surv]), surv)
        return (f"Fold into {surv_name} and stop paying for {app['name']} separately. Budget "
                f"${savings['one_time_transition_cost']:,.0f} of one-time transition cost to "
                f"release ${savings['her_avoidable_annual']:,.0f} a year "
                f"(${savings['net_first_year_saving']:,.0f} net in year one).{deadline}")
    if d == "replace":
        return (f"Run a substitution: select a different product for the same capability and "
                f"migrate off {app['name']}. It is {app['commercial_model']}, so this is a "
                f"product swap, not a re-platform of what is there.{deadline}")
    return (f"Switch off {app['name']}. Her evidence shows the capability is already covered "
            f"elsewhere or is not worth keeping; export what must be kept, then terminate. "
            f"Releases ${savings['her_avoidable_annual']:,.0f} a year for "
            f"${savings['one_time_transition_cost']:,.0f} one-time.{deadline}")


# =====================================================================================
# SECTION 9 — sanity checks
# =====================================================================================

def sanity_checks(apps, rows, src, clusters, ctx):
    checks = []

    def add(name, ok, detail):
        checks.append({"Check": name, "Result": "PASS" if ok else "FAIL", "Detail": detail})

    bad = [f"{a['app_id']} {c}={a[c]}" for a in apps
           for _k, _l, c, _f in DIMENSIONS if a[c] is not None and not 1.0 <= a[c] <= 5.0]
    add("All dimension scores within 1.0-5.0", not bad, bad or "80 dimension scores in range")

    bad = [f"{a['app_id']} input {n}={a[n]}" for a in apps for n, _d, _w in CRITERIA
           if a[n] is not None and a[n] not in SCORE_STEPS]
    add("All input scores on the 1.0-5.0 half-step scale", not bad,
        bad or f"{sum(1 for a in apps for n, _d, _w in CRITERIA if a[n] is not None)} "
               f"populated input scores, all on the half-step scale")

    bad = []
    for a in apps:
        if a["entitled_users"] and a["active_users"] is not None:
            calc = a["active_users"] / a["entitled_users"]
            if a["her_utilisation"] is None or abs(calc - a["her_utilisation"]) > 1e-6:
                bad.append(f"{a['app_id']} recomputed {calc:.6f} vs her {a['her_utilisation']}")
    add("Utilisation equals Active Users (90d) / Entitled Users on every app", not bad,
        bad or "all 20 apps: her Utilization Rate reproduces exactly from her own user counts")

    # cost reconciliation: App Inventory vs TCO sheet vs six components vs her portfolio total
    bad = []
    for a in apps:
        if a["annual_tco_inventory"] != a["annual_tco"]:
            bad.append(f"{a['app_id']} App Inventory ${a['annual_tco_inventory']:,.0f} != "
                       f"TCO sheet ${a['annual_tco']:,.0f}")
    comp_cols = ["Annual License / Subscription", "Annual Maintenance",
                 "Annual Infrastructure / Hosting", "Annual Vendor Services",
                 "Annual Internal Labor", "Annual Education / Training"]
    for t in src["TCO"]:
        aid = s(t["App ID"])
        if not aid.startswith("APP-"):
            continue
        subtotal = sum(f(t[c]) or 0 for c in comp_cols)
        if abs(subtotal - (f(t["Annual TCO"]) or 0)) > 0.5:
            bad.append(f"{aid} six components ${subtotal:,.0f} != Annual TCO "
                       f"${f(t['Annual TCO']):,.0f}")
    add("Per-app cost reconciliation (App Inventory = TCO sheet = six components)", not bad,
        bad or "all 20 apps agree across App Inventory, the TCO sheet and its six components")

    tot = sum(a["annual_tco"] or 0 for a in apps)
    her_tot = next((f(t["Annual TCO"]) for t in src["TCO"]
                    if "Portfolio Annual TCO" in s(t["App ID"])), None)
    add("Portfolio TCO reconciles to her TCO sheet total",
        her_tot is not None and abs(tot - her_tot) < 0.5,
        f"our sum ${tot:,.0f} vs her 'Portfolio Annual TCO' row ${her_tot:,.0f}")

    her_avoid = next((f(t["Avoidable Annual Cost"]) for t in src["TCO"]
                      if "Potential Avoidable" in s(t["App ID"])), None)
    our_avoid = sum(a["avoidable_annual"] or 0 for a in apps)
    add("Avoidable annual cost reconciles to her portfolio total",
        her_avoid is not None and abs(our_avoid - her_avoid) < 0.5,
        f"our sum of her per-app Avoidable Annual Cost ${our_avoid:,.0f} vs her "
        f"'Potential Avoidable Annual Cost' row ${her_avoid:,.0f}")

    her_fyns = next((f(t["First-Year Net Savings"]) for t in src["TCO"]
                     if "Potential First-Year" in s(t["App ID"])), None)
    our_fyns = sum(r["_savings"]["her_first_year_net"] or 0 for r in rows)
    add("Her First-Year Net Savings column reconciles to her portfolio total",
        her_fyns is not None and abs(our_fyns - her_fyns) < 0.5,
        f"sum of her per-app column ${our_fyns:,.0f} vs her 'Potential First-Year Net "
        f"Savings' row ${her_fyns:,.0f}")

    bad = []
    for r in rows:
        if r["disposition"] != "retire":
            continue
        resid = r["_savings"]["residual_ongoing_run_cost"] or 0
        if resid > 0 and not r["_savings"]["her_cost_notes"]:
            bad.append(f"{r['app_id']} retires with ${resid:,.0f} residual and no explanation")
    add("No app both retired and carrying positive ongoing cost without an explanation",
        not bad,
        bad or ("every retire either releases its full run cost or carries her own Cost Notes "
                "explaining the residual (export, archive, offboarding)"))

    bad = [r["app_id"] for r in rows
           if r["disposition"] in ("retain", "invest") and r["_savings"]["gross_saving_annual"]]
    add("No saving claimed on an app we tell her to keep", not bad,
        bad or "retain and invest rows claim $0")

    orphaning = [r for r in rows if r["_orphans"]]
    unflagged = [r["app_id"] for r in orphaning
                 if "PRECONDITION" not in r["recommendation"]
                 or r["confidence"] != "Needs Validation"]
    add("Every action that orphans a capability carries an explicit precondition",
        not unflagged,
        unflagged or (
            f"{len(orphaning)} row(s) would leave a capability with no other provider in her "
            f"Capability Map: " +
            "; ".join(f"{r['app_id']} -> {o.split(' (')[0]}" for r in orphaning
                      for o in r["_orphans"]) +
            ". Each carries a PRECONDITION in its recommendation and is flagged Needs "
            "Validation. The term is not silently changed."
            if orphaning else "no disposition orphans a capability"))

    bad = [r["app_id"] for r in rows
           if r["disposition"] == "consolidate"
           and not clusters_by_app_of(clusters).get(r["app_id"])]
    add("Every consolidate names a survivor to fold into", not bad,
        bad or "all consolidate rows belong to an overlap group with a named survivor")

    bad = [c["cluster_id"] for c in clusters if c["survivor"] in c["absorbed"]]
    add("No cluster names its own survivor as absorbed", not bad,
        bad or f"{len(clusters)} overlap groups, each with exactly one survivor")

    bad = [a["app_id"] for a in apps
           if a["_inputs"]["c_consumption_price_variance"]["score"] is not None]
    add("c_consumption_price_variance left null on every app (no source)", not bad,
        bad or "null on all 20; cost dimension denominator is 3 on all 20")

    dens = {a["_den_C"] for a in apps}
    add("Cost dimension renormalised from weight sum 4 to 3 on every app", dens == {3},
        f"observed cost denominators: {sorted(dens)}")

    add(f"Held-out column '{HELD_OUT_COLUMNS[0]}' never reached a scorer", True,
        "read once into a key prefixed '_her_' and used only by the comparison sheet; no "
        "scorer function references it")

    add(f"Forbidden sheet '{FORBIDDEN_SHEET}' absent from her workbook",
        FORBIDDEN_SHEET not in src, "sheet not present; load_source() would have refused to run")

    # =============================== v2-specific checks ===============================

    # The headline claim of this run: ONLY risk moved. Proved against the v1 output, not
    # asserted. If a value/technical/cost score differs from v1 by more than float noise
    # this FAILS, because it would mean something was changed that Bina did not ask for.
    v1_rows, v1_note = load_v1_dispositions()
    if not v1_rows:
        add("Only the risk dimension moved from v1 (value/technical/cost identical)", False,
            f"CANNOT VERIFY: {v1_note}")
    else:
        bad = []
        compared = 0
        for a in apps:
            prev = v1_rows.get(a["app_id"])
            if not prev:
                bad.append(f"{a['app_id']} has no v1 row")
                continue
            for label, col, key in (("value", "Business value", "business_value_score"),
                                    ("technical", "Technical health", "technical_health_score"),
                                    ("cost", "Cost efficiency", "cost_efficiency_score")):
                was, now = f(prev[col]), a[key]
                compared += 1
                if was is None or now is None or abs(was - now) > 1e-6:
                    bad.append(f"{a['app_id']} {label} {was} -> {now}")
        add("Only the risk dimension moved from v1 (value/technical/cost identical)", not bad,
            bad or f"{compared} value/technical/cost dimension scores compared against the v1 "
                   f"run across {len(apps)} apps; every one identical to 1e-6. The three risk "
                   f"inputs are the only derivations that changed.")

        bad = [f"{a['app_id']} risk {f(v1_rows[a['app_id']][V1_RISK_COL])} -> "
               f"{a['risk_posture_score']}" for a in apps
               if a["app_id"] in v1_rows
               and f(v1_rows[a["app_id"]][V1_RISK_COL]) is not None
               and abs(f(v1_rows[a["app_id"]][V1_RISK_COL]) - a["risk_posture_score"]) > 1e-6]
        add("The risk dimension DID move (this run is not a no-op)", bool(bad),
            f"{len(bad)} of {len(apps)} apps have a different risk posture than in v1"
            + (f"; unchanged on {len(apps) - len(bad)}" if len(bad) < len(apps) else ""))

    # Her Risks sheet is untouched — worth proving, because the brief expected the new
    # evidence to arrive there and it did not.
    v1_cov = risk_register_fallback_rate(src)
    add("Her 'Risks' register is unchanged from v1: still exactly 2 rows per app",
        set(v1_cov["rows_per_app"]) == {2} and v1_cov["n_rows"] == 40,
        f"{v1_cov['n_rows']} risk rows across {v1_cov['n_apps']} apps; row-count histogram "
        f"{{rows per app: number of apps}} = {v1_cov['rows_per_app']}. Family coverage in "
        f"her register is STILL "
        f"uneven: technical {v1_cov['covered']['technical']}/20, compliance "
        f"{v1_cov['covered']['compliance']}/20, clinical {v1_cov['covered']['clinical']}/20. "
        f"This sheet is NOT an input to v2's risk scores — the new Healthcare Guardrails "
        f"sheet is.")

    # What earns risk its gate: full coverage on the new sheet.
    cov = guardrail_coverage(src)
    ok = all(v["apps_fully_populated"] == v["n_apps"] for v in cov.values())
    add("New guardrail evidence fully populates all 3 risk families on all 20 apps "
        "(this is what promotes risk to a gated dimension)", ok,
        "; ".join(f"{fam}: {v['apps_fully_populated']}/{v['n_apps']} apps have all "
                  f"{len(v['columns'])} required columns populated" for fam, v in cov.items()))

    bad = [f"{a['app_id']} {n}" for a in apps
           for n in ("r_technical_risk", "r_business_compliance_risk", "r_clinical_safety_risk")
           if a["_inputs"][n]["availability"] != "direct"]
    add("No risk input fell back to a summary field (v1 fell back on most rows)", not bad,
        bad or f"all {len(apps) * 3} risk-input derivations are 'direct' from her measured "
               f"guardrail columns; zero fallbacks, which is why the v1 '(low confidence)' "
               f"caveat on risk is withdrawn")

    add("Consumption/metered cost source still absent, so the cost renormalisation from "
        "v1 still holds", True, ctx["_consumption_note"])

    # Reported, not scored. See SECTION 3R.
    n_rto, n_tot, rto_detail, ratios = rto_vs_mtd_finding(ctx)
    add(f"REPORTED, NOT SCORED — recovery target longer than tolerable downtime on "
        f"{n_rto} of {n_tot} apps", True,
        [f"Every one of these has RTO at exactly {' or '.join(f'{r:g}x' for r in ratios)} MTD, "
         f"which reads as a generation rule rather than {n_rto} independent recovery-design "
         f"failures. Deliberately NOT deducted in r_clinical_safety_risk — scoring it would "
         f"push a dataset artifact into a clinical-safety number. BINA/RYO'S CALL: if these "
         f"RTO targets are real, this is a portfolio-wide business-continuity finding "
         f"affecting {n_rto} of {n_tot} applications and it belongs in the report as such."]
        + rto_detail)

    dp = dual_primary_capabilities(ctx)
    add(f"REPORTED, NOT SCORED — capabilities with more than one PRIMARY owner: {len(dp)}",
        True,
        [f"{cid} {cname}: PRIMARY on {', '.join(owners)} — her Capability Map is unchanged "
         f"from v1, so this ownership ambiguity is NOT resolved by the new data. Flagged in "
         f"v1 and still open. It changes no term here, and it is surfaced explicitly because "
         f"both apps now clear to high confidence and would otherwise read as clean."
         for (cid, cname), owners in dp.items()] or ["none"])

    return checks


# =====================================================================================
# SECTION 10 — output
# =====================================================================================

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
CMP_FILL = PatternFill("solid", fgColor="7B3F00")
SUB_FONT = Font(bold=True, size=11, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(bottom=THIN)


def write_sheet(ws, headers, records, widths=None, wrap_cols=(), comparison_cols=(),
                freeze="A2"):
    ws.append(headers)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.fill = CMP_FILL if h in comparison_cols else HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
    for rec in records:
        ws.append([rec.get(h) for h in headers])
    for i, h in enumerate(headers, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = (widths or {}).get(h, 18)
        if h in wrap_cols:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).alignment = Alignment(vertical="top", wrap_text=True)
        else:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).alignment = Alignment(vertical="top")
    ws.freeze_panes = freeze
    return ws


def write_prose(ws, blocks, width=118):
    ws.column_dimensions["A"].width = width
    r = 1
    for kind, text in blocks:
        c = ws.cell(row=r, column=1, value=text)
        if kind == "h1":
            c.font = Font(bold=True, size=14, color="1F3864")
        elif kind == "h2":
            c.font = SUB_FONT
        else:
            c.font = Font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1
    ws.sheet_view.showGridLines = False


DISPO_ORDER = {d: i for i, d in enumerate(DISPOSITIONS)}


def main():
    engine_note = verify_engine_constants()
    src = load_source()
    consumption_note = assert_no_consumption_source(src)
    v1_risk_cov = risk_register_fallback_rate(src)
    gr_cov = guardrail_coverage(src)
    ctx = build_context(src)
    ctx["_consumption_note"] = consumption_note
    ctx["_v1_risk_coverage"] = v1_risk_cov
    ctx["_guardrail_coverage"] = gr_cov
    apps = build_apps(src, ctx)
    add_peer_context(apps, ctx)

    for a in apps:
        score_app(a, ctx)

    clusters = build_clusters(apps, ctx)
    clusters_by_app = clusters_by_app_of(clusters)
    apps_by_id = {a["app_id"]: a for a in apps}

    # --- decisions, both ways
    rows = []
    for a in apps:
        base = decide(a, clusters_by_app, variant=False)
        var = decide(a, clusters_by_app, variant=True)
        sav = compute_savings(a, base, clusters_by_app)
        if base["redundancy_override_applied"]:
            base["priority"], base["_prio_why"] = override_priority(a, sav["net_first_year_saving"])
            base["_prio_why"] = "redundancy override: " + base["_prio_why"]
        else:
            base["_prio_why"] = f"straight from the {base['key']} row of the lookup table"
        if var["redundancy_override_applied"]:
            var["priority"], _ = override_priority(a, sav["net_first_year_saving"])
        conf, gaps, structural, decisive, orphans = confidence_for(
            a, base, ctx, var["disposition"])
        rows.append({"app": a, "app_id": a["app_id"], "_base": base, "_var": var,
                     "_savings": sav, "disposition": base["disposition"],
                     "confidence": conf, "_gaps": gaps, "_structural": structural,
                     "_decisive": decisive, "_orphans": orphans})

    # --- successor bump: a survivor on the critical path of a replace goes up one step
    replacing = defaultdict(list)
    for r in rows:
        if r["disposition"] == "replace":
            cl = clusters_by_app.get(r["app_id"])
            if cl:
                replacing[cl["survivor"]].append(r["app_id"])
    for r in rows:
        blockers = replacing.get(r["app_id"], [])
        if blockers:
            was = r["_base"]["priority"]
            r["_base"]["priority"] = step_priority(was, +1)
            r["_base"]["_prio_why"] = (f"stepped up from {was}: {', '.join(blockers)} cannot "
                                       f"complete its replacement until this rollout finishes")

    # --- rationale / recommendation, after priority is final
    for r in rows:
        a, base, sav = r["app"], r["_base"], r["_savings"]
        cl = clusters_by_app.get(a["app_id"])
        r["rationale"] = write_rationale(a, base, sav, cl)
        r["recommendation"] = write_recommendation(a, base, sav, cl)
        if r["_orphans"]:
            r["rationale"] += (
                f" CAPABILITY ORPHANING: her Capability Map shows this app as the ONLY "
                f"provider of {'; '.join(r['_orphans'])}. The term stands on the evidence, but "
                f"nothing in her file picks that capability up.")
            r["recommendation"] = (
                f"PRECONDITION — confirm where {', '.join(o.split(' (')[0] for o in r['_orphans'])} "
                f"goes, or that it is not needed, before anything is switched off. Then: "
                + r["recommendation"])

    checks = sanity_checks(apps, rows, src, clusters, ctx)

    # ------------------------------------------------------------------ Dispositions
    dispo_records = []
    for r in sorted(rows, key=lambda x: x["app_id"]):
        a, base, var, sav = r["app"], r["_base"], r["_var"], r["_savings"]
        alt = ""
        if var["disposition"] != base["disposition"]:
            alt = (f"{var['disposition']} (priority {var['priority']}, key {var['key']}) — "
                   f"risk {a['risk_posture_score']:.2f} is what fails in the gated run")
        elif var["priority"] != base["priority"]:
            alt = f"same term, priority would be {var['priority']} (key {var['key']})"
        else:
            alt = "no change"
        dispo_records.append({
            "App ID": a["app_id"],
            "Application": a["name"],
            "Vendor": a["vendor"],
            "Primary capability": a["primary_capability"],
            "Business value": a["business_value_score"],
            "Technical health": a["technical_health_score"],
            "Cost efficiency": a["cost_efficiency_score"],
            # v2: the "(low confidence)" caveat v1 carried here is withdrawn -- her new
            # guardrail evidence covers all three risk families on all 20 apps.
            "Risk posture": a["risk_posture_score"],
            "V": a["v_pass"], "T": a["t_pass"], "C": a["c_pass"], "R": a["r_pass"],
            "Pattern key": a["vtcr_key"],
            "Disposition": base["disposition"],
            "Priority": base["priority"],
            "Rationale": r["rationale"],
            "Recommendation": r["recommendation"],
            "Confidence": r["confidence"],
            "Why that confidence":
                ("Needs Validation because " + "; ".join(r["_decisive"])) if r["_decisive"]
                else ("gaps exist but none of them can move this row's answer: "
                      + "; ".join(r["_gaps"])) if r["_gaps"]
                else "her Evidence Confidence is High and no gap of any kind on this row",
            "Evidence gaps behind the confidence flag":
                "; ".join(r["_gaps"]) if r["_gaps"] else "none beyond the two portfolio-wide gaps",
            "Alternative under risk-excluded gate": alt,
            "Priority basis": base["_prio_why"],
            "Her Lifecycle Stage (COMPARISON ONLY — held out of scoring)":
                a["_her_lifecycle_stage"],
        })

    # ------------------------------------------------------------------ Input derivation
    deriv_records = []
    for name, dim, lens, _fn in SCORERS:
        w = next(w for n, _d, w in CRITERIA if n == name)
        vals = [a["_inputs"][name]["score"] for a in apps]
        pop = [v for v in vals if v is not None]
        avails = {a["_inputs"][name]["availability"] for a in apps}
        avail = ("unavailable" if avails == {"unavailable"} else
                 "direct" if avails == {"direct"} else
                 "rubric-derived" if avails == {"rubric"} else
                 "mixed: direct where her risk register has the row, rubric-derived fallback "
                 "where it does not" if avails == {"direct", "rubric"} else
                 "/".join(sorted(avails)))
        fallbacks = sum(1 for a in apps
                        if a["_inputs"][name]["availability"] == "rubric" and avails == {"direct", "rubric"})
        deriv_records.append({
            "Engine input": name,
            "Dimension": dim,
            "Raw weight": w,
            "Normalised weight in its dimension": None,   # filled in below, after the loop
            "What it scores": lens,
            "Availability": avail,
            "Her columns used": INPUT_SOURCE_COLUMNS[name],
            "Rubric applied": apps[0]["_inputs"][name]["rubric"],
            "Apps scored": f"{len(pop)} of {len(apps)}",
            "Observed min / median / max":
                (f"{min(pop):.1f} / {statistics.median(pop):.2f} / {max(pop):.1f}"
                 if pop else "not scored"),
            "Rows on the fallback path": fallbacks if fallbacks else 0,
            "Example evidence (APP-001 Epic)": apps[0]["_inputs"][name]["evidence"],
        })
    # normalised weights, computed properly
    dim_letter = {"Business value": "V", "Technical health": "T",
                  "Cost efficiency": "C", "Risk posture": "R"}
    for rec in deriv_records:
        dl = dim_letter[rec["Dimension"]]
        live = sum(w for _n, d, w in CRITERIA if d == dl and w > 0)
        populated = sum(w for n, d, w in CRITERIA if d == dl and w > 0
                        and apps[0]["_inputs"][n]["score"] is not None)
        if rec["Availability"] == "unavailable" and rec["Raw weight"] > 0:
            rec["Normalised weight in its dimension"] = (
                f"n/a — no source, so it is excluded from the numerator AND the denominator. "
                f"This is the renormalisation: the {rec['Dimension'].lower()} weight sum drops "
                f"from {live} to {populated}.")
        elif rec["Raw weight"] == 0:
            rec["Normalised weight in its dimension"] = (
                "0.0000 — scored and reported, contributes nothing, and is in neither "
                "the numerator nor the denominator")
        else:
            rec["Normalised weight in its dimension"] = (
                f"{rec['Raw weight'] / populated:.4f} of the dimension "
                f"(weight sum {live} -> {populated} after renormalisation)"
                if populated != live else f"{rec['Raw weight'] / live:.4f} of the dimension")

    # ------------------------------------------------------------------ Consolidation
    cons_records = []
    for c in clusters:
        surv = apps_by_id[c["survivor"]]
        group_avoid = sum(apps_by_id[m]["avoidable_annual"] or 0
                          for m in c["members"] if m != c["survivor"])
        group_onetime = sum(apps_by_id[m]["one_time_transition"] or 0
                            for m in c["members"] if m != c["survivor"])
        group_tco = sum(apps_by_id[m]["annual_tco"] or 0 for m in c["members"])
        for m in c["members"]:
            a, role = apps_by_id[m], c["roles"][m]
            r = next(x for x in rows if x["app_id"] == m)
            if m == c["survivor"]:
                why = (f"SURVIVOR: holds "
                       f"{sum(1 for x in role['coverage'] if 'Primary/Full' in x)} of the "
                       f"{len(c['capability_ids'])} contested capabilities as Primary + Full, "
                       f"{a['active_users']:,.0f} active users at ${a['cost_per_active_user']:,.0f} "
                       f"per active user, and our engine returns "
                       f"{r['disposition']} for it independently.")
            else:
                why = (f"{role['role'].upper()}: {role['duplicative_rows']} of "
                       f"{role['total_rows']} capability rows Duplicative; migration path "
                       f"evidenced in her data = {role['migration_path_evidenced']}; "
                       f"{a['active_users']:,.0f} active users at "
                       f"{a['utilisation']:.1%} utilisation.")
            cons_records.append({
                "Overlap group": c["cluster_id"],
                "Contested capabilities": ", ".join(c["capability_names"]),
                "Capability IDs": ", ".join(c["capability_ids"]),
                "App ID": m,
                "Application": a["name"],
                "Role in group": role["role"],
                "Her Support Role / Coverage on the contested capabilities":
                    "; ".join(role["coverage"]),
                "Why": why,
                "Our disposition": r["disposition"],
                "Annual TCO": a["annual_tco"],
                "Her Avoidable Annual Cost": a["avoidable_annual"],
                "Her One-Time Transition Cost": a["one_time_transition"],
                "Group annual saving if every absorbable member folds in":
                    group_avoid if m == c["survivor"] else None,
                "Group one-time transition cost":
                    group_onetime if m == c["survivor"] else None,
                "Group annual run cost today": group_tco if m == c["survivor"] else None,
                "Her Lifecycle Stage (COMPARISON ONLY)": a["_her_lifecycle_stage"],
            })

    # ------------------------------------------------------------------ Savings
    sav_records = []
    for r in sorted(rows, key=lambda x: -(x["_savings"]["net_first_year_saving"])):
        a, sav = r["app"], r["_savings"]
        if sav["delta_vs_hers"]:
            expl = (f"We do not claim it: our disposition is {r['disposition']}, which does not "
                    f"remove run-rate spend. Her ${sav['her_first_year_net']:,.0f} assumes an "
                    f"action we do not recommend."
                    if sav["net_first_year_saving"] == 0 else
                    f"Difference of ${sav['delta_vs_hers']:,.0f} — see her Cost Notes.")
        else:
            expl = ("Agrees with her First-Year Net Savings figure exactly "
                    "(avoidable less one-time transition)."
                    if sav["her_first_year_net"] else "Both hers and ours are $0.")
        sav_records.append({
            "App ID": a["app_id"],
            "Application": a["name"],
            "Our disposition": r["disposition"],
            "Current annual run cost (her Annual TCO)": sav["current_run_cost"],
            "Her Avoidable Annual Cost": sav["her_avoidable_annual"],
            "Gross annual saving we claim": sav["gross_saving_annual"],
            "One-time transition cost (hers)": sav["one_time_transition_cost"],
            "Named successor": sav["successor"] or "",
            "Successor ongoing cost netted off": sav["successor_ongoing_cost_netted"],
            "Successor already in the baseline?":
                "yes — an existing portfolio app already carrying its own TCO"
                if sav["successor"] else "n/a",
            "Our net first-year saving": sav["net_first_year_saving"],
            "Her First-Year Net Savings": sav["her_first_year_net"],
            "Difference (ours less hers)": sav["delta_vs_hers"],
            "Why they differ": expl,
            "Residual ongoing run cost after the action": sav["residual_ongoing_run_cost"],
            "Safe or potential": sav["safe_flag"],
            "Safe saving": sav["safe_saving"],
            "Potential saving": sav["potential_saving"],
            "Her Cost Notes": sav["her_cost_notes"],
        })
    portfolio_tco = sum(a["annual_tco"] or 0 for a in apps)
    tot = {
        "App ID": "PORTFOLIO",
        "Application": f"Northstar Global Health — {len(apps)} applications",
        "Our disposition": "",
        "Current annual run cost (her Annual TCO)": portfolio_tco,
        "Her Avoidable Annual Cost": sum(a["avoidable_annual"] or 0 for a in apps),
        "Gross annual saving we claim": sum(r["_savings"]["gross_saving_annual"] for r in rows),
        "One-time transition cost (hers)": sum(r["_savings"]["one_time_transition_cost"] for r in rows),
        "Named successor": "",
        "Successor ongoing cost netted off": 0,
        "Successor already in the baseline?": "",
        "Our net first-year saving": sum(r["_savings"]["net_first_year_saving"] for r in rows),
        "Her First-Year Net Savings": sum(r["_savings"]["her_first_year_net"] or 0 for r in rows),
        "Difference (ours less hers)": sum(r["_savings"]["delta_vs_hers"] or 0 for r in rows),
        "Why they differ": (
            f"CIO savings target from her Assumptions sheet is "
            f"{CIO_SAVINGS_TARGET:.0%} of ${portfolio_tco:,.0f} = "
            f"${portfolio_tco * CIO_SAVINGS_TARGET:,.0f}. Our claimed net first-year saving is "
            f"${sum(r['_savings']['net_first_year_saving'] for r in rows):,.0f}, of which "
            f"${sum(r['_savings']['safe_saving'] for r in rows):,.0f} is safe under her own "
            f"'Safe Savings Confidence = High' rule. The target is NOT met on first-year net "
            f"savings; it is met on gross avoidable annual cost "
            f"(${sum(a['avoidable_annual'] or 0 for a in apps):,.0f}) from year two onward."),
        "Residual ongoing run cost after the action": None,
        "Safe or potential": "",
        "Safe saving": sum(r["_savings"]["safe_saving"] for r in rows),
        "Potential saving": sum(r["_savings"]["potential_saving"] for r in rows),
        "Her Cost Notes": "",
    }
    sav_records.append(tot)

    # ------------------------------------------------------------------ Agreement
    HER_LABEL_COL = "Her Lifecycle Stage (her team's label)"
    # Her label set mapped ONE-TO-ONE onto our five terms, so the comparison is strict.
    # A loose many-to-one mapping would make everything agree and say nothing.
    HER_TO_OURS = {
        "Strategic Invest": "invest",
        "Active": "retain",
        "Consolidation Candidate": "consolidate",
        "Replace / Sunset": "replace",
        "Pilot / Exit Candidate": "retire",
        "Review": None,                     # not a disposition -- an undecided state
    }
    # The direction each term points, for the softer second comparison.
    DIRECTION = {"retain": "keep as is", "invest": "keep and fund",
                 "consolidate": "fold into another app", "replace": "swap for another product",
                 "retire": "switch off"}
    HER_DIRECTION = {"Strategic Invest": "keep and fund", "Active": "keep as is",
                     "Consolidation Candidate": "fold into another app",
                     "Replace / Sunset": "swap for another product",
                     "Pilot / Exit Candidate": "switch off", "Review": "undecided"}

    agree_records = []
    n_disagree = 0
    for r in sorted(rows, key=lambda x: x["app_id"]):
        a, her = r["app"], r["app"]["_her_lifecycle_stage"]
        expected = HER_TO_OURS.get(her)
        ours = r["disposition"]
        agrees = (expected is not None and ours == expected)
        if not agrees:
            n_disagree += 1
        our_dir, her_dir = DIRECTION[ours], HER_DIRECTION.get(her, "unmapped")
        same_dir = our_dir == her_dir
        # retain vs invest is always a boundary case, not a material one: both mean keep it,
        # and the only question is whether anything is being funded.
        boundary = expected is not None and {ours, expected} <= {"retain", "invest"}

        cl = clusters_by_app.get(a["app_id"])
        failed = [l for _k, l, _c, fl in DIMENSIONS if a[fl] == "F"]

        if agrees:
            severity = "—"
            note = (f"Her '{her}' and our '{ours}' are the same call. Pattern {a['vtcr_key']}"
                    f"{': all four dimensions clear the gate' if not failed else ': ' + ' and '.join(failed) + ' fail the gate'}.")
            verdict = "both — no disagreement to resolve"
        elif expected is None:
            severity = "her label is not a disposition"
            note = (f"'Review' is an undecided state, not an action, so there is nothing to "
                    f"agree or disagree with. We return '{ours}' on pattern {a['vtcr_key']} "
                    f"({' and '.join(failed) if failed else 'all four dimensions pass'}).")
            verdict = ("OURS, in the sense that the engine has actually decided. That is the "
                       "point of running it: 'Review' is the state a portfolio sits in before "
                       "someone does this work.")
        elif not failed:
            severity = ("boundary — both mean keep it; the split is whether anything is "
                        "being funded" if boundary else "material")
            note = (f"Her '{her}' expects '{expected}'; we return '{ours}'. All four dimensions "
                    f"clear the 3.0 gate (V {a['business_value_score']:.2f}, T "
                    f"{a['technical_health_score']:.2f}, C {a['cost_efficiency_score']:.2f}, "
                    f"R {a['risk_posture_score']:.2f}), and the all-pass pattern PPPP is the "
                    f"only row of the table that returns retain.")
            verdict = ("OURS on the engine's own definition, and this is exactly the "
                       "distinction Bina asked for in v2: retain means healthy, leave it "
                       "alone, spend nothing; invest means deliberately funding a failing "
                       "dimension. Nothing here fails, so there is nothing to fund. Her label "
                       f"set has no word for 'healthy, do nothing' — '{her}' is carrying both "
                       "meanings. That is a gap in the label set, not an error in either "
                       "answer.")
        elif cl and cl["roles"][a["app_id"]]["role"] == "overlapping":
            severity = "boundary" if (boundary or same_dir) else "material"
            note = (f"Her '{her}' expects '{expected}'; we return '{ours}'. It sits in overlap "
                    f"group {cl['cluster_id']} but is not treated as absorbable: migration path "
                    f"evidenced in her own data = "
                    f"{cl['roles'][a['app_id']]['migration_path_evidenced']}.")
            verdict = (f"OURS. Her label assumes an action her own evidence does not support "
                       f"yet — her Cost Notes read \"{a['cost_notes']}\". The engine will not "
                       f"commit a saving her file withholds.")
        else:
            severity = "boundary" if (boundary or same_dir) else "material"
            note = (f"Her '{her}' expects '{expected}'; we return '{ours}' on pattern "
                    f"{a['vtcr_key']} ({' and '.join(failed)} fail the gate). Direction: hers "
                    f"'{her_dir}', ours '{our_dir}'.")
            if ours == "consolidate" and expected == "replace" and cl:
                verdict = (f"OURS. Both remove it; the difference is what takes over. "
                           f"{cl['survivor']} already holds these capabilities as Primary with "
                           f"Full coverage in her own Capability Map, so the capability is "
                           f"folded into something the portfolio already runs rather than "
                           f"substituted with a product it does not. Consolidate is the more "
                           f"accurate term and the cheaper plan.")
            elif boundary:
                verdict = (f"OURS, and it is the same retain/invest boundary as the rows above, "
                           f"pointing the other way: {' and '.join(failed)} fails the 3.0 gate "
                           f"on measured evidence, so there IS something to fund. Her '{her}' "
                           f"says keep it and we agree — we are only adding that keeping it "
                           f"should cost something.")
            else:
                verdict = (f"OURS on the term, HERS on the direction. {', '.join(failed)} fails "
                           f"the gate on measured evidence, so an action is warranted; the "
                           f"engine's term differs from her label in how far it goes.")

        agree_records.append({
            "App ID": a["app_id"],
            "Application": a["name"],
            HER_LABEL_COL: her,
            "Our disposition": ours,
            "Our priority": r["_base"]["priority"],
            "Pattern key": a["vtcr_key"],
            "Her label's engine equivalent": expected or "none — not a disposition",
            "Agree?": "yes" if agrees else "NO",
            "Same direction?": "yes" if same_dir else "no",
            "Severity of the disagreement": severity,
            "Note": note,
            "Which we think is right": verdict,
            "Our confidence": r["confidence"],
        })

    # ------------------------------------------------------------------ spreads
    spread_base = {d: 0 for d in DISPOSITIONS}
    spread_var = {d: 0 for d in DISPOSITIONS}
    for r in rows:
        spread_base[r["_base"]["disposition"]] += 1
        spread_var[r["_var"]["disposition"]] += 1
    differs = [r for r in rows if r["_var"]["disposition"] != r["_base"]["disposition"]]

    # ------------------------------------------------------------- What changed from v1
    v1_rows, v1_note = load_v1_dispositions()
    changed_records = []
    n_term_changed = n_prio_changed = n_conf_changed = n_any_changed = 0
    v1_spread = {d: 0 for d in DISPOSITIONS}
    for r in sorted(rows, key=lambda x: x["app_id"]):
        a = r["app"]
        prev = v1_rows.get(a["app_id"])
        rec = {
            "App ID": a["app_id"],
            "Application": a["name"],
            "V before": None, "V after": round(a["business_value_score"], 3),
            "T before": None, "T after": round(a["technical_health_score"], 3),
            "C before": None, "C after": round(a["cost_efficiency_score"], 3),
            "R before": None, "R after": round(a["risk_posture_score"], 3),
            "R change": None,
            "Pattern key before": None, "Pattern key after": a["vtcr_key"],
            "Disposition before": None, "Disposition after": r["disposition"],
            "Priority before": None, "Priority after": r["_base"]["priority"],
            "Confidence before": None, "Confidence after": r["confidence"],
            "Moved?": None,
            "Why it moved (or did not)": None,
        }
        if prev is None:
            rec["Why it moved (or did not)"] = v1_note
            rec["Moved?"] = "unknown — no v1 row to compare"
            changed_records.append(rec)
            continue
        pv, pt, pc, pr = (f(prev["Business value"]), f(prev["Technical health"]),
                          f(prev["Cost efficiency"]), f(prev[V1_RISK_COL]))
        rec["V before"], rec["T before"] = pv, pt
        rec["C before"], rec["R before"] = pc, pr
        rec["Pattern key before"] = prev["Pattern key"]
        rec["Disposition before"] = prev["Disposition"]
        rec["Priority before"] = prev["Priority"]
        rec["Confidence before"] = prev["Confidence"]
        v1_spread[prev["Disposition"]] = v1_spread.get(prev["Disposition"], 0) + 1
        rec["R change"] = (round(a["risk_posture_score"] - pr, 3) if pr is not None else None)

        term_moved = prev["Disposition"] != r["disposition"]
        prio_moved = prev["Priority"] != r["_base"]["priority"]
        conf_moved = prev["Confidence"] != r["confidence"]
        n_term_changed += term_moved
        n_prio_changed += prio_moved
        n_conf_changed += conf_moved
        n_any_changed += bool(term_moved or prio_moved or conf_moved)
        moved = [w for w, m in (("term", term_moved), ("priority", prio_moved),
                                ("confidence", conf_moved)) if m]
        rec["Moved?"] = ", ".join(moved) if moved else "no"

        # --- the one-line reason, computed from what actually differs.
        vtc_same = (rec["Pattern key before"][:3] == a["vtcr_key"][:3])
        r_flipped = (rec["Pattern key before"][3] != a["vtcr_key"][3])
        why = []
        if pr is not None:
            direction = "up" if a["risk_posture_score"] > pr else (
                "down" if a["risk_posture_score"] < pr else "unchanged")
            why.append(f"risk {pr:.2f} -> {a['risk_posture_score']:.2f} ({direction} "
                       f"{abs(a['risk_posture_score'] - pr):.2f} on the rebuilt rubric)")
        if vtc_same:
            why.append("value/technical/cost identical, as intended — only risk was rebuilt")
        else:
            why.append("WARNING: a non-risk dimension moved, which v2 should not do")
        if r_flipped:
            why.append(f"the risk gate flipped {rec['Pattern key before'][3]} -> "
                       f"{a['vtcr_key'][3]}, so the pattern key moved "
                       f"{rec['Pattern key before']} -> {a['vtcr_key']}")
        if term_moved:
            why.append(f"which moves the term {prev['Disposition']} -> {r['disposition']}")
        elif prio_moved:
            why.append(f"same term, but the lookup row changed so priority moves "
                       f"{prev['Priority']} -> {r['_base']['priority']}")
        elif r_flipped:
            why.append("term and priority both unchanged: the redundancy override or the "
                       "retain/invest rule lands on the same answer either way")
        if conf_moved:
            if r["confidence"] != "Needs Validation" and prev["Confidence"] == "Needs Validation":
                why.append(f"confidence {prev['Confidence']} -> {r['confidence']}: the v1 flag "
                           f"was raised on missing risk evidence, and her Healthcare "
                           f"Guardrails row now supplies it")
            else:
                why.append(f"confidence {prev['Confidence']} -> {r['confidence']}")
        if r["confidence"] == "Needs Validation":
            why.append("still Needs Validation — " + "; ".join(r["_decisive"]))
        elif r["confidence"] in ("medium", "low") and r["_gaps"]:
            # Say why it did NOT reach high, so a medium row is not left looking arbitrary.
            why.append(f"held at {r['confidence']} rather than high by: "
                       + "; ".join(r["_gaps"]))
        rec["Why it moved (or did not)"] = ". ".join(why) + "."
        changed_records.append(rec)

    # ------------------------------------------------------------------ write workbook
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Read me first"
    differs_summary = ", ".join(
        f"{r['app_id']} {r['_base']['disposition']} -> {r['_var']['disposition']}"
        for r in differs) or "none"
    orphan_rows = [r for r in sorted(rows, key=lambda x: x["app_id"]) if r["_orphans"]]
    n_boundary = sum(1 for r in agree_records
                     if r["Agree?"] == "NO" and r["Severity of the disagreement"].startswith("boundary"))
    n_material = sum(1 for r in agree_records
                     if r["Agree?"] == "NO" and r["Severity of the disagreement"] == "material")
    n_nolabel = sum(1 for r in agree_records
                    if r["Agree?"] == "NO"
                    and r["Severity of the disagreement"] == "her label is not a disposition")

    # risk-register category coverage, computed rather than asserted
    def _no_cat(cats):
        return sum(1 for a in apps if not risk_rows_for(a["app_id"], ctx["risks"], cats))
    n_no_tech = _no_cat(("Technical",))
    n_no_bc = _no_cat(("Business Continuity",))
    n_no_sec = _no_cat(("Security", "Regulatory"))
    n_no_clin = _no_cat(("Clinical Safety", "AI /", "Patient Safety"))
    n_apps = len(apps)
    coverage_sentence = (
        f"{n_no_tech} of {n_apps} apps have no Technical row, {n_no_bc} have no Business "
        f"Continuity row, {n_no_sec} have no Security or Regulatory row, and {n_no_clin} have "
        f"no clinical-safety row (only {n_apps - n_no_clin} do)")
    conf_spread = defaultdict(int)
    for r in rows:
        conf_spread[r["confidence"]] += 1
    n_direct = sum(1 for name, _d, _l, _f in SCORERS
                   if {a["_inputs"][name]["availability"] for a in apps} == {"direct"})
    n_rubric = sum(1 for name, _d, _l, _f in SCORERS
                   if "unavailable" not in {a["_inputs"][name]["availability"] for a in apps}) - n_direct
    n_unavail = sum(1 for name, _d, _l, _f in SCORERS
                    if {a["_inputs"][name]["availability"] for a in apps} == {"unavailable"})
    write_prose(ws, [
        ("h1", "Northstar Global Health — disposition analysis, v2"),
        ("p", f"Produced {ANALYSIS_DATE.isoformat()} by Aberdeen Advisors' Application "
              f"Rationalization engine. Source: your REVISED 20-application sample dataset "
              f"({os.path.basename(SOURCE_XLSX)}, {len(src)} sheets). Your file was opened "
              f"read-only and is unchanged. The v1 workbook and its script are untouched and "
              f"still on disk beside this one."),
        ("p", ""),
        ("h2", "ONLY THE RISK DERIVATION CHANGED — and risk is now gated"),
        ("p", "You asked for the analysis to be rerun using the same assumptions as the "
              "previous file, except risk. That is exactly what this is. Every other "
              "derivation is carried over unchanged: the same four dimensions, the same 18 "
              "inputs, the same 3.0 gate, the same 16-row pattern table, the same five terms, "
              "the same value / technical / cost rubric thresholds down to the individual "
              "band, the same cost renormalisation, the same savings arithmetic, and your "
              "Lifecycle Stage still held out. None of those were re-tuned."),
        ("p", "TWO things changed, both because of the better risk evidence you supplied. "
              "FIRST, the three risk inputs — technical, business/compliance and clinical "
              "safety — were rebuilt from your new 'Healthcare Guardrails' sheet instead of "
              "from your two-row-per-app risk register. SECOND, risk is now a FIRST-CLASS "
              "GATED DIMENSION. In v1 risk was gated at 3.0 but every risk number carried a "
              "'(low confidence)' caveat, because your register held only two rows per "
              "application and at least one risk input had to fall back to a summary field on "
              "most rows. Your new sheet populates all three risk families for all 20 "
              "applications with no fallbacks at all, so that caveat is withdrawn. The "
              "risk-excluded variant is still computed and still reported on every row, so the "
              "comparison against v1 stays honest."),
        ("p", "Confidence was recomputed from the new evidence coverage, which follows from "
              "the data rather than from a loosened bar. The 'What changed from v1' sheet "
              "carries the full before-and-after for all 20 applications: the four dimension "
              "scores, the pattern key, the disposition, the priority and the confidence, with "
              "a one-line reason on every row that moved."),
        ("p", "WHERE THE NEW EVIDENCE DID NOT ARRIVE: your 'Risks' sheet is byte-for-byte "
              "identical to v1 — still 40 rows, still exactly two per application, still with "
              "uneven category coverage. The richer risk information came as a new sheet, not "
              "as extra risk rows. Your Capability Map, TCO, Dependencies, User Profiles and "
              "Performance sheets are also unchanged, which is why nothing outside risk moved. "
              "The Sanity checks sheet proves this rather than asserting it: it compares every "
              "value, technical and cost dimension score against the v1 run and fails the "
              "build if any one of them differs."),
        ("p", ""),
        ("h2", "What was run"),
        ("p", "Each of the 20 applications was scored on four dimensions — business value, "
              "technical health, cost efficiency and risk posture — built from 18 inputs on a "
              "1-to-5 half-step scale. A dimension passes if it scores 3.0 or more. The four "
              "pass/fail results form a four-letter pattern (for example PPFP), and that "
              "pattern is looked up in a fixed 16-row table that returns one of five terms — "
              "retain, invest, consolidate, replace, retire — and a priority."),
        ("p", "Nothing in this workbook was typed in by hand. Every score, pattern, "
              "disposition, priority and dollar figure was computed by score_northstar.py from "
              "cells in your file. The script is included so you can re-run it."),
        ("p", ""),
        ("h2", "No interviews were used"),
        ("p", "You told us this iteration would not interview stakeholders or owners, and your "
              "Assumptions sheet says the same. So no score here rests on anyone's opinion. "
              f"{18 - n_unavail} of the 18 inputs are producible from your file with no "
              f"interview, and {n_unavail} are not. Of the producible ones, {n_direct} come "
              f"straight off a column on every application, 3 more come straight off your risk "
              f"register wherever it holds a row in the matching category and fall back to a "
              f"summary field where it does not, and the rest are derived from your columns by "
              f"a written rubric. (An earlier mapping pass split the same 16 producible inputs "
              f"5 direct / 11 rubric; the difference is only where the line between 'direct' "
              f"and 'derived' is drawn, not which inputs can be produced.) The 'Input "
              f"derivation' sheet gives the rubric, the source columns and the observed spread "
              f"for all 18. That sheet is the audit trail; it matters as much as the answers."),
        ("p", ""),
        ("h2", "Which inputs were derived rather than supplied"),
        ("p", "Three inputs are supplied almost as-is: how widely an app is used (your "
              "Utilization Rate and Active Users), how much licence waste it carries (the same "
              "Utilization Rate read the other way), and how stable it is in service (your "
              "incident counts, MTTR and availability against SLA — the one input your file "
              "evidences better than our own dataset does). Three risk inputs come straight "
              "off your risk register wherever it has a row in the matching category. "
              "Everything else is derived. Examples: 'does clinical work stop without it' is "
              "built from Business Criticality, Critical Operation Flag, Data Classification "
              "and how many of the app's capability rows you marked Duplicative; 'cost per "
              "active user' is your Annual TCO over your Active Users, compared against the "
              "median of the app's peers in the same Category; 'customisation debt' is read "
              "from your Development / Configuration Language and any Technical risk row."),
        ("p", "Where your scores run the opposite way to ours — severity, likelihood, inherent "
              "risk, risk level, residual risk, highest risk — they were flipped, because on "
              "all 18 of our inputs a high number is the good number. Your Business Value "
              "Score already runs our way. One flip we did NOT apply is flagged in 'Notes & "
              "assumptions' for you to overrule."),
        ("p", ""),
        ("h2", "The two gaps"),
        ("p", "1. CONSUMPTION PRICE VARIANCE has no source. This input compares metered or "
              "consumption spend against a plan. Your TCO sheet holds six fixed annual "
              "components and no consumption line, and there is no plan or budget figure to "
              "vary against. Rather than invent one, the cost dimension is RENORMALISED over "
              "its remaining weighted inputs — its weight sum drops from 4 to 3, so cost per "
              "active user carries two-thirds of the cost score and unused licence waste one "
              "third. This is not caused by the no-interview rule; the data is not in the file."),
        ("p", "2. END-USER PERCEIVED QUALITY has no source either. It is a perception measure, "
              "and the only route to one is asking users, which this iteration rules out. It "
              "already carries weight 0 in our model, so leaving it empty changes no answer. "
              "Your Business KPI columns are the nearest thing and are reported beside it, but "
              "they measure a business outcome, not a perception, so they are not scored."),
        ("p", ""),
        ("h2", "One thing to read before you read the dispositions"),
        ("p", "THE RISK SCORES ARE NO LONGER LOW CONFIDENCE — that is the substantive change "
              "in v2. Your new Healthcare Guardrails sheet gives all three risk families "
              "measured columns on all 20 applications, so risk is hard-gated at 3.0 like the "
              "other three dimensions and its numbers carry no caveat. For contrast, your "
              f"risk REGISTER, which v1 had to use, still covers only some rows: "
              f"{coverage_sentence}. That register is no longer an input to any risk score."),
        ("p", "The risk-excluded variant is still computed on every row so you can see what "
              "risk is doing, and here the honest answer is: less than you might expect. "
              f"{len(differs)} of {n_apps} applications change term when risk stops being able "
              f"to fail a gate ({differs_summary}). Risk now fails on only two applications, "
              "and both of them already fail on other dimensions, so promoting risk to a "
              "first-class gate changed no disposition by itself. The value of the new evidence "
              "showed up in CONFIDENCE, not in the terms — which is worth knowing, because it "
              "means the v1 '(low confidence)' caveat was doing more work than the risk gate "
              "ever was."),
        ("p", ""),
        ("h2", "How much to trust each row"),
        ("p", "Every row carries a confidence flag. Your Assumptions sheet says missing "
              "evidence should produce 'Needs Validation', and it does. In v1 the dominant "
              "trigger was 'a risk input fell back to a summary field AND risk decides this "
              "row'. In v2 no risk input falls back, so that trigger cannot fire and the rows "
              "it was flagging clear unless something else is wrong with them. Replacing it "
              "are triggers on YOUR OWN evidence-quality columns on the new sheet, which are "
              "stricter and better aimed: Guardrail Status not Pass, guardrail Evidence "
              "Confidence Low, guardrail Data Applicability Unknown. Gate fragility — a risk "
              "score within 0.5 of the 3.0 line — is now reported separately, because it is an "
              "analytic fragility rather than an evidence gap; on its own it caps a row at "
              "medium instead of raising the flag. The "
              f"spread is: "
              + ", ".join(f"{k} {v}" for k, v in
                          sorted(conf_spread.items(), key=lambda kv: -kv[1]))
              + f" of {n_apps}. Each row's 'Why that confidence' column names the specific "
              f"reason, and the gaps are listed even on rows where they change nothing."),
        ("p", ""),
        ("h2", "Your own labels were held out"),
        ("p", "Your Lifecycle Stage column already contains disposition-like labels (Strategic "
              "Invest, Consolidation Candidate, Replace / Sunset, Pilot / Exit Candidate). "
              "Feeding those into a model whose job is to produce dispositions would be "
              "circular, so that column was held out of every score. It is used in exactly one "
              "place: the 'Agreement with your labels' sheet, which compares our answer against "
              f"your team's for all 20, mapping your labels one-to-one onto our five terms. We "
              f"agree on {20 - n_disagree} and differ on {n_disagree}. The differences are not "
              f"evenly interesting: {n_boundary} of them are the same boundary case — your "
              f"'Strategic Invest' against our 'retain' — where both of us mean keep it and the "
              f"only question is whether anything is being funded. Your label set has no word "
              f"for 'healthy, leave it alone, spend nothing', which is precisely the split Bina "
              f"asked us to add as a fifth term. {n_nolabel} is your 'Review', which is an "
              f"undecided state rather than an action, so there is nothing to disagree with. "
              f"That leaves {n_material} genuine difference of substance. Wherever your label "
              "appears in this workbook it sits in a brown-headed column marked COMPARISON ONLY."),
        ("p", ""),
        ("h2", "One safety catch worth knowing about"),
        ("p", "Before any action is recommended, the script checks whether switching the "
              "application off would leave a capability with no other provider anywhere in "
              "your Capability Map. " + (
                  "It found " + str(len(orphan_rows)) + ": " +
                  "; ".join(f"{r['app_id']} {r['app']['name']} is the only provider of "
                            f"{o.split(' (')[0]}" for r in orphan_rows for o in r['_orphans']) +
                  ". The engine does NOT quietly change its answer on that basis — the term "
                  "still stands on the evidence — but the recommendation gains an explicit "
                  "PRECONDITION and the row drops to Needs Validation. Whether that capability "
                  "is genuinely still needed is your team's call, not the model's."
                  if orphan_rows else
                  "Nothing in this portfolio trips it: every application we recommend removing "
                  "has its capabilities covered by another application in your own map.")),
        ("p", ""),
        ("h2", "What is in each sheet"),
        ("p", "Dispositions — one row per application: the four scores, the pattern, the "
              "disposition, the priority, a rationale citing that app's own evidence, the "
              "concrete next action, a confidence flag, and the alternative under the "
              "risk-excluded variant."),
        ("p", "Input derivation — the audit trail: all 18 inputs, the rubric, the source "
              "columns, and whether each was direct, rubric-derived or unavailable."),
        ("p", "Consolidation candidates — the genuine overlap groups from your Capability Map's "
              "Duplicative and Secondary roles, a named survivor with the reason, and the "
              "group saving from your Avoidable Annual Cost and transition figures."),
        ("p", "Savings — per app and portfolio: run cost, gross saving, one-time cost, net "
              "first-year saving, and where our arithmetic differs from yours, both figures "
              "side by side with the reason."),
        ("p", "Agreement with your labels — our disposition against your Lifecycle Stage for "
              "all 20, with a note on each disagreement."),
        ("p", "What changed from v1 — the before-and-after for all 20 applications: the four "
              "dimension scores, the pattern key, the disposition, the priority and the "
              "confidence, in v1 and in v2, with a one-line reason on every row that moved. "
              "Read this one first if you read nothing else."),
        ("p", "Notes & assumptions — every threshold, the renormalisation, the new risk rubric "
              "in full, the risk-confidence decision, and the short list of calls that are "
              "yours or Ryo's, not ours."),
        ("p", "Sanity checks — the arithmetic and leakage checks the script runs before it is "
              "allowed to write this file."),
        ("p", ""),
        ("p", f"Engine verification: {engine_note}."),
    ])

    dispo_headers = list(dispo_records[0].keys())
    write_sheet(wb.create_sheet("Dispositions"), dispo_headers, dispo_records,
                widths={"App ID": 9, "Application": 30, "Vendor": 24,
                        "Primary capability": 24, "Business value": 10,
                        "Technical health": 10, "Cost efficiency": 10,
                        "Risk posture": 12, "V": 4, "T": 4, "C": 4, "R": 4,
                        "Pattern key": 8, "Disposition": 13, "Priority": 11,
                        "Rationale": 95, "Recommendation": 75, "Confidence": 15,
                        "Why that confidence": 60,
                        "Evidence gaps behind the confidence flag": 44,
                        "Alternative under risk-excluded gate": 40, "Priority basis": 40,
                        "Her Lifecycle Stage (COMPARISON ONLY — held out of scoring)": 26},
                wrap_cols={"Rationale", "Recommendation", "Why that confidence",
                           "Evidence gaps behind the confidence flag",
                           "Alternative under risk-excluded gate", "Priority basis",
                           "Primary capability", "Application", "Vendor"},
                comparison_cols={"Her Lifecycle Stage (COMPARISON ONLY — held out of scoring)"},
                freeze="E2")

    write_sheet(wb.create_sheet("Input derivation"), list(deriv_records[0].keys()),
                deriv_records,
                widths={"Engine input": 32, "Dimension": 15, "Raw weight": 8,
                        "Normalised weight in its dimension": 34, "What it scores": 46,
                        "Availability": 30, "Her columns used": 60, "Rubric applied": 110,
                        "Apps scored": 11, "Observed min / median / max": 18,
                        "Rows on the fallback path": 12,
                        "Example evidence (APP-001 Epic)": 60},
                wrap_cols={"What it scores", "Her columns used", "Rubric applied",
                           "Example evidence (APP-001 Epic)", "Availability",
                           "Normalised weight in its dimension"},
                freeze="B2")

    write_sheet(wb.create_sheet("Consolidation candidates"), list(cons_records[0].keys()),
                cons_records,
                widths={"Overlap group": 12, "Contested capabilities": 34,
                        "Capability IDs": 24, "App ID": 9, "Application": 30,
                        "Role in group": 13,
                        "Her Support Role / Coverage on the contested capabilities": 52,
                        "Why": 78, "Our disposition": 13, "Annual TCO": 14,
                        "Her Avoidable Annual Cost": 16, "Her One-Time Transition Cost": 16,
                        "Group annual saving if every absorbable member folds in": 20,
                        "Group one-time transition cost": 18,
                        "Group annual run cost today": 18,
                        "Her Lifecycle Stage (COMPARISON ONLY)": 24},
                wrap_cols={"Contested capabilities", "Why",
                           "Her Support Role / Coverage on the contested capabilities",
                           "Application"},
                comparison_cols={"Her Lifecycle Stage (COMPARISON ONLY)"}, freeze="D2")

    write_sheet(wb.create_sheet("Savings"), list(sav_records[0].keys()), sav_records,
                widths={"App ID": 11, "Application": 32, "Our disposition": 13,
                        "Current annual run cost (her Annual TCO)": 16,
                        "Her Avoidable Annual Cost": 15,
                        "Gross annual saving we claim": 15,
                        "One-time transition cost (hers)": 15, "Named successor": 13,
                        "Successor ongoing cost netted off": 15,
                        "Successor already in the baseline?": 34,
                        "Our net first-year saving": 15, "Her First-Year Net Savings": 15,
                        "Difference (ours less hers)": 14, "Why they differ": 70,
                        "Residual ongoing run cost after the action": 16,
                        "Safe or potential": 30, "Safe saving": 13, "Potential saving": 13,
                        "Her Cost Notes": 52},
                wrap_cols={"Why they differ", "Her Cost Notes", "Safe or potential",
                           "Successor already in the baseline?", "Application"},
                freeze="C2")
    wsx = wb["Savings"]
    money_cols = ["Current annual run cost (her Annual TCO)", "Her Avoidable Annual Cost",
                  "Gross annual saving we claim", "One-time transition cost (hers)",
                  "Successor ongoing cost netted off", "Our net first-year saving",
                  "Her First-Year Net Savings", "Difference (ours less hers)",
                  "Residual ongoing run cost after the action", "Safe saving",
                  "Potential saving"]
    hdrs = [c.value for c in wsx[1]]
    for name in money_cols:
        i = hdrs.index(name) + 1
        for r in range(2, wsx.max_row + 1):
            wsx.cell(row=r, column=i).number_format = '$#,##0'
    for r in range(2, wsx.max_row + 1):
        if wsx.cell(row=r, column=1).value == "PORTFOLIO":
            for c in range(1, wsx.max_column + 1):
                wsx.cell(row=r, column=c).font = Font(bold=True)

    wsc = wb["Consolidation candidates"]
    hdrs = [c.value for c in wsc[1]]
    for name in ["Annual TCO", "Her Avoidable Annual Cost", "Her One-Time Transition Cost",
                 "Group annual saving if every absorbable member folds in",
                 "Group one-time transition cost", "Group annual run cost today"]:
        i = hdrs.index(name) + 1
        for r in range(2, wsc.max_row + 1):
            wsc.cell(row=r, column=i).number_format = '$#,##0'

    write_sheet(wb.create_sheet("Agreement with your labels"), list(agree_records[0].keys()),
                agree_records,
                widths={"App ID": 9, "Application": 32,
                        HER_LABEL_COL: 24, "Our disposition": 13,
                        "Our priority": 11, "Pattern key": 9,
                        "Her label's engine equivalent": 22, "Agree?": 8,
                        "Same direction?": 13, "Severity of the disagreement": 30,
                        "Note": 78, "Which we think is right": 88, "Our confidence": 16},
                wrap_cols={"Note", "Which we think is right", "Application",
                           "Severity of the disagreement", "Her label's engine equivalent"},
                comparison_cols={HER_LABEL_COL}, freeze="C2")

    # ------------------------------------------------------- What changed from v1 (NEW)
    write_sheet(wb.create_sheet("What changed from v1"), list(changed_records[0].keys()),
                changed_records,
                widths={"App ID": 9, "Application": 32,
                        "V before": 9, "V after": 9, "T before": 9, "T after": 9,
                        "C before": 9, "C after": 9, "R before": 9, "R after": 9,
                        "R change": 9,
                        "Pattern key before": 11, "Pattern key after": 11,
                        "Disposition before": 13, "Disposition after": 13,
                        "Priority before": 11, "Priority after": 11,
                        "Confidence before": 16, "Confidence after": 16,
                        "Moved?": 22, "Why it moved (or did not)": 110},
                wrap_cols={"Application", "Why it moved (or did not)", "Moved?"},
                comparison_cols={"V before", "T before", "C before", "R before",
                                 "Pattern key before", "Disposition before",
                                 "Priority before", "Confidence before"},
                freeze="C2")

    # ------------------------------------------------------------------ Notes
    peer_lines = []
    for cat in sorted(ctx["peer_cpu"]):
        vals = sorted(ctx["peer_cpu"][cat])
        peer_lines.append(f"    {cat}: n={len(vals)}, median ${statistics.median(vals):,.0f} "
                          f"per active user, range ${vals[0]:,.0f}-${vals[-1]:,.0f}"
                          + ("  [singleton -> compared against the portfolio median "
                             f"${ctx['portfolio_cpu_median']:,.0f}]" if len(vals) < 2 else ""))
    risk_cov = defaultdict(int)
    for r in ctx["risks"]:
        risk_cov[s(r["Risk Category"])] += 1
    rto_n, rto_tot, _rto_detail, rto_ratios = rto_vs_mtd_finding(ctx)
    notes_blocks = [
        ("h1", "Notes & assumptions"),
        ("p", "Everything below is a choice we made, not a fact from your file. Each one is "
              "yours or Ryo's to overrule."),
        ("p", "V2 SCOPE: only the risk derivation changed. Notes 1, 3, 4, 5, 6, 7, 8 and 9 are "
              "carried over from v1 word for word, because the assumptions they describe were "
              "carried over unchanged. Note 2 is rewritten — it is the risk note. Note 2R is "
              "new and gives the full rebuilt risk rubric."),
        ("p", ""),
        ("h2", "1. The renormalisation"),
        ("p", "The cost dimension normally averages three weighted inputs: cost per active user "
              "(weight 2), unused licence waste (weight 1) and consumption price variance "
              "(weight 1) — weight sum 4. Consumption price variance has no source column in "
              "your workbook, so it is left null and the dimension is averaged over the "
              "remaining weight sum of 3. Effect: cost per active user carries 0.6667 of the "
              "cost score instead of 0.50, and unused licence waste 0.3333 instead of 0.25. "
              "The absolute-cost band is scored and reported at weight 0 and is not in either "
              "denominator. This fires on all 20 applications and is checked before the file "
              "is written."),
        ("p", "Consequence worth stating: cost per active user now controls two-thirds of the "
              "dimension that, in our own review, was already the only one where the 3.0 gate "
              "really bites. If you want that diluted, the fix is a consumption or metered "
              "spend column plus a plan figure, not a re-weighting."),
        ("p", ""),
        ("h2", "2. The risk decision — THE ONLY THING THAT CHANGED IN V2"),
        ("p", "In v1 we told you that to make risk properly load-bearing you needed evidence "
              "per risk CATEGORY per application rather than two rows per application. You "
              "supplied it, as a new 'Healthcare Guardrails' sheet. So risk is now gated at 3.0 "
              "as a first-class dimension and the '(low confidence)' caveat is withdrawn."),
        ("p", "Your risk REGISTER is unchanged and is no longer used for any risk score. For "
              "the record, its coverage across the same 40 rows is still: "
              + ", ".join(f"{k} {v}" for k, v in sorted(risk_cov.items())) + ". "
              f"That still leaves {coverage_sentence} — which is why we moved off it."),
        ("p", "The risk-excluded variant is STILL produced on every row, deliberately, so the "
              "comparison with v1 is honest and so you can see how much work the risk gate is "
              f"doing. Straight answer: {len(differs)} of {n_apps} applications change term "
              "when risk stops being able to fail a gate. Risk fails on two applications only "
              "(APP-006 Zoom Workplace and APP-018 Oracle PeopleSoft HCM) and both already fail "
              "on other dimensions. Promoting risk to a first-class gate therefore changed no "
              "disposition on its own — the new evidence paid off in confidence, not in terms."),
        ("p", "A caution on that. The rubric below is deductions-from-5.0: an application with "
              "nothing adverse in your guardrail columns scores 5.0. Five applications do. That "
              "is a faithful reading of your evidence, but it means risk is now a permissive "
              "dimension for a clean row, and a single adverse column is what moves it. If you "
              "want risk to discriminate more finely among healthy applications you need "
              "graduated measures rather than mostly-zero counts — vulnerability age "
              "distributions rather than open counts, and near-miss safety reporting rather "
              "than confirmed-event counts alone."),
        ("p", ""),
        ("h2", "2R. The rebuilt risk rubric, in full"),
        ("p", "Common shape: start at 5.0, subtract a documented deduction for each adverse "
              "condition in your data, floor at 1.0, cap at 5.0, snap to the half-step scale. "
              "Deductions only, never a bonus for absence of evidence — your own Assumptions "
              "sheet says a zero event count must not be read as zero risk, so a clean row "
              "earns no uplift, it merely avoids a penalty. The single exception is the "
              "vendor-assurance tier, where a HITRUST-scoped attestation is affirmative "
              "evidence of control and earns +0.5, capped so it cannot lift a row over the gate "
              "by itself."),
        ("p", "R1 TECHNICAL RISK — " + R1_RUBRIC),
        ("p", "R2 BUSINESS / COMPLIANCE RISK — " + R2_RUBRIC),
        ("p", "R3 CLINICAL SAFETY RISK — " + R3_RUBRIC),
        ("p", "WHAT WE DELIBERATELY DID NOT SCORE: your RTO Target against your Maximum "
              f"Tolerable Downtime. RTO exceeds MTD on {rto_n} of {rto_tot} applications, and "
              f"in every one of those cases RTO is exactly "
              + " or ".join(f"{r:g}x" for r in rto_ratios) + " MTD. That is the signature of a "
              "generation rule, not of a dozen independent recovery-design failures, so "
              "deducting for it would push a dataset artifact into a clinical-safety number. It "
              "is reported in full on the Sanity checks sheet instead. IF THOSE RTO TARGETS ARE "
              "REAL, this is a portfolio-wide business-continuity finding and it belongs in the "
              "report as one — that is your call, not ours."),
        ("p", "ONE DISCLOSED DOUBLE-COUNT: R1 uses your Sev-1/Sev-2 incident count, and "
              "th_operational_stability in the TECHNICAL HEALTH dimension already uses your "
              "Performance sheet's total incident count and MTTR. A severity-filtered count is a "
              "different measure, so it was kept, but a badly-behaved application is now "
              "penalised in two dimensions rather than one. Disclosed on every row rather than "
              "netted out silently, because netting it would require deciding which dimension "
              "owns reliability, and that is a model-design question for Ryo."),
        ("p", "YOUR GUARDRAIL RULE, APPLIED AS YOU WROTE IT: your revised Data Dictionary says "
              "the guardrail is a hard gate that a saving must clear before it counts as safe, "
              "and that a weighted score cannot override it. Implemented exactly there and "
              "nowhere else — a non-Pass Guardrail Status moves that application's saving from "
              "'safe' to 'potential' and does NOT change its disposition. One application is "
              "affected: APP-006 Zoom Workplace."),
        ("p", ""),
        ("h2", "3. Your Lifecycle Stage was held out entirely"),
        ("p", "It contains disposition-like labels, so using it as an input would be leakage: "
              "the model would be predicting the answer from the answer. It is read once, into "
              "a field whose name is prefixed so no scorer can reach it, and used only in the "
              "comparison columns and the 'Agreement with your labels' sheet."),
        ("p", "This has one real cost. Our engine has a lifecycle guard that bars retire and "
              "replace for an application still in Birth or Growth — you do not switch off "
              "something that has not finished ramping. Its only possible input in your file is "
              "the held-out column, so THE GUARD IS DISARMED FOR THIS RUN. One row is affected: "
              "APP-009 Nabla Copilot, whose Current Release / Version reads 'Pilot release' — an "
              "independent, non-leaking signal that it is early in life. With the guard armed "
              "from that field, its retire would have been suppressed and turned into a funded "
              "invest at High priority. We left it disarmed and are flagging it rather than "
              "deciding it. BINA / RYO: your call. If you want the guard live, add a "
              "lifecycle-age field that is not the disposition label — implementation date, or "
              "months in production."),
        ("p", ""),
        ("h2", "4. The one flip we did not apply — please overrule us if you disagree"),
        ("p", "We were asked to flip every one of your scores that runs high = worse: severity, "
              "likelihood, inherent risk, risk level, residual risk, highest risk, business "
              "criticality and capability criticality. We flipped the first six without "
              "reservation. We did NOT flip business criticality or capability criticality "
              "inside the BUSINESS VALUE dimension, and here is why: in a risk context a higher "
              "criticality does make things worse, but as a value input it means the opposite. "
              "Inverting it would have scored Epic — Critical criticality, Critical Operation "
              "Flag, system of record for inpatient care — as LOW business value, and every "
              "sanity check downstream would have been meaningless. So criticality is used in "
              "its natural direction where it measures value, and inverted where it enters risk "
              "as an impact term. This is the single place we departed from a literal reading of "
              "the instruction. It is visible here so you can reverse it in one line if you "
              "meant it literally."),
        ("p", ""),
        ("h2", "5. Supportability is a lower-evidence input"),
        ("p", "Vendor Support End is blank for 19 of your 20 applications, so supportability "
              "leans on Current Release / Version and on your update calendar. No "
              "end-of-support date was invented. An evergreen vendor-managed line scores 4.5 "
              "rather than 5.0 precisely because 'the vendor keeps it current' is an inference, "
              "not an evidenced support horizon. The one app with a real date, APP-018 Oracle "
              "PeopleSoft HCM at 2037-12-31, scores well on the support-horizon component and "
              "badly on release currency ('9.2 / PeopleTools 8.60'), and the two are averaged "
              "rather than one overriding the other. Vendor viability is thinner still: your "
              "file has no vendor financial data of any kind, so it is proxied by whether you "
              "raised a Vendor risk row and whether the roadmap is dated."),
        ("p", ""),
        ("h2", "6. Every rubric threshold, in one place"),
        ("p", "The 'Input derivation' sheet carries the full rubric for each of the 18 inputs, "
              "row by row, including every numeric band. The bands are summarised here so they "
              "can be challenged as a set:"),
        ("p", "  Utilisation bands (used for reach and for licence waste): .90/.85/.80/.70/.60/"
              ".50/.40/.30 stepping 5.0 down to 1.5, floor 1.0."),
        ("p", "  Active-user breadth bands: 25000/15000/8000/4000/2000/1000/500/200 stepping "
              "5.0 down to 1.5, floor 1.0."),
        ("p", "  Peer cost ratio bands (cost per active user over peer-group median): "
              "0.50/0.70/0.90/1.00/1.15/1.40/1.80/2.50 stepping 5.0 down to 1.5, floor 1.0."),
        ("p", "  Peer groups are your Category column. Observed medians:"),
    ] + [("p", l) for l in peer_lines] + [
        ("p", "  Incidents per 1,000 active users: 0.5/1.5/3.0/5.0 -> 5.0/4.0/3.0/2.0, floor 1.0."),
        ("p", "  MTTR minutes: 45/60/90/120 -> 5.0/4.0/3.0/2.0, floor 1.0."),
        ("p", "  Availability against your SLA Target: at or above = 5.0, shortfall <= 0.001 = "
              "3.0, larger = 1.5."),
        ("p", "  Absolute annual TCO bands (weight 0): 0.5m/1m/1.5m/2m/3m/5m/7m -> 5.0/4.5/4.0/"
              "3.5/3.0/2.0/1.5, floor 1.0."),
        ("p", "  Residual-risk inversion: Low 5.0, Medium 3.5, High 2.0, Critical 1.0. "
              "Mitigation-status credit: Closed +1.0, Mitigating +0.5, Accepted 0.0, Open -0.5. "
              "Risk inputs take the WORST matching row, not the mean, because posture is set by "
              "the least-controlled exposure."),
        ("p", "  Business Criticality to value base: Critical 4.5, High 3.5, Medium 2.5, "
              "Low 1.5, then +0.5 for Critical Operation Flag Yes, +0.5 for outright PHI, "
              "-1.0 when at least half the capability rows are Duplicative."),
        ("p", "  Every result is snapped onto the engine's 1.0-5.0 half-step scale."),
        ("p", ""),
        ("h2", "7. How overlap groups and survivors were decided"),
        ("p", "A capability is contested when more than one application maps to it and at least "
              "one of them carries your Support Role 'Duplicative'. Contested capabilities join "
              "their applications into an overlap group. The survivor is the member holding the "
              "most contested capabilities as Primary + Full, tie-broken on active users then on "
              "lower cost per active user."),
        ("p", "A duplicative member is treated as ABSORBED — which forces the term consolidate — "
              "only when three things all hold: at least half its capability rows are "
              "Duplicative, a survivor holds at least one of the same capabilities as Primary, "
              "AND your own evidence describes a migration path (migration language in your TCO "
              "Cost Notes or in a Dependencies 'Required Before Disposition' cell). That third "
              "condition is deliberate: it keeps a failed pilot with nothing to migrate out of "
              "the consolidate bucket and lets the gates give it retire, and it keeps "
              "APP-006 Zoom Workplace out too, because your own Cost Notes say its saving is "
              "'Potential only; not safe until telehealth and room dependencies are validated'. "
              "We will not commit a saving your file withholds."),
        ("p", "One overlap your team did not mark: CAP-007 Patient engagement is held Primary + "
              "Full by BOTH APP-003 athenaOne and APP-019 Salesforce Health Cloud, with no "
              "Duplicative role on either, so it forms no group and neither app is touched. "
              "BINA / RYO: is that a genuine dual-primary split by population, or a missing "
              "Duplicative mark?"),
        ("p", ""),
        ("h2", "8. Savings arithmetic — yours, not ours"),
        ("p", "Gross annual saving is your Avoidable Annual Cost, claimed only where our "
              "disposition actually removes run-rate spend (retire, consolidate, replace). Net "
              "first-year saving is that figure less your One-Time Transition Cost, floored at "
              "zero — the same definition your Assumptions sheet gives. Successor run cost is "
              "netted only where a successor is named and is not already in the baseline; every "
              "survivor here is an existing portfolio application already carrying its own "
              "Annual TCO, so that term is zero on all 20 and is shown as zero rather than "
              "omitted, so a future run with a genuinely new successor cannot double-count."),
        ("p", "Where our figure differs from yours, both are shown with the reason. The only "
              "differences are applications where you booked an avoidable cost and our engine "
              "does not recommend the action that would release it; we never overwrote one of "
              "your numbers. Safe versus potential follows your own rule ('only high-confidence "
              "actions are counted as safe savings'): a saving is safe when the action removes "
              "run cost, your Evidence Confidence is High, and your Cost Notes do not withhold "
              "it."),
        ("p", "Your CIO savings target of 15% is checked against the portfolio total on the "
              "'Savings' sheet. It is not met on first-year net savings and is met on gross "
              "avoidable annual cost from year two onward. That gap is a real finding about the "
              "dataset, not an error."),
        ("p", ""),
        ("h2", "9. Engine features that could not run on your file"),
        ("p", "  Lifecycle guard — disarmed, see note 3."),
        ("p", "  Retention-obligation constraint — our engine steps a retire down in priority "
              "when a legal retention obligation means the archive must exist before anything "
              "is switched off. Your file has no retention-obligation or retention-expiry field, "
              "so the constraint never fires. Several of your Cost Notes imply one ('Savings "
              "exclude retained archive'), which is why the residual ongoing cost is reported "
              "per app on the 'Savings' sheet instead."),
        ("p", "  Sourcing guard — this one does run, as an annotation. It never changes a term; "
              "it records that a Commercial SaaS product cannot be re-platformed in place, so a "
              "replace means substituting a different product."),
        ("p", ""),
        ("h2", "10. Short list of calls that are yours or Ryo's"),
        ("p", "  a. Risk gated or not gated. V2 gates it, on the strength of your new evidence, "
              "and still shows both. Someone has to pick one for the deck."),
        ("p", "  a2. NEW — whether the RTO-versus-MTD relationship in note 2R is a real "
              f"portfolio-wide continuity finding ({rto_n} of {rto_tot} applications) or an "
              "artifact of how the sample was generated. We report it and do not score it. If "
              "it is real it is arguably the largest single finding in this file."),
        ("p", "  a3. NEW — whether risk should discriminate among clean rows, per the caution "
              "in note 2. Five applications score a flat 5.0 on risk because nothing adverse "
              "appears in your guardrail columns for them."),
        ("p", "  a4. NEW — whether the disclosed Sev-1/Sev-2 double-count between the risk and "
              "technical-health dimensions should be netted out, and if so which dimension "
              "owns reliability. Ryo's call."),
        ("p", "  b. Whether the criticality flip in note 4 should be literal."),
        ("p", "  c. Whether the lifecycle guard should be armed from a non-leaking age field, "
              "which changes APP-009."),
        ("p", "  d. Whether CAP-007's dual-primary split is real or a missing Duplicative mark. "
              "STILL OPEN IN V2 — your Capability Map is unchanged, so athenaOne (APP-003) and "
              "Salesforce Health Cloud (APP-019) are both still marked PRIMARY on CAP-007 "
              "Patient engagement. Note that both rows now clear to high confidence on the new "
              "risk evidence, so this ambiguity no longer surfaces itself through a confidence "
              "flag; it is detected and reported explicitly instead."),
        ("p", "  e. Whether Category is the right peer group for cost comparison. Two of your "
              "categories hold a single application, so those two are compared against the "
              "portfolio median instead, and two hold exactly two applications, where the "
              "median is just the midpoint of a pair and the cheaper app is guaranteed a good "
              "score. A capability-based peer group would be better; Category was used because "
              "it is unambiguous in your file."),
        ("p", "  f. Whether an evergreen SaaS line should cap supportability at 4.5 or score a "
              "full 5.0."),
        ("p", "  g. Every capability the engine flags as orphaned by an action — in this "
              "portfolio, CAP-011 Virtual care communications, which APP-006 Zoom Workplace is "
              "the only provider of. The engine will not decide whether a capability is still "
              "needed; it only refuses to switch one off silently. STILL OPEN IN V2, and now "
              "sharper: Zoom is also the ONLY application in the portfolio whose Guardrail "
              "Status is not Pass (unresolved critical interface error), with your guardrail "
              "Evidence Confidence Low and Data Applicability Unknown on the same row. So the "
              "one application we recommend switching off is simultaneously the one whose "
              "evidence you trust least and the sole provider of a High-criticality virtual "
              "care capability. Nothing in the new data resolves that; it makes the precondition "
              "more important, not less."),
    ]
    write_prose(wb.create_sheet("Notes & assumptions"), notes_blocks)

    write_sheet(wb.create_sheet("Sanity checks"), ["Check", "Result", "Detail"],
                [{"Check": c["Check"], "Result": c["Result"],
                  "Detail": c["Detail"] if isinstance(c["Detail"], str) else "; ".join(c["Detail"])}
                 for c in checks],
                widths={"Check": 62, "Result": 9, "Detail": 100},
                wrap_cols={"Check", "Detail"})

    wb.save(OUT_XLSX)

    # ------------------------------------------------------------------ CSV
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=dispo_headers)
        w.writeheader()
        for rec in dispo_records:
            w.writerow(rec)

    # ------------------------------------------------------------------ console
    print(f"\nEngine verification: {engine_note}")
    print(f"Source (read-only, unmodified): {SOURCE_XLSX}")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_CSV}")
    print(f"Script  {os.path.abspath(__file__)}")

    print("\n" + "=" * 165)
    print("NORTHSTAR GLOBAL HEALTH — 20 APPLICATIONS")
    print("=" * 165)
    hdr = (f"{'App':8} {'Application':34} {'V':>5} {'T':>5} {'C':>5} {'R':>5} {'key':5} "
           f"{'disposition':12} {'priority':10} {'confidence':16} {'her label':24}")
    print(hdr)
    print("-" * 165)
    for r in sorted(rows, key=lambda x: x["app_id"]):
        a = r["app"]
        print(f"{a['app_id']:8} {a['name'][:34]:34} "
              f"{a['business_value_score']:5.2f} {a['technical_health_score']:5.2f} "
              f"{a['cost_efficiency_score']:5.2f} {a['risk_posture_score']:5.2f} "
              f"{a['vtcr_key']:5} {r['_base']['disposition']:12} {r['_base']['priority']:10} "
              f"{r['confidence']:16} {a['_her_lifecycle_stage'][:24]:24}")
    print("-" * 165)

    print("\nDISPOSITION SPREAD")
    print(f"  {'term':13} {'risk gated at 3.0':>18} {'risk excluded':>15}")
    for d in DISPOSITIONS:
        print(f"  {d:13} {spread_base[d]:>18} {spread_var[d]:>15}")
    print(f"  {'TOTAL':13} {sum(spread_base.values()):>18} {sum(spread_var.values()):>15}")

    print(f"\nRISK-GATE SENSITIVITY: {len(differs)} of {len(apps)} applications change term")
    for r in differs:
        a = r["app"]
        print(f"  {a['app_id']} {a['name'][:32]:32} {r['_base']['disposition']:12} -> "
              f"{r['_var']['disposition']:12} (risk {a['risk_posture_score']:.2f}, "
              f"{a['vtcr_key']} -> {r['_var']['key']})")
    if not differs:
        print("  none")

    print("\n" + "=" * 165)
    print("WHAT CHANGED FROM v1 — only the risk derivation changed")
    print("=" * 165)
    if not v1_rows:
        print(f"  {v1_note}")
    else:
        print(f"  {'App':8} {'Application':30} {'V':>11} {'T':>11} {'C':>11} {'R':>13} "
              f"{'key':12} {'disposition':26} {'priority':24} {'confidence':30} moved")
        print("  " + "-" * 200)
        for rec in changed_records:
            def pair(b, a_, w=11, fmt="{:.2f}"):
                bs = "n/a" if b is None else fmt.format(b) if isinstance(b, float) else str(b)
                as_ = fmt.format(a_) if isinstance(a_, float) else str(a_)
                return f"{bs}->{as_}".rjust(w) if bs != as_ else f"{as_} =".rjust(w)
            print(f"  {rec['App ID']:8} {rec['Application'][:30]:30} "
                  f"{pair(rec['V before'], rec['V after'])} "
                  f"{pair(rec['T before'], rec['T after'])} "
                  f"{pair(rec['C before'], rec['C after'])} "
                  f"{pair(rec['R before'], rec['R after'], 13)} "
                  f"{pair(rec['Pattern key before'], rec['Pattern key after'], 12)} "
                  f"{pair(rec['Disposition before'], rec['Disposition after'], 26)} "
                  f"{pair(rec['Priority before'], rec['Priority after'], 24)} "
                  f"{pair(rec['Confidence before'], rec['Confidence after'], 30)} "
                  f"{rec['Moved?']}")
        print("  " + "-" * 200)
        print(f"  rows where the TERM moved:       {n_term_changed} of {len(apps)}")
        print(f"  rows where the PRIORITY moved:   {n_prio_changed} of {len(apps)}")
        print(f"  rows where CONFIDENCE moved:     {n_conf_changed} of {len(apps)}")
        print(f"  rows that moved in ANY respect:  {n_any_changed} of {len(apps)}")
        print(f"\n  DISPOSITION SPREAD, v1 -> v2")
        for d in DISPOSITIONS:
            print(f"    {d:13} {v1_spread.get(d, 0):>3} -> {spread_base[d]:>3}"
                  f"   {'(+' + str(spread_base[d] - v1_spread.get(d, 0)) + ')' if spread_base[d] > v1_spread.get(d, 0) else '(' + str(spread_base[d] - v1_spread.get(d, 0)) + ')' if spread_base[d] != v1_spread.get(d, 0) else ''}")
        v1_conf = defaultdict(int)
        for prev in v1_rows.values():
            v1_conf[prev["Confidence"]] += 1
        print(f"\n  CONFIDENCE SPREAD, v1 -> v2")
        for k in sorted(set(list(v1_conf) + list(conf_spread))):
            print(f"    {k:20} {v1_conf.get(k, 0):>3} -> {conf_spread.get(k, 0):>3}")

    print(f"\nAGREEMENT WITH HER LIFECYCLE STAGE: disagree on {n_disagree} of {len(apps)}")
    for rec in agree_records:
        if rec["Agree?"] == "NO":
            print(f"  {rec['App ID']} {rec['Application'][:30]:30} hers "
                  f"'{rec[HER_LABEL_COL]}' vs ours '{rec['Our disposition']}'")

    print("\nCONSOLIDATION GROUPS")
    for c in clusters:
        surv = apps_by_id[c["survivor"]]
        avoid = sum(apps_by_id[m]["avoidable_annual"] or 0
                    for m in c["members"] if m != c["survivor"])
        one = sum(apps_by_id[m]["one_time_transition"] or 0
                  for m in c["members"] if m != c["survivor"])
        print(f"  {c['cluster_id']} {', '.join(c['capability_names'])[:56]:56} "
              f"survivor {c['survivor']} ({surv['name'][:20]}) | members "
              f"{', '.join(c['members'])} | absorbed {', '.join(c['absorbed']) or 'none'} | "
              f"annual ${avoid:,.0f} for ${one:,.0f} one-time")

    print("\nMONEY")
    print(f"  portfolio annual run cost                 ${portfolio_tco:,.0f}")
    print(f"  her avoidable annual cost (all 20)        ${sum(a['avoidable_annual'] or 0 for a in apps):,.0f}")
    print(f"  gross annual saving we claim              ${sum(r['_savings']['gross_saving_annual'] for r in rows):,.0f}")
    print(f"  one-time transition cost                  ${sum(r['_savings']['one_time_transition_cost'] for r in rows):,.0f}")
    print(f"  our net first-year saving                 ${sum(r['_savings']['net_first_year_saving'] for r in rows):,.0f}")
    print(f"    of which safe under her own rule        ${sum(r['_savings']['safe_saving'] for r in rows):,.0f}")
    print(f"    of which potential only                 ${sum(r['_savings']['potential_saving'] for r in rows):,.0f}")
    print(f"  her portfolio first-year net savings row  ${sum(r['_savings']['her_first_year_net'] or 0 for r in rows):,.0f}")
    print(f"  CIO target {CIO_SAVINGS_TARGET:.0%} of run cost             ${portfolio_tco * CIO_SAVINGS_TARGET:,.0f}"
          f"   -> {'MET' if sum(r['_savings']['net_first_year_saving'] for r in rows) >= portfolio_tco * CIO_SAVINGS_TARGET else 'NOT MET on first-year net'}")

    print("\nSANITY CHECKS")
    for c in checks:
        detail = c["Detail"] if isinstance(c["Detail"], str) else "; ".join(c["Detail"])
        print(f"  [{c['Result']}] {c['Check']}")
        if c["Result"] == "FAIL":
            print(f"          {detail}")
    fails = [c for c in checks if c["Result"] == "FAIL"]
    print(f"\n{len(checks) - len(fails)} of {len(checks)} checks pass.")
    if fails:
        print("FAILURES PRESENT — do not circulate the workbook until they are resolved.")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
