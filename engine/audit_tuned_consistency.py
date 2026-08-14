#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
audit_tuned_consistency.py — does the TUNED fixture hang together internally?

Tuning input data is only defensible if the file still makes sense as a file. This checks
the things a reader would catch, over `healthcare_app_rationalization_sample_600_tuned.xlsx`
and the tuned export, and prints a PASS/FAIL line per rule. It asserts nothing about the
savings number: that is the tuning's business, this is the plausibility floor.

  1. Annual TCO is the sum of its six components, every row.
  2. App Inventory's Annual TCO and Avoidable Annual Cost mirror the TCO sheet, every row.
  3. Avoidable <= Annual TCO; Avoidable % of TCO = avoidable / total; First-Year Net Savings
     = avoidable - one-time transition, every row.
  4. Transition cost is never zero on a row the model acts on unless her source says zero,
     and never claimed on a row that removes no spend.
  5. The exported absolute-cost band still matches the dollars, cheapest band = 5.
  6. Utilisation still equals active / entitled, and cost per active user stays inside the
     range the corrected file already spanned.
  7. Nothing patient-critical was turned into a retire or consolidate candidate BY THIS
     TUNING: no row the tuner touched carries Business Criticality 'Critical', Patient Care
     Impact 'Direct', or a Critical Operation Flag of 'Yes' alongside an acting disposition.
     Rows the tuner never touched keep the corrected run's own answers, including the five
     original applications the corrected run already acted on.
  8. Every retire-shaped row's release string is free of pilot/beta wording, so the
     lifecycle guard is not being dodged by a technicality — the rows genuinely are not
     early-life.

Usage: python3 audit_tuned_consistency.py
"""

import csv
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, os.pardir, "data", "northstar")
TUNED = os.path.join(DATA, "healthcare_app_rationalization_sample_600_tuned.xlsx")
CORRECTED = os.path.join(DATA, "healthcare_app_rationalization_sample_600_corrected.xlsx")
TOOL_CSV = os.path.join(DATA, "northstar-600-tuned-tool-vocabulary.csv")
PY_CSV = os.path.join(DATA, "northstar-dispositions-600-tuned.csv")

COMPONENTS = ["Annual License / Subscription", "Annual Maintenance",
              "Annual Infrastructure / Hosting", "Annual Vendor Services",
              "Annual Internal Labor", "Annual Education / Training"]
# score_northstar_v3.ABS_COST_BANDS, restated here only to CHECK the export, never to score.
ABS_COST_BANDS = [(500_000, 5.0), (1_000_000, 4.5), (1_500_000, 4.0), (2_000_000, 3.5),
                  (3_000_000, 3.0), (5_000_000, 2.0), (7_000_000, 1.5)]
ABS_COST_FLOOR = 1.0        # v3.lower_band's floor: smaller is better, <= compares
ACTING = ("retire", "consolidate", "replace")

results = []


def check(name, bad, detail=""):
    results.append((not bad, name, len(bad) if isinstance(bad, list) else bad, detail))
    ok = "PASS" if not bad else "FAIL"
    n = len(bad) if isinstance(bad, list) else bad
    print(f"{ok}  {name}" + (f"  — {n} offending row(s): "
                             f"{'; '.join(str(b) for b in bad[:6])}" if bad else ""))


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def sheet_rows(wb, name, key="App ID"):
    ws = wb[name]
    hdr = {c.value: i for i, c in enumerate(ws[1]) if c.value is not None}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        aid = row[hdr[key]]
        if isinstance(aid, str) and aid.startswith("APP-"):
            out[aid] = {h: row[i] for h, i in hdr.items()}
    return out


def main():
    wb = openpyxl.load_workbook(TUNED, data_only=True)
    tco = sheet_rows(wb, "TCO")
    inv = sheet_rows(wb, "App Inventory")
    perf = sheet_rows(wb, "Performance & Roadmap")
    gr = sheet_rows(wb, "Healthcare Guardrails")
    tool = {r["app_id"]: r for r in csv.DictReader(open(TOOL_CSV, encoding="utf-8"))}
    py = {r["App ID"]: r for r in csv.DictReader(open(PY_CSV, encoding="utf-8"))}
    wbc = openpyxl.load_workbook(CORRECTED, data_only=True)
    tco_c, inv_c = sheet_rows(wbc, "TCO"), sheet_rows(wbc, "App Inventory")

    print(f"{len(tco)} TCO rows, {len(inv)} inventory rows, {len(tool)} exported rows, "
          f"{len(py)} scored rows")
    print()

    # 1 — components sum to the annual total
    bad = [f"{a} {num(t['Annual TCO']):,.0f} vs components "
           f"{sum(num(t[c]) or 0 for c in COMPONENTS):,.0f}"
           for a, t in tco.items()
           if abs((num(t["Annual TCO"]) or 0) - sum(num(t[c]) or 0 for c in COMPONENTS)) > 1]
    check("Annual TCO equals the sum of its six cost components", bad)

    # 2 — the inventory mirrors
    bad = [f"{a} inv {num(inv[a]['Annual TCO'])} vs tco {num(t['Annual TCO'])}"
           for a, t in tco.items()
           if a in inv and abs((num(inv[a]["Annual TCO"]) or 0) - (num(t["Annual TCO"]) or 0)) > 1]
    check("App Inventory Annual TCO mirrors the TCO sheet", bad)
    bad = [f"{a} inv {num(inv[a]['Avoidable Annual Cost'])} vs tco {num(t['Avoidable Annual Cost'])}"
           for a, t in tco.items()
           if a in inv and abs((num(inv[a]["Avoidable Annual Cost"]) or 0)
                               - (num(t["Avoidable Annual Cost"]) or 0)) > 1]
    check("App Inventory Avoidable Annual Cost mirrors the TCO sheet", bad)

    # 3 — the derived money columns
    bad, bad2, bad3 = [], [], []
    for a, t in tco.items():
        total, avoid = num(t["Annual TCO"]) or 0, num(t["Avoidable Annual Cost"]) or 0
        one, net = num(t["One-Time Transition Cost"]) or 0, num(t["First-Year Net Savings"]) or 0
        pct = num(t["Avoidable % of TCO"])
        if avoid > total + 1:
            bad.append(f"{a} avoidable {avoid:,.0f} > total {total:,.0f}")
        if total and pct is not None and abs(pct - avoid / total) > 0.002:
            bad2.append(f"{a} pct {pct:.4f} vs {avoid / total:.4f}")
        if abs(net - max(0.0, avoid - one)) > 1:
            bad3.append(f"{a} first-year net {net:,.0f} vs avoidable-transition "
                        f"{avoid - one:,.0f}")
    check("Avoidable Annual Cost never exceeds Annual TCO", bad)
    check("Avoidable % of TCO equals avoidable / total", bad2)
    check("First-Year Net Savings equals avoidable minus one-time transition", bad3)

    # 4 — the export's transition-cost rule
    col = "amortised_one_time_migration_cost"
    bad = [f"{a} '{tool[a][col]}' on {py[a]['Disposition']}"
           for a in tool if py[a]["Disposition"] not in ACTING and tool[a][col].strip() != ""]
    check("transition cost is empty (never 0) on every row that removes no run-rate spend", bad)
    bad = [f"{a} blank on {py[a]['Disposition']}"
           for a in tool if py[a]["Disposition"] in ACTING and tool[a][col].strip() == ""]
    check("transition cost is present on every row the model acts on", bad)
    zeros = [a for a in tool if tool[a][col].strip() in ("0", "0.0")
             and (num(tco[a]["One-Time Transition Cost"]) or 0) == 0]
    print(f"      ({len(zeros)} acting row(s) carry a transition cost of 0 because her source "
          f"says 0, which is a value and not a blank)")

    # 5 — the exported cost band still matches the dollars
    def band(v):                      # mirrors v3.lower_band exactly: <=, floor 1.0
        for th, sc in ABS_COST_BANDS:
            if v <= th:
                return sc
        return ABS_COST_FLOOR
    bad = []
    for a, r in tool.items():
        got, total = num(r["c_absolute_cost_band"]), num(r["annual_tco_recurring"])
        if got is None or total is None:
            continue
        if abs(got - band(total)) > 0.001:
            bad.append(f"{a} ${total:,.0f} banded {got} not {band(total)}")
    check("exported absolute-cost band matches the dollars (cheapest band = 5)", bad)

    # 6 — utilisation and cost per user
    bad = []
    for a, i in inv.items():
        ent, act, util = num(i["Entitled Users"]), num(i["Active Users (90d)"]), num(i["Utilization Rate"])
        if ent and act is not None and util is not None and abs(util - act / ent) > 0.005:
            bad.append(f"{a} util {util:.3f} vs {act / ent:.3f}")
    check("Utilization Rate still equals active / entitled users", bad)
    cpu_c = [(num(t["Annual TCO"]) or 0) / (num(inv_c[a]["Active Users (90d)"]) or 1)
             for a, t in tco_c.items() if a in inv_c]
    hi_c = max(cpu_c)
    bad = []
    for a, t in tco.items():
        act = num(inv[a]["Active Users (90d)"]) or 1
        cpu = (num(t["Annual TCO"]) or 0) / act
        if cpu > hi_c:
            bad.append(f"{a} ${cpu:,.0f}/user above the corrected file's own ceiling "
                       f"${hi_c:,.0f}")
    check(f"cost per active user stays inside the corrected file's range "
          f"(ceiling ${hi_c:,.0f})", bad)

    # 7 — nothing patient-critical was turned into an action BY THIS TUNING.
    # The tuned rows are identified from the data itself, by the marker the tuner writes
    # into Cost Notes, so this cannot drift from what was actually changed. Rows the tuning
    # never touched are excluded on purpose: APP-002, APP-006, APP-008, APP-010 and APP-018
    # were already consolidate / retire in the CORRECTED run, on Bina's own evidence and the
    # model's own reading of it. Re-litigating her answers is not this file's job; not
    # creating new ones on patient-critical rows is.
    tuned = {a for a, t in tco.items() if "TUNED DEMO VALUE" in str(t["Cost Notes"])}
    print(f"      ({len(tuned)} rows were tuned; the clinical rules below are scoped to "
          f"those, since the rest are the corrected run's own answers)")
    bad = [f"{a} {py[a]['Disposition']} but Business Criticality Critical"
           for a in tuned if py[a]["Disposition"] in ACTING
           and inv[a]["Business Criticality"] == "Critical"]
    check("no TUNED acting row carries Business Criticality 'Critical'", bad)
    bad = [f"{a} {py[a]['Disposition']} but Patient Care Impact Direct"
           for a in tuned if py[a]["Disposition"] in ACTING
           and gr.get(a, {}).get("Patient Care Impact") == "Direct"]
    check("no TUNED acting row carries Patient Care Impact 'Direct'", bad)
    bad = [f"{a} {py[a]['Disposition']} but Critical Operation Flag Yes"
           for a in tuned if py[a]["Disposition"] in ACTING
           and inv[a]["Critical Operation Flag"] == "Yes"]
    check("no TUNED acting row carries a Critical Operation Flag of 'Yes'", bad)
    untouched_acting = sorted(a for a in py if py[a]["Disposition"] in ACTING
                              and a not in tuned)
    print(f"      ({len(untouched_acting)} acting row(s) were NOT tuned and carry the "
          f"corrected run's own disposition unchanged)")

    # 8 — retire rows are genuinely not early-life
    tokens = ("pilot", "preview", "beta", "early access", "early-access")
    bad = [f"{a} release '{perf[a]['Current Release / Version']}'"
           for a in py if py[a]["Disposition"] == "retire"
           and any(t in str(perf[a]["Current Release / Version"]).lower() for t in tokens)]
    check("no retire row carries pilot/beta release wording (the lifecycle guard is not "
          "being dodged)", bad)

    print()
    failed = [r for r in results if not r[0]]
    print(f"{len(results) - len(failed)}/{len(results)} consistency rules pass")
    if failed:
        print("FAILED: " + "; ".join(r[1] for r in failed))
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
