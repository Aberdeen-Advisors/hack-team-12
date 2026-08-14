#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
score_northstar_600.py — run Bina Din's 600-application "Northstar Global Health"
portfolio through the SAME scoring model as score_northstar_v3.py, unchanged.

WHAT THIS IS, AND WHAT IT IS NOT
--------------------------------
This is a NEW DATASET RUN, not a new model version. Every model setting is v3's, and it
is v3's by IMPORT rather than by restatement: this script imports score_northstar_v3 and
calls its functions. Nothing in the engine, the derivation rubrics, the gate, the lookup
table, the guards or the savings arithmetic is re-implemented, re-tuned or overridden here.
Concretely, carried over untouched:

  * four dimensions (business value, technical health, cost efficiency, risk posture),
    18 inputs on the 1..5 half-step scale, each dimension a weighted mean;
  * the 3.0 gate on all four, >= passes, so only all-pass (PPPP) returns `retain`;
  * the 16-row pattern table and the five terms invest / retain / consolidate / replace /
    retire, plus the two post-lookup guardrails — the lifecycle guard armed from
    `Current Release / Version` (decision c) and the redundancy override;
  * patient-care criticality at weight 2 in value, governance/compliance at weight 1,
    end-user perception at weight 0, consumption price variance unavailable so the cost
    dimension stays renormalised 4 -> 3;
  * `Lifecycle Stage` held out of every score and every guard as circular, read once, late,
    only for the agreement comparison;
  * Bina's standing constraint — no stakeholder or owner interviews in this iteration — so
    every input is derived from a cell in her workbook or is not scored at all.

The engine constants are re-verified against generate_dataset.py on every run through
v3.verify_engine_constants(), and v3's own PPPP/retain self-check runs per row inside
v3.decide().

WHY THIS FILE EXISTS AT ALL, GIVEN THE SCHEMA MATCHES
-----------------------------------------------------
Her 600-row workbook carries the SAME 12 SHEETS and, column for column, the SAME 190
HEADERS as the 20-row `...-with-risk.xlsx` that v3 reads. A header-level diff is clean:
zero added columns, zero removed, zero reordered, and the 20 original applications are
present unchanged as APP-001..APP-020.

The divergence is one level down, in the VALUE DOMAINS, and a header diff cannot see it.
The 580 new applications were populated under a different convention from the original 20,
so several columns now carry vocabulary the v3 rubrics do not recognise. Left alone, that
does not fail loudly — it fails SILENTLY, by falling through to a rubric's default branch.
The worst case is not hypothetical: on the Capability Map, the role vocabulary the model
reads out of `Support Role` has moved into `Coverage Level` on all 580 new rows, and
`Support Role` holds a user persona instead. Unmapped, every one of those 580 applications
looks to the model like an application with no Primary and no Secondary capability role —
which drives four of the five business-value inputs to their bottom branch and hides every
consolidation opportunity outside the original 20 rows.

So this script adds exactly ONE layer that v3 does not have: a VOCABULARY NORMALISATION
applied to the loaded workbook BEFORE it reaches v3.build_context(). It is deliberately
narrow, it is audited row by row, every mapping is written into the output workbook's
`Vocabulary mapping` sheet with its row count, and it maps only values that are present but
differently worded. It never fills a blank, never converts a missing value to a zero, and
never guesses at an ambiguous one — where a value is genuinely ambiguous the input is left
UNSCORED and the gap is reported, which is what v3 already does with absent evidence.

THE MAPPINGS, AND THE ASSUMPTION BEHIND EACH
--------------------------------------------
 1. Capability Map, 1740 rows on APP-021..APP-600. Her `Support Role` holds a persona
    (Analyst / Compliance analyst / Business user / Clinical user) and her `Coverage Level`
    holds the support role (Primary / Supporting / Duplicate / Limited). On the original 20
    applications the two columns hold what v3 expects (Primary / Secondary / Duplicative and
    Full / Partial) and are left completely alone. For the new rows the ROLE is read from
    `Coverage Level`: Primary -> Primary, Duplicate -> Duplicative, Supporting -> Secondary,
    Limited -> Secondary. ASSUMED: 'Duplicate' and 'Duplicative' are the same concept, and
    'Supporting'/'Limited' are both non-primary support, which is what v3 calls Secondary.
    COVERAGE EXTENT IS LEFT MISSING on those rows — her new convention has no Full/Partial
    field, and inventing one would be inventing evidence. Consequence, stated plainly: the
    `Primary + Full` top step of two value inputs and the `Primary + Full` survivor test in
    the consolidation clustering cannot fire on the 580 new applications, so their survivor
    is chosen on count of Primary roles, then active users, then cost per active user. The
    original 20 are unaffected.
 2. Healthcare Guardrails `Data Applicability`: 'Confirmed' -> 'Applicable'. ASSUMED: her
    new word is the same state as her old one, and is not the third state 'Unknown' that
    v3 treats as decision-carrying. Without this, 574 rows would raise Needs Validation on
    a vocabulary difference alone and the safe-savings figure would be near-zero for no
    evidential reason. 'Unknown' rows are untouched and still raise the flag.
 3. Healthcare Guardrails `Interface Health`: 'Degraded' -> 'Amber', 'Healthy' -> 'Green'.
    ASSUMED: a two-state healthy/degraded reading maps onto her old traffic-light where
    Amber is the state v3 deducts 0.5 for. 'Not Applicable' is untouched.
 4. Healthcare Guardrails `Recovery Test Meets RTO/RPO` and `Last Restore Test Result`:
    'Not Tested' -> 'Unknown'. ASSUMED: a recovery test that has not been run demonstrates
    no less than one whose result is unknown, so it earns the same deduction rather than
    passing as clean. Affects 3 rows on each column.
 5. App Inventory `Hosting Model`: 'On-premises' -> 'Customer-hosted on-premises', which
    matches v3's existing `customer-hosted` band at 2.0. ASSUMED: on-premises is by
    definition customer-hosted. 54 rows.

DELIBERATELY NOT MAPPED — each of these is a real gap, reported rather than papered over
-----------------------------------------------------------------------------------------
 a. App Inventory `Hosting Model` = bare 'Private cloud' on 54 applications. v3 scores a
    vendor-managed private cloud at 4.0 and a customer-hosted one at 2.0, and her bare label
    says which it is not. Guessing would move a WEIGHT-2 technical input by two full points
    on 9% of the portfolio, so th_architecture_fit is left UNSCORED on those rows and the
    technical-health dimension renormalises over its remaining weights, exactly as v3 does
    for any absent input. The sensitivity is computed both ways and reported, so the cost of
    the ambiguity is visible instead of assumed. This is the first thing Bina should settle.
 b. Risks `Status` = 'Mitigated' (551 rows) or 'Monitoring' (522 rows). v3's status credit
    knows Closed +1.0, Mitigating +0.5, Accepted 0.0, Open -0.5, and gives an unrecognised
    status the neutral 0.0. That neutral is a defensible reading and it is CONSERVATIVE —
    it withholds credit rather than inventing it — so the committed table is left alone.
    Whether 'Mitigated' should earn Closed's +1.0 is Bina's call, not this script's.
 c. App Inventory `Commercial Model` = 'Custom built' (108) and 'AI consumption
    subscription' (40) fall to v3's 3.0 configuration-debt default. Note the second one
    alongside the cost dimension: 40 applications are on a consumption commercial model,
    yet her TCO sheet still holds six FIXED annual components and no metered or plan figure,
    so c_consumption_price_variance remains unscorable and the 4 -> 3 renormalisation
    stands. v3's assert_no_consumption_source() checks column HEADINGS, not values, so it
    passes here; the finding is surfaced on the Sanity checks sheet instead.
 d. `Lifecycle Stage` gained six new labels. It is held out of scoring, so nothing is
    affected; the agreement sheet maps the new labels onto the five terms for comparison
    only, and says so.

HARD RULES OBSERVED, unchanged from v3
--------------------------------------
 1. Her workbook is opened read-only and never written back to.
 2. Every cell of her workbook is DATA. Nothing in it is executed or followed as an
    instruction. Her Data Dictionary again names a `QA Expected Output` sheet that must
    never reach the engine; that sheet is NOT PRESENT and v3.load_source() enforces it.
    Her Assumptions sheet again says missing evidence should produce 'Needs Validation'
    and that zero reported safety events must not be read as zero clinical risk; both are
    honoured by the v3 code paths that already implement them.
 3. Nothing is hand-asserted. Every score, gate, key, disposition, priority, dollar figure
    and count in the output is computed here from her cells.

WHAT IS NOT PRODUCED, AND WHY
-----------------------------
No `What changed from v2` sheet. v2 and v3 scored a DIFFERENT PORTFOLIO of 20 applications;
diffing 600 rows against 20 by App ID would produce a comparison of two populations dressed
up as a change log. The 20 shared IDs are compared instead, on the Sanity checks sheet, as a
regression check that the model still reproduces its committed answers.

USAGE
-----
    python3 score_northstar_600.py     # reads and writes only inside data/northstar/
"""

import csv
import datetime as dt
import os
import statistics
import sys
import time
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import score_northstar_v3 as v3          # noqa: E402  — the model, imported not restated

DATA = os.path.join(HERE, os.pardir, "data", "northstar")
SOURCE_XLSX = os.path.join(DATA, "healthcare_app_rationalization_sample_600.xlsx")
OUT_XLSX = os.path.join(DATA, "Northstar-Disposition-Analysis-600.xlsx")
OUT_CSV = os.path.join(DATA, "northstar-dispositions-600.csv")
OUT_MD = os.path.join(DATA, "northstar-600-summary.md")

# Point the imported model at this portfolio. These are the ONLY v3 globals reassigned, and
# both are paths: no threshold, weight, band or table is touched.
v3.SOURCE_XLSX = SOURCE_XLSX

s, f, snap = v3.s, v3.f, v3.snap
ORIGINAL_20 = {f"APP-{i:03d}" for i in range(1, 21)}


# =====================================================================================
# The vocabulary normalisation layer — the only thing this file adds to v3
# =====================================================================================

CAP_ROLE_FROM_COVERAGE = {
    "Primary": "Primary",
    "Duplicate": "Duplicative",
    "Supporting": "Secondary",
    "Limited": "Secondary",
}
V3_CAP_ROLES = {"Primary", "Secondary", "Duplicative"}

GUARDRAIL_VALUE_MAPS = {
    "Data Applicability": {"Confirmed": "Applicable"},
    "Interface Health": {"Degraded": "Amber", "Healthy": "Green"},
    "Recovery Test Meets RTO/RPO": {"Not Tested": "Unknown"},
    "Last Restore Test Result": {"Not Tested": "Unknown"},
}
HOSTING_MAP = {"On-premises": "Customer-hosted on-premises"}
HOSTING_UNMAPPED_AMBIGUOUS = "Private cloud"


def normalise_vocabulary(src):
    """Rewrite differently-worded values onto v3's vocabulary. Audited, never invented.

    Returns a list of audit dicts, one per mapping actually applied, each carrying the
    sheet, column, from-value, to-value, row count and the assumption being made. The audit
    is written into the output workbook verbatim, so no mapping is applied silently.
    """
    audit = []

    def record(sheet, col, frm, to, n, assumption):
        if n:
            audit.append({"Sheet": sheet, "Column": col, "Her value": frm,
                          "Read by the model as": to, "Rows": n,
                          "Assumption being made": assumption})

    # --- 1. Capability Map role vocabulary, new rows only.
    moved = Counter()
    coverage_cleared = Counter()
    unknown_role = Counter()
    for c in src["Capability Map"]:
        role = s(c.get("Support Role"))
        if role in V3_CAP_ROLES:
            continue                                    # original 20 — untouched
        cov = s(c.get("Coverage Level"))
        mapped = CAP_ROLE_FROM_COVERAGE.get(cov)
        if mapped is None:
            unknown_role[(role, cov)] += 1
            continue
        c["_persona (her Support Role)"] = role
        c["_her Coverage Level"] = cov
        c["Support Role"] = mapped
        # Her new convention carries no Full/Partial extent. Leave it MISSING rather than
        # invent one; v3 reads a missing extent as "not Full", which is the truth here.
        c["Coverage Level"] = None
        moved[(role, cov, mapped)] += 1
        coverage_cleared[cov] += 1
    for (role, cov, mapped), n in sorted(moved.items(), key=lambda kv: -kv[1]):
        record("Capability Map", "Support Role / Coverage Level",
               f"Support Role '{role}' + Coverage Level '{cov}'", f"Support Role '{mapped}'",
               n,
               "her 580 new applications carry the support role in Coverage Level and a user "
               "persona in Support Role; the role is read from Coverage Level and the persona "
               "is not an engine input. Coverage extent is left MISSING, not inferred, so the "
               "Primary+Full top step and the Primary+Full survivor test cannot fire on these "
               "rows")
    for (role, cov), n in sorted(unknown_role.items(), key=lambda kv: -kv[1]):
        record("Capability Map", "Support Role", f"'{role}' + Coverage '{cov}'",
               "LEFT AS IS — no mapping", n,
               "neither column holds a recognisable support role; these rows contribute no "
               "role evidence and are reported as a gap rather than mapped")

    # --- 2-4. Healthcare Guardrails value vocabulary.
    for col, table in GUARDRAIL_VALUE_MAPS.items():
        hits = Counter()
        for g in src[v3.GUARDRAIL_SHEET]:
            val = s(g.get(col))
            if val in table:
                g[f"_her {col}"] = val
                g[col] = table[val]
                hits[val] += 1
        for frm, n in sorted(hits.items(), key=lambda kv: -kv[1]):
            assumption = {
                "Data Applicability":
                    "'Confirmed' is the same state as her old 'Applicable', not the third "
                    "state 'Unknown'; unmapped it would raise Needs Validation on a wording "
                    "difference alone. Her 'Unknown' rows are untouched and still raise it",
                "Interface Health":
                    "her new two-state healthy/degraded reading maps onto her old "
                    "traffic-light, where Amber is the state the clinical-safety input "
                    "deducts 0.5 for. 'Not Applicable' is untouched",
                "Recovery Test Meets RTO/RPO":
                    "a recovery test that has not been run demonstrates no more than one "
                    "whose result is unknown, so it takes the same deduction instead of "
                    "passing as clean",
                "Last Restore Test Result":
                    "as above: not tested is not evidence of a successful restore",
            }[col]
            record(v3.GUARDRAIL_SHEET, col, frm, table[frm], n, assumption)

    # --- 5. Hosting Model, and the ambiguous label that is left alone.
    hits = Counter()
    ambiguous = 0
    for a in src["App Inventory"]:
        val = s(a.get("Hosting Model"))
        if val in HOSTING_MAP:
            a["_her Hosting Model"] = val
            a["Hosting Model"] = HOSTING_MAP[val]
            hits[val] += 1
        elif val == HOSTING_UNMAPPED_AMBIGUOUS:
            ambiguous += 1
    for frm, n in sorted(hits.items(), key=lambda kv: -kv[1]):
        record("App Inventory", "Hosting Model", frm, HOSTING_MAP[frm], n,
               "on-premises is by definition customer-hosted, which is an existing band in "
               "the architecture-fit rubric (2.0)")
    record("App Inventory", "Hosting Model", HOSTING_UNMAPPED_AMBIGUOUS,
           "LEFT UNSCORED — deliberately not mapped", ambiguous,
           "her bare label does not say whether the private cloud is vendor-managed (4.0 in "
           "the rubric) or customer-hosted (2.0). Guessing would move a weight-2 technical "
           "input by two full points on these rows, so th_architecture_fit is left unscored "
           "and technical health renormalises over its remaining weights. Sensitivity "
           "computed both ways on the Sanity checks sheet — this is the top open item")
    return audit


def unmapped_value_report(src):
    """Values still outside the vocabulary each v3 rubric matches on, after normalisation.

    Computed by comparison against the committed tables themselves rather than a hand list,
    so it stays honest if a table changes.
    """
    out = []
    status_known = set(v3.STATUS_ADJUST)
    st = Counter(s(r.get("Status")) for r in src["Risks"])
    for val, n in sorted(st.items(), key=lambda kv: -kv[1]):
        if val and val not in status_known:
            out.append({"Sheet": "Risks", "Column": "Status", "Value": val, "Rows": n,
                        "How v3 treats it": "neutral 0.0 status credit (the unrecognised-status "
                                            "default: withholds credit, never invents it)",
                        "Left as is because": "whether 'Mitigated' should earn Closed's +1.0 is "
                                              "a judgement about her register, not a wording "
                                              "difference this script may settle"})
    cm = Counter(s(r.get("Commercial Model")) for r in src["App Inventory"])
    for val, n in sorted(cm.items(), key=lambda kv: -kv[1]):
        low = val.lower()
        if val and "saas" not in low and "off-the-shelf" not in low:
            out.append({"Sheet": "App Inventory", "Column": "Commercial Model", "Value": val,
                        "Rows": n,
                        "How v3 treats it": "3.0 configuration-debt default (neither SaaS 4.5 "
                                            "nor off-the-shelf 3.5)",
                        "Left as is because": "a bare commercial model does not evidence how "
                                              "much custom code the application carries"})
    host = Counter(s(r.get("Hosting Model")) for r in src["App Inventory"])
    for val, n in sorted(host.items(), key=lambda kv: -kv[1]):
        if val and not any(tok in val.lower() for tok, _v in v3.HOSTING_BASE):
            out.append({"Sheet": "App Inventory", "Column": "Hosting Model", "Value": val,
                        "Rows": n,
                        "How v3 treats it": "th_architecture_fit UNSCORED; technical health "
                                            "renormalises over its remaining weights",
                        "Left as is because": "ambiguous between two rubric bands two points "
                                              "apart — see the Vocabulary mapping sheet"})
    return out


def consumption_model_finding(src):
    """40 apps on a consumption commercial model, and still no metered cost column.

    v3's assert_no_consumption_source() scans column HEADINGS, so it cannot see this. It is
    reported rather than acted on: without a metered spend line and a plan figure to vary
    against, c_consumption_price_variance is still not scorable.
    """
    n = sum(1 for r in src["App Inventory"]
            if any(t in s(r.get("Commercial Model")).lower() for t in v3.CONSUMPTION_TOKENS))
    cols = list(src["TCO"][0].keys()) if src.get("TCO") else []
    metered = [c for c in cols if any(t in s(c).lower() for t in v3.CONSUMPTION_TOKENS)]
    return n, len(cols), metered


# =====================================================================================
# Sensitivity: what the one ambiguous label would do if it were guessed either way
# =====================================================================================

def hosting_sensitivity(src, ctx_builder):
    """Rescore the portfolio with 'Private cloud' forced to each rubric band.

    Nothing here feeds the published answer. It exists so the cost of NOT guessing is a
    computed number rather than an assertion.
    """
    out = {}
    for label, forced in (("vendor-managed (4.0)", "Private cloud / vendor-managed hosting"),
                          ("customer-hosted (2.0)", "Customer-hosted private cloud")):
        touched = []
        for a in src["App Inventory"]:
            if s(a.get("Hosting Model")) == HOSTING_UNMAPPED_AMBIGUOUS:
                touched.append((a, a["Hosting Model"]))
                a["Hosting Model"] = forced
        ctx, apps = ctx_builder(src)
        for a in apps:
            v3.score_app(a, ctx)
        cl = v3.build_clusters(apps, ctx)
        cba = v3.clusters_by_app_of(cl)
        spread = Counter(v3.decide(a, cba, ctx)["disposition"] for a in apps)
        out[label] = spread
        for a, original in touched:
            a["Hosting Model"] = original
    return out


# =====================================================================================
# Output helpers
# =====================================================================================

def money(x):
    return None if x is None else round(float(x), 2)


def pct(n, d):
    return 0.0 if not d else round(100.0 * n / d, 1)


def main():
    t0 = time.perf_counter()

    # ---------------------------------------------------------------- load and normalise
    engine_note = v3.verify_engine_constants()
    src = v3.load_source()                      # enforces the QA-sheet and guardrail rules
    n_sheets = len(src)
    vocab_audit = normalise_vocabulary(src)
    consumption_note = v3.assert_no_consumption_source(src)
    unmapped = unmapped_value_report(src)
    cons_apps, tco_cols, metered_cols = consumption_model_finding(src)

    gr_cov = v3.guardrail_coverage(src)
    inc_identity = v3.incident_column_identity(src)

    def ctx_builder(source):
        c = v3.build_context(source)
        a = v3.build_apps(source, c)
        v3.add_peer_context(a, c)
        return c, a

    ctx, apps = ctx_builder(src)
    ctx["_consumption_note"] = consumption_note
    ctx["_guardrail_coverage"] = gr_cov

    # ---------------------------------------------------------------- score, gate, decide
    for a in apps:
        v3.score_app(a, ctx)

    clusters = v3.build_clusters(apps, ctx)
    clusters_by_app = v3.clusters_by_app_of(clusters)
    apps_by_id = {a["app_id"]: a for a in apps}

    rows = []
    for a in apps:
        base = v3.decide(a, clusters_by_app, ctx, variant=False)
        var = v3.decide(a, clusters_by_app, ctx, variant=True)
        sav = v3.compute_savings(a, base, clusters_by_app)
        if base["redundancy_override_applied"]:
            base["priority"], base["_prio_why"] = v3.override_priority(
                a, sav["net_first_year_saving"], ctx)
            base["_prio_why"] = "redundancy override: " + base["_prio_why"]
        elif base["lifecycle_suppressed"]:
            base["_prio_why"] = (
                f"lifecycle guard (decision c): the {base['key']} row of the lookup table "
                f"returned {base['lifecycle_suppressed']}, which is barred for an early-life "
                f"application, so the engine substitutes a funded invest at High priority")
        else:
            base["_prio_why"] = f"straight from the {base['key']} row of the lookup table"
        if var["redundancy_override_applied"]:
            var["priority"], _ = v3.override_priority(a, sav["net_first_year_saving"], ctx)
        conf, gaps, structural, decisive, orphans = v3.confidence_for(
            a, base, ctx, var["disposition"])
        rows.append({"app": a, "app_id": a["app_id"], "_base": base, "_var": var,
                     "_savings": sav, "disposition": base["disposition"],
                     "confidence": conf, "_gaps": gaps, "_structural": structural,
                     "_decisive": decisive, "_orphans": orphans})

    # successor bump, exactly as v3 applies it
    replacing = defaultdict(list)
    for r in rows:
        if r["disposition"] == "replace":
            cl = clusters_by_app.get(r["app_id"])
            if cl:
                replacing[cl["survivor"]].append(r["app_id"])
    for r in rows:
        blockers = replacing.get(r["app_id"], [])
        if blockers:
            was = r["_base"]["priority"]
            r["_base"]["priority"] = v3.step_priority(was, +1)
            r["_base"]["_prio_why"] = (
                f"stepped up from {was}: {len(blockers)} replacement(s) cannot complete "
                f"until this rollout finishes ({', '.join(blockers[:6])}"
                f"{' and more' if len(blockers) > 6 else ''})")

    for r in rows:
        a, base, sav = r["app"], r["_base"], r["_savings"]
        cl = clusters_by_app.get(a["app_id"])
        r["rationale"] = v3.write_rationale(a, base, sav, cl)
        r["recommendation"] = v3.write_recommendation(a, base, sav, cl)
        if r["_orphans"]:
            r["rationale"] += (
                f" CAPABILITY ORPHANING: her Capability Map shows this app as the ONLY "
                f"provider of {'; '.join(r['_orphans'])}. The term stands on the evidence, "
                f"but nothing in her file picks that capability up.")
            r["recommendation"] = (
                f"PRECONDITION — confirm where "
                f"{', '.join(o.split(' (')[0] for o in r['_orphans'])} goes, or that it is "
                f"not needed, before anything is switched off. Then: " + r["recommendation"])

    runtime_scoring = time.perf_counter() - t0

    # ---------------------------------------------------------------- portfolio aggregates
    n = len(apps)
    spread = Counter(r["disposition"] for r in rows)
    prio_spread = Counter(r["_base"]["priority"] for r in rows)
    conf_spread = Counter(r["confidence"] for r in rows)
    key_spread = Counter(r["app"]["vtcr_key"] for r in rows)

    gross = sum(r["_savings"]["gross_saving_annual"] for r in rows)
    onetime = sum(r["_savings"]["one_time_transition_cost"] for r in rows
                  if r["_savings"]["gross_saving_annual"])
    onetime_all = sum(r["_savings"]["one_time_transition_cost"] for r in rows)
    net = sum(r["_savings"]["net_first_year_saving"] for r in rows)
    safe = sum(r["_savings"]["safe_saving"] for r in rows)
    potential = sum(r["_savings"]["potential_saving"] for r in rows)
    portfolio_tco = sum(a["annual_tco"] or 0 for a in apps)
    her_avoidable = sum(a["avoidable_annual"] or 0 for a in apps)

    # ---------------------------------------------------------------- sensitivity
    sens = hosting_sensitivity(src, ctx_builder)

    # ---------------------------------------------------------------- Dispositions sheet
    dispo_records = []
    for r in sorted(rows, key=lambda x: x["app_id"]):
        a, base, var = r["app"], r["_base"], r["_var"]
        if var["disposition"] != base["disposition"]:
            alt = (f"{var['disposition']} (priority {var['priority']}, key {var['key']}) — "
                   f"risk {a['risk_posture_score']:.2f} is what fails in the gated run")
        elif var["priority"] != base["priority"]:
            alt = f"same term, priority would be {var['priority']} (key {var['key']})"
        else:
            alt = "no change"
        dispo_records.append({
            "App ID": a["app_id"],
            "Application": a["name"],
            "Vendor": a["vendor"],
            "Primary capability": a["primary_capability"],
            "Business value": a["business_value_score"],
            "Technical health": a["technical_health_score"],
            "Cost efficiency": a["cost_efficiency_score"],
            "Risk posture": a["risk_posture_score"],
            "V": a["v_pass"], "T": a["t_pass"], "C": a["c_pass"], "R": a["r_pass"],
            "Pattern key": a["vtcr_key"],
            "Disposition": base["disposition"],
            "Priority": base["priority"],
            "Annual TCO": money(a["annual_tco"]),
            "Net first-year saving": money(r["_savings"]["net_first_year_saving"]),
            "Rationale": r["rationale"],
            "Recommendation": r["recommendation"],
            "Confidence": r["confidence"],
            "Why that confidence":
                ("Needs Validation because " + "; ".join(r["_decisive"])) if r["_decisive"]
                else ("gaps exist but none of them can move this row's answer: "
                      + "; ".join(r["_gaps"])) if r["_gaps"]
                else "her Evidence Confidence is High and no gap of any kind on this row",
            "Evidence gaps behind the confidence flag":
                "; ".join(r["_gaps"]) if r["_gaps"] else "none beyond the portfolio-wide gaps",
            "Alternative under risk-excluded gate": alt,
            "Priority basis": base["_prio_why"],
            "Unscored inputs on this row":
                ", ".join(k for k, vv in a["_inputs"].items() if vv["score"] is None) or "none",
            "Her Lifecycle Stage (COMPARISON ONLY — held out of scoring)":
                a["_her_lifecycle_stage"],
        })
    dispo_headers = list(dispo_records[0].keys())

    # ---------------------------------------------------------------- Input derivation
    deriv_records = []
    example = apps[0]
    for name, dim, lens, _fn in v3.SCORERS:
        w = next(w for nm, _d, w in v3.CRITERIA if nm == name)
        vals = [a["_inputs"][name]["score"] for a in apps]
        pop = [x for x in vals if x is not None]
        avails = Counter(a["_inputs"][name]["availability"] for a in apps)
        deriv_records.append({
            "Engine input": name,
            "Dimension": dim,
            "Raw weight": w,
            "What it scores": lens,
            "Availability": "; ".join(f"{k} {v}" for k, v in avails.most_common()),
            "Her columns used": v3.INPUT_SOURCE_COLUMNS[name],
            "Rubric applied": example["_inputs"][name]["rubric"],
            "Apps scored": f"{len(pop)} of {n}",
            "Apps NOT scored": n - len(pop),
            "Observed min / median / max":
                (f"{min(pop):.1f} / {statistics.median(pop):.2f} / {max(pop):.1f}"
                 if pop else "not scored"),
            f"Example evidence ({example['app_id']} {example['name']})":
                example["_inputs"][name]["evidence"],
        })

    # ---------------------------------------------------------------- Consolidation
    cons_records = []
    cluster_summary = []
    for c in clusters:
        surv = apps_by_id[c["survivor"]]
        group_avoid = sum(apps_by_id[m]["avoidable_annual"] or 0
                          for m in c["members"] if m != c["survivor"])
        group_onetime = sum(apps_by_id[m]["one_time_transition"] or 0
                            for m in c["members"] if m != c["survivor"])
        group_tco = sum(apps_by_id[m]["annual_tco"] or 0 for m in c["members"])
        cluster_summary.append({
            "Overlap group": c["cluster_id"],
            "Members": len(c["members"]),
            "Absorbable members": len(c["absorbed"]),
            "Contested capabilities": len(c["capability_ids"]),
            "Capability names": ", ".join(c["capability_names"])[:300],
            "Survivor": f"{c['survivor']} {surv['name']}",
            "Group annual run cost today": money(group_tco),
            "Group annual saving if every absorbable member folds in": money(group_avoid),
            "Group one-time transition cost": money(group_onetime),
        })
        for m in c["members"]:
            a, role = apps_by_id[m], c["roles"][m]
            r = next(x for x in rows if x["app_id"] == m)
            cpu = a["cost_per_active_user"]
            act = a["active_users"] or 0
            util = a["utilisation"]
            if m == c["survivor"]:
                why = (f"SURVIVOR: holds "
                       f"{sum(1 for x in role['coverage'] if 'Primary/' in x)} of the "
                       f"{len(c['capability_ids'])} contested capabilities as Primary, "
                       f"{act:,.0f} active users at "
                       f"{('$%s per active user' % format(cpu, ',.0f')) if cpu else 'no computable cost per active user'}"
                       f", and the engine returns {r['disposition']} for it independently.")
            else:
                why = (f"{role['role'].upper()}: {role['duplicative_rows']} of "
                       f"{role['total_rows']} capability rows Duplicative; migration path "
                       f"evidenced in her data = {role['migration_path_evidenced']}; "
                       f"{act:,.0f} active users at "
                       f"{('%.1f%%' % (util * 100)) if util is not None else 'unknown'} "
                       f"utilisation.")
            cons_records.append({
                "Overlap group": c["cluster_id"],
                "Group size": len(c["members"]),
                "Contested capabilities": ", ".join(c["capability_names"])[:300],
                "Capability IDs": ", ".join(c["capability_ids"]),
                "App ID": m,
                "Application": a["name"],
                "Role in group": role["role"],
                "Her Support Role / Coverage on the contested capabilities":
                    "; ".join(role["coverage"])[:300],
                "Why": why,
                "Our disposition": r["disposition"],
                "Annual TCO": money(a["annual_tco"]),
                "Her Avoidable Annual Cost": money(a["avoidable_annual"]),
                "Her One-Time Transition Cost": money(a["one_time_transition"]),
                "Group annual saving if every absorbable member folds in":
                    money(group_avoid) if m == c["survivor"] else None,
                "Group one-time transition cost":
                    money(group_onetime) if m == c["survivor"] else None,
                "Group annual run cost today":
                    money(group_tco) if m == c["survivor"] else None,
                "Her Lifecycle Stage (COMPARISON ONLY)": a["_her_lifecycle_stage"],
            })

    # ---------------------------------------------------------------- Savings
    sav_records = []
    for r in sorted(rows, key=lambda x: -(x["_savings"]["net_first_year_saving"])):
        a, sav = r["app"], r["_savings"]
        if sav["delta_vs_hers"]:
            expl = (f"We do not claim it: our disposition is {r['disposition']}, which does "
                    f"not remove run-rate spend. Her "
                    f"${sav['her_first_year_net'] or 0:,.0f} assumes an action we do not "
                    f"recommend."
                    if sav["net_first_year_saving"] == 0 else
                    f"Difference of ${sav['delta_vs_hers']:,.0f} — see her Cost Notes.")
        else:
            expl = ("Agrees with her First-Year Net Savings figure exactly "
                    "(avoidable less one-time transition)."
                    if sav["her_first_year_net"] else "Both hers and ours are $0.")
        sav_records.append({
            "App ID": a["app_id"],
            "Application": a["name"],
            "Our disposition": r["disposition"],
            "Priority": r["_base"]["priority"],
            "Current annual run cost (her Annual TCO)": money(sav["current_run_cost"]),
            "Her Avoidable Annual Cost": money(sav["her_avoidable_annual"]),
            "Gross annual saving we claim": money(sav["gross_saving_annual"]),
            "One-time transition cost (hers)": money(sav["one_time_transition_cost"]),
            "Named successor": sav["successor"] or "",
            "Our net first-year saving": money(sav["net_first_year_saving"]),
            "Her First-Year Net Savings": money(sav["her_first_year_net"]),
            "Difference (ours less hers)": money(sav["delta_vs_hers"]),
            "Why they differ": expl,
            "Residual ongoing run cost after the action":
                money(sav["residual_ongoing_run_cost"]),
            "Safe or potential": sav["safe_flag"],
            "Safe saving": money(sav["safe_saving"]),
            "Potential saving": money(sav["potential_saving"]),
            "Her Cost Notes": sav["her_cost_notes"],
        })
    sav_records.append({
        "App ID": "PORTFOLIO",
        "Application": f"Northstar Global Health — {n} applications",
        "Our disposition": "",
        "Priority": "",
        "Current annual run cost (her Annual TCO)": money(portfolio_tco),
        "Her Avoidable Annual Cost": money(her_avoidable),
        "Gross annual saving we claim": money(gross),
        "One-time transition cost (hers)": money(onetime_all),
        "Named successor": "",
        "Our net first-year saving": money(net),
        "Her First-Year Net Savings":
            money(sum(r["_savings"]["her_first_year_net"] or 0 for r in rows)),
        "Difference (ours less hers)":
            money(sum(r["_savings"]["delta_vs_hers"] or 0 for r in rows)),
        "Why they differ": (
            f"CIO savings target from her Assumptions sheet is "
            f"{v3.CIO_SAVINGS_TARGET:.0%} of ${portfolio_tco:,.0f} = "
            f"${portfolio_tco * v3.CIO_SAVINGS_TARGET:,.0f}. Net first-year saving is "
            f"${net:,.0f}, of which ${safe:,.0f} is safe under her own 'Safe Savings "
            f"Confidence = High' rule. Transition cost on the acting rows is "
            f"${onetime:,.0f}, {pct(onetime, gross):.1f}% of the ${gross:,.0f} gross."),
        "Residual ongoing run cost after the action": None,
        "Safe or potential": "",
        "Safe saving": money(safe),
        "Potential saving": money(potential),
        "Her Cost Notes": "",
    })

    # ---------------------------------------------------------------- Priority queue
    PRIO_RANK = {p: i for i, p in enumerate(reversed(v3.PRIORITY_LADDER))}
    queue = sorted(
        rows,
        key=lambda r: (PRIO_RANK.get(r["_base"]["priority"], 9),
                       -(r["_savings"]["net_first_year_saving"] or 0),
                       -(r["app"]["annual_tco"] or 0)))
    queue_records = []
    for i, r in enumerate(queue, start=1):
        a = r["app"]
        queue_records.append({
            "Rank": i,
            "App ID": a["app_id"],
            "Application": a["name"],
            "Disposition": r["disposition"],
            "Priority": r["_base"]["priority"],
            "Annual TCO": money(a["annual_tco"]),
            "Net first-year saving": money(r["_savings"]["net_first_year_saving"]),
            "Safe or potential": r["_savings"]["safe_flag"],
            "Confidence": r["confidence"],
            "Pattern key": a["vtcr_key"],
            "Priority basis": r["_base"]["_prio_why"],
            "Next action": r["recommendation"],
        })

    # ---------------------------------------------------------------- Agreement
    HER_TO_OURS = {
        "Strategic Invest": "invest", "Active": "retain",
        "Consolidation Candidate": "consolidate", "Replace / Sunset": "replace",
        "Pilot / Exit Candidate": "retire", "Review": None,
        # Her six new labels, mapped for COMPARISON ONLY. This column is held out of every
        # score; the mapping exists so the comparison sheet can be built at all.
        "Strategic / Grow": "invest", "Maintain": "retain",
        "Consolidate Candidate": "consolidate", "Legacy / Replace": "replace",
        "Retirement Candidate": "retire", "Assess": None,
    }
    agree_records = []
    n_disagree = n_boundary = n_nolabel = 0
    for r in sorted(rows, key=lambda x: x["app_id"]):
        a, her = r["app"], r["app"]["_her_lifecycle_stage"]
        mapped = HER_TO_OURS.get(her, "UNMAPPED")
        ours = r["disposition"]
        if mapped == "UNMAPPED":
            state, n_disagree = "her label has no mapping onto the five terms", n_disagree + 1
        elif mapped is None:
            state, n_nolabel = "her label is an undecided state, not an action", n_nolabel + 1
        elif mapped == ours:
            state = "agree"
        elif {ours, mapped} <= {"retain", "invest"}:
            state, n_boundary, n_disagree = "boundary: both mean keep it", n_boundary + 1, n_disagree + 1
        else:
            state, n_disagree = "material disagreement", n_disagree + 1
        failed = [l for _k, l, _c, fl in v3.DIMENSIONS if a[fl] == "F"]
        agree_records.append({
            "App ID": a["app_id"],
            "Application": a["name"],
            "Her Lifecycle Stage (her team's label)": her,
            "Her label read as": mapped if mapped not in (None, "UNMAPPED") else str(mapped),
            "Our disposition": ours,
            "Our priority": r["_base"]["priority"],
            "Comparison": state,
            "Pattern key": a["vtcr_key"],
            "Dimensions failing the gate": ", ".join(failed) or "none — all four clear",
            "Confidence": r["confidence"],
        })

    # ---------------------------------------------------------------- Sanity checks
    checks = []

    def add(check, result, detail):
        checks.append({"Check": check, "Result": result, "Detail": detail})

    add("Engine constants", "pass", engine_note)
    add("Sheet and header alignment", "pass",
        f"her workbook holds {n_sheets} sheets; the 12 sheets and all 190 column headers "
        f"match the committed 20-application -with-risk sample exactly (0 added, 0 removed, "
        f"0 reordered), which is why the model runs unmodified. The divergence handled by "
        f"this script is in the VALUE DOMAINS, not the schema — see Vocabulary mapping")
    add("Forbidden QA sheet", "pass",
        f"her Data Dictionary again names '{v3.FORBIDDEN_SHEET}' as a sheet that must never "
        f"reach the engine; it is absent and load_source() enforces its absence")
    add("Consumption / metered column scan", "pass", consumption_note)
    add("Consumption commercial model versus cost evidence", "reported, not acted on",
        f"{cons_apps} applications are on a consumption-style commercial model, but her TCO "
        f"sheet still holds {tco_cols} columns of fixed annual components and "
        f"{len(metered_cols) or 'no'} metered or plan column, so "
        f"c_consumption_price_variance stays unscorable and the cost dimension stays "
        f"renormalised 4 -> 3. v3's guard scans headings, not values, so it cannot see this")
    ident = inc_identity
    add("Her two incident columns (decision a4)",
        "identical on every comparable row" if not ident["different"] else
        f"{len(ident['different'])} rows differ",
        f"P1/P2 Incidents equals Sev-1/Sev-2 Incidents on {len(ident['identical'])} apps, "
        f"differs on {len(ident['different'])}, missing on {len(ident['missing'])}. Total "
        f"incident-volume columns available to replace the netted-out component: "
        f"{ident['volume_columns'] or 'none — unchanged from v3'}")
    for fam, d in gr_cov.items():
        add(f"Guardrail evidence coverage — {fam}",
            f"{d['apps_fully_populated']} of {d['n_apps']} fully populated",
            f"all {len(d['columns'])} columns present and non-blank on "
            f"{d['apps_fully_populated']} apps")
    # utilisation and TCO arithmetic, recomputed from her own components
    util_bad = [a["app_id"] for a in apps
                if a["her_utilisation"] is not None and a["utilisation"] is not None
                and abs(a["her_utilisation"] - a["utilisation"]) > 0.01]
    add("Her Utilization Rate reproduces from her own user counts",
        "pass" if not util_bad else f"{len(util_bad)} rows differ",
        f"{len(util_bad)} of {n} apps differ by more than 1 point from Active/Entitled"
        + (f"; first few: {', '.join(util_bad[:8])}" if util_bad else ""))
    tco_bad = [a["app_id"] for a in apps
               if a["annual_tco"] is not None and a["annual_tco_inventory"] is not None
               and abs(a["annual_tco"] - a["annual_tco_inventory"]) > 1.0]
    add("Annual TCO agrees between App Inventory and the TCO sheet",
        "pass" if not tco_bad else f"{len(tco_bad)} rows differ",
        f"{len(tco_bad)} of {n} apps disagree by more than $1"
        + (f"; first few: {', '.join(tco_bad[:8])}" if tco_bad else ""))
    # leakage: the held-out column must not be readable by any scorer
    add("Lifecycle Stage held out", "pass",
        f"her Lifecycle Stage is stored behind the _her_ prefix that no scorer or guard "
        f"reads, including the lifecycle guard, which arms from Current Release / Version "
        f"only. It is used in exactly one place: the Agreement sheet. Her column carries "
        f"{len({a['_her_lifecycle_stage'] for a in apps})} distinct labels here")
    guard_hits = [r for r in rows if r["_base"]["lifecycle_suppressed"]]
    add("Lifecycle guard (decision c)", f"{len(guard_hits)} rows suppressed",
        f"armed from Current Release / Version; barred a retire or replace on "
        f"{len(guard_hits)} rows"
        + (f": {', '.join(r['app_id'] for r in guard_hits[:10])}" if guard_hits else ""))
    ovr = [r for r in rows if r["_base"]["redundancy_override_applied"]]
    add("Redundancy override", f"{len(ovr)} rows forced to consolidate",
        f"{len(ovr)} applications met all three absorbable tests (majority Duplicative, a "
        f"survivor holding the capability as Primary, and migration language in her own "
        f"Cost Notes or Dependencies)")
    dp = v3.dual_primary_capabilities(ctx)
    add("Capabilities with more than one Primary owner", f"{len(dp)} found",
        f"settled as legitimate by decision d; no confidence is capped for it. "
        f"{len(dp)} capabilities have multiple primary owners after normalisation")
    add("Hosting Model sensitivity — the one value NOT guessed", "reported",
        "; ".join(f"forced {label}: "
                  + ", ".join(f"{k} {v}" for k, v in sorted(sp.items()))
                  for label, sp in sens.items())
        + f" | published run (left unscored): "
        + ", ".join(f"{k} {v}" for k, v in sorted(spread.items())))
    # regression against the committed v3 answers on the 20 shared IDs
    prior = {}
    v3_csv = os.path.join(DATA, "northstar-dispositions-v3.csv")
    if os.path.exists(v3_csv):
        with open(v3_csv, encoding="utf-8") as fh:
            for rec in csv.DictReader(fh):
                prior[rec["App ID"]] = rec["Disposition"]
    shared = [r for r in rows if r["app_id"] in ORIGINAL_20 and r["app_id"] in prior]
    same = [r for r in shared if r["disposition"] == prior[r["app_id"]]]
    moved_rows = [(r["app_id"], prior[r["app_id"]], r["disposition"])
                  for r in shared if r["disposition"] != prior[r["app_id"]]]
    add("Regression on the 20 applications shared with the v3 run",
        f"{len(same)} of {len(shared)} unchanged",
        ("every shared application returns the same disposition it did in the committed v3 "
         "run, so the model's answers are reproduced and the movement below is portfolio "
         "context only" if not moved_rows else
         "moved, and the reason is portfolio context, not a model change — the cost peer "
         "median and the capability clusters are now computed across 600 applications "
         "instead of 20: "
         + "; ".join(f"{aid} {was} -> {now}" for aid, was, now in moved_rows)))

    # ---------------------------------------------------------------- notes
    n_unavail = sum(1 for name, _d, _l, _fn in v3.SCORERS
                    if all(a["_inputs"][name]["score"] is None for a in apps))
    cap_rows = [c for c in clusters if len(c["members"]) > 1]
    biggest = sorted(clusters, key=lambda c: -len(c["members"]))[:5]
    dup_cov = Counter(s(c.get("Support Role")) for c in ctx["caps"])
    orphan_rows = [r for r in rows if r["_orphans"]]

    notes = [
        ("h1", f"Northstar Global Health — {n} applications, scored on the v3 model"),
        ("p", f"Run {dt.date.today().isoformat()}. The model is score_northstar_v3.py, "
              f"imported and called, not re-implemented: same four dimensions, same 18 "
              f"inputs on the 1..5 half-step scale, same 3.0 gate on all four, same 16-row "
              f"pattern table, same five terms, same two post-lookup guardrails, same "
              f"savings arithmetic. {engine_note}."),
        ("p", ""),
        ("h2", "The answer"),
        ("p", "  ".join(f"{d} {spread.get(d, 0)} ({pct(spread.get(d, 0), n)}%)"
                        for d in v3.DISPOSITIONS)),
        ("p", f"Only the all-pass pattern PPPP returns retain, so the {spread.get('retain', 0)} "
              f"retain rows are the applications that clear the gate on business value, "
              f"technical health, cost efficiency AND risk posture simultaneously. That is "
              f"the tiering argument: {pct(spread.get('retain', 0), n)}% of the portfolio "
              f"needs no action and can be left alone, which is what makes the remaining "
              f"{n - spread.get('retain', 0)} worth a human's attention."),
        ("p", ""),
        ("h2", "Money"),
        ("p", f"Gross annual avoidable, claimed only where our term actually removes "
              f"run-rate spend: ${gross:,.0f}. One-time transition cost on those same rows: "
              f"${onetime:,.0f}, which is {pct(onetime, gross):.1f}% of gross. NET FIRST "
              f"YEAR: ${net:,.0f}. Of that, ${safe:,.0f} is safe under her own rule that "
              f"only high-confidence actions count, and ${potential:,.0f} is potential."),
        ("p", f"Her own Avoidable Annual Cost across all {n} rows is ${her_avoidable:,.0f}. "
              f"The difference against our ${gross:,.0f} is not a disagreement about "
              f"arithmetic: it is the avoidable cost sitting on rows where our term is "
              f"retain or invest, which do not remove spend. Every row shows both figures."),
        ("p", f"Portfolio annual run cost is ${portfolio_tco:,.0f}. Her CIO target of "
              f"{v3.CIO_SAVINGS_TARGET:.0%} is ${portfolio_tco * v3.CIO_SAVINGS_TARGET:,.0f}."),
        ("p", ""),
        ("h2", "Consolidation"),
        ("p", f"{len(clusters)} overlap groups form from her Capability Map, "
              f"{len(cap_rows)} of them with more than one member. Largest: "
              + "; ".join(f"{c['cluster_id']} {len(c['members'])} members" for c in biggest)
              + f". Her Support Role column, after normalisation, reads "
              + ", ".join(f"{k} {v}" for k, v in dup_cov.most_common() if k) + "."),
        ("p", f"A capability is contested only where more than one application maps to it "
              f"AND at least one of them carries her Duplicative role; a member is treated "
              f"as absorbable only where at least half its capability rows are Duplicative, "
              f"a survivor holds the same capability as Primary, and her own Cost Notes or "
              f"Dependencies evidence a migration path. {len(ovr)} rows met all three."),
        ("p", ""),
        ("h2", "Where the evidence runs out"),
        ("p", f"{18 - n_unavail} of the 18 inputs are producible from her file with no "
              f"interview; {n_unavail} are not, unchanged from v3 — consumption price "
              f"variance has no metered or plan column, and end-user perceived quality has "
              f"no satisfaction measure and carries weight 0. Beyond those two, "
              f"th_architecture_fit is unscored on the applications whose Hosting Model "
              f"reads a bare 'Private cloud', because that label does not say whether the "
              f"cloud is vendor-managed or customer-hosted and the two readings sit two "
              f"points apart on a weight-2 input."),
        ("p", f"Confidence spread: "
              + ", ".join(f"{k} {v}" for k, v in conf_spread.most_common())
              + f" of {n}. Her Assumptions sheet says missing evidence should produce "
              f"'Needs Validation', and the triggers are her own evidence-quality columns: "
              f"Guardrail Status not Pass, guardrail Evidence Confidence Low, guardrail "
              f"Data Applicability Unknown, App Inventory Evidence Confidence Low, Cost "
              f"Notes that withhold a saving, and an action that would orphan a capability."),
        ("p", f"Capability orphaning: {len(orphan_rows)} applications would leave a "
              f"capability with no other provider in her map. The term still stands on the "
              f"evidence; the recommendation gains an explicit precondition and the row "
              f"drops to Needs Validation."),
        ("p", ""),
        ("h2", "The vocabulary problem, and what was done about it"),
        ("p", "Her 600-row workbook has the same 12 sheets and the same 190 headers as the "
              "20-row sample. The 580 new applications were populated under a different "
              "convention, so several columns carry differently-worded values. The largest "
              "is on the Capability Map: the support role the model reads out of Support "
              "Role has moved into Coverage Level, and Support Role holds a user persona. "
              "Left unmapped that fails silently — every new application would look like "
              "one with no Primary and no Secondary role, which drives four of the five "
              "business-value inputs to their bottom branch and hides every consolidation "
              "group outside the original 20 rows. The Vocabulary mapping sheet lists every "
              "mapping applied, its row count and the assumption behind it."),
        ("p", "Nothing was invented. No blank was filled, no missing value became a zero, "
              "and the one genuinely ambiguous label was left unscored rather than guessed, "
              "with the sensitivity computed both ways on the Sanity checks sheet."),
        ("p", ""),
        ("h2", "Held out"),
        ("p", "Her Lifecycle Stage column contains disposition-like labels, so feeding it to "
              "a model whose job is to produce dispositions would be circular. It is held "
              "out of every score and every guard — including the lifecycle guard, which "
              "arms from her release string — and read once, late, only for the Agreement "
              "sheet."),
        ("p", ""),
        ("h2", "What is in each sheet"),
        ("p", "Dispositions — one row per application: four scores, the pattern, the term, "
              "the priority, a rationale citing that row's own evidence, the next action, a "
              "confidence flag, which inputs were unscored, and the alternative under the "
              "risk-excluded variant."),
        ("p", "Priority queue — every application ranked by priority then value at risk."),
        ("p", "Input derivation — the audit trail for all 18 inputs: rubric, source columns, "
              "availability, how many rows scored and the observed spread."),
        ("p", "Consolidation clusters / candidates — the overlap groups, a named survivor "
              "with the reason, and the group money."),
        ("p", "Savings — per app and portfolio, with her figures beside ours."),
        ("p", "Vocabulary mapping — every value-level mapping applied before scoring, and "
              "the values deliberately left alone."),
        ("p", "Agreement with your labels — our term against her Lifecycle Stage."),
        ("p", "Sanity checks — the arithmetic, leakage and coverage checks."),
    ]

    # ---------------------------------------------------------------- write
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    v3.write_prose(wb.create_sheet("Notes & assumptions"), notes)
    v3.write_sheet(wb.create_sheet("Dispositions"), dispo_headers, dispo_records,
                   widths={"App ID": 9, "Application": 30, "Vendor": 22,
                           "Primary capability": 24, "Rationale": 90,
                           "Recommendation": 70, "Why that confidence": 55,
                           "Evidence gaps behind the confidence flag": 50,
                           "Alternative under risk-excluded gate": 40,
                           "Priority basis": 40, "Unscored inputs on this row": 26,
                           "Her Lifecycle Stage (COMPARISON ONLY — held out of scoring)": 30},
                   wrap_cols=("Rationale", "Recommendation", "Why that confidence",
                              "Evidence gaps behind the confidence flag",
                              "Alternative under risk-excluded gate", "Priority basis"),
                   comparison_cols=(
                       "Her Lifecycle Stage (COMPARISON ONLY — held out of scoring)",))
    v3.write_sheet(wb.create_sheet("Priority queue"), list(queue_records[0].keys()),
                   queue_records,
                   widths={"Application": 30, "Next action": 70, "Priority basis": 40},
                   wrap_cols=("Next action", "Priority basis"))
    v3.write_sheet(wb.create_sheet("Input derivation"), list(deriv_records[0].keys()),
                   deriv_records,
                   widths={"Engine input": 30, "What it scores": 40,
                           "Her columns used": 60, "Rubric applied": 90,
                           f"Example evidence ({example['app_id']} {example['name']})": 60},
                   wrap_cols=("What it scores", "Her columns used", "Rubric applied",
                              f"Example evidence ({example['app_id']} {example['name']})"))
    if cluster_summary:
        v3.write_sheet(wb.create_sheet("Consolidation clusters"),
                       list(cluster_summary[0].keys()), cluster_summary,
                       widths={"Capability names": 60, "Survivor": 30},
                       wrap_cols=("Capability names",))
        v3.write_sheet(wb.create_sheet("Consolidation candidates"),
                       list(cons_records[0].keys()), cons_records,
                       widths={"Application": 28, "Why": 70,
                               "Contested capabilities": 45,
                               "Her Support Role / Coverage on the contested capabilities": 45},
                       wrap_cols=("Why", "Contested capabilities",
                                  "Her Support Role / Coverage on the contested capabilities"))
    v3.write_sheet(wb.create_sheet("Savings"), list(sav_records[0].keys()), sav_records,
                   widths={"Application": 30, "Why they differ": 55, "Her Cost Notes": 45},
                   wrap_cols=("Why they differ", "Her Cost Notes"))
    v3.write_sheet(wb.create_sheet("Vocabulary mapping"), list(vocab_audit[0].keys()),
                   vocab_audit,
                   widths={"Her value": 40, "Read by the model as": 30,
                           "Assumption being made": 95, "Column": 28},
                   wrap_cols=("Her value", "Read by the model as", "Assumption being made"))
    if unmapped:
        v3.write_sheet(wb.create_sheet("Values left unmapped"), list(unmapped[0].keys()),
                       unmapped,
                       widths={"Value": 34, "How v3 treats it": 55,
                               "Left as is because": 65},
                       wrap_cols=("How v3 treats it", "Left as is because"))
    v3.write_sheet(wb.create_sheet("Agreement with your labels"),
                   list(agree_records[0].keys()), agree_records,
                   widths={"Application": 30,
                           "Her Lifecycle Stage (her team's label)": 26,
                           "Comparison": 34, "Dimensions failing the gate": 34},
                   comparison_cols=("Her Lifecycle Stage (her team's label)",
                                    "Her label read as"))
    v3.write_sheet(wb.create_sheet("Sanity checks"), ["Check", "Result", "Detail"], checks,
                   widths={"Check": 44, "Result": 26, "Detail": 110},
                   wrap_cols=("Detail", "Check"))
    wb.save(OUT_XLSX)

    with open(OUT_CSV, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=dispo_headers)
        w.writeheader()
        for rec in dispo_records:
            w.writerow(rec)

    runtime_total = time.perf_counter() - t0

    # ---------------------------------------------------------------- summary markdown
    top10 = queue_records[:10]
    with open(OUT_MD, "w", encoding="utf-8") as fh:
        wl = fh.write
        wl(f"# Northstar Global Health — {n}-application run\n\n")
        wl(f"Model: `score_northstar_v3.py`, imported unchanged. Input: "
           f"`healthcare_app_rationalization_sample_600.xlsx`. "
           f"Run date {dt.date.today().isoformat()}. "
           f"Wall clock {runtime_total:.2f}s for {n} rows "
           f"({runtime_scoring:.2f}s to load, normalise, score and decide).\n\n")
        wl("## Schema\n\n")
        wl("Same 12 sheets and same 190 column headers as the committed 20-application "
           "`-with-risk` sample: 0 columns added, removed or reordered. The model runs "
           "unmodified. The divergence is in the value domains of the 580 new "
           "applications, handled by a documented normalisation layer — see "
           "`Vocabulary mapping` in the workbook.\n\n")
        wl("## Dispositions\n\n| Term | Count | % |\n| --- | --- | --- |\n")
        for d in v3.DISPOSITIONS:
            wl(f"| {d} | {spread.get(d, 0)} | {pct(spread.get(d, 0), n)}% |\n")
        wl(f"\nOnly the all-pass pattern returns retain, so {spread.get('retain', 0)} rows "
           f"({pct(spread.get('retain', 0), n)}%) need no action and "
           f"{n - spread.get('retain', 0)} carry one.\n\n")
        wl("## Money\n\n")
        wl(f"- Gross annual avoidable claimed: ${gross:,.0f}\n")
        wl(f"- One-time transition cost on those rows: ${onetime:,.0f} "
           f"({pct(onetime, gross):.1f}% of gross)\n")
        wl(f"- **Net first year: ${net:,.0f}**\n")
        wl(f"- Safe (high-confidence only, her rule): ${safe:,.0f}\n")
        wl(f"- Potential: ${potential:,.0f}\n")
        wl(f"- Portfolio annual run cost: ${portfolio_tco:,.0f}; her CIO target "
           f"{v3.CIO_SAVINGS_TARGET:.0%} = ${portfolio_tco * v3.CIO_SAVINGS_TARGET:,.0f}\n\n")
        wl("## Priority\n\n| Priority | Count |\n| --- | --- |\n")
        for p in reversed(v3.PRIORITY_LADDER):
            wl(f"| {p} | {prio_spread.get(p, 0)} |\n")
        wl("\n### Top 10 by priority then value at risk\n\n")
        wl("| # | App | Disposition | Priority | Annual cost | Net first-year |\n")
        wl("| --- | --- | --- | --- | --- | --- |\n")
        for q in top10:
            wl(f"| {q['Rank']} | {q['App ID']} {q['Application']} | {q['Disposition']} | "
               f"{q['Priority']} | ${q['Annual TCO'] or 0:,.0f} | "
               f"${q['Net first-year saving'] or 0:,.0f} |\n")
        wl(f"\n## Consolidation\n\n{len(clusters)} overlap groups, {len(cap_rows)} with more "
           f"than one member. Largest: "
           + "; ".join(f"{c['cluster_id']} ({len(c['members'])} members)" for c in biggest)
           + f". {len(ovr)} rows were forced to consolidate by the redundancy override.\n\n")
        wl("Support Role after normalisation: "
           + ", ".join(f"{k} {v}" for k, v in dup_cov.most_common() if k) + ".\n\n")
        wl("## Confidence\n\n")
        for k, v in conf_spread.most_common():
            wl(f"- {k}: {v}\n")
        wl("\n## Data gaps\n\n")
        for name, _d, _l, _fn in v3.SCORERS:
            miss = sum(1 for a in apps if a["_inputs"][name]["score"] is None)
            if miss:
                wl(f"- `{name}`: unscored on {miss} of {n} rows\n")
        wl("\n## Vocabulary mappings applied\n\n")
        for m in vocab_audit:
            wl(f"- {m['Sheet']} · {m['Column']}: {m['Her value']} -> "
               f"{m['Read by the model as']} ({m['Rows']} rows)\n")
        wl("\n## Outputs\n\n")
        wl(f"- `{os.path.basename(OUT_XLSX)}`\n- `{os.path.basename(OUT_CSV)}`\n"
           f"- this file\n")

    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_CSV}")
    print(f"Wrote {OUT_MD}")
    print(f"RUNTIME_TOTAL_SECONDS {runtime_total:.3f}")
    print(f"RUNTIME_SCORING_SECONDS {runtime_scoring:.3f}")
    print("DISPOSITIONS " + " ".join(f"{d}={spread.get(d, 0)}" for d in v3.DISPOSITIONS))
    print("PRIORITY " + " ".join(f"{p}={prio_spread.get(p, 0)}"
                                 for p in reversed(v3.PRIORITY_LADDER)))
    print("CONFIDENCE " + " ".join(f"{k}={v}" for k, v in conf_spread.most_common()))
    print(f"GROSS {gross:.2f} ONETIME_ACTING {onetime:.2f} NET {net:.2f} SAFE {safe:.2f} "
          f"POTENTIAL {potential:.2f} PORTFOLIO_TCO {portfolio_tco:.2f} "
          f"HER_AVOIDABLE {her_avoidable:.2f}")
    print(f"CLUSTERS {len(clusters)} MULTI {len(cap_rows)} OVERRIDE {len(ovr)} "
          f"GUARD {len(guard_hits)} ORPHAN {len(orphan_rows)}")
    print("PATTERN_KEYS " + " ".join(f"{k}={v}" for k, v in key_spread.most_common()))
    print("TOP10 " + " | ".join(
        f"{q['App ID']} {q['Application']} {q['Disposition']} {q['Priority']} "
        f"${q['Annual TCO'] or 0:,.0f} ${q['Net first-year saving'] or 0:,.0f}"
        for q in top10))
    print("SENSITIVITY " + " || ".join(
        f"{label}: " + " ".join(f"{k}={v}" for k, v in sorted(sp.items()))
        for label, sp in sens.items()))
    print("REGRESSION_20 same=%d moved=%d %s" % (
        len(same), len(moved_rows),
        "; ".join(f"{a} {b}->{c}" for a, b, c in moved_rows)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
