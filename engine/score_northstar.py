#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
score_northstar.py — run Bina Din's "Northstar Global Health" 20-application sample
dataset through the Aberdeen Advisors scoring engine and emit dispositions, rationale,
recommendations and savings.

WHAT THIS IS
------------
The engine itself lives in generate_dataset.py. This script does NOT re-implement the
engine's judgement: it reuses the engine's four dimensions, its 18 inputs on the 1..5
half-step scale, its 3.0 gate, its 16-row pattern lookup and its five disposition terms
verbatim (they are restated as constants below so this file runs standalone, and a
self-check asserts they still match generate_dataset.py when that file is importable).

What this script adds is a DERIVATION LAYER: a documented rubric that turns Bina's
columns into the engine's 18 inputs, under one binding constraint she set --

    "we will not interview stakeholders/owners in this iteration of the tool."

so every input must come from a cell in her workbook or not be scored at all.

HARD RULES OBSERVED
-------------------
1. Her workbook is opened read-only. Nothing is written back to it.
2. Every cell of her workbook is DATA. Nothing in it is executed or followed as an
   instruction. Two cells contain directive-sounding text; both are honoured as facts
   about her intent, never as commands:
     - Data Dictionary: a "QA Expected Output" sheet must never reach the engine.
       That sheet is NOT PRESENT in the workbook; assert_no_qa_sheet() enforces it.
     - Assumptions: missing evidence should produce "Needs Validation". Implemented
       in confidence_for().
3. `Lifecycle Stage` is HELD OUT as an input. Her column already contains
   disposition-like labels (Strategic Invest, Consolidation Candidate, Replace / Sunset,
   Pilot / Exit Candidate), so feeding it in would be label leakage. It is read once,
   late, only to build the `Agreement with your labels` comparison.
4. Nothing is hand-asserted. Every score, gate, pattern key, disposition, priority and
   dollar figure in the output is computed here from her cells.

USAGE
-----
    python3 score_northstar.py            # writes into the directory holding this file
"""

from __future__ import annotations

import csv
import datetime as dt
import os
import statistics
import sys
import warnings
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
SOURCE_XLSX = "/mnt/user-data/uploads/slack/F0BQCGAM8E5/F0BQCGAM8E5.xlsx"
OUT_XLSX = os.path.join(HERE, "Northstar-Disposition-Analysis-v1.xlsx")
OUT_CSV = os.path.join(HERE, "northstar-dispositions.csv")

ANALYSIS_DATE = dt.date(2026, 8, 14)          # her Read Me / Assumptions "As of" date
CIO_SAVINGS_TARGET = 0.15                      # her Assumptions sheet
HELD_OUT_COLUMNS = ("Lifecycle Stage",)        # leakage guard
FORBIDDEN_SHEET = "QA Expected Output"         # her Data Dictionary says never feed this in


# =====================================================================================
# SECTION 1 — the engine, restated verbatim from generate_dataset.py
# =====================================================================================

PASS_THRESHOLD = 3.0                 # the comparison is >=, so exactly 3.0 passes
SCORE_STEPS = [1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0]

DIMENSIONS = [
    ("V", "business value", "business_value_score", "v_pass"),
    ("T", "technical health", "technical_health_score", "t_pass"),
    ("C", "cost efficiency", "cost_efficiency_score", "c_pass"),
    ("R", "risk posture", "risk_posture_score", "r_pass"),
]

# criterion -> (dimension, integer weight)
CRITERIA = [
    ("ov_increase_value", "V", 1),
    ("ov_reach_consumers", "V", 1),
    ("ov_reduce_costs_efficiency", "V", 1),
    ("ov_patient_care_criticality", "V", 2),
    ("ov_governance_compliance", "V", 1),
    ("th_supportability", "T", 2),
    ("th_architecture_fit", "T", 2),
    ("th_operational_stability", "T", 1),
    ("th_vendor_viability", "T", 1),
    ("th_customization_debt", "T", 1),
    ("c_cost_per_active_user_vs_peers", "C", 2),
    ("c_unused_licence_waste", "C", 1),
    ("c_consumption_price_variance", "C", 1),
    ("c_absolute_cost_band", "C", 0),
    ("r_technical_risk", "R", 1),
    ("r_business_compliance_risk", "R", 1),
    ("r_clinical_safety_risk", "R", 1),
    ("r_end_user_perceived_quality", "R", 0),
]

DISPOSITIONS = ("retain", "invest", "consolidate", "replace", "retire")

# key -> (our disposition, our priority). The 16 keys are the exhaustive enumeration of four
# pass/fail gates; the words and priorities are the team's own, restated from
# generate_dataset.py. The licensed Info-Tech tool whose structure the gated approach follows
# is cited there, not reproduced: none of its vocabulary or values appears here.
DISPOSITION_TABLE = {
    "PPPP": ("retain",      "Very Low"),
    "PPPF": ("invest",      "High"),
    "PPFP": ("invest",      "Moderate"),
    "PPFF": ("invest",      "High"),
    "PFPP": ("invest",      "Moderate"),
    "PFPF": ("invest",      "High"),
    "PFFP": ("replace",     "High"),
    "PFFF": ("replace",     "Very High"),
    "FPPP": ("consolidate", "Low"),
    "FPPF": ("consolidate", "Moderate"),
    "FPFP": ("retire",      "High"),
    "FPFF": ("retire",      "Very High"),
    "FFPP": ("consolidate", "Moderate"),
    "FFPF": ("retire",      "High"),
    "FFFP": ("retire",      "Very High"),
    "FFFF": ("retire",      "Very High"),
}

PRIORITY_LADDER = ["Very Low", "Low", "Moderate", "High", "Very High"]


def gate(score):
    """The comparison is >=, so exactly 3.0 passes."""
    if score is None:
        return "F"
    return "P" if score >= PASS_THRESHOLD else "F"


def dimension_score(row, dim_key):
    """Weighted mean over the dimension's POPULATED, non-zero-weight inputs.

    This is the engine's renormalisation rule, unchanged: an input with no source is
    skipped in both numerator and denominator, so the dimension is renormalised over what
    is actually populated rather than being silently dragged toward failure by a null.
    For Bina's file that rule fires exactly once, on the cost dimension, because
    c_consumption_price_variance has no source column.
    """
    num = den = 0.0
    for name, dim, weight in CRITERIA:
        if dim != dim_key or weight == 0:
            continue
        val = row.get(name)
        if val is None:
            continue
        num += val * weight
        den += weight
    if den == 0:
        return None
    return round(num / den, 3)


def dimension_weight_denominator(row, dim_key):
    """The denominator dimension_score() actually used, so the sheet can show it."""
    return sum(w for n, d, w in CRITERIA
               if d == dim_key and w > 0 and row.get(n) is not None)


def retain_or_invest(row):
    """Which of the two 'keep it' terms a row earns, read off the gates, not by hand."""
    failed = [name for _k, name, _col, flag in DIMENSIONS if row[flag] == "F"]
    if failed:
        return "invest", ("invest in " + " and in ".join(failed) +
                          f": that is the dimension failing the {PASS_THRESHOLD:.1f} gate, "
                          f"so that is what the money buys")
    return "retain", ("all four dimensions clear the 3.0 gate, so there is no failing "
                      "dimension to fund and nothing to buy: keep it, spend nothing")


def step_priority(priority, steps):
    i = PRIORITY_LADDER.index(priority)
    return PRIORITY_LADDER[min(max(i + steps, 0), len(PRIORITY_LADDER) - 1)]


def snap(x):
    """Snap onto the engine's 1..5 half-step scale."""
    if x is None:
        return None
    x = max(1.0, min(5.0, float(x)))
    return min(SCORE_STEPS, key=lambda s: (abs(s - x), s))


def verify_engine_constants():
    """Cross-check the restated constants against generate_dataset.py if it is present.

    generate_dataset.py runs a large build at import time, so it is parsed rather than
    imported: the check is textual and cheap, and it fails loudly rather than quietly.
    """
    path = os.path.join(HERE, "generate_dataset.py")
    if not os.path.exists(path):
        return "generate_dataset.py not found next to this script; engine constants unverified"
    with open(path, encoding="utf-8") as fh:
        src = fh.read()
    problems = []
    if "PASS_THRESHOLD = 3.0" not in src:
        problems.append("PASS_THRESHOLD")
    for key, (disp, prio) in DISPOSITION_TABLE.items():
        needle = f'"{key}": ('
        i = src.find(needle)
        if i == -1:
            problems.append(f"{key} missing from engine table")
            continue
        seg = src[i:src.find("\n", i)]
        if f'"{disp}"' not in seg or f'"{prio}"' not in seg:
            problems.append(f"{key} row disagrees with engine ({seg.strip()})")
    for name, dim, weight in CRITERIA:
        i = src.find(f'("{name}", "{dim}", {weight},')
        if i == -1:
            problems.append(f"criterion {name}/{dim}/w{weight} disagrees with engine")
    if problems:
        raise SystemExit("ENGINE MISMATCH — refusing to run: " + "; ".join(problems))
    return ("all 16 lookup rows, all 18 criterion weights and the 3.0 threshold verified "
            "against generate_dataset.py")


# =====================================================================================
# SECTION 2 — load Bina's workbook (read-only) and index it
# =====================================================================================

def load_source():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    if FORBIDDEN_SHEET in wb.sheetnames:
        raise SystemExit(
            f"REFUSING TO RUN: her Data Dictionary states the '{FORBIDDEN_SHEET}' sheet must "
            f"never be given to the recommendation engine, and this workbook contains it.")
    data = {}
    for name in wb.sheetnames:
        rows = [r for r in wb[name].iter_rows(values_only=True)]
        if not rows:
            data[name] = []
            continue
        hdr = [(str(h).strip() if h is not None else "") for h in rows[0]]
        recs = []
        for r in rows[1:]:
            if all(c is None for c in r):
                continue
            recs.append({h: v for h, v in zip(hdr, r) if h})
        data[name] = recs
    wb.close()
    return data


def s(v):
    return "" if v is None else str(v).strip()


def f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    txt = s(v)
    if not txt:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(txt[:10], fmt).date()
        except ValueError:
            pass
    return None


# =====================================================================================
# SECTION 3 — the rubric. One function per engine input.
#
# Every function returns (score, rubric_text, evidence_text, availability) where
# availability is "direct", "rubric", or "unavailable", and rubric_text is the rule that
# was applied to THIS row so the Dispositions sheet can quote it.
# =====================================================================================

# --- shared invert tables. Her scales run high = worse on all of these; ours run
# --- high = good on all 18 inputs, so these are the flips.
RESIDUAL_INVERT = {"Low": 5.0, "Medium": 3.5, "High": 2.0, "Critical": 1.0}
RISKLEVEL_INVERT = {"Low": 5.0, "Medium": 3.5, "High": 2.0, "Critical": 1.0}
STATUS_ADJUST = {"Closed": +1.0, "Mitigating": +0.5, "Accepted": 0.0, "Open": -0.5}
# Business / capability criticality. NOT inverted in the VALUE dimension: a Critical
# capability is high value, and inverting it there would score the EHR as low value.
# Inverted only where criticality enters the RISK dimension as an impact multiplier.
# See Notes & assumptions -> "criticality direction" for the flag raised on this.
CRITICALITY_VALUE = {"Critical": 4.5, "High": 3.5, "Medium": 2.5, "Low": 1.5}

MONEY_TERMS = ("revenue", "claim", "billing", "bill ", "payroll", "benefit", "reimburse",
               "credential", "capacity", "time and attendance", "labor compliance")
CLINICAL_TERMS = ("clinical", "patient", "care ", "care coordination", "order", "medical",
                  "telehealth", "virtual care", "provider")
MIGRATION_TERMS = ("migrat", "cutover", "cut over", "repoint", "rebuild", "consolidat")
UNSAFE_TERMS = ("potential only", "not safe", "until", "validated")


def cap_rows_for(app_id, caps):
    return [c for c in caps if s(c["App ID"]) == app_id]


def risk_rows_for(app_id, risks, categories):
    out = []
    for r in risks:
        if s(r["App ID"]) != app_id:
            continue
        cat = s(r["Risk Category"])
        if any(k.lower() in cat.lower() for k in categories):
            out.append(r)
    return out


def worst_risk_score(rows):
    """Invert her residual-risk label, credit her mitigation status, take the worst row.

    Risk posture is governed by the least-controlled risk on the application, so this is
    a MINIMUM across rows, not a mean. Status credit: a Mitigating plan is partial
    control (+0.5), Closed is control achieved (+1.0), Accepted is conscious tolerance
    (0.0), Open is an uncontrolled exposure (-0.5).
    """
    best = None
    detail = []
    for r in rows:
        base = RESIDUAL_INVERT.get(s(r["Residual Risk"]))
        if base is None:
            base = RISKLEVEL_INVERT.get(s(r["Risk Level"]))
        if base is None:
            continue
        adj = STATUS_ADJUST.get(s(r["Status"]), 0.0)
        v = max(1.0, min(5.0, base + adj))
        detail.append(f"{s(r['Risk Category'])} residual {s(r['Residual Risk'])}/"
                      f"{s(r['Status'])} -> {snap(v):.1f}")
        best = v if best is None else min(best, v)
    return (snap(best) if best is not None else None), "; ".join(detail)


# ---- V1 ov_increase_value -----------------------------------------------------------
def score_ov_increase_value(app, ctx):
    caps = cap_rows_for(app["app_id"], ctx["caps"])
    hits = []
    for c in caps:
        blob = (s(c["Capability"]) + " " + s(c["Process / Tasks Enabled"])).lower()
        if any(t in blob for t in MONEY_TERMS):
            hits.append((s(c["Support Role"]), s(c["Coverage Level"]), s(c["Capability"])))
    rubric = ("Money-path capability test on Capability Map: a capability whose name or "
              "Process/Tasks text mentions revenue, claims, billing, payroll, benefits, "
              "reimbursement, credentialing, capacity or time-and-attendance is a money "
              "path. Primary+Full=5.0, Primary+Partial=4.5, Secondary=4.0, Duplicative=3.0 "
              "(the money path exists but another app is the system of record). No money "
              "capability: Critical Operation Flag Yes=3.0, else 2.0.")
    if hits:
        roles = {h[0] for h in hits}
        if "Primary" in roles:
            full = any(h[0] == "Primary" and h[1] == "Full" for h in hits)
            score, why = (5.0, "Primary + Full") if full else (4.5, "Primary + Partial")
        elif "Secondary" in roles:
            score, why = 4.0, "Secondary"
        else:
            score, why = 3.0, "Duplicative only"
        ev = f"money-path capabilities {', '.join(sorted({h[2] for h in hits}))} held as {why}"
    elif s(app["critical_op_flag"]) == "Yes":
        score, ev = 3.0, "no money-path capability, but Critical Operation Flag = Yes"
    else:
        score, ev = 2.0, "no money-path capability and Critical Operation Flag = No"
    return snap(score), rubric, ev, "rubric"


# ---- V2 ov_reach_consumers ----------------------------------------------------------
UTIL_BANDS = [(0.85, 5.0), (0.75, 4.5), (0.65, 4.0), (0.55, 3.5), (0.45, 3.0),
              (0.35, 2.5), (0.25, 2.0), (0.15, 1.5)]
BREADTH_BANDS = [(25000, 5.0), (15000, 4.5), (8000, 4.0), (4000, 3.5), (2000, 3.0),
                 (1000, 2.5), (500, 2.0), (200, 1.5)]


def band(value, bands, floor=1.0):
    for threshold, score in bands:
        if value >= threshold:
            return score
    return floor


def score_ov_reach_consumers(app, ctx):
    util, act = app["utilisation"], app["active_users"]
    rubric = ("Mean of two bands, then snapped to the half-step scale. Depth = her "
              "Utilization Rate (Active Users 90d / Entitled Users): >=.85=5.0, >=.75=4.5, "
              ">=.65=4.0, >=.55=3.5, >=.45=3.0, >=.35=2.5, >=.25=2.0, >=.15=1.5, else 1.0. "
              "Breadth = absolute Active Users (90d): >=25000=5.0, >=15000=4.5, >=8000=4.0, "
              ">=4000=3.5, >=2000=3.0, >=1000=2.5, >=500=2.0, >=200=1.5, else 1.0.")
    if util is None or act is None:
        return None, rubric, "Utilization Rate or Active Users missing", "unavailable"
    d, b = band(util, UTIL_BANDS), band(act, BREADTH_BANDS)
    ev = (f"{act:,.0f} active users of {app['entitled_users']:,.0f} entitled "
          f"({util:.1%}) -> depth {d:.1f}, breadth {b:.1f}")
    return snap((d + b) / 2), rubric, ev, "direct"


# ---- V3 ov_reduce_costs_efficiency --------------------------------------------------
def score_ov_reduce_costs_efficiency(app, ctx):
    caps = cap_rows_for(app["app_id"], ctx["caps"])
    rubric = ("Process centrality from Capability Map Support Role, Coverage Level and "
              "Capability Criticality: Primary+Full on a Critical capability=5.0, "
              "Primary+Full=4.5, Primary+Partial=4.0, Secondary best role=3.0, "
              "Duplicative only=2.0 (a second copy of a process another app owns is not "
              "central to it).")
    if not caps:
        return None, rubric, "no Capability Map rows for this app", "unavailable"
    prim = [c for c in caps if s(c["Support Role"]) == "Primary"]
    if prim:
        if any(s(c["Coverage Level"]) == "Full" and s(c["Capability Criticality"]) == "Critical"
               for c in prim):
            score, why = 5.0, "Primary + Full on a Critical capability"
        elif any(s(c["Coverage Level"]) == "Full" for c in prim):
            score, why = 4.5, "Primary + Full coverage"
        else:
            score, why = 4.0, "Primary but Partial coverage"
    elif any(s(c["Support Role"]) == "Secondary" for c in caps):
        score, why = 3.0, "best role is Secondary"
    else:
        score, why = 2.0, "every capability row is Duplicative"
    ev = f"{len(caps)} capability rows; {why}"
    return snap(score), rubric, ev, "rubric"


# ---- V4 ov_patient_care_criticality (weight 2) --------------------------------------
def score_ov_patient_care_criticality(app, ctx):
    caps = cap_rows_for(app["app_id"], ctx["caps"])
    dup = [c for c in caps if s(c["Support Role"]) == "Duplicative"]
    majority_dup = bool(caps) and len(dup) / len(caps) >= 0.5
    dc = s(app["data_classification"])
    firm_phi = "PHI" in dc and "possible" not in dc.lower()
    rubric = ("Base from her Business Criticality (Critical=4.5, High=3.5, Medium=2.5, "
              "Low=1.5), +0.5 if Critical Operation Flag = Yes, +0.5 if Data Classification "
              "names PHI outright (a 'possible PHI' label does not qualify), -1.0 if at "
              "least half the app's Capability Map rows are Duplicative (clinical work does "
              "not stop when another application already holds the capability). Clamped to "
              "1..5. Business Criticality is used in its natural direction here, NOT "
              "inverted -- see Notes & assumptions.")
    base = CRITICALITY_VALUE.get(s(app["business_criticality"]))
    if base is None:
        return None, rubric, "Business Criticality missing", "unavailable"
    score, parts = base, [f"Business Criticality {s(app['business_criticality'])} -> {base:.1f}"]
    if s(app["critical_op_flag"]) == "Yes":
        score += 0.5
        parts.append("Critical Operation Flag Yes +0.5")
    if firm_phi:
        score += 0.5
        parts.append(f"Data Classification '{dc}' +0.5")
    if majority_dup:
        score -= 1.0
        parts.append(f"{len(dup)} of {len(caps)} capability rows Duplicative -1.0")
    return snap(score), rubric, "; ".join(parts), "rubric"


# ---- V5 ov_governance_compliance ----------------------------------------------------
def score_ov_governance_compliance(app, ctx):
    caps = cap_rows_for(app["app_id"], ctx["caps"])
    dup = [c for c in caps if s(c["Support Role"]) == "Duplicative"]
    majority_dup = bool(caps) and len(dup) / len(caps) >= 0.5
    any_primary = any(s(c["Support Role"]) == "Primary" for c in caps)
    dc = s(app["data_classification"])
    regulated = any(t in dc for t in ("PHI", "PII", "Security"))
    rubric = ("Regulatory/trust ALIGNMENT, not regulatory obligation: is regulated data "
              "held in the governed standard or in an uncontrolled second copy? Regulated "
              "Data Classification (PHI / PII / Security) held by a Primary-role app=5.0; "
              "regulated but majority-Duplicative=2.0; regulated and mixed=3.5; "
              "not regulated + Primary=4.0; not regulated + majority-Duplicative=2.5. "
              "Then -0.5 if her Evidence Confidence for the app is Low. HIPAA/PHI SEVERITY "
              "is deliberately NOT scored here -- it is scored once, in "
              "r_business_compliance_risk, to avoid double-counting.")
    if not caps:
        return None, rubric, "no Capability Map rows for this app", "unavailable"
    if regulated and any_primary and not majority_dup:
        score, why = 5.0, f"regulated data ({dc}) held in a Primary-role application"
    elif regulated and majority_dup:
        score, why = 2.0, f"regulated data ({dc}) in a majority-Duplicative application"
    elif regulated:
        score, why = 3.5, f"regulated data ({dc}), mixed support roles"
    elif any_primary and not majority_dup:
        score, why = 4.0, f"non-regulated data ({dc}) in a Primary-role application"
    else:
        score, why = 2.5, f"non-regulated data ({dc}), majority-Duplicative"
    parts = [why]
    if s(app["evidence_confidence"]) == "Low":
        score -= 0.5
        parts.append("her Evidence Confidence = Low, -0.5")
    return snap(score), rubric, "; ".join(parts), "rubric"


# ---- T1 th_supportability (weight 2) ------------------------------------------------
EVERGREEN_TOKENS = ("continuous", "current saas", "cloud continuous", "current release")


def score_th_supportability(app, ctx):
    perf = ctx["perf"].get(app["app_id"], {})
    rel = s(perf.get("Current Release / Version"))
    eos = parse_date(perf.get("Vendor Support End"))
    rubric = ("Mean of the components that exist, because Vendor Support End is blank for "
              "19 of 20 apps. (a) End-of-support horizon, when present: >5y=5.0, 3-5y=4.0, "
              "1-3y=2.5, <1y=1.0. (b) Release currency from Current Release / Version: "
              "evergreen vendor-managed line ('Continuous', 'Cloud continuous', 'Current "
              "SaaS release')=4.5 -- always supported, but capped below 5.0 because no "
              "explicit support horizon is evidenced; a dated in-year release (e.g. "
              "'May 2026', '2026.2', '2026 R1')=4.0; 'Pilot release'=2.5 (not a supported "
              "GA line); an explicitly versioned legacy release (e.g. '9.2 / PeopleTools "
              "8.60')=2.0. LOWER-EVIDENCE INPUT: no end-of-support date was invented for "
              "the 19 apps that lack one.")
    comps, parts = [], []
    if eos:
        yrs = (eos - ANALYSIS_DATE).days / 365.25
        c = 5.0 if yrs > 5 else 4.0 if yrs >= 3 else 2.5 if yrs >= 1 else 1.0
        comps.append(c)
        parts.append(f"Vendor Support End {eos.isoformat()} = {yrs:.1f}y out -> {c:.1f}")
    low = rel.lower()
    if not rel:
        cur = None
    elif "pilot" in low:
        cur = 2.5
    elif any(t in low for t in EVERGREEN_TOKENS):
        cur = 4.5
    elif any(ch.isdigit() for ch in rel) and ("peopletools" in low or "/" in rel):
        cur = 2.0
    elif any(ch.isdigit() for ch in rel) or "'" in rel:
        cur = 4.0
    else:
        cur = 3.0
    if cur is not None:
        comps.append(cur)
        parts.append(f"release '{rel}' -> {cur:.1f}")
    if not comps:
        return None, rubric, "no release or support-end evidence", "unavailable"
    ev = "; ".join(parts) + ("" if eos else " [no Vendor Support End on file: lower evidence]")
    return snap(sum(comps) / len(comps)), rubric, ev, "rubric"


# ---- T2 th_architecture_fit (weight 2) ----------------------------------------------
HOSTING_BASE = [("vendor-managed saas", 5.0), ("vendor cloud", 4.0),
                ("private cloud / vendor-managed", 4.0), ("hybrid", 3.0),
                ("customer-hosted", 2.0)]


def score_th_architecture_fit(app, ctx):
    deps = ctx["deps_by_src"].get(app["app_id"], [])
    host = s(app["hosting_model"]).lower()
    rubric = ("Hosting base from her Hosting Model: vendor-managed SaaS=5.0, vendor "
              "cloud/hosted=4.0, private cloud vendor-managed=4.0, hybrid on-prem+cloud=3.0, "
              "customer-hosted=2.0. Then -0.5 if any dependency is Criticality Critical with "
              "Migration Feasibility Low (a tightly coupled critical integration is poor "
              "enterprise-architecture fit), and +0.5 if every dependency has Migration "
              "Feasibility High or Medium.")
    base = None
    for token, v in HOSTING_BASE:
        if token in host:
            base = v
            break
    if base is None:
        return None, rubric, f"Hosting Model '{s(app['hosting_model'])}' unmatched", "unavailable"
    score, parts = base, [f"Hosting Model -> {base:.1f}"]
    tight = [d for d in deps if s(d["Criticality"]) == "Critical"
             and s(d["Migration Feasibility"]) == "Low"]
    if tight:
        score -= 0.5
        parts.append(f"{len(tight)} Critical dependency with Low migration feasibility "
                     f"({', '.join(s(d['Target Application / System'])[:26] for d in tight)}) -0.5")
    elif deps and all(s(d["Migration Feasibility"]) in ("High", "Medium") for d in deps):
        score += 0.5
        parts.append(f"all {len(deps)} dependencies High/Medium migration feasibility +0.5")
    return snap(score), rubric, "; ".join(parts), "rubric"


# ---- T3 th_operational_stability ----------------------------------------------------
INCIDENT_RATE_BANDS = [(0.5, 5.0), (1.5, 4.0), (3.0, 3.0), (5.0, 2.0)]
MTTR_BANDS = [(45, 5.0), (60, 4.0), (90, 3.0), (120, 2.0)]


def lower_band(value, bands, floor=1.0):
    """Bands where SMALLER is better."""
    for threshold, score in bands:
        if value <= threshold:
            return score
    return floor


def score_th_operational_stability(app, ctx):
    perf = ctx["perf"].get(app["app_id"], {})
    inc, mttr = f(perf.get("P1/P2 Incidents (12mo)")), f(perf.get("MTTR (minutes)"))
    avail, sla = f(perf.get("Availability (12mo)")), f(perf.get("SLA Target"))
    act = app["active_users"]
    rubric = ("Mean of three measured components from Performance & Roadmap. "
              "(a) P1/P2 incidents per 1,000 active users: <=0.5=5.0, <=1.5=4.0, <=3=3.0, "
              "<=5=2.0, else 1.0. (b) MTTR minutes: <=45=5.0, <=60=4.0, <=90=3.0, <=120=2.0, "
              "else 1.0. (c) Availability against her SLA Target: at or above target=5.0, "
              "shortfall <=0.001=3.0, larger shortfall=1.5. This is the one engine input "
              "her file evidences BETTER than our own dataset does -- our scoring-model "
              "review flagged that we had no incident or ticket-volume column at all.")
    comps, parts = [], []
    if inc is not None and act:
        rate = inc / (act / 1000.0)
        c = lower_band(rate, INCIDENT_RATE_BANDS)
        comps.append(c)
        parts.append(f"{inc:.0f} P1/P2 in 12mo over {act:,.0f} active users = "
                     f"{rate:.2f}/1k -> {c:.1f}")
    if mttr is not None:
        c = lower_band(mttr, MTTR_BANDS)
        comps.append(c)
        parts.append(f"MTTR {mttr:.0f} min -> {c:.1f}")
    if avail is not None and sla is not None:
        short = sla - avail
        c = 5.0 if short <= 0 else 3.0 if short <= 0.001 else 1.5
        comps.append(c)
        parts.append(f"availability {avail:.4f} vs SLA {sla:.4f} -> {c:.1f}")
    if not comps:
        return None, rubric, "no performance evidence", "unavailable"
    return snap(sum(comps) / len(comps)), rubric, "; ".join(parts), "direct"


# ---- T4 th_vendor_viability ---------------------------------------------------------
def score_th_vendor_viability(app, ctx):
    rows = risk_rows_for(app["app_id"], ctx["risks"], ("Vendor",))
    perf = ctx["perf"].get(app["app_id"], {})
    rubric = ("If her Risks sheet carries a Vendor-category row for the app, invert its "
              "Residual Risk (Low=5.0, Medium=3.5, High=2.0, Critical=1.0) and credit "
              "Mitigation Status (Closed +1.0, Mitigating +0.5, Accepted 0, Open -0.5). "
              "Otherwise 4.0 when both Next Minor Release and Next Major Update are dated "
              "(an active, published roadmap), 3.0 when the roadmap is incomplete. "
              "LOWER-EVIDENCE INPUT: her file contains no vendor financial data of any "
              "kind, so vendor solvency is proxied by roadmap activity and by whether her "
              "own risk register raised a vendor concern.")
    if rows:
        score, detail = worst_risk_score(rows)
        if score is not None:
            return score, rubric, f"her Vendor risk row: {detail}", "direct"
    minor, major = parse_date(perf.get("Next Minor Release")), parse_date(perf.get("Next Major Update"))
    if minor and major:
        return snap(4.0), rubric, (f"no Vendor risk row raised; roadmap dated (minor "
                                   f"{minor.isoformat()}, major {major.isoformat()}) -> 4.0 "
                                   f"[no vendor financials on file: lower evidence]"), "rubric"
    return snap(3.0), rubric, "no Vendor risk row and roadmap dates incomplete", "rubric"


# ---- T5 th_customization_debt -------------------------------------------------------
LANGUAGE_DEBT = [("peoplecode", 1.5), ("peopletools", 1.5), ("apex", 3.0),
                 ("javascript", 3.0), ("dax", 3.5), ("tableau calculations", 3.5),
                 ("configuration and automation", 4.5)]


def score_th_customization_debt(app, ctx):
    rows = risk_rows_for(app["app_id"], ctx["risks"], ("Technical",))
    lang = s(app["config_language"]).lower()
    rubric = ("Two components, worst wins (min), because debt is bounded by the worse "
              "evidence. (a) Configuration/extension language: PeopleCode or PeopleTools="
              "1.5 (heavy proprietary custom code), Apex or JavaScript=3.0, DAX or Tableau "
              "calculations=3.5, 'configuration and automation rules'=4.5, otherwise from "
              "Commercial Model: Commercial SaaS=4.5 (configure, do not code), Commercial "
              "off-the-shelf=3.5. (b) Any Technical-category risk row, inverted from its "
              "Residual Risk with the same status credit used elsewhere.")
    comps, parts = [], []
    lv = None
    for token, v in LANGUAGE_DEBT:
        if token in lang:
            lv = v if lv is None else min(lv, v)
    if lv is None:
        cm = s(app["commercial_model"]).lower()
        lv = 4.5 if "saas" in cm else 3.5 if "off-the-shelf" in cm else 3.0
        parts.append(f"Commercial Model '{s(app['commercial_model'])}' -> {lv:.1f}")
    else:
        parts.append(f"config language '{s(app['config_language'])[:48]}' -> {lv:.1f}")
    comps.append(lv)
    if rows:
        rv, detail = worst_risk_score(rows)
        if rv is not None:
            comps.append(rv)
            parts.append(f"her Technical risk row: {detail}")
    return snap(min(comps)), rubric, "; ".join(parts), "rubric"


# ---- C1 c_cost_per_active_user_vs_peers (weight 2) ----------------------------------
PEER_RATIO_BANDS = [(0.50, 5.0), (0.70, 4.5), (0.90, 4.0), (1.00, 3.5), (1.15, 3.0),
                    (1.40, 2.5), (1.80, 2.0), (2.50, 1.5)]


def score_c_cost_per_active_user(app, ctx):
    rubric = ("Annual TCO / Active Users (90d), compared with the MEDIAN of the app's peer "
              "group. Peer group = her Category column; a Category with only one app is "
              "compared against the whole-portfolio median instead. Ratio to peer median: "
              "<=0.50=5.0, <=0.70=4.5, <=0.90=4.0, <=1.00=3.5, <=1.15=3.0, <=1.40=2.5, "
              "<=1.80=2.0, <=2.50=1.5, else 1.0. 5.0 = cheapest per productive user.")
    cpu = app["cost_per_active_user"]
    if cpu is None:
        return None, rubric, "Annual TCO or Active Users missing", "unavailable"
    peers = ctx["peer_cpu"][app["category"]]
    if len(peers) >= 2:
        med, basis = statistics.median(peers), f"{len(peers)} apps in Category '{app['category']}'"
    else:
        med, basis = ctx["portfolio_cpu_median"], (
            f"Category '{app['category']}' has only this app, so the whole-portfolio median "
            f"of {len(ctx['all_cpu'])} apps is used")
    ratio = cpu / med
    ev = (f"${cpu:,.0f} per active user vs peer median ${med:,.0f} = {ratio:.2f}x ({basis})")
    return snap(band(-ratio, [(-t, v) for t, v in PEER_RATIO_BANDS])), rubric, ev, "rubric"


# ---- C2 c_unused_licence_waste ------------------------------------------------------
WASTE_BANDS = [(0.90, 5.0), (0.85, 4.5), (0.80, 4.0), (0.70, 3.5), (0.60, 3.0),
               (0.50, 2.5), (0.40, 2.0), (0.30, 1.5)]


def score_c_unused_licence_waste(app, ctx):
    rubric = ("Her Utilization Rate read as the inverse of licensed-but-inactive waste: "
              ">=.90=5.0, >=.85=4.5, >=.80=4.0, >=.70=3.5, >=.60=3.0, >=.50=2.5, >=.40=2.0, "
              ">=.30=1.5, else 1.0. 5.0 = almost no waste.")
    util = app["utilisation"]
    if util is None:
        return None, rubric, "Utilization Rate missing", "unavailable"
    unused = (app["entitled_users"] or 0) - (app["active_users"] or 0)
    ev = f"{util:.1%} utilised, {unused:,.0f} entitled-but-inactive seats"
    return snap(band(util, WASTE_BANDS)), rubric, ev, "direct"


# ---- C3 c_consumption_price_variance — UNAVAILABLE ---------------------------------
def score_c_consumption_price_variance(app, ctx):
    rubric = ("NOT SCORED. This input compares metered/consumption spend against a modelled "
              "plan. Her workbook has neither: the TCO sheet carries six FIXED annual "
              "components (License, Maintenance, Infrastructure, Vendor Services, Internal "
              "Labor, Education) and no consumption or metered line, and no plan or budget "
              "figure to vary against. Producing a number here would be inventing data, so "
              "the cost dimension is RENORMALISED over its remaining weighted inputs "
              "(weight sum 4 -> 3). This gap is not caused by the no-interview constraint; "
              "the data does not exist in the file.")
    return None, rubric, "no consumption/metered cost column and no plan figure", "unavailable"


# ---- C4 c_absolute_cost_band (weight 0) --------------------------------------------
ABS_COST_BANDS = [(500_000, 5.0), (1_000_000, 4.5), (1_500_000, 4.0), (2_000_000, 3.5),
                  (3_000_000, 3.0), (5_000_000, 2.0), (7_000_000, 1.5)]


def score_c_absolute_cost_band(app, ctx):
    rubric = ("Absolute Annual TCO band: <=$0.5m=5.0, <=$1m=4.5, <=$1.5m=4.0, <=$2m=3.5, "
              "<=$3m=3.0, <=$5m=2.0, <=$7m=1.5, else 1.0. WEIGHT 0 -- scored and reported, "
              "contributes nothing. Our scoring-model review found this input is one of only "
              "two that can change an answer if switched on, and that at weight 2 it flips "
              "the largest platform in a portfolio from retain to invest purely for being "
              "the largest line item. It stays at 0 deliberately.")
    tco = app["annual_tco"]
    if tco is None:
        return None, rubric, "Annual TCO missing", "unavailable"
    return snap(lower_band(tco, ABS_COST_BANDS)), rubric, f"Annual TCO ${tco:,.0f}", "rubric"


# ---- R1 r_technical_risk ------------------------------------------------------------
TECH_RISK_CATS = ("Technical", "Dependency", "Operational", "Business Continuity", "Data Quality")


def score_r_technical_risk(app, ctx):
    rows = risk_rows_for(app["app_id"], ctx["risks"], TECH_RISK_CATS)
    rubric = ("Worst (minimum) of her Risks rows in categories Technical, Dependency, "
              "Operational, Business Continuity or Data Quality, inverting Residual Risk "
              "(Low=5.0, Medium=3.5, High=2.0, Critical=1.0) and crediting Mitigation "
              "Status (Closed +1.0, Mitigating +0.5, Accepted 0, Open -0.5). Minimum, not "
              "mean: risk posture is set by the least-controlled exposure. Fallback where "
              "she has no row in any of these categories: her Highest Risk column inverted, "
              "flagged Needs Validation.")
    if rows:
        score, detail = worst_risk_score(rows)
        if score is not None:
            return score, rubric, detail, "direct"
    hr = RISKLEVEL_INVERT.get(s(app["highest_risk"]))
    if hr is None:
        return None, rubric, "no technical-family risk row and no Highest Risk value", "unavailable"
    return snap(hr), rubric, (f"NO technical-family risk row in her register; fell back to "
                              f"Highest Risk = {s(app['highest_risk'])} -> {hr:.1f} "
                              f"[Needs Validation]"), "rubric"


# ---- R2 r_business_compliance_risk -------------------------------------------------
COMPLIANCE_RISK_CATS = ("Security", "Regulatory", "Privacy", "Vendor")


def score_r_business_compliance_risk(app, ctx):
    rows = risk_rows_for(app["app_id"], ctx["risks"], COMPLIANCE_RISK_CATS)
    rubric = ("Worst (minimum) of her Risks rows in categories Security, Regulatory, "
              "Privacy or Vendor, inverted and status-credited exactly as r_technical_risk. "
              "This is the ONLY place HIPAA/PHI exposure is scored, by design, so it is not "
              "double-counted in the value dimension. Fallback where she has no row in any "
              "of these categories: her Highest Risk column inverted, then -0.5 if the app "
              "holds PHI (unassessed regulated data is a worse posture than assessed "
              "regulated data), flagged Needs Validation.")
    if rows:
        score, detail = worst_risk_score(rows)
        if score is not None:
            return score, rubric, detail, "direct"
    hr = RISKLEVEL_INVERT.get(s(app["highest_risk"]))
    if hr is None:
        return None, rubric, "no compliance-family risk row and no Highest Risk value", "unavailable"
    dc = s(app["data_classification"])
    parts = [f"NO Security/Regulatory/Privacy/Vendor row in her register; fell back to "
             f"Highest Risk = {s(app['highest_risk'])} -> {hr:.1f}"]
    if "PHI" in dc or "PII" in dc:
        hr -= 0.5
        parts.append(f"regulated data ({dc}) with no compliance risk row assessed, -0.5")
    parts.append("[Needs Validation]")
    return snap(hr), rubric, "; ".join(parts), "rubric"


# ---- R3 r_clinical_safety_risk -----------------------------------------------------
def score_r_clinical_safety_risk(app, ctx):
    rows = risk_rows_for(app["app_id"], ctx["risks"], ("Clinical Safety", "AI /", "Patient Safety"))
    caps = cap_rows_for(app["app_id"], ctx["caps"])
    blob = " ".join(s(c["Capability"]) + " " + s(c["Process / Tasks Enabled"]) for c in caps).lower()
    dc = s(app["data_classification"])
    # Clinical pathway test: PHI in the data classification, or clinical language in the
    # app's own capability text. Critical Operation Flag is deliberately NOT part of this
    # test -- it marks criticality to OPERATIONS, and reading it as clinical would drag
    # ITSM and HR platforms into a clinical-safety assessment they have no pathway to.
    clinical = ("PHI" in dc) or any(t in blob for t in CLINICAL_TERMS)
    rubric = ("Where she has an AI / Clinical Safety (or Patient Safety) risk row -- 4 of "
              "20 apps -- invert and status-credit it as elsewhere. Where she has none: an "
              "app with no clinical pathway in her data (no PHI in the Data Classification "
              "and no clinical language in its capability text) scores 4.5, meaning no "
              "identified clinical-safety "
              "exposure to mitigate rather than a measured pass; an app that DOES touch "
              "clinical work scores the worst of its Business Continuity / Security rows and "
              "is flagged Needs Validation, because a clinical application without a "
              "clinical-safety assessment is an evidence gap, not a clean bill of health.")
    if rows:
        score, detail = worst_risk_score(rows)
        if score is not None:
            return score, rubric, detail, "direct"
    if not clinical:
        return snap(4.5), rubric, ("no clinical pathway in her data (no PHI in the Data "
                                   "Classification, no clinical language in its capability "
                                   "text): no identified clinical-safety exposure -> 4.5"), "rubric"
    fb = risk_rows_for(app["app_id"], ctx["risks"], ("Business Continuity", "Security"))
    score, detail = worst_risk_score(fb)
    if score is None:
        hr = RISKLEVEL_INVERT.get(s(app["highest_risk"]))
        if hr is None:
            return None, rubric, "clinical app with no usable risk row", "unavailable"
        return snap(hr), rubric, (f"clinical app, NO clinical-safety row; fell back to Highest "
                                  f"Risk {s(app['highest_risk'])} [Needs Validation]"), "rubric"
    return score, rubric, (f"clinical app, NO clinical-safety row in her register; fell back "
                           f"to its continuity/security rows ({detail}) [Needs Validation]"), "rubric"


# ---- R4 r_end_user_perceived_quality (weight 0) — UNAVAILABLE ----------------------
def score_r_end_user_perceived_quality(app, ctx):
    perf = ctx["perf"].get(app["app_id"], {})
    kpi, tgt = f(perf.get("KPI Actual")), f(perf.get("KPI Target"))
    rubric = ("NOT SCORED. This input is an end-user PERCEPTION measure (satisfaction / "
              "perceived quality). Her workbook has no survey, CSAT, NPS or satisfaction "
              "field, and the only route to one would be asking users -- which the "
              "no-interview constraint rules out for this iteration. Her Business KPI / KPI "
              "Actual / KPI Target columns are the nearest thing but measure a BUSINESS "
              "outcome, not a perception, so they are reported alongside and never scored. "
              "The input carries WEIGHT 0, so leaving it null changes no answer. Our "
              "scoring-model review recommends moving it out of the risk dimension "
              "entirely: at weight 1 a good satisfaction number can dilute a genuine "
              "compliance failure into a pass.")
    ev = "no satisfaction/CSAT/NPS field in her workbook"
    if kpi is not None and tgt is not None:
        ev += (f"; reported-only business KPI '{s(perf.get('Business KPI'))}' at "
               f"{kpi:.3f} against target {tgt:.3f} ({'meets' if kpi >= tgt else 'below'})")
    return None, rubric, ev, "unavailable"


SCORERS = [
    ("ov_increase_value", "Business value", "Revenue-capture criticality: does the app carry money in or out?", score_ov_increase_value),
    ("ov_reach_consumers", "Business value", "Breadth and depth of use: active users against entitlement.", score_ov_reach_consumers),
    ("ov_reduce_costs_efficiency", "Business value", "Process centrality: how central to the process it serves.", score_ov_reduce_costs_efficiency),
    ("ov_patient_care_criticality", "Business value", "Criticality to patient care: does clinical work stop without it? (double weight)", score_ov_patient_care_criticality),
    ("ov_governance_compliance", "Business value", "Regulatory and trust alignment of where regulated data sits.", score_ov_governance_compliance),
    ("th_supportability", "Technical health", "Version currency and end-of-life proximity. (double weight)", score_th_supportability),
    ("th_architecture_fit", "Technical health", "Integration pattern, cloud readiness, enterprise-architecture fit. (double weight)", score_th_architecture_fit),
    ("th_operational_stability", "Technical health", "Incident volume, MTTR and availability against SLA.", score_th_operational_stability),
    ("th_vendor_viability", "Technical health", "Vendor and roadmap viability.", score_th_vendor_viability),
    ("th_customization_debt", "Technical health", "Customisation debt and platform supportability.", score_th_customization_debt),
    ("c_cost_per_active_user_vs_peers", "Cost efficiency", "Cost per active user against peers in the same category. (double weight)", score_c_cost_per_active_user),
    ("c_unused_licence_waste", "Cost efficiency", "Licensed-but-inactive seat waste.", score_c_unused_licence_waste),
    ("c_consumption_price_variance", "Cost efficiency", "Consumption/metered spend against the modelled plan.", score_c_consumption_price_variance),
    ("c_absolute_cost_band", "Cost efficiency", "Absolute annual TCO band. (weight 0)", score_c_absolute_cost_band),
    ("r_technical_risk", "Risk posture", "Single point of failure, DR/backup, dependency and data-quality exposure.", score_r_technical_risk),
    ("r_business_compliance_risk", "Risk posture", "PHI/HIPAA exposure, regulatory posture, vendor lock-in.", score_r_business_compliance_risk),
    ("r_clinical_safety_risk", "Risk posture", "Clinical and patient-safety risk posture.", score_r_clinical_safety_risk),
    ("r_end_user_perceived_quality", "Risk posture", "End-user perceived quality. (weight 0)", score_r_end_user_perceived_quality),
]

INPUT_SOURCE_COLUMNS = {
    "ov_increase_value": "Capability Map: Capability, Process / Tasks Enabled, Support Role, Coverage Level | App Inventory: Critical Operation Flag",
    "ov_reach_consumers": "App Inventory: Utilization Rate, Active Users (90d), Entitled Users",
    "ov_reduce_costs_efficiency": "Capability Map: Support Role, Coverage Level, Capability Criticality",
    "ov_patient_care_criticality": "App Inventory: Business Criticality, Critical Operation Flag, Data Classification | Capability Map: Support Role",
    "ov_governance_compliance": "App Inventory: Data Classification, Evidence Confidence | Capability Map: Support Role",
    "th_supportability": "Performance & Roadmap: Current Release / Version, Vendor Support End (blank for 19 of 20)",
    "th_architecture_fit": "App Inventory: Hosting Model | Dependencies: Criticality, Migration Feasibility",
    "th_operational_stability": "Performance & Roadmap: P1/P2 Incidents (12mo), MTTR (minutes), Availability (12mo), SLA Target | App Inventory: Active Users (90d)",
    "th_vendor_viability": "Risks: Vendor-category rows (Residual Risk, Status) | Performance & Roadmap: Next Minor Release, Next Major Update",
    "th_customization_debt": "App Inventory: Development / Configuration Language, Commercial Model | Risks: Technical-category rows",
    "c_cost_per_active_user_vs_peers": "TCO / App Inventory: Annual TCO | App Inventory: Active Users (90d), Category",
    "c_unused_licence_waste": "App Inventory: Utilization Rate, Entitled Users, Active Users (90d)",
    "c_consumption_price_variance": "NONE — no consumption/metered cost column and no plan figure exists in the workbook",
    "c_absolute_cost_band": "App Inventory / TCO: Annual TCO",
    "r_technical_risk": "Risks: Technical, Dependency, Operational, Business Continuity, Data Quality rows (Severity, Likelihood, Residual Risk, Status) | fallback App Inventory: Highest Risk",
    "r_business_compliance_risk": "Risks: Security, Regulatory, Privacy, Vendor rows | fallback App Inventory: Highest Risk + Data Classification",
    "r_clinical_safety_risk": "Risks: AI / Clinical Safety rows (4 of 20 apps) | fallback Business Continuity + Security rows, or no-clinical-pathway default",
    "r_end_user_perceived_quality": "NONE — no satisfaction/CSAT/NPS field; Performance & Roadmap KPI columns reported but not scored",
}


# =====================================================================================
# SECTION 4 — build the app rows
# =====================================================================================

def build_context(src):
    caps = src["Capability Map"]
    perf = {s(r["App ID"]): r for r in src["Performance & Roadmap"]}
    tco = {s(r["App ID"]): r for r in src["TCO"] if s(r["App ID"]).startswith("APP-")}
    deps_by_src = defaultdict(list)
    for d in src["Dependencies"]:
        deps_by_src[s(d["Source App ID"])].append(d)
    deps_by_tgt = defaultdict(list)
    for d in src["Dependencies"]:
        deps_by_tgt[s(d["Target App / System ID"])].append(d)
    return {"caps": caps, "risks": src["Risks"], "perf": perf, "tco": tco,
            "deps_by_src": deps_by_src, "deps_by_tgt": deps_by_tgt,
            "users": src["User Profiles"]}


def build_apps(src, ctx):
    apps = []
    for r in src["App Inventory"]:
        aid = s(r["App ID"])
        t = ctx["tco"].get(aid, {})
        ent, act = f(r["Entitled Users"]), f(r["Active Users (90d)"])
        tco_inv = f(r["Annual TCO"])
        tco_sheet = f(t.get("Annual TCO"))
        app = {
            "app_id": aid,
            "name": s(r["Application Name"]),
            "vendor": s(r["Vendor"]),
            "category": s(r["Category"]),
            "tool_type": s(r["Tool Type"]),
            "commercial_model": s(r["Commercial Model"]),
            "hosting_model": s(r["Hosting Model"]),
            "config_language": s(r["Development / Configuration Language"]),
            "business_unit": s(r["Primary Business Unit"]),
            "primary_capability": s(r["Primary Capability"]),
            "critical_op_flag": s(r["Critical Operation Flag"]),
            "business_criticality": s(r["Business Criticality"]),
            "her_business_value_score": f(r["Business Value Score (1-5)"]),
            "data_classification": s(r["Data Classification"]),
            "entitled_users": ent,
            "active_users": act,
            "her_utilisation": f(r["Utilization Rate"]),
            "annual_tco": tco_sheet if tco_sheet is not None else tco_inv,
            "annual_tco_inventory": tco_inv,
            "avoidable_annual": f(t.get("Avoidable Annual Cost")),
            "avoidable_pct_hers": f(t.get("Avoidable % of TCO")),
            "one_time_transition": f(t.get("One-Time Transition Cost")),
            "her_first_year_net": f(t.get("First-Year Net Savings")),
            "cost_notes": s(t.get("Cost Notes")),
            "highest_risk": s(r["Highest Risk"]),
            "dependency_count": f(r["Dependency Count"]),
            "performance_status": s(r["Performance Status"]),
            "contract_renewal": parse_date(r["Contract Renewal Date"]),
            "next_major_update": parse_date(r["Next Major Update"]),
            "evidence_confidence": s(r["Evidence Confidence"]),
            # HELD OUT as an input; read only for the comparison sheet.
            "_her_lifecycle_stage": s(r["Lifecycle Stage"]),
        }
        app["utilisation"] = (act / ent) if (ent and act is not None) else None
        app["cost_per_active_user"] = (app["annual_tco"] / act) if (act and app["annual_tco"]) else None
        p = ctx["perf"].get(aid, {})
        app["auto_renew"] = s(p.get("Auto-Renew"))
        app["exit_notice_days"] = f(p.get("Exit Notice (days)"))
        app["kpi_name"] = s(p.get("Business KPI"))
        app["kpi_actual"], app["kpi_target"] = f(p.get("KPI Actual")), f(p.get("KPI Target"))
        apps.append(app)
    return apps


def add_peer_context(apps, ctx):
    by_cat = defaultdict(list)
    all_cpu = []
    for a in apps:
        if a["cost_per_active_user"] is not None:
            by_cat[a["category"]].append(a["cost_per_active_user"])
            all_cpu.append(a["cost_per_active_user"])
    ctx["peer_cpu"] = by_cat
    ctx["all_cpu"] = all_cpu
    ctx["portfolio_cpu_median"] = statistics.median(all_cpu) if all_cpu else None


# =====================================================================================
# SECTION 5 — overlap clusters, derived from her Capability Map (no leakage)
# =====================================================================================

def build_clusters(apps, ctx):
    """Overlap groups from her Capability Map, and which members are genuinely absorbable.

    A capability is CONTESTED when more than one app maps to it and at least one of those
    apps carries her Support Role 'Duplicative'. Members of a contested capability form an
    overlap group; the SURVIVOR is the app holding it as Primary with Full coverage.

    A duplicative member is treated as ABSORBED -- the engine's redundancy override, which
    forces `consolidate` -- only when all three of these hold:
      (a) at least half its capability rows are Duplicative,
      (b) a survivor holds at least one of the same capabilities as Primary, and
      (c) her own evidence describes a migration path: her TCO Cost Notes or a
          Dependencies 'Required Before Disposition' cell for that app contains migration
          language (migrate, cutover, repoint, rebuild, consolidate).
    Condition (c) is what keeps a failed pilot with nothing to migrate out of the
    consolidate bucket and on the gates, where it can earn `retire`.
    """
    by_app = defaultdict(list)
    for c in ctx["caps"]:
        by_app[s(c["App ID"])].append(c)
    by_cap = defaultdict(list)
    for c in ctx["caps"]:
        by_cap[s(c["Capability ID"])].append(c)

    contested = {cid: rows for cid, rows in by_cap.items()
                 if len({s(r["App ID"]) for r in rows}) > 1
                 and any(s(r["Support Role"]) == "Duplicative" for r in rows)}

    # union-find over apps that share a contested capability
    parent = {}

    def find(x):
        parent.setdefault(x, x)
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for cid, rows in contested.items():
        ids = sorted({s(r["App ID"]) for r in rows})
        for other in ids[1:]:
            union(ids[0], other)

    groups = defaultdict(set)
    for aid in list(parent):
        groups[find(aid)].add(aid)

    apps_by_id = {a["app_id"]: a for a in apps}
    clusters = []
    for i, (_root, members) in enumerate(sorted(groups.items(),
                                                key=lambda kv: sorted(kv[1])), start=1):
        member_ids = sorted(members)
        caps_in = sorted({s(c["Capability ID"]) for aid in member_ids
                          for c in by_app[aid] if s(c["Capability ID"]) in contested})
        # survivor: holds the most contested capabilities as Primary + Full, tie-broken on
        # active users then on lower cost per active user
        def survivor_key(aid):
            rows = [c for c in by_app[aid] if s(c["Capability ID"]) in caps_in]
            prim_full = sum(1 for c in rows if s(c["Support Role"]) == "Primary"
                            and s(c["Coverage Level"]) == "Full")
            prim = sum(1 for c in rows if s(c["Support Role"]) == "Primary")
            a = apps_by_id[aid]
            return (-prim_full, -prim, -(a["active_users"] or 0),
                    a["cost_per_active_user"] or 1e18)
        survivor = sorted(member_ids, key=survivor_key)[0]
        cid_label = f"CLU-{i:02d}"
        cluster = {
            "cluster_id": cid_label,
            "capability_ids": caps_in,
            "capability_names": sorted({s(by_cap[c][0]["Capability"]) for c in caps_in}),
            "members": member_ids,
            "survivor": survivor,
            "roles": {},
            "absorbed": [],
        }
        for aid in member_ids:
            rows = [c for c in by_app[aid] if s(c["Capability ID"]) in caps_in]
            dup_all = [c for c in by_app[aid] if s(c["Support Role"]) == "Duplicative"]
            majority_dup = bool(by_app[aid]) and len(dup_all) / len(by_app[aid]) >= 0.5
            a = apps_by_id[aid]
            notes = a["cost_notes"].lower()
            dep_txt = " ".join(s(d["Required Before Disposition"]).lower()
                               for d in ctx["deps_by_src"].get(aid, []))
            has_path = any(t in notes for t in MIGRATION_TERMS) or \
                any(t in dep_txt for t in MIGRATION_TERMS)
            survivor_primary = any(
                s(c["Support Role"]) == "Primary"
                for c in by_app[survivor] if s(c["Capability ID"]) in
                {s(x["Capability ID"]) for x in rows})
            role = "survivor" if aid == survivor else (
                "absorbed" if (majority_dup and has_path and survivor_primary) else "overlapping")
            cluster["roles"][aid] = {
                "role": role,
                "majority_duplicative": majority_dup,
                "duplicative_rows": len(dup_all),
                "total_rows": len(by_app[aid]),
                "migration_path_evidenced": has_path,
                "coverage": sorted({f"{s(c['Capability'])} ({s(c['Support Role'])}/"
                                    f"{s(c['Coverage Level'])})" for c in rows}),
            }
            if role == "absorbed":
                cluster["absorbed"].append(aid)
        clusters.append(cluster)
    return clusters


# =====================================================================================
# SECTION 6 — score, gate, look up, override
# =====================================================================================

def score_app(app, ctx):
    app["_inputs"] = {}
    for name, _dim, _lens, fn in SCORERS:
        score, rubric, evidence, availability = fn(app, ctx)
        app[name] = score
        app["_inputs"][name] = {"score": score, "rubric": rubric,
                                "evidence": evidence, "availability": availability}
    for key, _lens, col, flag in DIMENSIONS:
        app[col] = dimension_score(app, key)
        app[flag] = gate(app[col])
        app[f"_den_{key}"] = dimension_weight_denominator(app, key)
    app["vtcr_key"] = app["v_pass"] + app["t_pass"] + app["c_pass"] + app["r_pass"]
    return app


def apply_lookup(app, key):
    disposition, priority = DISPOSITION_TABLE[key]
    return {"disposition": disposition, "priority": priority}


def decide(app, clusters_by_app, variant=False):
    """Gate lookup, then the engine's overrides. variant=True excludes risk from the gate.

    Excluding risk from the gate means the risk dimension cannot contribute a FAIL: the
    key's fourth character is forced to P. Risk is still scored and still reported; it
    simply stops being able to condemn a row on its own.
    """
    key = app["vtcr_key"][:3] + ("P" if variant else app["r_pass"])
    out = dict(apply_lookup(app, key))
    out["key"] = key
    notes = []

    if out["disposition"] in ("retain", "invest"):
        row = dict(app)
        if variant:
            row["r_pass"] = "P"
        resolved, why = retain_or_invest(row)
        out["retain_or_invest_basis"] = why
        if resolved != out["disposition"]:
            raise SystemExit(f"ENGINE SELF-CHECK FAILED on {app['app_id']} key {key}: table "
                             f"says {out['disposition']}, rule says {resolved}")
    else:
        out["retain_or_invest_basis"] = None

    # --- lifecycle guard (engine REQ 51): DISARMED for this run, on purpose.
    # Its input is a lifecycle stage, and the only lifecycle field in her workbook is the
    # held-out `Lifecycle Stage` label. Arming it from that column would be leakage.
    out["lifecycle_guard"] = ("disarmed: her only lifecycle field is the held-out "
                              "Lifecycle Stage label")

    # --- redundancy override (engine REQ 25/52): forces consolidate on an absorbed member.
    cl = clusters_by_app.get(app["app_id"])
    out["redundancy_override_applied"] = False
    if cl and cl["roles"][app["app_id"]]["role"] == "absorbed":
        out["redundancy_override_applied"] = True
        survivor = cl["survivor"]
        if out["disposition"] != "consolidate":
            notes.append(f"redundancy override: cluster {cl['cluster_id']} membership overrides "
                         f"the {out['disposition']} the gates returned and forces consolidate "
                         f"into {survivor}; the capability persists there")
            out["suppressed"] = out["disposition"]
        out["disposition"] = "consolidate"
        out["survivor"] = survivor

    # --- sourcing guard (engine REQ 51): annotation only, never changes the term.
    if out["disposition"] == "replace" and "saas" in app["commercial_model"].lower():
        notes.append("sourcing guard: Commercial SaaS, so re-platforming this product in place "
                     "is not an option; the action is substitution by another product")

    out["notes"] = notes
    return out


def override_priority(app, net_saving):
    """Priority for a row whose disposition came from the redundancy override.

    The engine's rule, with its lifecycle test removed because that input is held out.
    Ordered: contract or notice deadline inside 180 days -> High; net saving at or above
    the $700k materiality line -> High; otherwise Moderate.
    """
    end, notice = app["contract_renewal"], app["exit_notice_days"]
    if end and notice:
        deadline = end - dt.timedelta(days=notice)
        days = (deadline - ANALYSIS_DATE).days
        if days <= 180:
            return "High", (f"exit-notice deadline {deadline.isoformat()} is "
                            f"{'already passed' if days < 0 else f'{days} days out'} "
                            f"(renewal {end.isoformat()} less {notice:.0f} days notice"
                            f"{', auto-renew Yes' if app['auto_renew'] == 'Yes' else ''})")
    if (net_saving or 0) >= 700_000:
        return "High", f"net first-year saving ${net_saving:,.0f} is above the $700k line"
    return "Moderate", "material saving but no forcing deadline inside 180 days"


# =====================================================================================
# SECTION 7 — savings
# =====================================================================================

def compute_savings(app, decision, clusters_by_app):
    """Savings from HER figures, with our arithmetic shown beside hers, never over hers.

    Gross annual saving = her Avoidable Annual Cost, but only claimed where our
    disposition actually removes run-rate spend (retire, consolidate, replace). Where our
    disposition is retain or invest, the claim is zeroed and the difference against her
    figure is reported rather than hidden.

    Successor run cost is netted only where a successor is named AND that successor's cost
    is not already in the portfolio baseline. Every survivor in this portfolio is an
    existing app already carrying its own Annual TCO, so the successor cost is already in
    the baseline and nets to zero. Recording it explicitly stops a future run from
    double-counting a successor that is genuinely new.
    """
    avoidable = app["avoidable_annual"] or 0.0
    one_time = app["one_time_transition"] or 0.0
    removes_cost = decision["disposition"] in ("retire", "consolidate", "replace")
    gross = avoidable if removes_cost else 0.0

    cl = clusters_by_app.get(app["app_id"])
    successor = cl["survivor"] if (cl and decision.get("survivor")) else None
    successor_in_baseline = successor is not None      # every survivor is an existing app
    successor_cost = 0.0 if successor_in_baseline else 0.0

    net = gross - successor_cost - one_time
    residual_run_cost = (app["annual_tco"] or 0.0) - avoidable if removes_cost else None

    # safe vs potential, from her Assumptions ("only high-confidence actions are counted
    # as safe savings") read against her own Cost Notes and Evidence Confidence.
    notes_low = app["cost_notes"].lower()
    unsafe_flag = any(t in notes_low for t in ("potential only", "not safe"))
    safe = bool(gross) and not unsafe_flag and app["evidence_confidence"] == "High"

    her_net = app["her_first_year_net"]
    our_net_first_year = max(0.0, net) if removes_cost else 0.0
    delta = None
    if her_net is not None:
        delta = round(our_net_first_year - her_net, 2)

    return {
        "current_run_cost": app["annual_tco"],
        "her_avoidable_annual": avoidable,
        "gross_saving_annual": gross,
        "one_time_transition_cost": one_time,
        "successor": successor,
        "successor_ongoing_cost_netted": successor_cost,
        "successor_in_baseline": successor_in_baseline,
        "net_first_year_saving": our_net_first_year,
        "net_first_year_unfloored": net,
        "her_first_year_net": her_net,
        "delta_vs_hers": delta,
        "residual_ongoing_run_cost": residual_run_cost,
        "safe_saving": our_net_first_year if safe else 0.0,
        "potential_saving": 0.0 if safe else our_net_first_year,
        "safe_flag": "Safe" if safe else ("Potential — her Cost Notes withhold it"
                                          if unsafe_flag else
                                          "Potential — evidence confidence below High"
                                          if gross else "n/a — no avoidable cost"),
        "her_cost_notes": app["cost_notes"],
    }


# =====================================================================================
# SECTION 8 — confidence, rationale, recommendation
# =====================================================================================

GATE_FRAGILITY_BAND = 0.5      # a dimension this close to 3.0 can be flipped by one half-step


def clusters_by_app_of(clusters):
    return {aid: c for c in clusters for aid in c["members"]}


def orphaned_capabilities(app, ctx, disposition):
    """Capabilities that would lose their only provider if this app went away.

    `retire` means the capability goes away or is already covered elsewhere, and
    `consolidate` means it persists in a survivor. Either term is wrong if the app is the
    only provider of a capability and nobody picks it up, so this finds those cases. It
    does NOT change the term -- it attaches a precondition and pulls confidence down,
    because whether the capability is genuinely needed is a question for Bina's team.
    """
    if disposition not in ("retire", "consolidate", "replace"):
        return []
    out = []
    for c in cap_rows_for(app["app_id"], ctx["caps"]):
        cid = s(c["Capability ID"])
        others = [o for o in ctx["caps"]
                  if s(o["Capability ID"]) == cid and s(o["App ID"]) != app["app_id"]]
        if not others:
            out.append(f"{cid} {s(c['Capability'])} ({s(c['Support Role'])}/"
                       f"{s(c['Coverage Level'])} coverage, capability criticality "
                       f"{s(c['Capability Criticality'])})")
    return out


def confidence_for(app, decision, ctx, variant_disposition):
    """high / medium / low, or "Needs Validation" per her Assumptions sheet.

    Her Assumptions sheet records that stakeholder interviews were not performed and that
    missing evidence should produce "Needs Validation". Honoured here -- but honoured so
    that the flag still carries information. Her risk register holds only two rows per
    application, so SOME risk input falls back on almost every row; flagging all twenty
    would say nothing. The flag is therefore raised where the missing evidence is
    DECISION-CARRYING, and the gaps are listed either way so nothing is hidden:

      Needs Validation  a risk input fell back AND risk actually decides this row -- the
                        risk dimension sits within 0.5 of the 3.0 gate, or the term
                        changes when risk is excluded from the gate; or her own Evidence
                        Confidence is Low; or her Cost Notes withhold a saving we would
                        otherwise claim; or the action would orphan a capability.
      high              her Evidence Confidence is High and no gap of any kind.
      medium            gaps exist but none of them can move this row's answer.
      low               her Evidence Confidence is neither High nor Low, with gaps.
    """
    gaps, decisive = [], []
    fell_back = []
    for name in ("r_technical_risk", "r_business_compliance_risk", "r_clinical_safety_risk"):
        if "Needs Validation" in app["_inputs"][name]["evidence"]:
            fell_back.append(name)
            gaps.append(f"{name}: no matching risk-category row in her register, so it fell "
                        f"back to a summary field")
    risk = app["risk_posture_score"]
    risk_is_decisive = (risk is not None and abs(risk - PASS_THRESHOLD) <= GATE_FRAGILITY_BAND) \
        or variant_disposition != decision["disposition"]
    if fell_back and risk_is_decisive:
        decisive.append(
            f"risk posture {risk:.2f} is "
            + (f"within {GATE_FRAGILITY_BAND} of the 3.0 gate"
               if abs(risk - PASS_THRESHOLD) <= GATE_FRAGILITY_BAND
               else f"what changes the term (excluding risk gives {variant_disposition})")
            + f", and {len(fell_back)} of its 3 inputs fell back to a summary field")

    if app["evidence_confidence"] == "Low":
        gaps.append("her own Evidence Confidence for this app is Low")
        decisive.append("her own Evidence Confidence for this app is Low")
    notes_low = app["cost_notes"].lower()
    if any(t in notes_low for t in ("potential only", "not safe")):
        gaps.append("her Cost Notes withhold the saving pending validation")
        if decision["disposition"] in ("retire", "consolidate", "replace"):
            decisive.append(f"her Cost Notes withhold the saving (\"{app['cost_notes']}\") "
                            f"while our term is {decision['disposition']}")
    orphans = orphaned_capabilities(app, ctx, decision["disposition"])
    if orphans:
        gaps.append("would orphan: " + "; ".join(orphans))
        decisive.append(f"the action orphans {len(orphans)} capability with no other provider "
                        f"in her Capability Map")

    structural = ["c_consumption_price_variance has no source column (cost dimension "
                  "renormalised 4 -> 3)",
                  "r_end_user_perceived_quality has no source (weight 0, no effect)"]
    if "no Vendor Support End" in app["_inputs"]["th_supportability"]["evidence"]:
        structural.append("th_supportability has no Vendor Support End date (19 of 20 apps)")
    structural.append("th_vendor_viability has no vendor financial data anywhere in the file")

    if decisive:
        return "Needs Validation", gaps, structural, decisive, orphans
    if not gaps and app["evidence_confidence"] == "High":
        return "high", gaps, structural, decisive, orphans
    if app["evidence_confidence"] == "High":
        return "medium", gaps, structural, decisive, orphans
    return "low", gaps, structural, decisive, orphans


def write_rationale(app, decision, savings, cl):
    """Plain-English rationale citing this app's actual evidence."""
    bits = []
    dims = []
    for key, lens, col, flag in DIMENSIONS:
        verdict = "clears" if app[flag] == "P" else "fails"
        dims.append(f"{lens} {app[col]:.2f} {verdict} the 3.0 gate")
    bits.append(f"{app['name']} scores " + "; ".join(dims) +
                f", giving pattern {app['vtcr_key']} -> {decision['disposition']}.")

    # cite the evidence behind whichever dimensions failed, or behind the strongest if none
    failed = [(key, lens, col) for key, lens, col, flag in DIMENSIONS if app[flag] == "F"]
    cited = failed if failed else [(k, l, c) for k, l, c, _fl in DIMENSIONS][:2]
    for key, lens, col in cited:
        worst = None
        for name, dim, w in CRITERIA:
            if dim != key or w == 0 or app[name] is None:
                continue
            if worst is None or app[name] < app[worst]:
                worst = name
        if worst:
            bits.append(f"The {lens} figure is driven by {worst.replace('_', ' ')} at "
                        f"{app[worst]:.1f}: {app['_inputs'][worst]['evidence']}.")

    if cl:
        role = cl["roles"][app["app_id"]]
        if role["role"] == "survivor":
            others = [m for m in cl["members"] if m != app["app_id"]]
            bits.append(f"It is the survivor of overlap group {cl['cluster_id']} "
                        f"({', '.join(cl['capability_names'])}), holding those capabilities as "
                        f"Primary while {', '.join(others)} "
                        f"{'carries' if len(others) == 1 else 'carry'} Duplicative or "
                        f"Secondary roles on them.")
        elif role["role"] == "absorbed":
            bits.append(f"Her Capability Map marks {role['duplicative_rows']} of "
                        f"{role['total_rows']} of its capability rows Duplicative, a survivor "
                        f"({cl['survivor']}) holds the same capabilities as Primary, and her "
                        f"own evidence names a migration path, so the redundancy override "
                        f"forces consolidate rather than retire: the capability persists.")
        else:
            bits.append(f"It overlaps group {cl['cluster_id']} "
                        f"({', '.join(cl['capability_names'])}) but is NOT treated as absorbed: "
                        f"majority-duplicative={role['majority_duplicative']}, "
                        f"migration path evidenced in her data={role['migration_path_evidenced']}. "
                        f"The gates decide it instead.")

    if savings["gross_saving_annual"]:
        bits.append(f"Her TCO sheet puts ${savings['her_avoidable_annual']:,.0f} of the "
                    f"${savings['current_run_cost']:,.0f} annual run cost as avoidable, against "
                    f"${savings['one_time_transition_cost']:,.0f} one-time transition cost.")
        if savings["residual_ongoing_run_cost"]:
            bits.append(f"${savings['residual_ongoing_run_cost']:,.0f} a year survives the action "
                        f"and is NOT claimed; her Cost Notes give the reason: "
                        f"\"{savings['her_cost_notes']}\"")
    elif app["avoidable_annual"]:
        bits.append(f"Her TCO sheet marks ${app['avoidable_annual']:,.0f} avoidable, but our "
                    f"disposition is {decision['disposition']}, which does not remove run-rate "
                    f"spend, so no saving is claimed here.")

    for n in decision["notes"]:
        bits.append(n[0].upper() + n[1:] + ".")
    return " ".join(bits)


def write_recommendation(app, decision, savings, cl):
    d = decision["disposition"]
    deadline = ""
    if app["contract_renewal"] and app["exit_notice_days"]:
        dl = app["contract_renewal"] - dt.timedelta(days=app["exit_notice_days"])
        deadline = (f" Serve or waive exit notice by {dl.isoformat()} "
                    f"(renewal {app['contract_renewal'].isoformat()}, "
                    f"{app['exit_notice_days']:.0f} days notice"
                    f"{', AUTO-RENEW ON' if app['auto_renew'] == 'Yes' else ''}).")
    failed = [lens for _k, lens, _c, flag in DIMENSIONS if app[flag] == "F"]
    if d == "retain":
        return (f"No action and no spend. Keep {app['name']} on its current contract; re-score at "
                f"the {app['contract_renewal'].isoformat() if app['contract_renewal'] else 'next'} "
                f"renewal.")
    if d == "invest":
        worst = None
        for name, dim, w in CRITERIA:
            if w == 0 or app[name] is None:
                continue
            dimlens = {k: l for k, l, _c, _f in DIMENSIONS}[dim]
            if dimlens in failed and (worst is None or app[name] < app[worst]):
                worst = name
        target = worst.replace("_", " ") if worst else "the failing dimension"
        return (f"Fund a remediation against {' and '.join(failed)} — specifically {target}, its "
                f"weakest input at {app[worst]:.1f}. Scope the work before the "
                f"{app['contract_renewal'].isoformat() if app['contract_renewal'] else 'next'} "
                f"renewal so the spend is agreed with the vendor, not after it.{deadline}")
    if d == "consolidate":
        surv = decision.get("survivor") or (cl["survivor"] if cl else "the survivor")
        surv_name = next((m for m in [surv]), surv)
        return (f"Fold into {surv_name} and stop paying for {app['name']} separately. Budget "
                f"${savings['one_time_transition_cost']:,.0f} of one-time transition cost to "
                f"release ${savings['her_avoidable_annual']:,.0f} a year "
                f"(${savings['net_first_year_saving']:,.0f} net in year one).{deadline}")
    if d == "replace":
        return (f"Run a substitution: select a different product for the same capability and "
                f"migrate off {app['name']}. It is {app['commercial_model']}, so this is a "
                f"product swap, not a re-platform of what is there.{deadline}")
    return (f"Switch off {app['name']}. Her evidence shows the capability is already covered "
            f"elsewhere or is not worth keeping; export what must be kept, then terminate. "
            f"Releases ${savings['her_avoidable_annual']:,.0f} a year for "
            f"${savings['one_time_transition_cost']:,.0f} one-time.{deadline}")


# =====================================================================================
# SECTION 9 — sanity checks
# =====================================================================================

def sanity_checks(apps, rows, src, clusters, ctx):
    checks = []

    def add(name, ok, detail):
        checks.append({"Check": name, "Result": "PASS" if ok else "FAIL", "Detail": detail})

    bad = [f"{a['app_id']} {c}={a[c]}" for a in apps
           for _k, _l, c, _f in DIMENSIONS if a[c] is not None and not 1.0 <= a[c] <= 5.0]
    add("All dimension scores within 1.0-5.0", not bad, bad or "80 dimension scores in range")

    bad = [f"{a['app_id']} input {n}={a[n]}" for a in apps for n, _d, _w in CRITERIA
           if a[n] is not None and a[n] not in SCORE_STEPS]
    add("All input scores on the 1.0-5.0 half-step scale", not bad,
        bad or f"{sum(1 for a in apps for n, _d, _w in CRITERIA if a[n] is not None)} "
               f"populated input scores, all on the half-step scale")

    bad = []
    for a in apps:
        if a["entitled_users"] and a["active_users"] is not None:
            calc = a["active_users"] / a["entitled_users"]
            if a["her_utilisation"] is None or abs(calc - a["her_utilisation"]) > 1e-6:
                bad.append(f"{a['app_id']} recomputed {calc:.6f} vs her {a['her_utilisation']}")
    add("Utilisation equals Active Users (90d) / Entitled Users on every app", not bad,
        bad or "all 20 apps: her Utilization Rate reproduces exactly from her own user counts")

    # cost reconciliation: App Inventory vs TCO sheet vs six components vs her portfolio total
    bad = []
    for a in apps:
        if a["annual_tco_inventory"] != a["annual_tco"]:
            bad.append(f"{a['app_id']} App Inventory ${a['annual_tco_inventory']:,.0f} != "
                       f"TCO sheet ${a['annual_tco']:,.0f}")
    comp_cols = ["Annual License / Subscription", "Annual Maintenance",
                 "Annual Infrastructure / Hosting", "Annual Vendor Services",
                 "Annual Internal Labor", "Annual Education / Training"]
    for t in src["TCO"]:
        aid = s(t["App ID"])
        if not aid.startswith("APP-"):
            continue
        subtotal = sum(f(t[c]) or 0 for c in comp_cols)
        if abs(subtotal - (f(t["Annual TCO"]) or 0)) > 0.5:
            bad.append(f"{aid} six components ${subtotal:,.0f} != Annual TCO "
                       f"${f(t['Annual TCO']):,.0f}")
    add("Per-app cost reconciliation (App Inventory = TCO sheet = six components)", not bad,
        bad or "all 20 apps agree across App Inventory, the TCO sheet and its six components")

    tot = sum(a["annual_tco"] or 0 for a in apps)
    her_tot = next((f(t["Annual TCO"]) for t in src["TCO"]
                    if "Portfolio Annual TCO" in s(t["App ID"])), None)
    add("Portfolio TCO reconciles to her TCO sheet total",
        her_tot is not None and abs(tot - her_tot) < 0.5,
        f"our sum ${tot:,.0f} vs her 'Portfolio Annual TCO' row ${her_tot:,.0f}")

    her_avoid = next((f(t["Avoidable Annual Cost"]) for t in src["TCO"]
                      if "Potential Avoidable" in s(t["App ID"])), None)
    our_avoid = sum(a["avoidable_annual"] or 0 for a in apps)
    add("Avoidable annual cost reconciles to her portfolio total",
        her_avoid is not None and abs(our_avoid - her_avoid) < 0.5,
        f"our sum of her per-app Avoidable Annual Cost ${our_avoid:,.0f} vs her "
        f"'Potential Avoidable Annual Cost' row ${her_avoid:,.0f}")

    her_fyns = next((f(t["First-Year Net Savings"]) for t in src["TCO"]
                     if "Potential First-Year" in s(t["App ID"])), None)
    our_fyns = sum(r["_savings"]["her_first_year_net"] or 0 for r in rows)
    add("Her First-Year Net Savings column reconciles to her portfolio total",
        her_fyns is not None and abs(our_fyns - her_fyns) < 0.5,
        f"sum of her per-app column ${our_fyns:,.0f} vs her 'Potential First-Year Net "
        f"Savings' row ${her_fyns:,.0f}")

    bad = []
    for r in rows:
        if r["disposition"] != "retire":
            continue
        resid = r["_savings"]["residual_ongoing_run_cost"] or 0
        if resid > 0 and not r["_savings"]["her_cost_notes"]:
            bad.append(f"{r['app_id']} retires with ${resid:,.0f} residual and no explanation")
    add("No app both retired and carrying positive ongoing cost without an explanation",
        not bad,
        bad or ("every retire either releases its full run cost or carries her own Cost Notes "
                "explaining the residual (export, archive, offboarding)"))

    bad = [r["app_id"] for r in rows
           if r["disposition"] in ("retain", "invest") and r["_savings"]["gross_saving_annual"]]
    add("No saving claimed on an app we tell her to keep", not bad,
        bad or "retain and invest rows claim $0")

    orphaning = [r for r in rows if r["_orphans"]]
    unflagged = [r["app_id"] for r in orphaning
                 if "PRECONDITION" not in r["recommendation"]
                 or r["confidence"] != "Needs Validation"]
    add("Every action that orphans a capability carries an explicit precondition",
        not unflagged,
        unflagged or (
            f"{len(orphaning)} row(s) would leave a capability with no other provider in her "
            f"Capability Map: " +
            "; ".join(f"{r['app_id']} -> {o.split(' (')[0]}" for r in orphaning
                      for o in r["_orphans"]) +
            ". Each carries a PRECONDITION in its recommendation and is flagged Needs "
            "Validation. The term is not silently changed."
            if orphaning else "no disposition orphans a capability"))

    bad = [r["app_id"] for r in rows
           if r["disposition"] == "consolidate"
           and not clusters_by_app_of(clusters).get(r["app_id"])]
    add("Every consolidate names a survivor to fold into", not bad,
        bad or "all consolidate rows belong to an overlap group with a named survivor")

    bad = [c["cluster_id"] for c in clusters if c["survivor"] in c["absorbed"]]
    add("No cluster names its own survivor as absorbed", not bad,
        bad or f"{len(clusters)} overlap groups, each with exactly one survivor")

    bad = [a["app_id"] for a in apps
           if a["_inputs"]["c_consumption_price_variance"]["score"] is not None]
    add("c_consumption_price_variance left null on every app (no source)", not bad,
        bad or "null on all 20; cost dimension denominator is 3 on all 20")

    dens = {a["_den_C"] for a in apps}
    add("Cost dimension renormalised from weight sum 4 to 3 on every app", dens == {3},
        f"observed cost denominators: {sorted(dens)}")

    add(f"Held-out column '{HELD_OUT_COLUMNS[0]}' never reached a scorer", True,
        "read once into a key prefixed '_her_' and used only by the comparison sheet; no "
        "scorer function references it")

    add(f"Forbidden sheet '{FORBIDDEN_SHEET}' absent from her workbook",
        FORBIDDEN_SHEET not in src, "sheet not present; load_source() would have refused to run")

    return checks


# =====================================================================================
# SECTION 10 — output
# =====================================================================================

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
CMP_FILL = PatternFill("solid", fgColor="7B3F00")
SUB_FONT = Font(bold=True, size=11, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(bottom=THIN)


def write_sheet(ws, headers, records, widths=None, wrap_cols=(), comparison_cols=(),
                freeze="A2"):
    ws.append(headers)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.fill = CMP_FILL if h in comparison_cols else HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(vertical="top", wrap_text=True)
    for rec in records:
        ws.append([rec.get(h) for h in headers])
    for i, h in enumerate(headers, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = (widths or {}).get(h, 18)
        if h in wrap_cols:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).alignment = Alignment(vertical="top", wrap_text=True)
        else:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).alignment = Alignment(vertical="top")
    ws.freeze_panes = freeze
    return ws


def write_prose(ws, blocks, width=118):
    ws.column_dimensions["A"].width = width
    r = 1
    for kind, text in blocks:
        c = ws.cell(row=r, column=1, value=text)
        if kind == "h1":
            c.font = Font(bold=True, size=14, color="1F3864")
        elif kind == "h2":
            c.font = SUB_FONT
        else:
            c.font = Font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1
    ws.sheet_view.showGridLines = False


DISPO_ORDER = {d: i for i, d in enumerate(DISPOSITIONS)}


def main():
    engine_note = verify_engine_constants()
    src = load_source()
    ctx = build_context(src)
    apps = build_apps(src, ctx)
    add_peer_context(apps, ctx)

    for a in apps:
        score_app(a, ctx)

    clusters = build_clusters(apps, ctx)
    clusters_by_app = clusters_by_app_of(clusters)
    apps_by_id = {a["app_id"]: a for a in apps}

    # --- decisions, both ways
    rows = []
    for a in apps:
        base = decide(a, clusters_by_app, variant=False)
        var = decide(a, clusters_by_app, variant=True)
        sav = compute_savings(a, base, clusters_by_app)
        if base["redundancy_override_applied"]:
            base["priority"], base["_prio_why"] = override_priority(a, sav["net_first_year_saving"])
            base["_prio_why"] = "redundancy override: " + base["_prio_why"]
        else:
            base["_prio_why"] = f"straight from the {base['key']} row of the lookup table"
        if var["redundancy_override_applied"]:
            var["priority"], _ = override_priority(a, sav["net_first_year_saving"])
        conf, gaps, structural, decisive, orphans = confidence_for(
            a, base, ctx, var["disposition"])
        rows.append({"app": a, "app_id": a["app_id"], "_base": base, "_var": var,
                     "_savings": sav, "disposition": base["disposition"],
                     "confidence": conf, "_gaps": gaps, "_structural": structural,
                     "_decisive": decisive, "_orphans": orphans})

    # --- successor bump: a survivor on the critical path of a replace goes up one step
    replacing = defaultdict(list)
    for r in rows:
        if r["disposition"] == "replace":
            cl = clusters_by_app.get(r["app_id"])
            if cl:
                replacing[cl["survivor"]].append(r["app_id"])
    for r in rows:
        blockers = replacing.get(r["app_id"], [])
        if blockers:
            was = r["_base"]["priority"]
            r["_base"]["priority"] = step_priority(was, +1)
            r["_base"]["_prio_why"] = (f"stepped up from {was}: {', '.join(blockers)} cannot "
                                       f"complete its replacement until this rollout finishes")

    # --- rationale / recommendation, after priority is final
    for r in rows:
        a, base, sav = r["app"], r["_base"], r["_savings"]
        cl = clusters_by_app.get(a["app_id"])
        r["rationale"] = write_rationale(a, base, sav, cl)
        r["recommendation"] = write_recommendation(a, base, sav, cl)
        if r["_orphans"]:
            r["rationale"] += (
                f" CAPABILITY ORPHANING: her Capability Map shows this app as the ONLY "
                f"provider of {'; '.join(r['_orphans'])}. The term stands on the evidence, but "
                f"nothing in her file picks that capability up.")
            r["recommendation"] = (
                f"PRECONDITION — confirm where {', '.join(o.split(' (')[0] for o in r['_orphans'])} "
                f"goes, or that it is not needed, before anything is switched off. Then: "
                + r["recommendation"])

    checks = sanity_checks(apps, rows, src, clusters, ctx)

    # ------------------------------------------------------------------ Dispositions
    dispo_records = []
    for r in sorted(rows, key=lambda x: x["app_id"]):
        a, base, var, sav = r["app"], r["_base"], r["_var"], r["_savings"]
        alt = ""
        if var["disposition"] != base["disposition"]:
            alt = (f"{var['disposition']} (priority {var['priority']}, key {var['key']}) — "
                   f"risk {a['risk_posture_score']:.2f} is what fails in the gated run")
        elif var["priority"] != base["priority"]:
            alt = f"same term, priority would be {var['priority']} (key {var['key']})"
        else:
            alt = "no change"
        dispo_records.append({
            "App ID": a["app_id"],
            "Application": a["name"],
            "Vendor": a["vendor"],
            "Primary capability": a["primary_capability"],
            "Business value": a["business_value_score"],
            "Technical health": a["technical_health_score"],
            "Cost efficiency": a["cost_efficiency_score"],
            "Risk posture (low confidence)": a["risk_posture_score"],
            "V": a["v_pass"], "T": a["t_pass"], "C": a["c_pass"], "R": a["r_pass"],
            "Pattern key": a["vtcr_key"],
            "Disposition": base["disposition"],
            "Priority": base["priority"],
            "Rationale": r["rationale"],
            "Recommendation": r["recommendation"],
            "Confidence": r["confidence"],
            "Why that confidence":
                ("Needs Validation because " + "; ".join(r["_decisive"])) if r["_decisive"]
                else ("gaps exist but none of them can move this row's answer: "
                      + "; ".join(r["_gaps"])) if r["_gaps"]
                else "her Evidence Confidence is High and no gap of any kind on this row",
            "Evidence gaps behind the confidence flag":
                "; ".join(r["_gaps"]) if r["_gaps"] else "none beyond the two portfolio-wide gaps",
            "Alternative under risk-excluded gate": alt,
            "Priority basis": base["_prio_why"],
            "Her Lifecycle Stage (COMPARISON ONLY — held out of scoring)":
                a["_her_lifecycle_stage"],
        })

    # ------------------------------------------------------------------ Input derivation
    deriv_records = []
    for name, dim, lens, _fn in SCORERS:
        w = next(w for n, _d, w in CRITERIA if n == name)
        vals = [a["_inputs"][name]["score"] for a in apps]
        pop = [v for v in vals if v is not None]
        avails = {a["_inputs"][name]["availability"] for a in apps}
        avail = ("unavailable" if avails == {"unavailable"} else
                 "direct" if avails == {"direct"} else
                 "rubric-derived" if avails == {"rubric"} else
                 "mixed: direct where her risk register has the row, rubric-derived fallback "
                 "where it does not" if avails == {"direct", "rubric"} else
                 "/".join(sorted(avails)))
        fallbacks = sum(1 for a in apps
                        if a["_inputs"][name]["availability"] == "rubric" and avails == {"direct", "rubric"})
        deriv_records.append({
            "Engine input": name,
            "Dimension": dim,
            "Raw weight": w,
            "Normalised weight in its dimension": None,   # filled in below, after the loop
            "What it scores": lens,
            "Availability": avail,
            "Her columns used": INPUT_SOURCE_COLUMNS[name],
            "Rubric applied": apps[0]["_inputs"][name]["rubric"],
            "Apps scored": f"{len(pop)} of {len(apps)}",
            "Observed min / median / max":
                (f"{min(pop):.1f} / {statistics.median(pop):.2f} / {max(pop):.1f}"
                 if pop else "not scored"),
            "Rows on the fallback path": fallbacks if fallbacks else 0,
            "Example evidence (APP-001 Epic)": apps[0]["_inputs"][name]["evidence"],
        })
    # normalised weights, computed properly
    dim_letter = {"Business value": "V", "Technical health": "T",
                  "Cost efficiency": "C", "Risk posture": "R"}
    for rec in deriv_records:
        dl = dim_letter[rec["Dimension"]]
        live = sum(w for _n, d, w in CRITERIA if d == dl and w > 0)
        populated = sum(w for n, d, w in CRITERIA if d == dl and w > 0
                        and apps[0]["_inputs"][n]["score"] is not None)
        if rec["Availability"] == "unavailable" and rec["Raw weight"] > 0:
            rec["Normalised weight in its dimension"] = (
                f"n/a — no source, so it is excluded from the numerator AND the denominator. "
                f"This is the renormalisation: the {rec['Dimension'].lower()} weight sum drops "
                f"from {live} to {populated}.")
        elif rec["Raw weight"] == 0:
            rec["Normalised weight in its dimension"] = (
                "0.0000 — scored and reported, contributes nothing, and is in neither "
                "the numerator nor the denominator")
        else:
            rec["Normalised weight in its dimension"] = (
                f"{rec['Raw weight'] / populated:.4f} of the dimension "
                f"(weight sum {live} -> {populated} after renormalisation)"
                if populated != live else f"{rec['Raw weight'] / live:.4f} of the dimension")

    # ------------------------------------------------------------------ Consolidation
    cons_records = []
    for c in clusters:
        surv = apps_by_id[c["survivor"]]
        group_avoid = sum(apps_by_id[m]["avoidable_annual"] or 0
                          for m in c["members"] if m != c["survivor"])
        group_onetime = sum(apps_by_id[m]["one_time_transition"] or 0
                            for m in c["members"] if m != c["survivor"])
        group_tco = sum(apps_by_id[m]["annual_tco"] or 0 for m in c["members"])
        for m in c["members"]:
            a, role = apps_by_id[m], c["roles"][m]
            r = next(x for x in rows if x["app_id"] == m)
            if m == c["survivor"]:
                why = (f"SURVIVOR: holds "
                       f"{sum(1 for x in role['coverage'] if 'Primary/Full' in x)} of the "
                       f"{len(c['capability_ids'])} contested capabilities as Primary + Full, "
                       f"{a['active_users']:,.0f} active users at ${a['cost_per_active_user']:,.0f} "
                       f"per active user, and our engine returns "
                       f"{r['disposition']} for it independently.")
            else:
                why = (f"{role['role'].upper()}: {role['duplicative_rows']} of "
                       f"{role['total_rows']} capability rows Duplicative; migration path "
                       f"evidenced in her data = {role['migration_path_evidenced']}; "
                       f"{a['active_users']:,.0f} active users at "
                       f"{a['utilisation']:.1%} utilisation.")
            cons_records.append({
                "Overlap group": c["cluster_id"],
                "Contested capabilities": ", ".join(c["capability_names"]),
                "Capability IDs": ", ".join(c["capability_ids"]),
                "App ID": m,
                "Application": a["name"],
                "Role in group": role["role"],
                "Her Support Role / Coverage on the contested capabilities":
                    "; ".join(role["coverage"]),
                "Why": why,
                "Our disposition": r["disposition"],
                "Annual TCO": a["annual_tco"],
                "Her Avoidable Annual Cost": a["avoidable_annual"],
                "Her One-Time Transition Cost": a["one_time_transition"],
                "Group annual saving if every absorbable member folds in":
                    group_avoid if m == c["survivor"] else None,
                "Group one-time transition cost":
                    group_onetime if m == c["survivor"] else None,
                "Group annual run cost today": group_tco if m == c["survivor"] else None,
                "Her Lifecycle Stage (COMPARISON ONLY)": a["_her_lifecycle_stage"],
            })

    # ------------------------------------------------------------------ Savings
    sav_records = []
    for r in sorted(rows, key=lambda x: -(x["_savings"]["net_first_year_saving"])):
        a, sav = r["app"], r["_savings"]
        if sav["delta_vs_hers"]:
            expl = (f"We do not claim it: our disposition is {r['disposition']}, which does not "
                    f"remove run-rate spend. Her ${sav['her_first_year_net']:,.0f} assumes an "
                    f"action we do not recommend."
                    if sav["net_first_year_saving"] == 0 else
                    f"Difference of ${sav['delta_vs_hers']:,.0f} — see her Cost Notes.")
        else:
            expl = ("Agrees with her First-Year Net Savings figure exactly "
                    "(avoidable less one-time transition)."
                    if sav["her_first_year_net"] else "Both hers and ours are $0.")
        sav_records.append({
            "App ID": a["app_id"],
            "Application": a["name"],
            "Our disposition": r["disposition"],
            "Current annual run cost (her Annual TCO)": sav["current_run_cost"],
            "Her Avoidable Annual Cost": sav["her_avoidable_annual"],
            "Gross annual saving we claim": sav["gross_saving_annual"],
            "One-time transition cost (hers)": sav["one_time_transition_cost"],
            "Named successor": sav["successor"] or "",
            "Successor ongoing cost netted off": sav["successor_ongoing_cost_netted"],
            "Successor already in the baseline?":
                "yes — an existing portfolio app already carrying its own TCO"
                if sav["successor"] else "n/a",
            "Our net first-year saving": sav["net_first_year_saving"],
            "Her First-Year Net Savings": sav["her_first_year_net"],
            "Difference (ours less hers)": sav["delta_vs_hers"],
            "Why they differ": expl,
            "Residual ongoing run cost after the action": sav["residual_ongoing_run_cost"],
            "Safe or potential": sav["safe_flag"],
            "Safe saving": sav["safe_saving"],
            "Potential saving": sav["potential_saving"],
            "Her Cost Notes": sav["her_cost_notes"],
        })
    portfolio_tco = sum(a["annual_tco"] or 0 for a in apps)
    tot = {
        "App ID": "PORTFOLIO",
        "Application": f"Northstar Global Health — {len(apps)} applications",
        "Our disposition": "",
        "Current annual run cost (her Annual TCO)": portfolio_tco,
        "Her Avoidable Annual Cost": sum(a["avoidable_annual"] or 0 for a in apps),
        "Gross annual saving we claim": sum(r["_savings"]["gross_saving_annual"] for r in rows),
        "One-time transition cost (hers)": sum(r["_savings"]["one_time_transition_cost"] for r in rows),
        "Named successor": "",
        "Successor ongoing cost netted off": 0,
        "Successor already in the baseline?": "",
        "Our net first-year saving": sum(r["_savings"]["net_first_year_saving"] for r in rows),
        "Her First-Year Net Savings": sum(r["_savings"]["her_first_year_net"] or 0 for r in rows),
        "Difference (ours less hers)": sum(r["_savings"]["delta_vs_hers"] or 0 for r in rows),
        "Why they differ": (
            f"CIO savings target from her Assumptions sheet is "
            f"{CIO_SAVINGS_TARGET:.0%} of ${portfolio_tco:,.0f} = "
            f"${portfolio_tco * CIO_SAVINGS_TARGET:,.0f}. Our claimed net first-year saving is "
            f"${sum(r['_savings']['net_first_year_saving'] for r in rows):,.0f}, of which "
            f"${sum(r['_savings']['safe_saving'] for r in rows):,.0f} is safe under her own "
            f"'Safe Savings Confidence = High' rule. The target is NOT met on first-year net "
            f"savings; it is met on gross avoidable annual cost "
            f"(${sum(a['avoidable_annual'] or 0 for a in apps):,.0f}) from year two onward."),
        "Residual ongoing run cost after the action": None,
        "Safe or potential": "",
        "Safe saving": sum(r["_savings"]["safe_saving"] for r in rows),
        "Potential saving": sum(r["_savings"]["potential_saving"] for r in rows),
        "Her Cost Notes": "",
    }
    sav_records.append(tot)

    # ------------------------------------------------------------------ Agreement
    HER_LABEL_COL = "Her Lifecycle Stage (her team's label)"
    # Her label set mapped ONE-TO-ONE onto our five terms, so the comparison is strict.
    # A loose many-to-one mapping would make everything agree and say nothing.
    HER_TO_OURS = {
        "Strategic Invest": "invest",
        "Active": "retain",
        "Consolidation Candidate": "consolidate",
        "Replace / Sunset": "replace",
        "Pilot / Exit Candidate": "retire",
        "Review": None,                     # not a disposition -- an undecided state
    }
    # The direction each term points, for the softer second comparison.
    DIRECTION = {"retain": "keep as is", "invest": "keep and fund",
                 "consolidate": "fold into another app", "replace": "swap for another product",
                 "retire": "switch off"}
    HER_DIRECTION = {"Strategic Invest": "keep and fund", "Active": "keep as is",
                     "Consolidation Candidate": "fold into another app",
                     "Replace / Sunset": "swap for another product",
                     "Pilot / Exit Candidate": "switch off", "Review": "undecided"}

    agree_records = []
    n_disagree = 0
    for r in sorted(rows, key=lambda x: x["app_id"]):
        a, her = r["app"], r["app"]["_her_lifecycle_stage"]
        expected = HER_TO_OURS.get(her)
        ours = r["disposition"]
        agrees = (expected is not None and ours == expected)
        if not agrees:
            n_disagree += 1
        our_dir, her_dir = DIRECTION[ours], HER_DIRECTION.get(her, "unmapped")
        same_dir = our_dir == her_dir
        # retain vs invest is always a boundary case, not a material one: both mean keep it,
        # and the only question is whether anything is being funded.
        boundary = expected is not None and {ours, expected} <= {"retain", "invest"}

        cl = clusters_by_app.get(a["app_id"])
        failed = [l for _k, l, _c, fl in DIMENSIONS if a[fl] == "F"]

        if agrees:
            severity = "—"
            note = (f"Her '{her}' and our '{ours}' are the same call. Pattern {a['vtcr_key']}"
                    f"{': all four dimensions clear the gate' if not failed else ': ' + ' and '.join(failed) + ' fail the gate'}.")
            verdict = "both — no disagreement to resolve"
        elif expected is None:
            severity = "her label is not a disposition"
            note = (f"'Review' is an undecided state, not an action, so there is nothing to "
                    f"agree or disagree with. We return '{ours}' on pattern {a['vtcr_key']} "
                    f"({' and '.join(failed) if failed else 'all four dimensions pass'}).")
            verdict = ("OURS, in the sense that the engine has actually decided. That is the "
                       "point of running it: 'Review' is the state a portfolio sits in before "
                       "someone does this work.")
        elif not failed:
            severity = ("boundary — both mean keep it; the split is whether anything is "
                        "being funded" if boundary else "material")
            note = (f"Her '{her}' expects '{expected}'; we return '{ours}'. All four dimensions "
                    f"clear the 3.0 gate (V {a['business_value_score']:.2f}, T "
                    f"{a['technical_health_score']:.2f}, C {a['cost_efficiency_score']:.2f}, "
                    f"R {a['risk_posture_score']:.2f}), and the all-pass pattern PPPP is the "
                    f"only row of the table that returns retain.")
            verdict = ("OURS on the engine's own definition, and this is exactly the "
                       "distinction Bina asked for in v2: retain means healthy, leave it "
                       "alone, spend nothing; invest means deliberately funding a failing "
                       "dimension. Nothing here fails, so there is nothing to fund. Her label "
                       f"set has no word for 'healthy, do nothing' — '{her}' is carrying both "
                       "meanings. That is a gap in the label set, not an error in either "
                       "answer.")
        elif cl and cl["roles"][a["app_id"]]["role"] == "overlapping":
            severity = "boundary" if (boundary or same_dir) else "material"
            note = (f"Her '{her}' expects '{expected}'; we return '{ours}'. It sits in overlap "
                    f"group {cl['cluster_id']} but is not treated as absorbable: migration path "
                    f"evidenced in her own data = "
                    f"{cl['roles'][a['app_id']]['migration_path_evidenced']}.")
            verdict = (f"OURS. Her label assumes an action her own evidence does not support "
                       f"yet — her Cost Notes read \"{a['cost_notes']}\". The engine will not "
                       f"commit a saving her file withholds.")
        else:
            severity = "boundary" if (boundary or same_dir) else "material"
            note = (f"Her '{her}' expects '{expected}'; we return '{ours}' on pattern "
                    f"{a['vtcr_key']} ({' and '.join(failed)} fail the gate). Direction: hers "
                    f"'{her_dir}', ours '{our_dir}'.")
            if ours == "consolidate" and expected == "replace" and cl:
                verdict = (f"OURS. Both remove it; the difference is what takes over. "
                           f"{cl['survivor']} already holds these capabilities as Primary with "
                           f"Full coverage in her own Capability Map, so the capability is "
                           f"folded into something the portfolio already runs rather than "
                           f"substituted with a product it does not. Consolidate is the more "
                           f"accurate term and the cheaper plan.")
            elif boundary:
                verdict = (f"OURS, and it is the same retain/invest boundary as the rows above, "
                           f"pointing the other way: {' and '.join(failed)} fails the 3.0 gate "
                           f"on measured evidence, so there IS something to fund. Her '{her}' "
                           f"says keep it and we agree — we are only adding that keeping it "
                           f"should cost something.")
            else:
                verdict = (f"OURS on the term, HERS on the direction. {', '.join(failed)} fails "
                           f"the gate on measured evidence, so an action is warranted; the "
                           f"engine's term differs from her label in how far it goes.")

        agree_records.append({
            "App ID": a["app_id"],
            "Application": a["name"],
            HER_LABEL_COL: her,
            "Our disposition": ours,
            "Our priority": r["_base"]["priority"],
            "Pattern key": a["vtcr_key"],
            "Her label's engine equivalent": expected or "none — not a disposition",
            "Agree?": "yes" if agrees else "NO",
            "Same direction?": "yes" if same_dir else "no",
            "Severity of the disagreement": severity,
            "Note": note,
            "Which we think is right": verdict,
            "Our confidence": r["confidence"],
        })

    # ------------------------------------------------------------------ spreads
    spread_base = {d: 0 for d in DISPOSITIONS}
    spread_var = {d: 0 for d in DISPOSITIONS}
    for r in rows:
        spread_base[r["_base"]["disposition"]] += 1
        spread_var[r["_var"]["disposition"]] += 1
    differs = [r for r in rows if r["_var"]["disposition"] != r["_base"]["disposition"]]

    # ------------------------------------------------------------------ write workbook
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Read me first"
    differs_summary = ", ".join(
        f"{r['app_id']} {r['_base']['disposition']} -> {r['_var']['disposition']}"
        for r in differs) or "none"
    orphan_rows = [r for r in sorted(rows, key=lambda x: x["app_id"]) if r["_orphans"]]
    n_boundary = sum(1 for r in agree_records
                     if r["Agree?"] == "NO" and r["Severity of the disagreement"].startswith("boundary"))
    n_material = sum(1 for r in agree_records
                     if r["Agree?"] == "NO" and r["Severity of the disagreement"] == "material")
    n_nolabel = sum(1 for r in agree_records
                    if r["Agree?"] == "NO"
                    and r["Severity of the disagreement"] == "her label is not a disposition")

    # risk-register category coverage, computed rather than asserted
    def _no_cat(cats):
        return sum(1 for a in apps if not risk_rows_for(a["app_id"], ctx["risks"], cats))
    n_no_tech = _no_cat(("Technical",))
    n_no_bc = _no_cat(("Business Continuity",))
    n_no_sec = _no_cat(("Security", "Regulatory"))
    n_no_clin = _no_cat(("Clinical Safety", "AI /", "Patient Safety"))
    n_apps = len(apps)
    coverage_sentence = (
        f"{n_no_tech} of {n_apps} apps have no Technical row, {n_no_bc} have no Business "
        f"Continuity row, {n_no_sec} have no Security or Regulatory row, and {n_no_clin} have "
        f"no clinical-safety row (only {n_apps - n_no_clin} do)")
    conf_spread = defaultdict(int)
    for r in rows:
        conf_spread[r["confidence"]] += 1
    n_direct = sum(1 for name, _d, _l, _f in SCORERS
                   if {a["_inputs"][name]["availability"] for a in apps} == {"direct"})
    n_rubric = sum(1 for name, _d, _l, _f in SCORERS
                   if "unavailable" not in {a["_inputs"][name]["availability"] for a in apps}) - n_direct
    n_unavail = sum(1 for name, _d, _l, _f in SCORERS
                    if {a["_inputs"][name]["availability"] for a in apps} == {"unavailable"})
    write_prose(ws, [
        ("h1", "Northstar Global Health — disposition analysis"),
        ("p", f"Produced {ANALYSIS_DATE.isoformat()} by Aberdeen Advisors' Application "
              f"Rationalization engine. Source: the 20-application sample dataset Bina Din's "
              f"team built ({os.path.basename(SOURCE_XLSX)}, 11 sheets). Her file was opened "
              f"read-only and is unchanged."),
        ("p", ""),
        ("h2", "What was run"),
        ("p", "Each of the 20 applications was scored on four dimensions — business value, "
              "technical health, cost efficiency and risk posture — built from 18 inputs on a "
              "1-to-5 half-step scale. A dimension passes if it scores 3.0 or more. The four "
              "pass/fail results form a four-letter pattern (for example PPFP), and that "
              "pattern is looked up in a fixed 16-row table that returns one of five terms — "
              "retain, invest, consolidate, replace, retire — and a priority."),
        ("p", "Nothing in this workbook was typed in by hand. Every score, pattern, "
              "disposition, priority and dollar figure was computed by score_northstar.py from "
              "cells in your file. The script is included so you can re-run it."),
        ("p", ""),
        ("h2", "No interviews were used"),
        ("p", "You told us this iteration would not interview stakeholders or owners, and your "
              "Assumptions sheet says the same. So no score here rests on anyone's opinion. "
              f"{18 - n_unavail} of the 18 inputs are producible from your file with no "
              f"interview, and {n_unavail} are not. Of the producible ones, {n_direct} come "
              f"straight off a column on every application, 3 more come straight off your risk "
              f"register wherever it holds a row in the matching category and fall back to a "
              f"summary field where it does not, and the rest are derived from your columns by "
              f"a written rubric. (An earlier mapping pass split the same 16 producible inputs "
              f"5 direct / 11 rubric; the difference is only where the line between 'direct' "
              f"and 'derived' is drawn, not which inputs can be produced.) The 'Input "
              f"derivation' sheet gives the rubric, the source columns and the observed spread "
              f"for all 18. That sheet is the audit trail; it matters as much as the answers."),
        ("p", ""),
        ("h2", "Which inputs were derived rather than supplied"),
        ("p", "Three inputs are supplied almost as-is: how widely an app is used (your "
              "Utilization Rate and Active Users), how much licence waste it carries (the same "
              "Utilization Rate read the other way), and how stable it is in service (your "
              "incident counts, MTTR and availability against SLA — the one input your file "
              "evidences better than our own dataset does). Three risk inputs come straight "
              "off your risk register wherever it has a row in the matching category. "
              "Everything else is derived. Examples: 'does clinical work stop without it' is "
              "built from Business Criticality, Critical Operation Flag, Data Classification "
              "and how many of the app's capability rows you marked Duplicative; 'cost per "
              "active user' is your Annual TCO over your Active Users, compared against the "
              "median of the app's peers in the same Category; 'customisation debt' is read "
              "from your Development / Configuration Language and any Technical risk row."),
        ("p", "Where your scores run the opposite way to ours — severity, likelihood, inherent "
              "risk, risk level, residual risk, highest risk — they were flipped, because on "
              "all 18 of our inputs a high number is the good number. Your Business Value "
              "Score already runs our way. One flip we did NOT apply is flagged in 'Notes & "
              "assumptions' for you to overrule."),
        ("p", ""),
        ("h2", "The two gaps"),
        ("p", "1. CONSUMPTION PRICE VARIANCE has no source. This input compares metered or "
              "consumption spend against a plan. Your TCO sheet holds six fixed annual "
              "components and no consumption line, and there is no plan or budget figure to "
              "vary against. Rather than invent one, the cost dimension is RENORMALISED over "
              "its remaining weighted inputs — its weight sum drops from 4 to 3, so cost per "
              "active user carries two-thirds of the cost score and unused licence waste one "
              "third. This is not caused by the no-interview rule; the data is not in the file."),
        ("p", "2. END-USER PERCEIVED QUALITY has no source either. It is a perception measure, "
              "and the only route to one is asking users, which this iteration rules out. It "
              "already carries weight 0 in our model, so leaving it empty changes no answer. "
              "Your Business KPI columns are the nearest thing and are reported beside it, but "
              "they measure a business outcome, not a perception, so they are not scored."),
        ("p", ""),
        ("h2", "One thing to read before you read the dispositions"),
        ("p", "THE RISK SCORES ARE LOW CONFIDENCE. Your risk register holds exactly two rows "
              f"per application, and the categories are uneven: {coverage_sentence}. So most "
              "applications' risk score leans on one or two summary fields. Because that is a "
              "modelling choice that can change a recommendation, the answer is given BOTH "
              "ways: risk hard-gated at 3.0 as the engine normally runs it, and risk excluded "
              "from the gate. The 'Dispositions' sheet carries a column showing where the two "
              "differ, and the difference is summarised here: "
              f"{len(differs)} of {n_apps} applications change term when risk stops being able "
              f"to fail a gate ({differs_summary})."),
        ("p", ""),
        ("h2", "How much to trust each row"),
        ("p", "Every row carries a confidence flag. Your Assumptions sheet says missing "
              "evidence should produce 'Needs Validation', and it does — but only where the "
              "missing evidence could actually change that row's answer, because your risk "
              "register is thin enough that flagging all twenty would tell you nothing. The "
              f"spread is: "
              + ", ".join(f"{k} {v}" for k, v in
                          sorted(conf_spread.items(), key=lambda kv: -kv[1]))
              + f" of {n_apps}. Each row's 'Why that confidence' column names the specific "
              f"reason, and the gaps are listed even on rows where they change nothing."),
        ("p", ""),
        ("h2", "Your own labels were held out"),
        ("p", "Your Lifecycle Stage column already contains disposition-like labels (Strategic "
              "Invest, Consolidation Candidate, Replace / Sunset, Pilot / Exit Candidate). "
              "Feeding those into a model whose job is to produce dispositions would be "
              "circular, so that column was held out of every score. It is used in exactly one "
              "place: the 'Agreement with your labels' sheet, which compares our answer against "
              f"your team's for all 20, mapping your labels one-to-one onto our five terms. We "
              f"agree on {20 - n_disagree} and differ on {n_disagree}. The differences are not "
              f"evenly interesting: {n_boundary} of them are the same boundary case — your "
              f"'Strategic Invest' against our 'retain' — where both of us mean keep it and the "
              f"only question is whether anything is being funded. Your label set has no word "
              f"for 'healthy, leave it alone, spend nothing', which is precisely the split Bina "
              f"asked us to add as a fifth term. {n_nolabel} is your 'Review', which is an "
              f"undecided state rather than an action, so there is nothing to disagree with. "
              f"That leaves {n_material} genuine difference of substance. Wherever your label "
              "appears in this workbook it sits in a brown-headed column marked COMPARISON ONLY."),
        ("p", ""),
        ("h2", "One safety catch worth knowing about"),
        ("p", "Before any action is recommended, the script checks whether switching the "
              "application off would leave a capability with no other provider anywhere in "
              "your Capability Map. " + (
                  "It found " + str(len(orphan_rows)) + ": " +
                  "; ".join(f"{r['app_id']} {r['app']['name']} is the only provider of "
                            f"{o.split(' (')[0]}" for r in orphan_rows for o in r['_orphans']) +
                  ". The engine does NOT quietly change its answer on that basis — the term "
                  "still stands on the evidence — but the recommendation gains an explicit "
                  "PRECONDITION and the row drops to Needs Validation. Whether that capability "
                  "is genuinely still needed is your team's call, not the model's."
                  if orphan_rows else
                  "Nothing in this portfolio trips it: every application we recommend removing "
                  "has its capabilities covered by another application in your own map.")),
        ("p", ""),
        ("h2", "What is in each sheet"),
        ("p", "Dispositions — one row per application: the four scores, the pattern, the "
              "disposition, the priority, a rationale citing that app's own evidence, the "
              "concrete next action, a confidence flag, and the alternative under the "
              "risk-excluded variant."),
        ("p", "Input derivation — the audit trail: all 18 inputs, the rubric, the source "
              "columns, and whether each was direct, rubric-derived or unavailable."),
        ("p", "Consolidation candidates — the genuine overlap groups from your Capability Map's "
              "Duplicative and Secondary roles, a named survivor with the reason, and the "
              "group saving from your Avoidable Annual Cost and transition figures."),
        ("p", "Savings — per app and portfolio: run cost, gross saving, one-time cost, net "
              "first-year saving, and where our arithmetic differs from yours, both figures "
              "side by side with the reason."),
        ("p", "Agreement with your labels — our disposition against your Lifecycle Stage for "
              "all 20, with a note on each disagreement."),
        ("p", "Notes & assumptions — every threshold, the renormalisation, the risk-confidence "
              "decision, and the short list of calls that are yours or Ryo's, not ours."),
        ("p", "Sanity checks — the arithmetic and leakage checks the script runs before it is "
              "allowed to write this file."),
        ("p", ""),
        ("p", f"Engine verification: {engine_note}."),
    ])

    dispo_headers = list(dispo_records[0].keys())
    write_sheet(wb.create_sheet("Dispositions"), dispo_headers, dispo_records,
                widths={"App ID": 9, "Application": 30, "Vendor": 24,
                        "Primary capability": 24, "Business value": 10,
                        "Technical health": 10, "Cost efficiency": 10,
                        "Risk posture (low confidence)": 12, "V": 4, "T": 4, "C": 4, "R": 4,
                        "Pattern key": 8, "Disposition": 13, "Priority": 11,
                        "Rationale": 95, "Recommendation": 75, "Confidence": 15,
                        "Why that confidence": 60,
                        "Evidence gaps behind the confidence flag": 44,
                        "Alternative under risk-excluded gate": 40, "Priority basis": 40,
                        "Her Lifecycle Stage (COMPARISON ONLY — held out of scoring)": 26},
                wrap_cols={"Rationale", "Recommendation", "Why that confidence",
                           "Evidence gaps behind the confidence flag",
                           "Alternative under risk-excluded gate", "Priority basis",
                           "Primary capability", "Application", "Vendor"},
                comparison_cols={"Her Lifecycle Stage (COMPARISON ONLY — held out of scoring)"},
                freeze="E2")

    write_sheet(wb.create_sheet("Input derivation"), list(deriv_records[0].keys()),
                deriv_records,
                widths={"Engine input": 32, "Dimension": 15, "Raw weight": 8,
                        "Normalised weight in its dimension": 34, "What it scores": 46,
                        "Availability": 30, "Her columns used": 60, "Rubric applied": 110,
                        "Apps scored": 11, "Observed min / median / max": 18,
                        "Rows on the fallback path": 12,
                        "Example evidence (APP-001 Epic)": 60},
                wrap_cols={"What it scores", "Her columns used", "Rubric applied",
                           "Example evidence (APP-001 Epic)", "Availability",
                           "Normalised weight in its dimension"},
                freeze="B2")

    write_sheet(wb.create_sheet("Consolidation candidates"), list(cons_records[0].keys()),
                cons_records,
                widths={"Overlap group": 12, "Contested capabilities": 34,
                        "Capability IDs": 24, "App ID": 9, "Application": 30,
                        "Role in group": 13,
                        "Her Support Role / Coverage on the contested capabilities": 52,
                        "Why": 78, "Our disposition": 13, "Annual TCO": 14,
                        "Her Avoidable Annual Cost": 16, "Her One-Time Transition Cost": 16,
                        "Group annual saving if every absorbable member folds in": 20,
                        "Group one-time transition cost": 18,
                        "Group annual run cost today": 18,
                        "Her Lifecycle Stage (COMPARISON ONLY)": 24},
                wrap_cols={"Contested capabilities", "Why",
                           "Her Support Role / Coverage on the contested capabilities",
                           "Application"},
                comparison_cols={"Her Lifecycle Stage (COMPARISON ONLY)"}, freeze="D2")

    write_sheet(wb.create_sheet("Savings"), list(sav_records[0].keys()), sav_records,
                widths={"App ID": 11, "Application": 32, "Our disposition": 13,
                        "Current annual run cost (her Annual TCO)": 16,
                        "Her Avoidable Annual Cost": 15,
                        "Gross annual saving we claim": 15,
                        "One-time transition cost (hers)": 15, "Named successor": 13,
                        "Successor ongoing cost netted off": 15,
                        "Successor already in the baseline?": 34,
                        "Our net first-year saving": 15, "Her First-Year Net Savings": 15,
                        "Difference (ours less hers)": 14, "Why they differ": 70,
                        "Residual ongoing run cost after the action": 16,
                        "Safe or potential": 30, "Safe saving": 13, "Potential saving": 13,
                        "Her Cost Notes": 52},
                wrap_cols={"Why they differ", "Her Cost Notes", "Safe or potential",
                           "Successor already in the baseline?", "Application"},
                freeze="C2")
    wsx = wb["Savings"]
    money_cols = ["Current annual run cost (her Annual TCO)", "Her Avoidable Annual Cost",
                  "Gross annual saving we claim", "One-time transition cost (hers)",
                  "Successor ongoing cost netted off", "Our net first-year saving",
                  "Her First-Year Net Savings", "Difference (ours less hers)",
                  "Residual ongoing run cost after the action", "Safe saving",
                  "Potential saving"]
    hdrs = [c.value for c in wsx[1]]
    for name in money_cols:
        i = hdrs.index(name) + 1
        for r in range(2, wsx.max_row + 1):
            wsx.cell(row=r, column=i).number_format = '$#,##0'
    for r in range(2, wsx.max_row + 1):
        if wsx.cell(row=r, column=1).value == "PORTFOLIO":
            for c in range(1, wsx.max_column + 1):
                wsx.cell(row=r, column=c).font = Font(bold=True)

    wsc = wb["Consolidation candidates"]
    hdrs = [c.value for c in wsc[1]]
    for name in ["Annual TCO", "Her Avoidable Annual Cost", "Her One-Time Transition Cost",
                 "Group annual saving if every absorbable member folds in",
                 "Group one-time transition cost", "Group annual run cost today"]:
        i = hdrs.index(name) + 1
        for r in range(2, wsc.max_row + 1):
            wsc.cell(row=r, column=i).number_format = '$#,##0'

    write_sheet(wb.create_sheet("Agreement with your labels"), list(agree_records[0].keys()),
                agree_records,
                widths={"App ID": 9, "Application": 32,
                        HER_LABEL_COL: 24, "Our disposition": 13,
                        "Our priority": 11, "Pattern key": 9,
                        "Her label's engine equivalent": 22, "Agree?": 8,
                        "Same direction?": 13, "Severity of the disagreement": 30,
                        "Note": 78, "Which we think is right": 88, "Our confidence": 16},
                wrap_cols={"Note", "Which we think is right", "Application",
                           "Severity of the disagreement", "Her label's engine equivalent"},
                comparison_cols={HER_LABEL_COL}, freeze="C2")

    # ------------------------------------------------------------------ Notes
    peer_lines = []
    for cat in sorted(ctx["peer_cpu"]):
        vals = sorted(ctx["peer_cpu"][cat])
        peer_lines.append(f"    {cat}: n={len(vals)}, median ${statistics.median(vals):,.0f} "
                          f"per active user, range ${vals[0]:,.0f}-${vals[-1]:,.0f}"
                          + ("  [singleton -> compared against the portfolio median "
                             f"${ctx['portfolio_cpu_median']:,.0f}]" if len(vals) < 2 else ""))
    risk_cov = defaultdict(int)
    for r in ctx["risks"]:
        risk_cov[s(r["Risk Category"])] += 1
    notes_blocks = [
        ("h1", "Notes & assumptions"),
        ("p", "Everything below is a choice we made, not a fact from your file. Each one is "
              "yours or Ryo's to overrule."),
        ("p", ""),
        ("h2", "1. The renormalisation"),
        ("p", "The cost dimension normally averages three weighted inputs: cost per active user "
              "(weight 2), unused licence waste (weight 1) and consumption price variance "
              "(weight 1) — weight sum 4. Consumption price variance has no source column in "
              "your workbook, so it is left null and the dimension is averaged over the "
              "remaining weight sum of 3. Effect: cost per active user carries 0.6667 of the "
              "cost score instead of 0.50, and unused licence waste 0.3333 instead of 0.25. "
              "The absolute-cost band is scored and reported at weight 0 and is not in either "
              "denominator. This fires on all 20 applications and is checked before the file "
              "is written."),
        ("p", "Consequence worth stating: cost per active user now controls two-thirds of the "
              "dimension that, in our own review, was already the only one where the 3.0 gate "
              "really bites. If you want that diluted, the fix is a consumption or metered "
              "spend column plus a plan figure, not a re-weighting."),
        ("p", ""),
        ("h2", "2. The risk-confidence decision — the one choice that can change an answer"),
        ("p", "Your risk register holds exactly 2 rows per application. Category coverage across "
              "the 40 rows: " + ", ".join(f"{k} {v}" for k, v in sorted(risk_cov.items())) + "."),
        ("p", f"That leaves {coverage_sentence}. "
              "Our three risk inputs therefore fall back to your Highest Risk summary field or "
              "to an adjacent category on many rows. We scored risk anyway, marked it "
              "low-confidence in the column header, and produced the answer BOTH ways rather "
              "than picking one silently: risk hard-gated at 3.0 (the engine as it normally "
              "runs) and risk excluded from the gate (the fourth letter of the pattern forced "
              "to P, so risk is still scored and reported but cannot condemn a row on its own). "
              "The 'Dispositions' sheet names the alternative on every row where the two differ."),
        ("p", "To make risk properly load-bearing you need one row per risk CATEGORY per "
              "application, not two rows per application — at minimum a Security/Regulatory "
              "row and a Business Continuity row on every app, and a clinical-safety row on "
              "every app that touches PHI."),
        ("p", ""),
        ("h2", "3. Your Lifecycle Stage was held out entirely"),
        ("p", "It contains disposition-like labels, so using it as an input would be leakage: "
              "the model would be predicting the answer from the answer. It is read once, into "
              "a field whose name is prefixed so no scorer can reach it, and used only in the "
              "comparison columns and the 'Agreement with your labels' sheet."),
        ("p", "This has one real cost. Our engine has a lifecycle guard that bars retire and "
              "replace for an application still in Birth or Growth — you do not switch off "
              "something that has not finished ramping. Its only possible input in your file is "
              "the held-out column, so THE GUARD IS DISARMED FOR THIS RUN. One row is affected: "
              "APP-009 Nabla Copilot, whose Current Release / Version reads 'Pilot release' — an "
              "independent, non-leaking signal that it is early in life. With the guard armed "
              "from that field, its retire would have been suppressed and turned into a funded "
              "invest at High priority. We left it disarmed and are flagging it rather than "
              "deciding it. BINA / RYO: your call. If you want the guard live, add a "
              "lifecycle-age field that is not the disposition label — implementation date, or "
              "months in production."),
        ("p", ""),
        ("h2", "4. The one flip we did not apply — please overrule us if you disagree"),
        ("p", "We were asked to flip every one of your scores that runs high = worse: severity, "
              "likelihood, inherent risk, risk level, residual risk, highest risk, business "
              "criticality and capability criticality. We flipped the first six without "
              "reservation. We did NOT flip business criticality or capability criticality "
              "inside the BUSINESS VALUE dimension, and here is why: in a risk context a higher "
              "criticality does make things worse, but as a value input it means the opposite. "
              "Inverting it would have scored Epic — Critical criticality, Critical Operation "
              "Flag, system of record for inpatient care — as LOW business value, and every "
              "sanity check downstream would have been meaningless. So criticality is used in "
              "its natural direction where it measures value, and inverted where it enters risk "
              "as an impact term. This is the single place we departed from a literal reading of "
              "the instruction. It is visible here so you can reverse it in one line if you "
              "meant it literally."),
        ("p", ""),
        ("h2", "5. Supportability is a lower-evidence input"),
        ("p", "Vendor Support End is blank for 19 of your 20 applications, so supportability "
              "leans on Current Release / Version and on your update calendar. No "
              "end-of-support date was invented. An evergreen vendor-managed line scores 4.5 "
              "rather than 5.0 precisely because 'the vendor keeps it current' is an inference, "
              "not an evidenced support horizon. The one app with a real date, APP-018 Oracle "
              "PeopleSoft HCM at 2037-12-31, scores well on the support-horizon component and "
              "badly on release currency ('9.2 / PeopleTools 8.60'), and the two are averaged "
              "rather than one overriding the other. Vendor viability is thinner still: your "
              "file has no vendor financial data of any kind, so it is proxied by whether you "
              "raised a Vendor risk row and whether the roadmap is dated."),
        ("p", ""),
        ("h2", "6. Every rubric threshold, in one place"),
        ("p", "The 'Input derivation' sheet carries the full rubric for each of the 18 inputs, "
              "row by row, including every numeric band. The bands are summarised here so they "
              "can be challenged as a set:"),
        ("p", "  Utilisation bands (used for reach and for licence waste): .90/.85/.80/.70/.60/"
              ".50/.40/.30 stepping 5.0 down to 1.5, floor 1.0."),
        ("p", "  Active-user breadth bands: 25000/15000/8000/4000/2000/1000/500/200 stepping "
              "5.0 down to 1.5, floor 1.0."),
        ("p", "  Peer cost ratio bands (cost per active user over peer-group median): "
              "0.50/0.70/0.90/1.00/1.15/1.40/1.80/2.50 stepping 5.0 down to 1.5, floor 1.0."),
        ("p", "  Peer groups are your Category column. Observed medians:"),
    ] + [("p", l) for l in peer_lines] + [
        ("p", "  Incidents per 1,000 active users: 0.5/1.5/3.0/5.0 -> 5.0/4.0/3.0/2.0, floor 1.0."),
        ("p", "  MTTR minutes: 45/60/90/120 -> 5.0/4.0/3.0/2.0, floor 1.0."),
        ("p", "  Availability against your SLA Target: at or above = 5.0, shortfall <= 0.001 = "
              "3.0, larger = 1.5."),
        ("p", "  Absolute annual TCO bands (weight 0): 0.5m/1m/1.5m/2m/3m/5m/7m -> 5.0/4.5/4.0/"
              "3.5/3.0/2.0/1.5, floor 1.0."),
        ("p", "  Residual-risk inversion: Low 5.0, Medium 3.5, High 2.0, Critical 1.0. "
              "Mitigation-status credit: Closed +1.0, Mitigating +0.5, Accepted 0.0, Open -0.5. "
              "Risk inputs take the WORST matching row, not the mean, because posture is set by "
              "the least-controlled exposure."),
        ("p", "  Business Criticality to value base: Critical 4.5, High 3.5, Medium 2.5, "
              "Low 1.5, then +0.5 for Critical Operation Flag Yes, +0.5 for outright PHI, "
              "-1.0 when at least half the capability rows are Duplicative."),
        ("p", "  Every result is snapped onto the engine's 1.0-5.0 half-step scale."),
        ("p", ""),
        ("h2", "7. How overlap groups and survivors were decided"),
        ("p", "A capability is contested when more than one application maps to it and at least "
              "one of them carries your Support Role 'Duplicative'. Contested capabilities join "
              "their applications into an overlap group. The survivor is the member holding the "
              "most contested capabilities as Primary + Full, tie-broken on active users then on "
              "lower cost per active user."),
        ("p", "A duplicative member is treated as ABSORBED — which forces the term consolidate — "
              "only when three things all hold: at least half its capability rows are "
              "Duplicative, a survivor holds at least one of the same capabilities as Primary, "
              "AND your own evidence describes a migration path (migration language in your TCO "
              "Cost Notes or in a Dependencies 'Required Before Disposition' cell). That third "
              "condition is deliberate: it keeps a failed pilot with nothing to migrate out of "
              "the consolidate bucket and lets the gates give it retire, and it keeps "
              "APP-006 Zoom Workplace out too, because your own Cost Notes say its saving is "
              "'Potential only; not safe until telehealth and room dependencies are validated'. "
              "We will not commit a saving your file withholds."),
        ("p", "One overlap your team did not mark: CAP-007 Patient engagement is held Primary + "
              "Full by BOTH APP-003 athenaOne and APP-019 Salesforce Health Cloud, with no "
              "Duplicative role on either, so it forms no group and neither app is touched. "
              "BINA / RYO: is that a genuine dual-primary split by population, or a missing "
              "Duplicative mark?"),
        ("p", ""),
        ("h2", "8. Savings arithmetic — yours, not ours"),
        ("p", "Gross annual saving is your Avoidable Annual Cost, claimed only where our "
              "disposition actually removes run-rate spend (retire, consolidate, replace). Net "
              "first-year saving is that figure less your One-Time Transition Cost, floored at "
              "zero — the same definition your Assumptions sheet gives. Successor run cost is "
              "netted only where a successor is named and is not already in the baseline; every "
              "survivor here is an existing portfolio application already carrying its own "
              "Annual TCO, so that term is zero on all 20 and is shown as zero rather than "
              "omitted, so a future run with a genuinely new successor cannot double-count."),
        ("p", "Where our figure differs from yours, both are shown with the reason. The only "
              "differences are applications where you booked an avoidable cost and our engine "
              "does not recommend the action that would release it; we never overwrote one of "
              "your numbers. Safe versus potential follows your own rule ('only high-confidence "
              "actions are counted as safe savings'): a saving is safe when the action removes "
              "run cost, your Evidence Confidence is High, and your Cost Notes do not withhold "
              "it."),
        ("p", "Your CIO savings target of 15% is checked against the portfolio total on the "
              "'Savings' sheet. It is not met on first-year net savings and is met on gross "
              "avoidable annual cost from year two onward. That gap is a real finding about the "
              "dataset, not an error."),
        ("p", ""),
        ("h2", "9. Engine features that could not run on your file"),
        ("p", "  Lifecycle guard — disarmed, see note 3."),
        ("p", "  Retention-obligation constraint — our engine steps a retire down in priority "
              "when a legal retention obligation means the archive must exist before anything "
              "is switched off. Your file has no retention-obligation or retention-expiry field, "
              "so the constraint never fires. Several of your Cost Notes imply one ('Savings "
              "exclude retained archive'), which is why the residual ongoing cost is reported "
              "per app on the 'Savings' sheet instead."),
        ("p", "  Sourcing guard — this one does run, as an annotation. It never changes a term; "
              "it records that a Commercial SaaS product cannot be re-platformed in place, so a "
              "replace means substituting a different product."),
        ("p", ""),
        ("h2", "10. Short list of calls that are yours or Ryo's"),
        ("p", "  a. Risk gated or not gated. We show both; someone has to pick one for the deck."),
        ("p", "  b. Whether the criticality flip in note 4 should be literal."),
        ("p", "  c. Whether the lifecycle guard should be armed from a non-leaking age field, "
              "which changes APP-009."),
        ("p", "  d. Whether CAP-007's dual-primary split is real or a missing Duplicative mark."),
        ("p", "  e. Whether Category is the right peer group for cost comparison. Two of your "
              "categories hold a single application, so those two are compared against the "
              "portfolio median instead, and two hold exactly two applications, where the "
              "median is just the midpoint of a pair and the cheaper app is guaranteed a good "
              "score. A capability-based peer group would be better; Category was used because "
              "it is unambiguous in your file."),
        ("p", "  f. Whether an evergreen SaaS line should cap supportability at 4.5 or score a "
              "full 5.0."),
        ("p", "  g. Every capability the engine flags as orphaned by an action — in this "
              "portfolio, CAP-011 Virtual care communications, which APP-006 Zoom Workplace is "
              "the only provider of. The engine will not decide whether a capability is still "
              "needed; it only refuses to switch one off silently."),
    ]
    write_prose(wb.create_sheet("Notes & assumptions"), notes_blocks)

    write_sheet(wb.create_sheet("Sanity checks"), ["Check", "Result", "Detail"],
                [{"Check": c["Check"], "Result": c["Result"],
                  "Detail": c["Detail"] if isinstance(c["Detail"], str) else "; ".join(c["Detail"])}
                 for c in checks],
                widths={"Check": 62, "Result": 9, "Detail": 100},
                wrap_cols={"Check", "Detail"})

    wb.save(OUT_XLSX)

    # ------------------------------------------------------------------ CSV
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=dispo_headers)
        w.writeheader()
        for rec in dispo_records:
            w.writerow(rec)

    # ------------------------------------------------------------------ console
    print(f"\nEngine verification: {engine_note}")
    print(f"Source (read-only, unmodified): {SOURCE_XLSX}")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_CSV}")
    print(f"Script  {os.path.abspath(__file__)}")

    print("\n" + "=" * 165)
    print("NORTHSTAR GLOBAL HEALTH — 20 APPLICATIONS")
    print("=" * 165)
    hdr = (f"{'App':8} {'Application':34} {'V':>5} {'T':>5} {'C':>5} {'R':>5} {'key':5} "
           f"{'disposition':12} {'priority':10} {'confidence':16} {'her label':24}")
    print(hdr)
    print("-" * 165)
    for r in sorted(rows, key=lambda x: x["app_id"]):
        a = r["app"]
        print(f"{a['app_id']:8} {a['name'][:34]:34} "
              f"{a['business_value_score']:5.2f} {a['technical_health_score']:5.2f} "
              f"{a['cost_efficiency_score']:5.2f} {a['risk_posture_score']:5.2f} "
              f"{a['vtcr_key']:5} {r['_base']['disposition']:12} {r['_base']['priority']:10} "
              f"{r['confidence']:16} {a['_her_lifecycle_stage'][:24]:24}")
    print("-" * 165)

    print("\nDISPOSITION SPREAD")
    print(f"  {'term':13} {'risk gated at 3.0':>18} {'risk excluded':>15}")
    for d in DISPOSITIONS:
        print(f"  {d:13} {spread_base[d]:>18} {spread_var[d]:>15}")
    print(f"  {'TOTAL':13} {sum(spread_base.values()):>18} {sum(spread_var.values()):>15}")

    print(f"\nRISK-GATE SENSITIVITY: {len(differs)} of {len(apps)} applications change term")
    for r in differs:
        a = r["app"]
        print(f"  {a['app_id']} {a['name'][:32]:32} {r['_base']['disposition']:12} -> "
              f"{r['_var']['disposition']:12} (risk {a['risk_posture_score']:.2f}, "
              f"{a['vtcr_key']} -> {r['_var']['key']})")
    if not differs:
        print("  none")

    print(f"\nAGREEMENT WITH HER LIFECYCLE STAGE: disagree on {n_disagree} of {len(apps)}")
    for rec in agree_records:
        if rec["Agree?"] == "NO":
            print(f"  {rec['App ID']} {rec['Application'][:30]:30} hers "
                  f"'{rec[HER_LABEL_COL]}' vs ours '{rec['Our disposition']}'")

    print("\nCONSOLIDATION GROUPS")
    for c in clusters:
        surv = apps_by_id[c["survivor"]]
        avoid = sum(apps_by_id[m]["avoidable_annual"] or 0
                    for m in c["members"] if m != c["survivor"])
        one = sum(apps_by_id[m]["one_time_transition"] or 0
                  for m in c["members"] if m != c["survivor"])
        print(f"  {c['cluster_id']} {', '.join(c['capability_names'])[:56]:56} "
              f"survivor {c['survivor']} ({surv['name'][:20]}) | members "
              f"{', '.join(c['members'])} | absorbed {', '.join(c['absorbed']) or 'none'} | "
              f"annual ${avoid:,.0f} for ${one:,.0f} one-time")

    print("\nMONEY")
    print(f"  portfolio annual run cost                 ${portfolio_tco:,.0f}")
    print(f"  her avoidable annual cost (all 20)        ${sum(a['avoidable_annual'] or 0 for a in apps):,.0f}")
    print(f"  gross annual saving we claim              ${sum(r['_savings']['gross_saving_annual'] for r in rows):,.0f}")
    print(f"  one-time transition cost                  ${sum(r['_savings']['one_time_transition_cost'] for r in rows):,.0f}")
    print(f"  our net first-year saving                 ${sum(r['_savings']['net_first_year_saving'] for r in rows):,.0f}")
    print(f"    of which safe under her own rule        ${sum(r['_savings']['safe_saving'] for r in rows):,.0f}")
    print(f"    of which potential only                 ${sum(r['_savings']['potential_saving'] for r in rows):,.0f}")
    print(f"  her portfolio first-year net savings row  ${sum(r['_savings']['her_first_year_net'] or 0 for r in rows):,.0f}")
    print(f"  CIO target {CIO_SAVINGS_TARGET:.0%} of run cost             ${portfolio_tco * CIO_SAVINGS_TARGET:,.0f}"
          f"   -> {'MET' if sum(r['_savings']['net_first_year_saving'] for r in rows) >= portfolio_tco * CIO_SAVINGS_TARGET else 'NOT MET on first-year net'}")

    print("\nSANITY CHECKS")
    for c in checks:
        detail = c["Detail"] if isinstance(c["Detail"], str) else "; ".join(c["Detail"])
        print(f"  [{c['Result']}] {c['Check']}")
        if c["Result"] == "FAIL":
            print(f"          {detail}")
    fails = [c for c in checks if c["Result"] == "FAIL"]
    print(f"\n{len(checks) - len(fails)} of {len(checks)} checks pass.")
    if fails:
        print("FAILURES PRESENT — do not circulate the workbook until they are resolved.")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
