#!/usr/bin/env python3
"""
Aberdeen Advisors / hack-team-12 - Application Rationalization demo dataset generator.

Builds a 20-row synthetic application portfolio with every data point needed to run an
application rationalization exercise end to end, and writes:

    App-Rationalization-Dummy-Dataset-v2.xlsx   (7 sheets)
    applications-v2.csv                         (the Applications sheet, flat)
    README.md                                   (plain-language cover note)
    CHANGELOG-v2.md                             (what changed from v1 and why)

WHAT CHANGED IN v2 (all of it from Bina Din's answers to the five v1 open questions)
    1. FIVE disposition terms, not four: retain, invest, consolidate, replace, retire.
       retain = healthy, leave alone, no spend. invest = deliberately fund a remediation or
       an enhancement. v1's invest carried both meanings, separated only by priority.
    2. The value dimension's double weight moves from governance and compliance to
       patient-care criticality, and ov_enhance_services is renamed
       ov_patient_care_criticality so the column says what it scores.
    3. Q1, Q2 and Q4 were confirmed as they stood; all five answers are recorded verbatim
       on the "Notes & assumptions" sheet with the decision each one produced.
    Same 20 applications, same product names, same clusters, same trap cases, same cost and
    savings arithmetic. Only the vocabulary, the value weights and three rows' dispositions
    move.

DATA PROVENANCE
    Vendor and product names are real, widely documented commercial products a US health
    system would plausibly run (Bina's instruction: "Generate common healthcare app names
    rest can be dummy data"). EVERYTHING ELSE IS INVENTED - costs, contract identifiers and
    dates, user and licence counts, versions, owner names, legal entities, business units,
    cost centres, scores, savings. No real organisation's portfolio, contract terms or
    internal figures appear anywhere. Owner names are obviously fictional placeholders.

DESIGN AUTHORITY (read before editing)
    reqs-fields.md      - the team's 65 requirements + the 276-attribute data dictionary
    infotech-scoring.md - the licensed Info-Tech scoring engine, cell level
    infotech-tco.md     - the licensed Info-Tech TCO calculator, cell level
    roster-design.md    - the agreed 20-app roster, intended dispositions and trap cases

HOW IT WORKS
    Row data below is explicit and readable (a list of dicts, one per application) so the
    team can hand-edit it or extend it to 600 apps. Everything derived - the four dimension
    scores, the pass/fail flags, the VTCR key, the disposition, the priority, the TCO totals,
    utilisation, and the savings arithmetic - is COMPUTED by the functions in section 4 and
    is never hardcoded on the row. Section 6 then verifies every computed disposition
    against the intended disposition in roster-design.md and fails loudly on a mismatch.

Run:  python3 generate_dataset.py
"""

from __future__ import annotations

import csv
import datetime as dt
import os
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_NAME = "App-Rationalization-Dummy-Dataset-v2.xlsx"
CSV_NAME = "applications-v2.csv"
README_NAME = "README.md"
CHANGELOG_NAME = "CHANGELOG-v2.md"

DATASET_VERSION = "v2 (2026-08-14)"
SCORING_MODEL_VERSION = "infotech-04-derived / weights v2 / thresholds 3.0 / five-term vocabulary"
ANALYSIS_DATE = dt.date(2026, 8, 13)          # the "as at" date every date field is read against
TARGET_REDUCTION_PCT = 0.15                   # the CIO's 15% spend-reduction target
AMORTISATION_YEARS = 5                        # one-time costs spread over the 5-year horizon
TCO_HORIZON_YEARS = 5                         # undiscounted, matching the Info-Tech convention

# ---------------------------------------------------------------------------------------
# 1. THE SCORING ENGINE AS CONFIGURATION (not code) - per REQ 50 / REQ 52
# ---------------------------------------------------------------------------------------
# Four dimensions, 18 scored inputs, 1-5 in 0.5 steps, integer weights, weight 0 disables.
# The four-dimension gated approach follows the structure of the licensed Info-Tech
# 04-Rationalize-Your-Application-Portfolio tool: a weighted arithmetic mean per dimension,
# a per-dimension pass threshold, a four-character P/F key and a 16-row lookup. That tool is
# CITED, NOT REPRODUCED - it is licensed third-party material and this repository is public,
# so its lens names, its criterion names, its disposition words and its priority values are
# deliberately absent from this file. Every name and value below is the team's own. Dimensions
# 3 and 4 carry Cost Efficiency (REQ 23) and Risk (REQ 24/65) rather than the two lenses the
# reference tool ships in those slots - see the "Scoring model" sheet for that decision.

DIMENSIONS = [
    # (key, our dimension name, output column, pass-flag column)
    ("V", "Business value", "business_value_score", "v_pass"),
    ("T", "Technical health", "technical_health_score", "t_pass"),
    ("C", "Cost efficiency", "cost_efficiency_score", "c_pass"),
    ("R", "Risk posture", "risk_posture_score", "r_pass"),
]

# criterion -> (dimension, integer weight, what it scores)
CRITERIA = [
    ("ov_increase_value", "V", 1,
     "REQ 21 criticality to revenue: does the app carry money in or out?"),
    ("ov_reach_consumers", "V", 1,
     "REQ 21 breadth of use: active users vs licences purchased, usage breadth band."),
    ("ov_reduce_costs_efficiency", "V", 1,
     "REQ 21 process centrality: how central to the process it serves."),
    ("ov_patient_care_criticality", "V", 2,
     "REQ 21 criticality to patient care: does clinical work stop without it? Carries the value "
     "dimension's DOUBLE WEIGHT from v2 - Bina's answer to Q3."),
    ("ov_governance_compliance", "V", 1,
     "REQ 21 regulatory/trust alignment plus owner-stated strategic importance. Weight 1 from v2 "
     "(it held the double weight in v1) - Bina's answer to Q3."),

    ("th_supportability", "T", 2,
     "REQ 22 version currency and EOL proximity. Sole owner of these facts."),
    ("th_architecture_fit", "T", 2,
     "REQ 22 integration pattern, cloud readiness, enterprise-architecture fit."),
    ("th_operational_stability", "T", 1,
     "REQ 22 incident and ticket volume, proactive vs reactive maintenance."),
    ("th_vendor_viability", "T", 1,
     "REQ 22 vendor financial and roadmap viability. Sole owner of this fact."),
    ("th_customization_debt", "T", 1,
     "REQ 22 customization debt and platform supportability."),

    ("c_cost_per_active_user_vs_peers", "C", 2,
     "REQ 23 cost per active user against peer apps in the same capability. 5 = cheapest."),
    ("c_unused_licence_waste", "C", 1,
     "REQ 23 licensed-but-inactive seat waste. 5 = almost no waste."),
    ("c_consumption_price_variance", "C", 1,
     "REQ 23 consumption/metered spend against the modelled plan. 5 = on or under plan."),
    ("c_absolute_cost_band", "C", 0,
     "Absolute-dollar cost band. Scored but WEIGHT 0 - see Scoring model."),

    ("r_technical_risk", "R", 1,
     "REQ 24 single point of failure, DR/backup, vendor concentration, hardening. 5 = controlled."),
    ("r_business_compliance_risk", "R", 1,
     "REQ 24 PHI/HIPAA exposure, data residency, SOC 2 / HITRUST, lock-in. 5 = controlled."),
    ("r_clinical_safety_risk", "R", 1,
     "REQ 65 clinical / patient-safety risk posture. 5 = consequence fully mitigated."),
    ("r_end_user_perceived_quality", "R", 0,
     "End-user perceived quality. Collected and stored at WEIGHT 0 - see Scoring model."),
]

PASS_THRESHOLD = 3.0            # >= passes, so exactly 3.0 passes
SCORE_STEPS = [1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0]

# Our FIVE agreed terms, from v2. Bina Din, 2026-08-14, answering the v1 open question about
# invest doing double duty: "No, separate by invest, retain, consolidate, replace, and retire."
#   retain       healthy, leave it alone, no spend
#   invest       deliberately inject money or effort - fund a remediation or an enhancement
#   consolidate  fold into another application that keeps the capability alive
#   replace      substitute a different product for the same capability
#   retire       switch it off; the capability goes away or is already covered elsewhere
DISPOSITIONS = ("retain", "invest", "consolidate", "replace", "retire")

# The 16-row lookup: pattern key -> OUR five agreed terms and OUR priority ladder position
# (REQ 52 default table, v2). This table IS the mapping decision. It is configuration, not code.
#
# The 16 keys are the exhaustive enumeration of four pass/fail gates, so the left column is
# arithmetic rather than anyone's property. What each key MEANS is ours: the words come from
# Bina's five-term ruling above and the priority comes from PRIORITY_LADDER. The licensed
# Info-Tech tool cited at the top of section 1 has a lookup of the same shape holding its own
# vocabulary; we do not reproduce it, and no column here is copied from it.
#
# v1 mapped these keys onto four terms; v2 maps them onto five, and the whole of the difference
# is one row: the all-pass key PPPP now returns retain instead of invest. Every other row is
# unchanged from v1, and it holds together because of a property of the table rather than a
# judgement call about any row - see retain_or_invest().
DISPOSITION_TABLE = {
    # key    disposition     priority
    "PPPP": ("retain",      "Very Low"),
    "PPPF": ("invest",      "High"),
    "PPFP": ("invest",      "Moderate"),
    "PPFF": ("invest",      "High"),
    "PFPP": ("invest",      "Moderate"),
    "PFPF": ("invest",      "High"),
    "PFFP": ("replace",     "High"),
    "PFFF": ("replace",     "Very High"),
    "FPPP": ("consolidate", "Low"),
    "FPPF": ("consolidate", "Moderate"),
    "FPFP": ("retire",      "High"),
    "FPFF": ("retire",      "Very High"),
    "FFPP": ("consolidate", "Moderate"),
    "FFPF": ("retire",      "High"),
    "FFFP": ("retire",      "Very High"),
    "FFFF": ("retire",      "Very High"),
}

# What v1's four-term table emitted for each key, so the "Scoring model" sheet can show the
# before and after side by side. PPPP is the only row that moved.
V1_DISPOSITION_MAP = {k: ("invest" if k == "PPPP" else v[0])
                      for k, v in DISPOSITION_TABLE.items()}

PRIORITY_LADDER = ["Very Low", "Low", "Moderate", "High", "Very High"]
# High/Medium/Low confidence fractions used by the savings arithmetic. The three-band shape
# follows the Info-Tech disposition-prioritization tool (cited, not reproduced).
HML = {"High": 1.0, "Medium": 0.75, "Low": 0.25, "None": 0.0}

# Fields the scoring model actually consumes - the completeness_score denominator (REQ 12).
# vendor_eos_date is deliberately NOT in this list: a SaaS product legitimately has none, so
# counting it as missing would penalise every SaaS row for a field that does not apply.
COMPLETENESS_FIELDS = [
    "vendor_name", "deployment_model", "sourcing_type", "lifecycle_stage", "implementation_date",
    "version_installed", "business_owner", "technical_owner", "business_unit",
    "cost_centre", "primary_capability", "licences_purchased", "active_users", "last_signin_date",
    "process_centrality", "owner_stated_strategic_importance", "cost_licence_subscription",
    "cost_maintenance_dev_labour", "cost_indirect_and_training", "contract_id",
    "annual_contract_value", "term_end", "auto_renewal_flag", "licence_metric",
    "has_downstream_dependents", "integration_pattern", "retention_obligation_flag",
    "information_classification", "ov_increase_value", "ov_reach_consumers",
    "ov_reduce_costs_efficiency", "ov_patient_care_criticality", "ov_governance_compliance",
    "th_supportability", "th_architecture_fit", "th_operational_stability", "th_vendor_viability",
    "th_customization_debt", "c_cost_per_active_user_vs_peers", "c_unused_licence_waste",
    "c_consumption_price_variance", "r_technical_risk", "r_business_compliance_risk",
    "r_clinical_safety_risk",
]

# ---------------------------------------------------------------------------------------
# 2. COLUMN SPECIFICATION - drives the Applications sheet order AND the Data dictionary
# ---------------------------------------------------------------------------------------
# (theme, column, plain-language definition, type / allowed values, motivating REQ IDs)
COLUMNS = [
    # -- identity and lifecycle ----------------------------------------------------------
    ("Identity & lifecycle", "app_id", "Stable key for the application record.", "str (APP-0nn)", "7, 13"),
    ("Identity & lifecycle", "app_name", "Canonical application name. Real product name; everything else on the row is invented.", "str", "4, 7, 13"),
    ("Identity & lifecycle", "vendor_name", "Legal vendor entity, not the marketing brand.", "str", "7, 15, 60"),
    ("Identity & lifecycle", "description", "What the application does, in one sentence. Used for capability inference.", "str", "19"),
    ("Identity & lifecycle", "is_ai_tool", "The record is an AI tool, copilot, chatbot or agent.", "bool", "11, 26"),
    ("Identity & lifecycle", "ai_delivery_form", "Bought as a standalone AI product, or an AI feature inside an app already licensed.", "enum(standalone, embedded_feature) or blank", "11, 26"),
    ("Identity & lifecycle", "ai_capability_class", "The AI capability delivered (scribe, coding assistant, chatbot, imaging triage...).", "str", "11, 26"),
    ("Identity & lifecycle", "ai_host_app_id", "For an embedded AI feature, the app whose licence already entitles it.", "str FK to app_id", "11, 26"),
    ("Identity & lifecycle", "ai_already_entitled_elsewhere", "This AI capability is already paid for inside another licence we hold.", "bool", "26"),
    ("Identity & lifecycle", "ai_entitled_alternative_app_id", "The already-paid-for alternative to name in the finding.", "str FK to app_id", "26"),
    ("Identity & lifecycle", "deployment_model", "How the software is delivered to us.", "enum(SaaS, hosted, on-prem)", "15, 55"),
    ("Identity & lifecycle", "sourcing_type", "Procurement or build shape. Gates which dispositions are even legal.", "enum(COTS, custom in-house, SaaS, hybrid)", "51, 55"),
    ("Identity & lifecycle", "lifecycle_stage", "Where the app is in its own life. Birth/Growth block retire and replace outright.", "enum(Birth, Growth, Mature, End of Life)", "51"),
    ("Identity & lifecycle", "implementation_date", "Go-live date. Blank means unknown, which is itself a finding.", "date ISO-8601", "51"),
    ("Identity & lifecycle", "version_installed", "Version we are actually running.", "str", "15, 22"),
    ("Identity & lifecycle", "version_vendor_supported", "Version the vendor currently supports.", "str", "15"),
    ("Identity & lifecycle", "vendor_eos_date", "Vendor end-of-support date for the installed version.", "date ISO-8601", "17, 22"),
    ("Identity & lifecycle", "technical_obsolescence_flag", "Obsolescence is material rather than merely notional.", "bool", "17"),

    # -- ownership and organisation ------------------------------------------------------
    ("Ownership & org", "business_owner", "Accountable business owner. Invented name. Blank means nobody could be resolved.", "str", "14, 18"),
    ("Ownership & org", "technical_owner", "Accountable technical owner. Invented name.", "str", "18"),
    ("Ownership & org", "legal_entity", "Invented legal entity carrying the app; also the M&A duplication key.", "str (LE-01..LE-03)", "59, 60"),
    ("Ownership & org", "business_unit", "Business unit.", "str", "59"),
    ("Ownership & org", "department", "Department.", "str", "59"),
    ("Ownership & org", "cost_centre", "Cost centre carrying the spend. The key a saving is allocated against.", "str", "18, 42, 59"),
    ("Ownership & org", "is_orphaned", "No accountable owner could be established. Orphaned plus low usage is the top target.", "bool", "18, 43, 47"),
    ("Ownership & org", "governance_visibility", "Governance/visibility class: is it under IT oversight, with an owner and a contract?", "enum(Managed, Unmanaged, Unknown, Unsanctioned)", "63"),
    ("Ownership & org", "is_shadow_it", "Paid for or in use but not governed.", "bool", "8, 63"),

    # -- capability ----------------------------------------------------------------------
    ("Capability", "primary_capability", "Primary business capability. Redundancy detection groups on this field.", "enum (health-system taxonomy)", "19, 25, 62"),
    ("Capability", "secondary_capabilities", "Other capabilities the app also covers.", "list[str], semicolon separated", "19, 25"),
    ("Capability", "capability_tag_confidence", "Confidence in the capability tag. Low means the tag was inferred, not stated.", "float 0-1", "19, 28"),

    # -- usage and entitlement -----------------------------------------------------------
    ("Usage & entitlement", "licences_purchased", "Entitlements bought. Deliberately distinct from actual usage.", "int", "21, 23, 42"),
    ("Usage & entitlement", "active_users", "Distinct users with real activity in the measurement window.", "int", "21, 23, 25"),
    ("Usage & entitlement", "licence_utilisation_rate", "COMPUTED: active_users / licences_purchased.", "float 0-1", "23"),
    ("Usage & entitlement", "unused_licence_count", "COMPUTED: licensed-but-inactive seats.", "int", "23"),
    ("Usage & entitlement", "last_signin_date", "Most recent sign-in seen in the SSO log.", "date ISO-8601", "8, 42"),
    ("Usage & entitlement", "process_centrality", "How central the app is to the process it serves.", "enum(High, Medium, Low, None)", "21"),
    ("Usage & entitlement", "owner_stated_strategic_importance", "What the owner says the app is worth strategically. Owner-attested, not measured.", "score 1-5 in 0.5 steps", "14, 21"),

    # -- cost ----------------------------------------------------------------------------
    ("Cost", "cost_licence_subscription", "Info-Tech category 1: licence, subscription and external maintenance, annual.", "money USD annual", "20, 55"),
    ("Cost", "cost_upgrade_and_modules", "Info-Tech category 2: version upgrades and additional modules, annual.", "money USD annual", "20, 55"),
    ("Cost", "cost_maintenance_dev_labour", "Info-Tech category 3: internal developer, support and service labour, annual.", "money USD annual", "20, 55"),
    ("Cost", "cost_infrastructure_peripherals", "Info-Tech category 4: storage, servers, workstations, network, annual.", "money USD annual", "20, 55, 57"),
    ("Cost", "cost_indirect_and_training", "Info-Tech category 5: help desk, training, skilled-staff premium, travel, annual.", "money USD annual", "20, 55"),
    ("Cost", "tco_five_category_subtotal", "COMPUTED: the five cost categories the Info-Tech TCO calculator structures its model around, and nothing else. Lets you see what that model covers and what our extensions add.", "money USD annual", "20, 54"),
    ("Cost", "consumption_based_cost", "OUR EXTENSION: metered / per-request / per-token spend. Info-Tech has no line for this.", "money USD annual", "20, 55"),
    ("Cost", "annual_tco_recurring", "COMPUTED: the five categories plus consumption. The run-rate figure. Excludes one-time.", "money USD annual", "20, 54"),
    ("Cost", "one_time_implementation_cost", "OUR EXTENSION: initial implementation. Held separately and never netted into the run-rate.", "money USD one-time", "20, 32"),
    ("Cost", "five_year_cumulative_tco", "COMPUTED: annual run-rate x 5, undiscounted (no NPV), matching the Info-Tech convention.", "money USD", "54"),
    ("Cost", "cost_per_active_user", "COMPUTED: annual_tco_recurring / active_users.", "money USD", "23"),
    ("Cost", "unused_licence_spend", "COMPUTED: licence spend sitting on inactive seats. The fastest credible saving.", "money USD annual", "23"),

    # -- contract ------------------------------------------------------------------------
    ("Contract", "contract_id", "Contract or subscription register identifier. Blank means no contract record found.", "str", "16, 61"),
    ("Contract", "annual_contract_value", "Annual contract value from the order form.", "money USD", "16"),
    ("Contract", "term_start", "Current term start date.", "date ISO-8601", "16"),
    ("Contract", "term_end", "Current term end date.", "date ISO-8601", "16, 29"),
    ("Contract", "auto_renewal_flag", "The contract renews itself unless notice is served.", "bool", "16, 41"),
    ("Contract", "renewal_notice_days", "How many days before term end notice must be served.", "int days", "16, 37"),
    ("Contract", "notice_deadline_date", "COMPUTED: term_end minus renewal_notice_days. The real deadline.", "date ISO-8601", "16, 29, 37"),
    ("Contract", "in_notice_window_now", "COMPUTED: the notice deadline is still ahead of us and within 180 days.", "bool", "45"),
    ("Contract", "licence_metric", "What the contract prices on.", "enum(per-user, per-bed, enterprise, consumption)", "16"),
    ("Contract", "early_termination_penalty", "Penalty for terminating early. Compared against the saving before recommending.", "money USD", "29"),
    ("Contract", "contract_runway_months", "COMPUTED: whole months from the analysis date to term end.", "int months", "60"),

    # -- dependencies --------------------------------------------------------------------
    ("Dependencies", "has_downstream_dependents", "Something else breaks if this is switched off. Gates retire and consolidate.", "bool", "10, 29"),
    ("Dependencies", "dependency_count", "Number of integration edges touching this app, either direction.", "int", "10, 57"),
    ("Dependencies", "integration_pattern", "How it integrates.", "enum(API, HL7, flat file, middleware, point-to-point, none)", "22, 39"),
    ("Dependencies", "data_types_held", "Kinds of data held. Sizes the archival obligation.", "list[str], semicolon separated", "57, 53"),

    # -- the 18 scored inputs ------------------------------------------------------------
    ("Scoring inputs (V)", "ov_increase_value", "V1. Criticality to revenue: does the app carry money in or out? 1 low, 5 high.", "score 1-5 in 0.5 steps", "21, 52"),
    ("Scoring inputs (V)", "ov_reach_consumers", "V2. Breadth of use, active vs licensed.", "score 1-5 in 0.5 steps", "21, 52"),
    ("Scoring inputs (V)", "ov_reduce_costs_efficiency", "V3. Process centrality: how central to the process it serves.", "score 1-5 in 0.5 steps", "21, 52"),
    ("Scoring inputs (V)", "ov_patient_care_criticality", "V4 Criticality to patient care - does clinical work stop without it? WEIGHT 2 from v2 (Bina's Q3 ruling). Named ov_enhance_services in v1, before the criterion was renamed to say what it measures.", "score 1-5 in 0.5 steps", "21, 52"),
    ("Scoring inputs (V)", "ov_governance_compliance", "V5. Regulatory and trust alignment plus owner-stated importance. WEIGHT 1 from v2; it held the double weight in v1.", "score 1-5 in 0.5 steps", "21, 52"),
    ("Scoring inputs (T)", "th_supportability", "T1 (weight 2). Version currency and end-of-support proximity. Scored here and nowhere else.", "score 1-5 in 0.5 steps", "22, 52"),
    ("Scoring inputs (T)", "th_architecture_fit", "T2 (weight 2). Integration pattern, cloud readiness, architecture fit.", "score 1-5 in 0.5 steps", "22, 52"),
    ("Scoring inputs (T)", "th_operational_stability", "T3. Incident and ticket volume; proactive vs reactive maintenance.", "score 1-5 in 0.5 steps", "22, 52"),
    ("Scoring inputs (T)", "th_vendor_viability", "T4. Vendor financial and roadmap viability. Scored here and nowhere else.", "score 1-5 in 0.5 steps", "22, 52"),
    ("Scoring inputs (T)", "th_customization_debt", "T5. Customization debt and platform supportability.", "score 1-5 in 0.5 steps", "22, 52"),
    ("Scoring inputs (C)", "c_cost_per_active_user_vs_peers", "C1 (weight 2). Cost per active user against capability peers. 5 = cheapest.", "score 1-5 in 0.5 steps", "23, 52"),
    ("Scoring inputs (C)", "c_unused_licence_waste", "C2. Licensed-but-inactive seat waste. 5 = almost none.", "score 1-5 in 0.5 steps", "23, 52"),
    ("Scoring inputs (C)", "c_consumption_price_variance", "C3. Consumption or unit-price spend against plan. 5 = on or under plan.", "score 1-5 in 0.5 steps", "23, 52"),
    ("Scoring inputs (C)", "c_absolute_cost_band", "C4, WEIGHT 0. Absolute-dollar cost band. Scored, disabled, kept visible.", "score 1-5 in 0.5 steps", "20, 52"),
    ("Scoring inputs (R)", "r_technical_risk", "R1. Technical risk posture: SPOF, DR/backup, hardening. 5 = controlled.", "score 1-5 in 0.5 steps", "24, 52"),
    ("Scoring inputs (R)", "r_business_compliance_risk", "R2. Compliance posture: PHI, residency, SOC 2 / HITRUST, lock-in. 5 = controlled.", "score 1-5 in 0.5 steps", "24, 52"),
    ("Scoring inputs (R)", "r_clinical_safety_risk", "R3. Clinical and patient-safety posture. 5 = consequence of outage fully mitigated.", "score 1-5 in 0.5 steps", "65, 24"),
    ("Scoring inputs (R)", "r_end_user_perceived_quality", "R4, WEIGHT 0. End-user perceived quality. Collected and stored, contributes nothing.", "score 1-5 in 0.5 steps", "52"),

    # -- gate outputs --------------------------------------------------------------------
    ("Gate output", "business_value_score", "COMPUTED dimension V: weighted mean of the five V inputs, renormalised over populated ones.", "float 1-5", "21, 52"),
    ("Gate output", "technical_health_score", "COMPUTED dimension T.", "float 1-5", "22, 52"),
    ("Gate output", "cost_efficiency_score", "COMPUTED dimension C.", "float 1-5", "23, 52"),
    ("Gate output", "risk_posture_score", "COMPUTED dimension R. High score = risk controlled, not risk absent.", "float 1-5", "24, 65, 52"),
    ("Gate output", "v_pass", "COMPUTED: business value >= 3.0.", "P or F", "52"),
    ("Gate output", "t_pass", "COMPUTED: technical health >= 3.0.", "P or F", "52"),
    ("Gate output", "c_pass", "COMPUTED: cost efficiency >= 3.0.", "P or F", "52"),
    ("Gate output", "r_pass", "COMPUTED: risk posture >= 3.0.", "P or F", "52"),
    ("Gate output", "vtcr_key", "COMPUTED: the four pass/fail flags concatenated. The skeleton of the rationale.", "str ^[PF]{4}$", "52, 27"),
    ("Gate output", "retain_or_invest_basis", "COMPUTED: on a retain or invest row, which dimension the invest is funding, or that no dimension fails at all. Blank on the other three terms. The audit trail for Bina's five-term vocabulary.", "str or blank", "52, 27"),
    ("Gate output", "disposition", "COMPUTED: what we recommend. Exactly five terms, emitted directly. retain = healthy, no spend; invest = fund a remediation or enhancement.", "enum(retain, invest, consolidate, replace, retire)", "4, 27, 52, 64"),
    ("Gate output", "priority", "COMPUTED: how urgent, held separately from the disposition.", "enum(Very High, High, Moderate, Low, Very Low)", "52"),

    # -- recommendation detail -----------------------------------------------------------
    ("Recommendation", "action", "The specific verb that will be done, distinct from the disposition. Set per row, validated against the enum.", "enum(none, monitor, remediate, upgrade, retrain, merge, absorb, decommission, re-platform)", "64"),
    ("Recommendation", "rationale", "COMPUTED: consultant-language reason citing the scores, the key and the evidence.", "str", "27"),
    ("Recommendation", "confidence", "COMPUTED from completeness and populated scores. Thin evidence is stated, not hidden.", "enum(high, medium, low)", "27, 28"),
    ("Recommendation", "redundancy_override_applied", "COMPUTED: cluster membership forced consolidate. Redundancy is an override, not a gate.", "bool", "25, 52"),
    ("Recommendation", "lifecycle_exclusion_applied", "COMPUTED: Birth/Growth guard is armed for this app, so retire and replace are barred.", "bool", "51"),
    ("Recommendation", "sourcing_exclusion_applied", "COMPUTED: sourcing type made a disposition or action illegal (SaaS cannot be re-platformed).", "bool", "51"),
    ("Recommendation", "retention_override_applied", "COMPUTED: a live data-retention obligation constrains a retire recommendation.", "bool", "53, 29"),
    ("Recommendation", "suppressed_recommendation", "COMPUTED: the disposition an exclusion suppressed, or the guard state. Never silent.", "str", "51"),
    ("Recommendation", "suppression_reason", "COMPUTED: why it was suppressed or constrained.", "str", "51, 29"),

    # -- clusters ------------------------------------------------------------------------
    ("Clusters", "overlap_cluster_id", "The overlap cluster this app belongs to, if any.", "str (CL-0n)", "25"),
    ("Clusters", "cluster_role", "Survivor (the app the capability lands on) or absorbed (the app folded into it).", "enum(survivor, absorbed) or blank", "25, 60"),
    ("Clusters", "consolidation_saving", "COMPUTED: net annual saving from absorbing THIS app. Blank on survivors, to avoid double counting.", "money USD annual", "25, 32"),

    # -- retention -----------------------------------------------------------------------
    ("Retention", "retention_obligation_flag", "Data must be kept after the app is switched off. Blank means we do not know yet.", "bool or blank", "53, 29"),
    ("Retention", "retention_expiry_date", "The date the data may actually be deleted. This gates retire, not the contract end date.", "date ISO-8601", "53, 29"),
    ("Retention", "residual_archival_cost", "Cost that survives retirement. Subtracted from the saving.", "money USD annual", "53, 32, 54"),
    ("Retention", "information_classification", "The most sensitive class of data the app holds. 'unknown' is itself a finding.", "enum(public, internal, confidential, PHI, PII, unknown)", "53, 24"),

    # -- replacement ---------------------------------------------------------------------
    ("Replacement", "replacement_app_id", "The named successor or survivor. Required for every consolidate and replace.", "str FK to app_id, or external label", "58"),
    ("Replacement", "replacement_ongoing_tco", "The successor's ongoing annual cost for carrying this workload. Netted off the saving.", "money USD annual", "58, 32, 54"),
    ("Replacement", "replacement_cost_already_in_baseline", "The successor is already a paid line in this portfolio, so nothing extra is netted off.", "bool", "58, 32"),

    # -- savings -------------------------------------------------------------------------
    ("Savings", "gross_saving_annual", "COMPUTED: the saving before anything is netted off.", "money USD annual", "32"),
    ("Savings", "amortised_one_time_migration_cost", "One-time migration or extraction cost, spread over five years.", "money USD annual", "32, 54"),
    ("Savings", "net_saving_annual", "COMPUTED: gross minus successor run cost, minus amortised one-time, minus residual archival.", "money USD annual", "32, 54"),
    ("Savings", "net_saving_five_year", "COMPUTED: net annual x 5, undiscounted. The 15% target is a run-rate figure, this is not.", "money USD", "54"),
    ("Savings", "realization_lag_months", "Months from decision to the first saved dollar.", "int months", "32"),
    ("Savings", "saving_type", "Recurring, one-time, or none.", "enum(recurring, one-time, none)", "32"),

    # -- effort and urgency --------------------------------------------------------------
    ("Effort & urgency", "urg_timeline_sensitivity", "How time-bound the decision is (renewal, EOS, notice window).", "enum(High, Medium, Low, None)", "56"),
    ("Effort & urgency", "urg_risk_pain_severity", "How much pain the current state is causing.", "enum(High, Medium, Low, None)", "56"),
    ("Effort & urgency", "urgency_score", "COMPUTED: mean of the two urgency inputs, 0-1. Feeds the priority rule on override rows.", "float 0-1", "56"),

    # -- provenance ----------------------------------------------------------------------
    ("Provenance", "data_source", "Which sources this row was assembled from. Absences are the shadow-IT signal.", "str, semicolon separated", "6, 8, 48, 61"),
    ("Provenance", "completeness_score", "COMPUTED: share of the fields the scoring model consumes that are actually populated.", "float 0-1", "12"),
    ("Provenance", "missing_fields", "COMPUTED: the model-consumed fields that are empty on this row.", "list[str], semicolon separated", "12, 14, 28"),
]

COLUMN_ORDER = [c[1] for c in COLUMNS]
COMPUTED_MARK = "COMPUTED"

# ---------------------------------------------------------------------------------------
# 3. THE 20 ROWS - explicit input data, one dict per application
# ---------------------------------------------------------------------------------------
# Only INPUTS live here. Anything the engine derives (dimension scores, pass flags, key,
# disposition, priority, rationale, TCO totals, utilisation, savings, completeness) is
# absent from these dicts on purpose and is computed in section 4.
#
# Keys beginning with "_" are working values used by the generator and the verification
# step; they are not columns in the deliverable.
#
# Legal entities, all invented:
#   LE-01 Lakeshore Health Partners      - parent / academic flagship
#   LE-02 Lakeshore Medical Group        - employed physician group (ambulatory)
#   LE-03 Riverbend Community Hospital   - community hospital acquired 2024

ROWS = [
    # ===== APP-001 ======================================================================
    dict(
        app_id="APP-001", app_name="Epic Hyperspace", vendor_name="Epic Systems Corporation",
        description="Core inpatient and ambulatory EHR: orders, clinical documentation, results and inpatient charge capture. Epic-embedded AI features exist but are inventoried as separate embedded_feature records, not on this row.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="hosted", sourcing_type="hybrid", lifecycle_stage="Mature",
        implementation_date="2019-03-02", version_installed="Feb 2026 IU",
        version_vendor_supported="Feb 2026 IU", vendor_eos_date="2029-02-28",
        technical_obsolescence_flag=False,
        business_owner="Dana Whitfield", technical_owner="Priya Raman", legal_entity="LE-01",
        business_unit="Clinical Enterprise", department="Clinical Informatics",
        cost_centre="CC-4410", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="clinical documentation",
        secondary_capabilities="patient access; revenue cycle; pharmacy", capability_tag_confidence=0.97,
        licences_purchased=18500, active_users=17240, last_signin_date="2026-08-13",
        process_centrality="High", owner_stated_strategic_importance=5.0,
        cost_licence_subscription=4150000, cost_upgrade_and_modules=620000,
        cost_maintenance_dev_labour=1850000, cost_infrastructure_peripherals=980000,
        cost_indirect_and_training=300000, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id="CTR-2019-0114", annual_contract_value=4150000, term_start="2023-01-01",
        term_end="2028-12-31", auto_renewal_flag=False, renewal_notice_days=180,
        licence_metric="enterprise", early_termination_penalty=0,
        has_downstream_dependents=True, dependency_count=42, integration_pattern="API",
        data_types_held="database; documents; images",
        ov_increase_value=5.0, ov_reach_consumers=5.0, ov_reduce_costs_efficiency=4.5,
        ov_patient_care_criticality=5.0, ov_governance_compliance=4.5,
        th_supportability=4.5, th_architecture_fit=4.0, th_operational_stability=4.0,
        th_vendor_viability=5.0, th_customization_debt=3.5,
        c_cost_per_active_user_vs_peers=3.5, c_unused_licence_waste=4.0,
        c_consumption_price_variance=4.0, c_absolute_cost_band=1.0,
        r_technical_risk=4.0, r_business_compliance_risk=4.5, r_clinical_safety_risk=4.0,
        r_end_user_perceived_quality=4.0,
        action="monitor", overlap_cluster_id="CL-03", cluster_role="survivor",
        retention_obligation_flag=True, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id=None, replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=0, realization_lag_months=None, saving_type="none",
        urg_timeline_sensitivity="Low", urg_risk_pain_severity="None",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="retain", _intended_priority="Very Low", _intended_key="PPPP",
        _gross_saving_basis="none",
        _evidence="Highest business value in the portfolio: enterprise breadth (17,240 of 18,500 seats active), patient-care criticality 5, process centrality High, current release on a supported stack. Cost per active user sits mid-band for its peer group despite the largest headline figure in the portfolio.",
    ),
    # ===== APP-002 ======================================================================
    dict(
        app_id="APP-002", app_name="Luma Health", vendor_name="Luma Health, Inc.",
        description="Patient engagement platform: self-scheduling, reminders and contact-centre deflection across the ambulatory network. Conversational scheduling AI is included in the subscription.",
        is_ai_tool=True, ai_delivery_form="embedded_feature",
        ai_capability_class="conversational scheduling / patient chat", ai_host_app_id="APP-002",
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Mature",
        implementation_date="2022-06-15", version_installed="2026.2 (SaaS)",
        version_vendor_supported="2026.2 (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner="Marcus Beauchamp", technical_owner="Ingrid Solheim", legal_entity="LE-02",
        business_unit="Patient Access & Contact Centre", department="Access Operations",
        cost_centre="CC-2130", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="patient access", secondary_capabilities="clinical communication",
        capability_tag_confidence=0.94,
        licences_purchased=900, active_users=742, last_signin_date="2026-08-13",
        process_centrality="High", owner_stated_strategic_importance=4.5,
        cost_licence_subscription=246000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=42000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=28000, consumption_based_cost=24000,
        one_time_implementation_cost=0,
        contract_id="CTR-2024-0231", annual_contract_value=246000, term_start="2024-07-01",
        term_end="2027-06-30", auto_renewal_flag=True, renewal_notice_days=60,
        licence_metric="per-user", early_termination_penalty=45000,
        has_downstream_dependents=True, dependency_count=9, integration_pattern="API",
        data_types_held="database; documents",
        ov_increase_value=4.5, ov_reach_consumers=4.0, ov_reduce_costs_efficiency=4.0,
        ov_patient_care_criticality=4.0, ov_governance_compliance=4.0,
        th_supportability=4.0, th_architecture_fit=4.0, th_operational_stability=4.0,
        th_vendor_viability=3.5, th_customization_debt=4.0,
        c_cost_per_active_user_vs_peers=3.5, c_unused_licence_waste=3.5,
        c_consumption_price_variance=3.5, c_absolute_cost_band=3.0,
        r_technical_risk=3.0, r_business_compliance_risk=1.5, r_clinical_safety_risk=3.5,
        r_end_user_perceived_quality=4.0,
        action="remediate", overlap_cluster_id="CL-05", cluster_role="survivor",
        retention_obligation_flag=True, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id=None, replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=0, realization_lag_months=None, saving_type="none",
        urg_timeline_sensitivity="Medium", urg_risk_pain_severity="High",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="invest", _intended_priority="High", _intended_key="PPPF",
        _gross_saving_basis="none",
        _evidence="Strong adoption and healthy technology, but the risk gate fails on posture, not on age: internet-facing PHI, SOC 2 lapsed pending re-audit, no data-residency attestation on file, and contractual lock-in rated High. The remediation funds the security work; the capability itself is not in question.",
    ),
    # ===== APP-003 ======================================================================
    dict(
        app_id="APP-003", app_name="MEDITECH Expanse",
        vendor_name="Medical Information Technology, Inc. (MEDITECH)",
        description="Full EHR running the acquired community hospital - the same capability Epic covers everywhere else in the group.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="hosted", sourcing_type="COTS", lifecycle_stage="Mature",
        implementation_date="2017-09-01", version_installed="Expanse 2025.1",
        version_vendor_supported="Expanse 2026.1", vendor_eos_date="2028-06-30",
        technical_obsolescence_flag=False,
        business_owner="Elena Vasquez-Ito", technical_owner="Tobias Nkemdirim", legal_entity="LE-03",
        business_unit="Community Hospital Division", department="Riverbend Clinical Systems",
        cost_centre="CC-7720", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="clinical documentation",
        secondary_capabilities="patient access; revenue cycle; laboratory", capability_tag_confidence=0.96,
        licences_purchased=2600, active_users=2285, last_signin_date="2026-08-13",
        process_centrality="High", owner_stated_strategic_importance=4.0,
        cost_licence_subscription=780000, cost_upgrade_and_modules=145000,
        cost_maintenance_dev_labour=395000, cost_infrastructure_peripherals=210000,
        cost_indirect_and_training=70000, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id="CTR-2021-0067", annual_contract_value=780000, term_start="2022-10-01",
        term_end="2027-09-30", auto_renewal_flag=False, renewal_notice_days=120,
        licence_metric="per-bed", early_termination_penalty=210000,
        has_downstream_dependents=True, dependency_count=18, integration_pattern="HL7",
        data_types_held="database; documents; images",
        ov_increase_value=4.5, ov_reach_consumers=4.0, ov_reduce_costs_efficiency=4.0,
        ov_patient_care_criticality=4.5, ov_governance_compliance=4.0,
        th_supportability=4.0, th_architecture_fit=3.5, th_operational_stability=4.0,
        th_vendor_viability=4.0, th_customization_debt=3.0,
        c_cost_per_active_user_vs_peers=3.5, c_unused_licence_waste=3.5,
        c_consumption_price_variance=3.5, c_absolute_cost_band=1.0,
        r_technical_risk=3.5, r_business_compliance_risk=3.5, r_clinical_safety_risk=3.5,
        r_end_user_perceived_quality=3.5,
        action="merge", overlap_cluster_id="CL-03", cluster_role="absorbed",
        retention_obligation_flag=True, retention_expiry_date="2037-12-31", residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id="APP-001", replacement_ongoing_tco=280000,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=220000, realization_lag_months=18, saving_type="recurring",
        urg_timeline_sensitivity="Medium", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="consolidate", _intended_priority="High", _intended_key="PPPP",
        _gross_saving_basis="run_rate",
        _evidence="On its own gates this is a healthy application - which is the point. It is a consolidation candidate only because it duplicates a capability already delivered in another legal entity (M&A pair MA-01). Survivor chosen on value, user count and contract runway, not on which entity acquired the other. Migration is real: HL7 interface rebuild and data migration rated High, so this lands in wave 3 and its one-time cost is amortised against the saving.",
    ),
    # ===== APP-004 ======================================================================
    dict(
        app_id="APP-004", app_name="Veradigm Sunrise Clinical Manager",
        vendor_name="Veradigm Inc. (formerly Allscripts)",
        description="Legacy inpatient EHR, read-only since the Epic cutover. Retained solely for historical record retrieval by Health Information Management.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="on-prem", sourcing_type="COTS", lifecycle_stage="End of Life",
        implementation_date="2008-11-01", version_installed="18.3",
        version_vendor_supported="21.2", vendor_eos_date="2024-12-31",
        technical_obsolescence_flag=True,
        business_owner="Saoirse Lindqvist", technical_owner="Desmond Achterberg", legal_entity="LE-01",
        business_unit="Health Information Management", department="Release of Information",
        cost_centre="CC-4185", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="clinical documentation", secondary_capabilities=None,
        capability_tag_confidence=0.92,
        licences_purchased=900, active_users=7, last_signin_date="2026-04-22",
        process_centrality="Low", owner_stated_strategic_importance=2.0,
        cost_licence_subscription=96000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=168000, cost_infrastructure_peripherals=132000,
        cost_indirect_and_training=24000, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id="CTR-2011-0009", annual_contract_value=96000, term_start="2024-01-01",
        term_end="2026-12-31", auto_renewal_flag=True, renewal_notice_days=90,
        licence_metric="enterprise", early_termination_penalty=0,
        has_downstream_dependents=False, dependency_count=2, integration_pattern="flat file",
        data_types_held="database; documents; images",
        ov_increase_value=1.5, ov_reach_consumers=1.0, ov_reduce_costs_efficiency=1.5,
        ov_patient_care_criticality=1.5, ov_governance_compliance=2.5,
        th_supportability=1.0, th_architecture_fit=1.5, th_operational_stability=2.5,
        th_vendor_viability=2.0, th_customization_debt=2.0,
        c_cost_per_active_user_vs_peers=1.0, c_unused_licence_waste=1.0,
        c_consumption_price_variance=3.0, c_absolute_cost_band=3.0,
        r_technical_risk=2.0, r_business_compliance_risk=2.5, r_clinical_safety_risk=3.0,
        r_end_user_perceived_quality=2.0,
        action="decommission", overlap_cluster_id=None, cluster_role=None,
        retention_obligation_flag=True, retention_expiry_date="2031-06-30",
        residual_archival_cost=48000, information_classification="PHI",
        replacement_app_id="ARCHIVE-COLD-01", replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=62000, realization_lag_months=12, saving_type="recurring",
        urg_timeline_sensitivity="High", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="retire", _intended_priority="High", _intended_key="FFFF",
        _gross_saving_basis="run_rate",
        _evidence="Seven active users against 900 licences, last sign-in months old, nothing downstream consumes it. Retire is correct, but it cannot be switched off yet: a 10-year retention obligation runs to 2031-06-30, release-of-information needs read-only access after retirement, and the application cannot purge its own data. So the saving is real but constrained - full run-rate minus a residual cold-storage archival cost that persists past shutdown, minus amortised extraction.",
    ),
    # ===== APP-005 ======================================================================
    dict(
        app_id="APP-005", app_name="Waystar", vendor_name="Waystar Holding Corp.",
        description="Claims clearinghouse, eligibility checking, denials management and patient payment estimation.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Mature",
        implementation_date="2021-02-01", version_installed="2026.3 (SaaS)",
        version_vendor_supported="2026.3 (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner="Rafael Okonkwo", technical_owner="Hazel Brightwater", legal_entity="LE-01",
        business_unit="Revenue Cycle", department="Patient Financial Services",
        cost_centre="CC-3050", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="revenue cycle", secondary_capabilities="finance",
        capability_tag_confidence=0.95,
        licences_purchased=420, active_users=386, last_signin_date="2026-08-13",
        process_centrality="High", owner_stated_strategic_importance=4.5,
        cost_licence_subscription=385000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=96000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=44000, consumption_based_cost=725000,
        one_time_implementation_cost=0,
        contract_id="CTR-2023-0188", annual_contract_value=385000, term_start="2023-04-01",
        term_end="2027-03-31", auto_renewal_flag=True, renewal_notice_days=90,
        licence_metric="consumption", early_termination_penalty=0,
        has_downstream_dependents=True, dependency_count=11, integration_pattern="API",
        data_types_held="database; documents",
        ov_increase_value=4.5, ov_reach_consumers=4.5, ov_reduce_costs_efficiency=4.5,
        ov_patient_care_criticality=4.0, ov_governance_compliance=4.5,
        th_supportability=4.0, th_architecture_fit=4.0, th_operational_stability=4.0,
        th_vendor_viability=4.5, th_customization_debt=4.0,
        c_cost_per_active_user_vs_peers=2.0, c_unused_licence_waste=3.0,
        c_consumption_price_variance=2.0, c_absolute_cost_band=1.0,
        r_technical_risk=4.0, r_business_compliance_risk=3.5, r_clinical_safety_risk=4.0,
        r_end_user_perceived_quality=4.0,
        action="remediate", overlap_cluster_id=None, cluster_role=None,
        retention_obligation_flag=True, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id=None, replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=0, realization_lag_months=None, saving_type="none",
        urg_timeline_sensitivity="Medium", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="invest", _intended_priority="Moderate", _intended_key="PPFP",
        _gross_saving_basis="none",
        _evidence="High value, healthy technology - but priced on consumption, and per-transaction charges are running well above the capability peer benchmark, so cost efficiency fails on both cost per active user and consumption variance. The remediation is a renegotiation at renewal, not a migration. Being SaaS, re-platforming is not a legal option for this row in any case.",
    ),
    # ===== APP-006 ======================================================================
    dict(
        app_id="APP-006", app_name="Solventum 360 Encompass",
        vendor_name="Solventum Corporation (health information systems, formerly part of 3M)",
        description="Computer-assisted coding and clinical documentation integrity: NLP over the chart to suggest codes and raise CDI queries.",
        is_ai_tool=True, ai_delivery_form="standalone",
        ai_capability_class="coding / CDI assistant", ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Mature",
        implementation_date="2020-05-04", version_installed="2024.4",
        version_vendor_supported="2026.2", vendor_eos_date="2027-12-31",
        technical_obsolescence_flag=True,
        business_owner="Yusuf Balogun", technical_owner="Marisol Cabrera", legal_entity="LE-01",
        business_unit="Revenue Cycle", department="HIM Coding",
        cost_centre="CC-3072", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="revenue cycle", secondary_capabilities="clinical documentation",
        capability_tag_confidence=0.93,
        licences_purchased=165, active_users=148, last_signin_date="2026-08-12",
        process_centrality="High", owner_stated_strategic_importance=4.0,
        cost_licence_subscription=690000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=148000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=62000, consumption_based_cost=250000,
        one_time_implementation_cost=0,
        contract_id="CTR-2022-0142", annual_contract_value=690000, term_start="2022-06-01",
        term_end="2027-05-31", auto_renewal_flag=False, renewal_notice_days=90,
        licence_metric="consumption", early_termination_penalty=0,
        has_downstream_dependents=True, dependency_count=6, integration_pattern="flat file",
        data_types_held="database; documents",
        ov_increase_value=4.5, ov_reach_consumers=4.0, ov_reduce_costs_efficiency=4.5,
        ov_patient_care_criticality=4.0, ov_governance_compliance=4.0,
        th_supportability=2.5, th_architecture_fit=2.0, th_operational_stability=3.0,
        th_vendor_viability=2.0, th_customization_debt=2.5,
        c_cost_per_active_user_vs_peers=2.0, c_unused_licence_waste=2.5,
        c_consumption_price_variance=3.0, c_absolute_cost_band=1.0,
        r_technical_risk=3.5, r_business_compliance_risk=3.5, r_clinical_safety_risk=3.5,
        r_end_user_perceived_quality=3.0,
        action="re-platform", overlap_cluster_id=None, cluster_role=None,
        retention_obligation_flag=True, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id="EXT-CAC-PLATFORM (outside the 20)", replacement_ongoing_tco=620000,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=130000, realization_lag_months=12, saving_type="recurring",
        urg_timeline_sensitivity="Medium", urg_risk_pain_severity="High",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="replace", _intended_priority="High", _intended_key="PFFP",
        _gross_saving_basis="run_rate",
        _evidence="Coding is high-value work, so value passes. Technical health fails on vendor roadmap uncertainty after the spin-off, a batch flat-file interface into the EHR where the architecture standard is API, and a version stalled two releases back. Cost fails on cost per coded chart against the peer band with flat coder productivity. The named successor sits outside these 20 and carries its own ongoing run cost, so the saving is the difference, not the incumbent's full bill.",
    ),
    # ===== APP-007 ======================================================================
    dict(
        app_id="APP-007", app_name="Microsoft Dragon Copilot",
        vendor_name="Microsoft Corporation (Nuance Communications)",
        description="Ambient AI clinical documentation: listens to the encounter and drafts the note into the EHR.",
        is_ai_tool=True, ai_delivery_form="standalone",
        ai_capability_class="ambient clinical documentation / scribe", ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Growth",
        implementation_date="2026-02-09", version_installed="2026.2 (SaaS)",
        version_vendor_supported="2026.2 (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner="Fenwick Odugbemi", technical_owner="Priya Raman", legal_entity="LE-01",
        business_unit="Clinical Enterprise", department="CMIO Office",
        cost_centre="CC-4432", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="clinical documentation", secondary_capabilities=None,
        capability_tag_confidence=0.96,
        licences_purchased=1400, active_users=905, last_signin_date="2026-08-13",
        process_centrality="High", owner_stated_strategic_importance=4.5,
        cost_licence_subscription=240000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=88000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=72000, consumption_based_cost=580000,
        one_time_implementation_cost=310000,
        contract_id="CTR-2025-0402", annual_contract_value=240000, term_start="2026-02-01",
        term_end="2029-01-31", auto_renewal_flag=False, renewal_notice_days=90,
        licence_metric="consumption", early_termination_penalty=0,
        has_downstream_dependents=False, dependency_count=4, integration_pattern="API",
        data_types_held="database; documents",
        ov_increase_value=4.5, ov_reach_consumers=4.0, ov_reduce_costs_efficiency=4.5,
        ov_patient_care_criticality=4.5, ov_governance_compliance=4.0,
        th_supportability=5.0, th_architecture_fit=4.5, th_operational_stability=4.0,
        th_vendor_viability=5.0, th_customization_debt=4.5,
        c_cost_per_active_user_vs_peers=2.5, c_unused_licence_waste=2.5,
        c_consumption_price_variance=1.5, c_absolute_cost_band=2.0,
        r_technical_risk=4.0, r_business_compliance_risk=3.5, r_clinical_safety_risk=3.5,
        r_end_user_perceived_quality=4.5,
        action="absorb", overlap_cluster_id="CL-04", cluster_role="survivor",
        retention_obligation_flag=True, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id=None, replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=0, realization_lag_months=None, saving_type="none",
        urg_timeline_sensitivity="Medium", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="invest", _intended_priority="Moderate", _intended_key="PPFP",
        _gross_saving_basis="none",
        _evidence="Clinician-attested value is high and rising and the technology is current, but per-encounter consumption spend is tracking well above the modelled plan and adoption is uneven by service line, so cost efficiency fails. Six months post go-live the app is in Growth, so the lifecycle guard bars retire and replace outright regardless of what the cost signal looks like - the guard against retiring a new platform because adoption has not ramped. It is also the enterprise-licensed survivor that gives the single-department pilot somewhere to go.",
    ),
    # ===== APP-008 ======================================================================
    dict(
        app_id="APP-008", app_name="Abridge", vendor_name="Abridge AI, Inc.",
        description="Ambient AI scribe piloted in one ambulatory service line - the same capability the enterprise ambient tool already delivers.",
        is_ai_tool=True, ai_delivery_form="standalone",
        ai_capability_class="ambient clinical documentation / scribe", ai_host_app_id=None,
        ai_already_entitled_elsewhere=True, ai_entitled_alternative_app_id="APP-007",
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Growth",
        implementation_date="2026-03-16", version_installed="2026.2 (SaaS)",
        version_vendor_supported="2026.2 (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner="Anneke Vandermolen", technical_owner=None, legal_entity="LE-02",
        business_unit="Ambulatory Orthopaedics", department="Orthopaedic Surgery",
        cost_centre="CC-5514", is_orphaned=False, governance_visibility="Unmanaged", is_shadow_it=False,
        primary_capability="clinical documentation", secondary_capabilities=None,
        capability_tag_confidence=0.95,
        licences_purchased=120, active_users=96, last_signin_date="2026-08-12",
        process_centrality="Medium", owner_stated_strategic_importance=4.0,
        cost_licence_subscription=186000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=22000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=16000, consumption_based_cost=16000,
        one_time_implementation_cost=40000,
        contract_id=None, annual_contract_value=None, term_start=None,
        term_end=None, auto_renewal_flag=None, renewal_notice_days=None,
        licence_metric="per-user", early_termination_penalty=None,
        has_downstream_dependents=False, dependency_count=2, integration_pattern="API",
        data_types_held="database; documents",
        ov_increase_value=4.0, ov_reach_consumers=3.5, ov_reduce_costs_efficiency=4.0,
        ov_patient_care_criticality=4.0, ov_governance_compliance=3.5,
        th_supportability=4.5, th_architecture_fit=4.0, th_operational_stability=4.0,
        th_vendor_viability=3.5, th_customization_debt=4.0,
        c_cost_per_active_user_vs_peers=2.0, c_unused_licence_waste=2.5,
        c_consumption_price_variance=2.5, c_absolute_cost_band=3.0,
        r_technical_risk=3.5, r_business_compliance_risk=3.0, r_clinical_safety_risk=3.5,
        r_end_user_perceived_quality=4.5,
        action="merge", overlap_cluster_id="CL-04", cluster_role="absorbed",
        retention_obligation_flag=True, retention_expiry_date="2033-03-31", residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id="APP-007", replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=True,
        amortised_one_time_migration_cost=6000, realization_lag_months=4, saving_type="recurring",
        urg_timeline_sensitivity="Medium", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger",
        _intended_disposition="consolidate", _intended_priority="Moderate", _intended_key="PPFP",
        _gross_saving_basis="run_rate",
        _evidence="A genuine duplicate AI capability inside the same capability tag, with a survivor that is already enterprise-licensed for every service line. Pilot clinicians hold both entitlements, so the user overlap is real rather than notional. Note the interaction the engine has to get right: Growth stage bars retire and replace, but it does not bar consolidate, so the lifecycle guard does not swallow the finding. Bought on a service-line decision with no IT contract record, hence Unmanaged.",
    ),
    # ===== APP-009 ======================================================================
    dict(
        app_id="APP-009", app_name="Aidoc", vendor_name="Aidoc Medical Ltd.",
        description="AI triage over CT studies, flagging suspected critical findings onto the radiology worklist.",
        is_ai_tool=True, ai_delivery_form="standalone",
        ai_capability_class="imaging triage / detection", ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="End of Life",
        implementation_date="2022-11-14", version_installed="2025.1 (SaaS)",
        version_vendor_supported="2026.1 (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner=None, technical_owner=None, legal_entity="LE-01",
        business_unit="Radiology & Imaging Services", department="Radiology Informatics",
        cost_centre="CC-4620", is_orphaned=True, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="imaging", secondary_capabilities=None, capability_tag_confidence=0.94,
        licences_purchased=26, active_users=3, last_signin_date="2026-06-30",
        process_centrality="Low", owner_stated_strategic_importance=2.0,
        cost_licence_subscription=210000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=34000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=12000, consumption_based_cost=54000,
        one_time_implementation_cost=0,
        contract_id="CTR-2023-0301", annual_contract_value=210000, term_start="2025-01-01",
        term_end="2026-12-15", auto_renewal_flag=False, renewal_notice_days=60,
        licence_metric="consumption", early_termination_penalty=0,
        has_downstream_dependents=False, dependency_count=2, integration_pattern="API",
        data_types_held="images",
        ov_increase_value=1.5, ov_reach_consumers=1.5, ov_reduce_costs_efficiency=2.0,
        ov_patient_care_criticality=2.0, ov_governance_compliance=2.5,
        th_supportability=4.0, th_architecture_fit=4.0, th_operational_stability=3.5,
        th_vendor_viability=3.0, th_customization_debt=4.0,
        c_cost_per_active_user_vs_peers=1.0, c_unused_licence_waste=1.0,
        c_consumption_price_variance=2.0, c_absolute_cost_band=3.0,
        r_technical_risk=3.5, r_business_compliance_risk=3.5, r_clinical_safety_risk=3.5,
        r_end_user_perceived_quality=3.0,
        action="decommission", overlap_cluster_id=None, cluster_role=None,
        retention_obligation_flag=False, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id=None, replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=0, realization_lag_months=4, saving_type="recurring",
        urg_timeline_sensitivity="High", urg_risk_pain_severity="Low",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="retire", _intended_priority="High", _intended_key="FPFP",
        _gross_saving_basis="run_rate",
        _evidence="The clean, genuinely retirable case: 3 active users of 26 licences, near-zero sign-in events, use confined to individuals, the informatics owner left and the assignment group has been disbanded, and nothing downstream consumes its output because findings are pushed to a worklist. The contract expires within four months and does not auto-renew, so notice is straightforward. The imaging record of truth stays in the PACS, so there is no retention obligation and no residual archival cost. Clinical risk is scored honestly: a clinical tool nobody uses carries low patient-safety consequence, which is precisely why it is safe to retire.",
    ),
    # ===== APP-010 ======================================================================
    dict(
        app_id="APP-010", app_name="Otter.ai", vendor_name="Otter.ai, Inc.",
        description="AI meeting transcription and summarisation, bought on a departmental card for committee minutes. Recordings are believed to contain PHI.",
        is_ai_tool=True, ai_delivery_form="standalone",
        ai_capability_class="meeting transcription / summarisation", ai_host_app_id=None,
        ai_already_entitled_elsewhere=True, ai_entitled_alternative_app_id="APP-011",
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Mature",
        implementation_date=None, version_installed="business tier (SaaS)",
        version_vendor_supported="business tier (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner=None, technical_owner=None, legal_entity="LE-01",
        business_unit="Quality & Patient Safety", department="Quality Committees",
        cost_centre="CC-4901", is_orphaned=True, governance_visibility="Unsanctioned", is_shadow_it=True,
        primary_capability="clinical documentation", secondary_capabilities="analytics",
        capability_tag_confidence=0.41,
        licences_purchased=None, active_users=None, last_signin_date="2026-08-11",
        process_centrality="Low", owner_stated_strategic_importance=None,
        cost_licence_subscription=14400, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=0, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=2600, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id=None, annual_contract_value=None, term_start=None,
        term_end=None, auto_renewal_flag=None, renewal_notice_days=None,
        licence_metric=None, early_termination_penalty=None,
        has_downstream_dependents=False, dependency_count=0, integration_pattern="none",
        data_types_held="documents",
        ov_increase_value=1.5, ov_reach_consumers=None, ov_reduce_costs_efficiency=2.0,
        ov_patient_care_criticality=2.0, ov_governance_compliance=1.5,
        th_supportability=3.0, th_architecture_fit=2.5, th_operational_stability=None,
        th_vendor_viability=2.5, th_customization_debt=3.0,
        c_cost_per_active_user_vs_peers=None, c_unused_licence_waste=None,
        c_consumption_price_variance=2.5, c_absolute_cost_band=5.0,
        r_technical_risk=2.5, r_business_compliance_risk=1.5, r_clinical_safety_risk=3.0,
        r_end_user_perceived_quality=3.5,
        action="decommission", overlap_cluster_id=None, cluster_role=None,
        retention_obligation_flag=None, retention_expiry_date=None, residual_archival_cost=2500,
        information_classification="unknown",
        replacement_app_id="APP-011", replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=True,
        amortised_one_time_migration_cost=0, realization_lag_months=1, saving_type="recurring",
        urg_timeline_sensitivity="High", urg_risk_pain_severity="High",
        data_source="AP ledger; SSO sign-in log",
        _intended_disposition="retire", _intended_priority="Very High", _intended_key="FFFF",
        _gross_saving_basis="run_rate",
        _evidence="Three findings on one row. Shadow IT: no CMDB record, no contract record, but an AP line and live SSO activity, and committee recordings that should be classified PHI are recorded as unknown. Thin evidence: licence count, contract, owner and go-live date are all absent, so several score inputs carry no value and are renormalised out rather than silently scored as zero toward retire - the recommendation stands but at reduced confidence, and the one field that would settle it is active_users (whether eight people rely on this or eighty). Already entitled elsewhere: the enterprise AI assistant the organisation already pays for covers meeting transcription and recap, so this is a saving with no functional loss.",
    ),
    # ===== APP-011 ======================================================================
    dict(
        app_id="APP-011", app_name="Microsoft 365 Copilot", vendor_name="Microsoft Corporation",
        description="Enterprise AI assistant across mail, documents, meetings and chat. Add-on to an existing enterprise agreement.",
        is_ai_tool=True, ai_delivery_form="standalone",
        ai_capability_class="general assistant / chat + meeting summarisation", ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Mature",
        implementation_date="2025-01-20", version_installed="current (SaaS)",
        version_vendor_supported="current (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner="Callum Thistlewood", technical_owner="Nadia Petrossian", legal_entity="LE-01",
        business_unit="IT Shared Services", department="Digital Workplace",
        cost_centre="CC-1020", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="analytics", secondary_capabilities="HR; finance",
        capability_tag_confidence=0.62,
        licences_purchased=3000, active_users=1215, last_signin_date="2026-08-13",
        process_centrality="Medium", owner_stated_strategic_importance=4.5,
        cost_licence_subscription=1188000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=66000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=96000, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id="CTR-2024-0455", annual_contract_value=1188000, term_start="2025-01-01",
        term_end="2027-12-31", auto_renewal_flag=True, renewal_notice_days=60,
        licence_metric="per-user", early_termination_penalty=0,
        has_downstream_dependents=False, dependency_count=5, integration_pattern="API",
        data_types_held="documents",
        ov_increase_value=4.0, ov_reach_consumers=3.5, ov_reduce_costs_efficiency=4.0,
        ov_patient_care_criticality=4.5, ov_governance_compliance=4.0,
        th_supportability=5.0, th_architecture_fit=4.5, th_operational_stability=4.5,
        th_vendor_viability=5.0, th_customization_debt=4.5,
        c_cost_per_active_user_vs_peers=2.0, c_unused_licence_waste=1.5,
        c_consumption_price_variance=3.0, c_absolute_cost_band=1.0,
        r_technical_risk=4.0, r_business_compliance_risk=4.0, r_clinical_safety_risk=4.0,
        r_end_user_perceived_quality=4.0,
        action="retrain", overlap_cluster_id=None, cluster_role=None,
        retention_obligation_flag=False, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="confidential",
        replacement_app_id=None, replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=0, realization_lag_months=6, saving_type="recurring",
        urg_timeline_sensitivity="Medium", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="invest", _intended_priority="Moderate", _intended_key="PPFP",
        _gross_saving_basis="licence_reharvest", _reharvest_share=0.60, _reharvest_reserve=9900,
        _evidence="The counterpart to the shadow transcription tool: this is the entitled alternative, so it has to survive. But it fails the cost gate on its own terms - utilisation around 0.41 and a large unused-licence count, which is the fastest credible saving in the portfolio because it needs no migration project. The right action is a seat reharvest and a retraining push, not a retirement. Note the deliberate tension for a reviewer: this row supplies a saving AND justifies retiring the shadow tool, so the reharvest figure explicitly reserves seats for the absorbed users instead of counting the same dollars twice.",
    ),
    # ===== APP-012 ======================================================================
    dict(
        app_id="APP-012", app_name="TigerConnect", vendor_name="TigerConnect, Inc.",
        description="Secure clinical messaging, on-call directory and care-team collaboration on mobile.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Mature",
        implementation_date="2021-08-02", version_installed="current (SaaS)",
        version_vendor_supported="current (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner="Oleander Marsh", technical_owner="Ingrid Solheim", legal_entity="LE-01",
        business_unit="Clinical Enterprise", department="Clinical Communications",
        cost_centre="CC-4415", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="clinical communication", secondary_capabilities=None,
        capability_tag_confidence=0.96,
        licences_purchased=6400, active_users=5910, last_signin_date="2026-08-13",
        process_centrality="High", owner_stated_strategic_importance=4.0,
        cost_licence_subscription=396000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=62000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=38000, consumption_based_cost=24000,
        one_time_implementation_cost=0,
        contract_id="CTR-2022-0119", annual_contract_value=396000, term_start="2024-01-01",
        term_end="2027-12-31", auto_renewal_flag=True, renewal_notice_days=90,
        licence_metric="per-user", early_termination_penalty=0,
        has_downstream_dependents=True, dependency_count=7, integration_pattern="API",
        data_types_held="database; documents",
        ov_increase_value=4.5, ov_reach_consumers=4.0, ov_reduce_costs_efficiency=4.5,
        ov_patient_care_criticality=4.5, ov_governance_compliance=4.0,
        th_supportability=4.5, th_architecture_fit=4.5, th_operational_stability=4.0,
        th_vendor_viability=4.0, th_customization_debt=4.5,
        c_cost_per_active_user_vs_peers=4.0, c_unused_licence_waste=3.5,
        c_consumption_price_variance=4.0, c_absolute_cost_band=2.0,
        r_technical_risk=4.0, r_business_compliance_risk=3.5, r_clinical_safety_risk=3.5,
        r_end_user_perceived_quality=4.0,
        action="absorb", overlap_cluster_id="CL-01", cluster_role="survivor",
        retention_obligation_flag=True, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id=None, replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=0, realization_lag_months=None, saving_type="none",
        urg_timeline_sensitivity="Low", urg_risk_pain_severity="Low",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="retain", _intended_priority="Low", _intended_key="PPPP",
        _gross_saving_basis="none",
        _evidence="Survivor of the messaging cluster: the highest active-user count and the broadest capability coverage of the group, the lowest cost per active user, and the on-call schedule integration already live. Its feature coverage matches the absorbed product everywhere except overhead paging, which the existing overhead system already covers - so the survivor choice is evidenced rather than asserted. It keeps its own gate disposition rather than being stamped consolidate.",
    ),
    # ===== APP-013 ======================================================================
    dict(
        app_id="APP-013", app_name="Spok Mobile", vendor_name="Spok Holdings, Inc.",
        description="Clinical secure messaging and paging, in place at the community hospital and in two flagship departments, with an on-prem paging gateway.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="SaaS", sourcing_type="COTS", lifecycle_stage="Mature",
        implementation_date="2016-04-11", version_installed="7.2",
        version_vendor_supported="8.1", vendor_eos_date="2027-06-30",
        technical_obsolescence_flag=True,
        business_owner="Bertram Quayle", technical_owner="Tobias Nkemdirim", legal_entity="LE-03",
        business_unit="Community Hospital Division", department="Perioperative Services",
        cost_centre="CC-7735", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="clinical communication", secondary_capabilities=None,
        capability_tag_confidence=0.95,
        licences_purchased=2100, active_users=1340, last_signin_date="2026-08-13",
        process_centrality="Medium", owner_stated_strategic_importance=3.0,
        cost_licence_subscription=198000, cost_upgrade_and_modules=14000,
        cost_maintenance_dev_labour=41000, cost_infrastructure_peripherals=28000,
        cost_indirect_and_training=14000, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id="CTR-2023-0155", annual_contract_value=198000, term_start="2023-10-01",
        term_end="2026-09-30", auto_renewal_flag=True, renewal_notice_days=90,
        licence_metric="per-user", early_termination_penalty=88000,
        has_downstream_dependents=True, dependency_count=5, integration_pattern="middleware",
        data_types_held="database",
        ov_increase_value=3.5, ov_reach_consumers=3.0, ov_reduce_costs_efficiency=3.5,
        ov_patient_care_criticality=3.5, ov_governance_compliance=3.0,
        th_supportability=3.0, th_architecture_fit=2.0, th_operational_stability=3.0,
        th_vendor_viability=2.5, th_customization_debt=3.0,
        c_cost_per_active_user_vs_peers=2.0, c_unused_licence_waste=2.0,
        c_consumption_price_variance=3.0, c_absolute_cost_band=3.0,
        r_technical_risk=3.5, r_business_compliance_risk=3.5, r_clinical_safety_risk=3.5,
        r_end_user_perceived_quality=3.0,
        action="merge", overlap_cluster_id="CL-01", cluster_role="absorbed",
        retention_obligation_flag=True, retention_expiry_date="2032-09-30", residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id="APP-012", replacement_ongoing_tco=46000,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=12000, realization_lag_months=13, saving_type="recurring",
        urg_timeline_sensitivity="High", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="consolidate", _intended_priority="High", _intended_key="PFFP",
        _gross_saving_basis="run_rate",
        _evidence="Real functional overlap with the survivor and more than 60% of its users already hold a survivor licence, so the consolidation saving is defensible. The row's job is the renewal calendar: the contract auto-renews, notice is 90 days, the term ends 2026-09-30 and the notice deadline of 2026-07-02 has ALREADY PASSED as at the analysis date. So the contract renews to 2027-09-30, the next servable deadline is 2027-07-02, and the saving cannot be booked in the current year however good the case looks. This is the auto-renewed-through-the-notice-window failure, pre-empted.",
    ),
    # ===== APP-014 ======================================================================
    dict(
        app_id="APP-014", app_name="Microsoft Power BI", vendor_name="Microsoft Corporation",
        description="Self-service BI and dashboarding for finance, operations and quality. Entitlement is bundled inside the existing enterprise agreement, so only marginal licence cost is carried.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Mature",
        implementation_date="2021-03-08", version_installed="current (SaaS)",
        version_vendor_supported="current (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner="Simone Adeyemi", technical_owner="Nadia Petrossian", legal_entity="LE-01",
        business_unit="Enterprise Data & Analytics", department="Business Intelligence",
        cost_centre="CC-1055", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="analytics", secondary_capabilities="finance", capability_tag_confidence=0.97,
        licences_purchased=2800, active_users=2455, last_signin_date="2026-08-13",
        process_centrality="Medium", owner_stated_strategic_importance=4.0,
        cost_licence_subscription=84000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=31000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=25000, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id="CTR-2024-0455-A", annual_contract_value=84000, term_start="2025-01-01",
        term_end="2027-12-31", auto_renewal_flag=True, renewal_notice_days=60,
        licence_metric="per-user", early_termination_penalty=0,
        has_downstream_dependents=True, dependency_count=14, integration_pattern="API",
        data_types_held="database",
        ov_increase_value=4.0, ov_reach_consumers=4.0, ov_reduce_costs_efficiency=4.5,
        ov_patient_care_criticality=4.5, ov_governance_compliance=4.0,
        th_supportability=5.0, th_architecture_fit=4.5, th_operational_stability=4.5,
        th_vendor_viability=5.0, th_customization_debt=4.5,
        c_cost_per_active_user_vs_peers=4.5, c_unused_licence_waste=4.0,
        c_consumption_price_variance=4.5, c_absolute_cost_band=4.0,
        r_technical_risk=4.0, r_business_compliance_risk=4.0, r_clinical_safety_risk=4.0,
        r_end_user_perceived_quality=4.0,
        action="absorb", overlap_cluster_id="CL-02", cluster_role="survivor",
        retention_obligation_flag=False, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="confidential",
        replacement_app_id=None, replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=0, realization_lag_months=None, saving_type="none",
        urg_timeline_sensitivity="Low", urg_risk_pain_severity="None",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="retain", _intended_priority="Very Low", _intended_key="PPPP",
        _gross_saving_basis="none",
        _evidence="The cheapest way to keep the analytics capability: marginal licence cost because the entitlement is already held inside the enterprise agreement, the largest active-user base of the three BI platforms, and no unique feature gap for the reports actually in use. Its low cost is what makes the analytics cluster saving large - and the absorbed tools' small but non-zero landing cost on this platform is still netted off.",
    ),
    # ===== APP-015 ======================================================================
    dict(
        app_id="APP-015", app_name="Tableau", vendor_name="Salesforce, Inc. (Tableau Software)",
        description="Visual analytics platform used by the service-line analytics team, running on Tableau Server in a private cloud.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="hosted", sourcing_type="COTS", lifecycle_stage="Mature",
        implementation_date="2019-07-22", version_installed="2024.2",
        version_vendor_supported="2026.1", vendor_eos_date="2027-09-30",
        technical_obsolescence_flag=False,
        business_owner="Kwame Lindgren", technical_owner="Desmond Achterberg", legal_entity="LE-01",
        business_unit="Service Line Analytics", department="Analytics Delivery",
        cost_centre="CC-1062", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="analytics", secondary_capabilities=None, capability_tag_confidence=0.96,
        licences_purchased=640, active_users=385, last_signin_date="2026-08-13",
        process_centrality="Medium", owner_stated_strategic_importance=4.0,
        cost_licence_subscription=452000, cost_upgrade_and_modules=36000,
        cost_maintenance_dev_labour=96000, cost_infrastructure_peripherals=62000,
        cost_indirect_and_training=34000, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id="CTR-2022-0207", annual_contract_value=452000, term_start="2024-04-01",
        term_end="2027-03-31", auto_renewal_flag=True, renewal_notice_days=90,
        licence_metric="per-user", early_termination_penalty=120000,
        has_downstream_dependents=False, dependency_count=8, integration_pattern="API",
        data_types_held="database",
        ov_increase_value=4.0, ov_reach_consumers=3.5, ov_reduce_costs_efficiency=4.0,
        ov_patient_care_criticality=4.0, ov_governance_compliance=3.5,
        th_supportability=4.0, th_architecture_fit=4.0, th_operational_stability=4.0,
        th_vendor_viability=4.5, th_customization_debt=3.5,
        c_cost_per_active_user_vs_peers=2.0, c_unused_licence_waste=2.0,
        c_consumption_price_variance=3.0, c_absolute_cost_band=2.0,
        r_technical_risk=3.5, r_business_compliance_risk=3.5, r_clinical_safety_risk=4.0,
        r_end_user_perceived_quality=4.5,
        action="merge", overlap_cluster_id="CL-02", cluster_role="absorbed",
        retention_obligation_flag=False, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="confidential",
        replacement_app_id="APP-014", replacement_ongoing_tco=96000,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=84000, realization_lag_months=9, saving_type="recurring",
        urg_timeline_sensitivity="Low", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="consolidate", _intended_priority="Moderate", _intended_key="PPFP",
        _gross_saving_basis="run_rate",
        _evidence="Genuine overlap with the survivor and a committed analyst community, so value passes - this is the row whose owner will fight for it and which needs a business-case one-pager. Cost efficiency fails on poor utilisation (385 of 640 seats) and material licence waste. Migration effort is real: a workbook inventory to rebuild and medium data-migration complexity, so it sits in wave 2 and its amortised migration cost comes off the saving.",
    ),
    # ===== APP-016 ======================================================================
    dict(
        app_id="APP-016", app_name="Qlik Sense Enterprise", vendor_name="QlikTech International AB",
        description="Legacy on-prem BI platform carrying a shrinking set of finance and supply-chain dashboards.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="on-prem", sourcing_type="COTS", lifecycle_stage="End of Life",
        implementation_date="2016-01-18", version_installed="2022.11",
        version_vendor_supported="2026.5", vendor_eos_date="2026-11-30",
        technical_obsolescence_flag=True,
        business_owner="Vera Holmquist", technical_owner=None, legal_entity="LE-01",
        business_unit="Finance", department="Financial Planning & Analysis",
        cost_centre="CC-1180", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="analytics", secondary_capabilities="supply chain; finance",
        capability_tag_confidence=0.95,
        licences_purchased=520, active_users=198, last_signin_date="2026-08-10",
        process_centrality="Low", owner_stated_strategic_importance=2.5,
        cost_licence_subscription=178000, cost_upgrade_and_modules=22000,
        cost_maintenance_dev_labour=92000, cost_infrastructure_peripherals=105000,
        cost_indirect_and_training=18000, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id="CTR-2019-0093", annual_contract_value=178000, term_start="2025-12-01",
        term_end="2026-11-30", auto_renewal_flag=False, renewal_notice_days=60,
        licence_metric="per-user", early_termination_penalty=0,
        has_downstream_dependents=False, dependency_count=4, integration_pattern="point-to-point",
        data_types_held="database",
        ov_increase_value=2.5, ov_reach_consumers=2.0, ov_reduce_costs_efficiency=2.5,
        ov_patient_care_criticality=2.5, ov_governance_compliance=3.0,
        th_supportability=3.0, th_architecture_fit=3.0, th_operational_stability=3.5,
        th_vendor_viability=3.5, th_customization_debt=3.0,
        c_cost_per_active_user_vs_peers=1.5, c_unused_licence_waste=1.5,
        c_consumption_price_variance=3.0, c_absolute_cost_band=3.0,
        r_technical_risk=3.0, r_business_compliance_risk=3.5, r_clinical_safety_risk=3.5,
        r_end_user_perceived_quality=3.0,
        action="absorb", overlap_cluster_id="CL-02", cluster_role="absorbed",
        retention_obligation_flag=False, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="internal",
        replacement_app_id="APP-014", replacement_ongoing_tco=38000,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=34000, realization_lag_months=3, saving_type="recurring",
        urg_timeline_sensitivity="High", urg_risk_pain_severity="High",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="consolidate", _intended_priority="High", _intended_key="FPFP",
        _gross_saving_basis="run_rate",
        _evidence="The wave-1 quick win: utilisation 0.38, a handful of dashboards still live and all of them reproducible on the survivor, nothing downstream, no auto-renewal, and a term ending inside the notice window so notice can actually be served. It is also the row that shows the override DIRECTION: on the gates alone this app scores as a retire, but a survivor exists and the capability must persist, so the correct output is consolidate. Retire would silently delete a live capability. On-prem infrastructure is non-trivial, so server reclamation is a real line in the runbook.",
    ),
    # ===== APP-017 ======================================================================
    dict(
        app_id="APP-017", app_name="Sunquest CoPath Plus",
        vendor_name="Clinisys Group Ltd. (Sunquest Information Systems)",
        description="Anatomic pathology LIS: specimen accessioning, synoptic reporting and case sign-out, with long-standing in-house report extensions.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="on-prem", sourcing_type="hybrid", lifecycle_stage="Mature",
        implementation_date="2009-10-05", version_installed="2020.1",
        version_vendor_supported="2025.2", vendor_eos_date="2027-03-31",
        technical_obsolescence_flag=True,
        business_owner="Padma Sundaresan", technical_owner="Gideon Frostwick", legal_entity="LE-01",
        business_unit="Laboratory Services", department="Anatomic Pathology",
        cost_centre="CC-4705", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="laboratory", secondary_capabilities="clinical documentation",
        capability_tag_confidence=0.94,
        licences_purchased=95, active_users=82, last_signin_date="2026-08-13",
        process_centrality="High", owner_stated_strategic_importance=4.5,
        cost_licence_subscription=62000, cost_upgrade_and_modules=9000,
        cost_maintenance_dev_labour=68000, cost_infrastructure_peripherals=38000,
        cost_indirect_and_training=8000, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id=None, annual_contract_value=None, term_start=None,
        term_end=None, auto_renewal_flag=None, renewal_notice_days=None,
        licence_metric="enterprise", early_termination_penalty=None,
        has_downstream_dependents=True, dependency_count=6, integration_pattern="HL7",
        data_types_held="database; documents; images",
        ov_increase_value=4.0, ov_reach_consumers=3.5, ov_reduce_costs_efficiency=4.0,
        ov_patient_care_criticality=4.5, ov_governance_compliance=4.0,
        th_supportability=1.5, th_architecture_fit=2.5, th_operational_stability=4.0,
        th_vendor_viability=3.0, th_customization_debt=2.0,
        c_cost_per_active_user_vs_peers=4.5, c_unused_licence_waste=4.0,
        c_consumption_price_variance=4.0, c_absolute_cost_band=4.0,
        r_technical_risk=3.0, r_business_compliance_risk=3.5, r_clinical_safety_risk=3.0,
        r_end_user_perceived_quality=2.5,
        action="upgrade", overlap_cluster_id=None, cluster_role=None,
        retention_obligation_flag=True, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id=None, replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=0, realization_lag_months=None, saving_type="none",
        urg_timeline_sensitivity="Medium", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger (allocation only)",
        _intended_disposition="invest", _intended_priority="Moderate", _intended_key="PFPP",
        _gross_saving_basis="none",
        _evidence="The deliberate trap. Every surface signal says retire: several releases behind, behind vendor support, a dated client UI, a small user population, a low absolute cost, and no contract record at all because a 2009 paper agreement was never loaded into the register. The correct answer is invest, and the data makes it inevitable: cancer diagnosis sign-out gives patient-care criticality 5 and process centrality High, an Epic Beaker interface plus the tumour registry and biobank feed off it, incident volume is low because it is stable, and its cost per active user is in the cheapest decile of the LIS peer band, so cost passes strongly. Only technical health fails, and its two detractors - version currency and platform supportability - ARE the remediation plan. Those same obsolescence facts must not be re-consumed as risk; if they are, this row flips to retire and the engine has exactly the defect the notes warn about.",
    ),
    # ===== APP-018 ======================================================================
    dict(
        app_id="APP-018", app_name="Infor Lawson", vendor_name="Infor, Inc. (Infor S3 / Lawson)",
        description="Legacy on-prem ERP covering HR, payroll and supply-chain procurement. Still runs payroll for two legal entities.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="on-prem", sourcing_type="COTS", lifecycle_stage="End of Life",
        implementation_date="2010-06-01", version_installed="S3 9.0.1",
        version_vendor_supported="CloudSuite 2026.1", vendor_eos_date="2025-09-30",
        technical_obsolescence_flag=True,
        business_owner="Marisol Cabrera", technical_owner="Desmond Achterberg", legal_entity="LE-01",
        business_unit="People & Culture", department="Payroll & HRIS",
        cost_centre="CC-1310", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="HR", secondary_capabilities="supply chain; finance",
        capability_tag_confidence=0.95,
        licences_purchased=9800, active_users=8420, last_signin_date="2026-08-13",
        process_centrality="High", owner_stated_strategic_importance=3.5,
        cost_licence_subscription=385000, cost_upgrade_and_modules=165000,
        cost_maintenance_dev_labour=810000, cost_infrastructure_peripherals=445000,
        cost_indirect_and_training=95000, consumption_based_cost=0,
        one_time_implementation_cost=0,
        contract_id="CTR-2014-0021", annual_contract_value=385000, term_start="2025-07-01",
        term_end="2027-06-30", auto_renewal_flag=True, renewal_notice_days=120,
        licence_metric="enterprise", early_termination_penalty=0,
        has_downstream_dependents=True, dependency_count=21, integration_pattern="point-to-point",
        data_types_held="database; documents",
        ov_increase_value=4.0, ov_reach_consumers=3.0, ov_reduce_costs_efficiency=4.0,
        ov_patient_care_criticality=3.5, ov_governance_compliance=4.0,
        th_supportability=1.0, th_architecture_fit=1.5, th_operational_stability=2.5,
        th_vendor_viability=3.0, th_customization_debt=1.5,
        c_cost_per_active_user_vs_peers=1.5, c_unused_licence_waste=2.0,
        c_consumption_price_variance=2.5, c_absolute_cost_band=1.0,
        r_technical_risk=1.5, r_business_compliance_risk=2.0, r_clinical_safety_risk=3.0,
        r_end_user_perceived_quality=2.0,
        action="re-platform", overlap_cluster_id=None, cluster_role=None,
        retention_obligation_flag=True, retention_expiry_date="2033-12-31",
        residual_archival_cost=55000, information_classification="PII",
        replacement_app_id="APP-019", replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=True,
        amortised_one_time_migration_cost=145000, realization_lag_months=9, saving_type="recurring",
        urg_timeline_sensitivity="High", urg_risk_pain_severity="High",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="replace", _intended_priority="Very High", _intended_key="PFFF",
        _gross_saving_basis="run_rate",
        _evidence="Payroll must run, so value passes. Everything else fails. Technical health: behind vendor support, end-of-support date already in the past, unsupported OS, database on extended support, heavy customization debt, point-to-point integrations. Cost: a large internal FTE load, an on-prem estate to keep alive, and upgrade fees on an unsupported release. Risk: single point of failure, no tested DR, unhardened configuration, and payroll PII. This is the strongest replace signal in the table. The netting is the real point: the successor is already live and dual-running as its own portfolio line, so its cost is ALREADY in the baseline and must not be subtracted again; what does come off is amortised migration and the residual payroll-retention archive.",
    ),
    # ===== APP-019 ======================================================================
    dict(
        app_id="APP-019", app_name="Workday HCM", vendor_name="Workday, Inc.",
        description="Core HR, payroll and talent. The named successor platform: live for one legal entity and mid-rollout for the others.",
        is_ai_tool=False, ai_delivery_form=None, ai_capability_class=None, ai_host_app_id=None,
        ai_already_entitled_elsewhere=False, ai_entitled_alternative_app_id=None,
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Growth",
        implementation_date="2026-04-06", version_installed="2026R1 (SaaS)",
        version_vendor_supported="2026R1 (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner="Marisol Cabrera", technical_owner="Nadia Petrossian", legal_entity="LE-01",
        business_unit="People & Culture", department="HR Transformation",
        cost_centre="CC-1315", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="HR", secondary_capabilities="finance", capability_tag_confidence=0.96,
        licences_purchased=12500, active_users=3180, last_signin_date="2026-08-13",
        process_centrality="High", owner_stated_strategic_importance=5.0,
        cost_licence_subscription=1560000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=320000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=220000, consumption_based_cost=0,
        one_time_implementation_cost=2900000,
        contract_id="CTR-2025-0388", annual_contract_value=1560000, term_start="2026-04-01",
        term_end="2031-03-31", auto_renewal_flag=False, renewal_notice_days=180,
        licence_metric="per-user", early_termination_penalty=0,
        has_downstream_dependents=True, dependency_count=13, integration_pattern="API",
        data_types_held="database; documents",
        ov_increase_value=4.0, ov_reach_consumers=3.5, ov_reduce_costs_efficiency=4.0,
        ov_patient_care_criticality=4.0, ov_governance_compliance=4.0,
        th_supportability=5.0, th_architecture_fit=4.5, th_operational_stability=4.0,
        th_vendor_viability=5.0, th_customization_debt=4.0,
        c_cost_per_active_user_vs_peers=1.5, c_unused_licence_waste=1.5,
        c_consumption_price_variance=3.0, c_absolute_cost_band=1.0,
        r_technical_risk=4.0, r_business_compliance_risk=4.0, r_clinical_safety_risk=4.0,
        r_end_user_perceived_quality=4.0,
        action="upgrade", overlap_cluster_id=None, cluster_role=None,
        retention_obligation_flag=True, retention_expiry_date=None, residual_archival_cost=0,
        information_classification="PII",
        replacement_app_id=None, replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=False,
        amortised_one_time_migration_cost=0, realization_lag_months=None, saving_type="none",
        urg_timeline_sensitivity="Medium", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="invest", _intended_priority="High", _intended_key="PPFP",
        _gross_saving_basis="none",
        _evidence="Cost efficiency fails on cost per active user BECAUSE the rollout is incomplete - one entity's population is live while the full subscription is already paid. A naive engine sees an expensive, under-used, brand-new platform and recommends retire or replace; the Growth lifecycle guard bars both outright with the suppression recorded. It is also the successor named by the legacy ERP row, so its go-live date and ongoing cost are what let the 'retirement scheduled before its replacement goes live' check run against real data.",
    ),
    # ===== APP-020 ======================================================================
    dict(
        app_id="APP-020", app_name="Hyro", vendor_name="Hyro, Inc.",
        description="Standalone conversational AI assistant for the patient contact centre: call and web deflection, scheduling intents and FAQ.",
        is_ai_tool=True, ai_delivery_form="standalone",
        ai_capability_class="conversational patient assistant / chatbot", ai_host_app_id=None,
        ai_already_entitled_elsewhere=True, ai_entitled_alternative_app_id="APP-002",
        deployment_model="SaaS", sourcing_type="SaaS", lifecycle_stage="Mature",
        implementation_date="2023-11-06", version_installed="current (SaaS)",
        version_vendor_supported="current (SaaS)", vendor_eos_date=None,
        technical_obsolescence_flag=False,
        business_owner="Marcus Beauchamp", technical_owner="Ingrid Solheim", legal_entity="LE-02",
        business_unit="Patient Access & Contact Centre", department="Contact Centre Operations",
        cost_centre="CC-2136", is_orphaned=False, governance_visibility="Managed", is_shadow_it=False,
        primary_capability="patient access", secondary_capabilities=None, capability_tag_confidence=0.93,
        licences_purchased=85, active_users=74, last_signin_date="2026-08-13",
        process_centrality="Medium", owner_stated_strategic_importance=3.5,
        cost_licence_subscription=165000, cost_upgrade_and_modules=0,
        cost_maintenance_dev_labour=32000, cost_infrastructure_peripherals=0,
        cost_indirect_and_training=14000, consumption_based_cost=54000,
        one_time_implementation_cost=0,
        contract_id="CTR-2025-0429", annual_contract_value=165000, term_start="2025-11-01",
        term_end="2026-10-31", auto_renewal_flag=True, renewal_notice_days=60,
        licence_metric="consumption", early_termination_penalty=0,
        has_downstream_dependents=False, dependency_count=3, integration_pattern="API",
        data_types_held="database; documents",
        ov_increase_value=3.5, ov_reach_consumers=4.0, ov_reduce_costs_efficiency=3.5,
        ov_patient_care_criticality=3.5, ov_governance_compliance=3.5,
        th_supportability=4.5, th_architecture_fit=4.0, th_operational_stability=4.0,
        th_vendor_viability=3.5, th_customization_debt=4.0,
        c_cost_per_active_user_vs_peers=2.5, c_unused_licence_waste=3.0,
        c_consumption_price_variance=2.0, c_absolute_cost_band=3.0,
        r_technical_risk=3.5, r_business_compliance_risk=3.5, r_clinical_safety_risk=3.5,
        r_end_user_perceived_quality=3.5,
        action="absorb", overlap_cluster_id="CL-05", cluster_role="absorbed",
        retention_obligation_flag=True, retention_expiry_date="2032-10-31", residual_archival_cost=0,
        information_classification="PHI",
        replacement_app_id="APP-002", replacement_ongoing_tco=0,
        replacement_cost_already_in_baseline=True,
        amortised_one_time_migration_cost=9000, realization_lag_months=4, saving_type="recurring",
        urg_timeline_sensitivity="High", urg_risk_pain_severity="Medium",
        data_source="CMDB export; SSO sign-in log; AP ledger; contract register",
        _intended_disposition="consolidate", _intended_priority="High", _intended_key="PPFP",
        _gross_saving_basis="run_rate",
        _evidence="The harder duplicate-AI finding. Nobody did anything wrong: it was procured properly, it has an owner, a contract and measurable deflection, and it is fully governed. The finding rests entirely on entitlement overlap inside a capability tag - the conversational scheduling and deflection capability is already included in the patient engagement subscription the organisation pays for. So the saving is the full run-rate minus configuration effort on the survivor, with no functional loss. A waste-signal-only engine misses this row completely; the notice window is open now, which is why it is urgent.",
    ),
]

# ---------------------------------------------------------------------------------------
# 3b. Cluster metadata (the "Overlap clusters" sheet) and the trap-case answer key
# ---------------------------------------------------------------------------------------
CLUSTERS = [
    dict(cluster_id="CL-01", capability="clinical communication",
         members=["APP-012", "APP-013"], survivor="APP-012", user_overlap_count=812,
         dependency_check_passed=True, wave=1,
         survivor_reason="Broader active adoption (5,910 active users vs 1,340), on-call directory "
                         "integration already live, and the lowest cost per active user in the cluster. "
                         "The absorbed product's only unique coverage is overhead paging, which the "
                         "survivor plus the existing overhead system already covers.",
         note="Saving is real but DEFERRED: the absorbed contract auto-renewed through its notice "
              "window on 2026-07-02, so the next servable deadline is 2027-07-02 and the saving lands FY28."),
    dict(cluster_id="CL-02", capability="analytics",
         members=["APP-014", "APP-015", "APP-016"], survivor="APP-014", user_overlap_count=1180,
         dependency_check_passed=True, wave="1 (Qlik) / 2 (Tableau)",
         survivor_reason="Already entitled inside the existing enterprise agreement, so only marginal "
                         "licence cost is carried; the largest active-user base of the three; and no "
                         "unique feature coverage in the other two for the reports actually in use.",
         note="The money slide. The survivor's low cost is what makes the cluster saving large. Landing "
              "cost on the survivor (premium capacity plus rebuild effort) is small but non-zero and is netted."),
    dict(cluster_id="CL-03", capability="clinical documentation (M&A pair MA-01)",
         members=["APP-001", "APP-003"], survivor="APP-001", user_overlap_count=0,
         dependency_check_passed=False, wave=3,
         survivor_reason="Same capability, different legal entity - the M&A duplication pattern. Survivor "
                         "chosen on business value, active user count and contract runway, not on which "
                         "entity acquired the other.",
         note="A naive dependency check FAILS here (18 HL7 interfaces at the absorbed site), which is why "
              "it sits in wave 3 and is excluded from wave 1 on clinical risk despite the largest single saving."),
    dict(cluster_id="CL-04", capability="clinical documentation (ambient AI)",
         members=["APP-007", "APP-008"], survivor="APP-007", user_overlap_count=96,
         dependency_check_passed=True, wave=2,
         survivor_reason="The enterprise agreement already covers every service line, so the absorbed "
                         "capability costs nothing incremental to land. The absorbed product is a "
                         "single-department pilot bought after the enterprise deal was signed.",
         note="Both apps are in Growth. Growth bars retire and replace but NOT consolidate, so the "
              "lifecycle guard does not swallow the finding."),
    dict(cluster_id="CL-05", capability="patient access (conversational AI)",
         members=["APP-002", "APP-020"], survivor="APP-002", user_overlap_count=74,
         dependency_check_passed=True, wave=1,
         survivor_reason="The conversational scheduling and deflection capability the absorbed product was "
                         "bought for is already entitled inside the patient engagement subscription the "
                         "organisation pays for. Retiring the standalone loses no function.",
         note="Entitlement overlap on a properly governed application: owner, contract, real usage, no "
              "waste signal. A waste-only engine misses this finding entirely."),
]

TRAP_CASES = [
    ("T1", "APP-017", "Old-but-adequate. Obsolescence is scored once, in technical health, and is not "
     "re-consumed by risk.", "invest / Moderate, key PFPP, action upgrade, detractors = version currency "
     "and platform supportability. Under five terms it must be invest and NOT retain: technical health "
     "genuinely fails, so there is something to fund.",
     "retire or replace - produced by letting one obsolescence fact fail "
     "two gates, or by reading 'old + cheap + small user base' as low value. NEW in v2: retain, produced "
     "by treating the five-term vocabulary as 'leave anything alone that is not being killed'."),
    ("T2", "APP-004", "The retention gate and residual cost that survives a retirement.",
     "retire / High but CONSTRAINED, retention obligation to 2031-06-30, residual archival cost netted off.",
     "Booking 100% of the run-rate as saving, or shutting it down before the archive exists."),
    ("T3", "APP-013", "Renewal calendar and timing deferral.",
     "consolidate / High, deferred, next servable notice 2027-07-02, saving lands FY28.",
     "Counting the saving this year because the gates and the overlap both say go - the "
     "'auto-renewed through the notice window' failure."),
    ("T4", "APP-010", "Shadow IT, plus null handling, plus confidence, plus entitlement overlap - all on "
     "one row.", "retire / Very High, is_shadow_it TRUE, Unsanctioned, alternative named as APP-011, "
     "confidence medium, and the missing field named (active_users).",
     "Scoring the nulls as zeros (right answer, wrong reason), or presenting it at the same confidence as "
     "a well-evidenced row."),
    ("T5", "APP-007 and APP-019", "Lifecycle exclusion - Birth/Growth cannot be retired or replaced.",
     "invest, lifecycle_exclusion_applied TRUE, and the guard state recorded in suppressed_recommendation.",
     "Retiring or replacing a new platform because adoption has not ramped and cost per user looks terrible."),
    ("T6", "APP-003", "A redundancy override beating a clean gate result; the M&A pair.",
     "consolidate / High on an app whose four gates all PASS, wave 3, migration cost netted.",
     "Leaving it as invest because all four gates pass, or claiming the full run-rate in wave 1."),
    ("T7", "APP-016", "Override DIRECTION - the capability must persist.",
     "consolidate (not retire), survivor named, dashboards rebuilt, on-prem infrastructure reclaimed.",
     "retire, which silently deletes a live capability."),
    ("T8", "APP-011", "Double-counting guard. The same row supplies an unused-licence saving AND justifies "
     "retiring APP-010.", "Both findings reported, neither double-counted: the reharvest explicitly reserves "
     "seats for the absorbed users, and unused licence spend is its own line.",
     "Adding APP-010's run-rate and APP-011's licence reharvest as if independent, or retiring the "
     "alternative that the other finding depends on."),
    ("T9", "APP-018 -> APP-019", "Netting against a successor that is already live and already in the baseline.",
     "Saving = incumbent run-rate minus amortised migration minus residual retention archive. The "
     "successor's own cost is NOT subtracted again.",
     "Subtracting the successor's full $2.1M (understating), or claiming the incumbent's full $1.9M (overstating)."),
    ("T10", "APP-006", "replace netting against a successor that does NOT exist in the portfolio yet.",
     "replace / High with the successor's own ongoing run cost populated and subtracted.",
     "Claiming the incumbent's full cost as saving because no successor row exists to net against."),
    ("T11", "APP-020", "Entitlement overlap on a properly governed application.",
     "consolidate / High on entitlement overlap alone.",
     "Missing it entirely because the row has an owner, a contract, real usage and no waste signal."),
    ("T12", "APP-001 and APP-002", "The retain/invest boundary, and that biggest-spend does not imply "
     "action. REWRITTEN FOR v2: in v1 both rows were invest and were told apart by priority; under "
     "Bina's five terms they land on DIFFERENT WORDS.",
     "APP-001 RETAIN / Very Low on PPPP - the largest line item in the portfolio, healthy, leave it "
     "alone, action monitor, no saving claimed. APP-002 INVEST / High on PPPF - fund a remediation of "
     "RISK, not of age, action remediate.",
     "Recommending action against the largest line item; or emitting invest for APP-001, which is the "
     "v1 answer and is now wrong; or still separating the two by priority alone."),
    ("T13", "APP-012 and APP-014", "That a cluster survivor keeps its own gate disposition (D2) and "
     "that the disposition is now retain. NEW in v2 - the vocabulary change is what makes this "
     "testable.", "retain / Very Low, cluster_role survivor, action absorb, consolidation_saving blank. "
     "The survivor is not stamped consolidate, and it is not an invest either: nothing about the "
     "survivor is being funded, the migration cost sits on the absorbed rows.",
     "'Consolidate Power BI into Power BI'; or invest / Very Low, the v1 answer, which reads to a "
     "steering committee as a funding request against an application nobody is proposing to spend on."),
    ("T14", "APP-005 and APP-011 and APP-019", "That a cost-efficiency-only failure is an invest and "
     "never a retire, and never a retain either. NEW in v2.",
     "All three are PPFP: value, technical health and risk pass, cost efficiency fails. All three are "
     "invest, and the invest names cost efficiency as the thing being funded.",
     "retire - reading an expensive application as a dead one, which Bina's Q2 answer explicitly rules "
     "out; or retain, ignoring the failing gate because the other three pass."),
]

# ---------------------------------------------------------------------------------------
# 3c. What v1 shipped, read out of App-Rationalization-Dummy-Dataset-v1.xlsx
# ---------------------------------------------------------------------------------------
# Held here only so the v1 -> v2 comparison table can be printed and the intended changes
# checked. Nothing in the engine reads it. If a row's v2 disposition differs from its v1 one
# and it is not in INTENDED_V2_CHANGES below, the run stops.
V1_OUTPUT = {
    "APP-001": ("invest", "Very Low"),   "APP-002": ("invest", "High"),
    "APP-003": ("consolidate", "High"),  "APP-004": ("retire", "High"),
    "APP-005": ("invest", "Moderate"),   "APP-006": ("replace", "High"),
    "APP-007": ("invest", "Moderate"),   "APP-008": ("consolidate", "Moderate"),
    "APP-009": ("retire", "High"),       "APP-010": ("retire", "Very High"),
    "APP-011": ("invest", "Moderate"),   "APP-012": ("invest", "Very Low"),
    "APP-013": ("consolidate", "High"),  "APP-014": ("invest", "Very Low"),
    "APP-015": ("consolidate", "Moderate"), "APP-016": ("consolidate", "High"),
    "APP-017": ("invest", "Moderate"),   "APP-018": ("replace", "Very High"),
    "APP-019": ("invest", "High"),       "APP-020": ("consolidate", "High"),
}

# The only disposition changes v2 is allowed to make: app_id -> (the term it must land on, why).
# The target term is part of the contract, so a row that changes to something ELSE still stops
# the run even though it is on this list.
INTENDED_V2_CHANGES = {
    "APP-001": ("retain",
                "PPPP: all four gates pass and nothing is being funded, so v1's invest / Very "
                "Low was really retain all along. The largest line item in the portfolio, and "
                "the right answer is still to leave it alone."),
    "APP-012": ("retain",
                "PPPP: cluster survivor, all four gates pass. Keeps its gate disposition under "
                "D2, and that disposition is now retain rather than a Very Low invest."),
    "APP-014": ("retain",
                "PPPP: cluster survivor, all four gates pass. Same as APP-012 - retain reads "
                "correctly where 'invest in Power BI, priority Very Low' did not."),
}

# ---------------------------------------------------------------------------------------
# 4. THE ENGINE - everything below is computed, nothing is asserted
# ---------------------------------------------------------------------------------------
ACTIONS = ("none", "monitor", "remediate", "upgrade", "retrain", "merge", "absorb",
           "decommission", "re-platform")

# Actions that by definition cost money, so they cannot sit on a retain row.
FUNDED_ACTIONS = ("remediate", "upgrade", "retrain", "re-platform")


def d(s):
    """ISO string -> date, passthrough None."""
    return dt.date.fromisoformat(s) if s else None


def dimension_score(row, dim_key):
    """Weighted arithmetic mean on the 1-5 scale, renormalised over the criteria that
    actually carry a value (REQ 52's correction to the reference tool's blank handling: a sparse
    row is reported as incomplete, never silently scored downward toward retire)."""
    num = den = 0.0
    for name, dim, weight, _desc in CRITERIA:
        if dim != dim_key or weight == 0:
            continue
        val = row.get(name)
        if val is None:
            continue
        num += val * weight
        den += weight
    if den == 0:
        return None
    return round(num / den, 3)


def gate(score):
    """The comparison is >=, so exactly 3.0 passes."""
    if score is None:
        return "F"
    return "P" if score >= PASS_THRESHOLD else "F"


def retain_or_invest(row):
    """Which of Bina's two "keep it" terms a row earns, and why. Returns (term, why).

    Bina's ruling separated invest into two words: retain means healthy, leave it alone, no
    spend; invest means a deliberate injection of money or effort. The rule that separates them
    is read off the pass/fail gates, not off any row by hand:

        all four dimensions pass  ->  retain    (there is no failing dimension to fund)
        any dimension fails       ->  invest    (the failing dimension IS what the money buys)

    That is exactly one key, PPPP, and it is why the other fifteen rows of the lookup table
    needed no change at all: every row that returned invest in v1 already fails at least one
    gate, so an invest recommendation can always name the dimension being funded. PPPF funds
    risk, PPFP funds cost efficiency, PFPP funds technical health. Priority no longer carries
    any part of the retain/invest distinction - it is only ever about urgency now.

    Cost is not special-cased here and must not be: a cost-efficiency-only failure (PPFP)
    resolves to invest, never to retire. Cost moves the queue, it does not kill an application
    (Bina's answer to Q2).
    """
    failed = [name for _k, name, _col, flag in DIMENSIONS if row[flag] == "F"]
    if failed:
        return "invest", (f"invest in {' and in '.join(failed)}: that is the dimension failing the "
                          f"{PASS_THRESHOLD:.1f} gate, so that is what the money buys")
    return "retain", ("all four dimensions clear the 3.0 gate, so there is no failing dimension "
                      "to fund and nothing to buy: keep it, spend nothing")


def step_priority(priority, steps):
    i = PRIORITY_LADDER.index(priority)
    return PRIORITY_LADDER[min(max(i + steps, 0), len(PRIORITY_LADDER) - 1)]


def urgency(row):
    a = HML.get(row.get("urg_timeline_sensitivity"))
    b = HML.get(row.get("urg_risk_pain_severity"))
    vals = [x for x in (a, b) if x is not None]
    return round(sum(vals) / len(vals), 3) if vals else None


def override_priority(row):
    """Priority for a row whose disposition came from the redundancy override rather than
    from the gate lookup. Documented rule, applied in order:
       1. Birth/Growth  -> Moderate  (do not rush an application that is still ramping)
       2. urgency >= 0.75 -> High    (a renewal, notice window or end-of-support forces the decision)
       3. net saving >= $700k -> High
       4. otherwise     -> Moderate
    """
    if row["lifecycle_stage"] in ("Birth", "Growth"):
        return "Moderate", "lifecycle is Growth, so the consolidation is not rushed"
    u = row.get("urgency_score") or 0
    if u >= 0.75:
        return "High", f"urgency {u:.2f} - a contract or support deadline forces the decision"
    if (row.get("net_saving_annual") or 0) >= 700000:
        return "High", "net saving above the $700k materiality line"
    return "Moderate", "material saving but no forcing deadline"


def money(x):
    return 0 if x is None else x


def compute(row):
    """Fill every derived field on one row, in the order the engine actually runs."""
    # ---- usage and cost arithmetic -----------------------------------------------------
    lic, act = row.get("licences_purchased"), row.get("active_users")
    row["licence_utilisation_rate"] = round(act / lic, 3) if lic and act is not None else None
    row["unused_licence_count"] = lic - act if lic and act is not None else None
    cats = ["cost_licence_subscription", "cost_upgrade_and_modules", "cost_maintenance_dev_labour",
            "cost_infrastructure_peripherals", "cost_indirect_and_training"]
    row["tco_five_category_subtotal"] = sum(money(row[c]) for c in cats)
    row["annual_tco_recurring"] = row["tco_five_category_subtotal"] + money(row["consumption_based_cost"])
    row["five_year_cumulative_tco"] = row["annual_tco_recurring"] * TCO_HORIZON_YEARS
    row["cost_per_active_user"] = round(row["annual_tco_recurring"] / act) if act else None
    row["unused_licence_spend"] = (
        round(money(row["cost_licence_subscription"]) * (1 - row["licence_utilisation_rate"]))
        if row["licence_utilisation_rate"] is not None else None)

    # ---- contract arithmetic -----------------------------------------------------------
    end, notice = d(row.get("term_end")), row.get("renewal_notice_days")
    if end and notice:
        deadline = end - dt.timedelta(days=notice)
        row["notice_deadline_date"] = deadline.isoformat()
        delta = (deadline - ANALYSIS_DATE).days
        row["in_notice_window_now"] = 0 <= delta <= 180
    else:
        row["notice_deadline_date"] = None
        row["in_notice_window_now"] = None
    row["contract_runway_months"] = (
        max(0, round((end - ANALYSIS_DATE).days / 30.44)) if end else None)

    # ---- the four dimension scores, the gates and the key ------------------------------
    row["business_value_score"] = dimension_score(row, "V")
    row["technical_health_score"] = dimension_score(row, "T")
    row["cost_efficiency_score"] = dimension_score(row, "C")
    row["risk_posture_score"] = dimension_score(row, "R")
    row["v_pass"] = gate(row["business_value_score"])
    row["t_pass"] = gate(row["technical_health_score"])
    row["c_pass"] = gate(row["cost_efficiency_score"])
    row["r_pass"] = gate(row["risk_posture_score"])
    row["vtcr_key"] = row["v_pass"] + row["t_pass"] + row["c_pass"] + row["r_pass"]

    disposition, priority = DISPOSITION_TABLE[row["vtcr_key"]]
    priority_reason = f"straight from the {row['vtcr_key']} row of the lookup table"

    # ---- the retain / invest split (v2) -------------------------------------------------
    # The table hardcodes the term for all 16 keys, so nothing is decided here. What runs here
    # is the CHECK that the table agrees with the stated rule: on every key the table maps to
    # retain or invest, retain_or_invest() must return the same word. If it ever did not, the
    # table and the rule would be telling different stories and sanity_checks refuses to write.
    # The reasoning is recorded on the row either way, so a reader can see which dimension an
    # invest is funding, or that a retain has no failing dimension at all.
    row["_split_conflict"] = None
    if disposition in ("retain", "invest"):
        resolved, why = retain_or_invest(row)
        row["retain_or_invest_basis"] = why
        if resolved != disposition:
            row["_split_conflict"] = (f"the lookup table maps {row['vtcr_key']} to {disposition} "
                                      f"but the retain/invest rule returns {resolved}")
    else:
        row["retain_or_invest_basis"] = None

    row["_gate_disposition"] = disposition
    row["_gate_priority"] = priority

    # ---- guard 1: lifecycle (REQ 51). Runs after the lookup, never silently. -----------
    row["lifecycle_exclusion_applied"] = row["lifecycle_stage"] in ("Birth", "Growth")
    suppressed, reasons = None, []
    if row["lifecycle_exclusion_applied"]:
        if disposition in ("retire", "replace"):
            suppressed = disposition
            disposition, priority = "invest", "High"
            priority_reason = "lifecycle guard replaced a retire/replace with a funded invest"
            reasons.append(f"lifecycle stage {row['lifecycle_stage']}: retire and replace are barred "
                           f"outright, so the {suppressed} the gates returned is suppressed")
        else:
            suppressed = f"none - guard armed, gates returned {disposition}"
            reasons.append(f"lifecycle stage {row['lifecycle_stage']}: the retire/replace bar is armed "
                           f"but the gates did not reach either, so nothing was suppressed. Consolidate "
                           f"is NOT barred by lifecycle")

    # ---- guard 2: sourcing (REQ 51) ----------------------------------------------------
    row["sourcing_exclusion_applied"] = (row["sourcing_type"] == "SaaS" and disposition == "replace")
    if row["sourcing_exclusion_applied"]:
        reasons.append("sourcing type SaaS: rebuilding or re-platforming the product in place is not a "
                       "legal option, so the action is substitution by another product, not a re-platform "
                       "of this one")

    # ---- override: redundancy (REQ 25/52). Not a gate - an override. -------------------
    row["redundancy_override_applied"] = row.get("cluster_role") == "absorbed"
    if row["redundancy_override_applied"]:
        if disposition != "consolidate":
            suppressed = f"{disposition} (gate result, overridden by cluster membership)"
            reasons.append(f"cluster {row['overlap_cluster_id']} membership overrides the gate result and "
                           f"forces consolidate into {row['replacement_app_id']}; the capability persists")
        disposition = "consolidate"

    row["disposition"], row["priority"] = disposition, priority
    row["_priority_reason"] = priority_reason
    row["_suppressed"] = suppressed
    row["_reasons"] = reasons
    return row


def compute_savings(row):
    """REQ 32/54 netting: never assume a retire returns 100% of an application's cost."""
    basis = row["_gross_saving_basis"]
    if basis == "run_rate" and row["disposition"] in ("retire", "consolidate", "replace"):
        gross = row["annual_tco_recurring"]
    elif basis == "licence_reharvest":
        gross = round(money(row["unused_licence_spend"]) * row["_reharvest_share"]) \
                - row["_reharvest_reserve"]
    else:
        gross = 0
    successor = 0 if row["replacement_cost_already_in_baseline"] else money(row["replacement_ongoing_tco"])
    net = gross - successor - money(row["amortised_one_time_migration_cost"]) \
        - money(row["residual_archival_cost"])
    row["gross_saving_annual"] = gross
    row["net_saving_annual"] = net
    row["net_saving_five_year"] = net * TCO_HORIZON_YEARS
    row["saving_type"] = row["saving_type"] if gross else "none"
    row["consolidation_saving"] = net if row.get("cluster_role") == "absorbed" else None
    return row


def apply_constraints_and_priority(rows):
    """Runs after savings, because the override priority rule reads the net saving, and after
    every row has a disposition, because the successor bump reads another row's."""
    by_id = {r["app_id"]: r for r in rows}

    for row in rows:
        # priority for override-derived consolidations
        if row["redundancy_override_applied"]:
            row["priority"], why = override_priority(row)
            row["_priority_reason"] = f"redundancy override: {why}"

        # retention gate (REQ 29/53): a retire with a live retention obligation is constrained
        expiry = d(row.get("retention_expiry_date"))
        row["retention_override_applied"] = bool(
            row["disposition"] == "retire" and row.get("retention_obligation_flag")
            and expiry and expiry > ANALYSIS_DATE)
        if row["retention_override_applied"]:
            was = row["priority"]
            row["priority"] = step_priority(was, -1)
            row["_priority_reason"] = (f"retention constraint stepped priority down from {was}: the "
                                       f"archive has to exist before anything can be switched off")
            row["_reasons"].append(
                f"retention obligation runs to {row['retention_expiry_date']} and the application cannot "
                f"purge its own data, so the retirement is CONSTRAINED, not cancelled; a residual archival "
                f"cost of ${money(row['residual_archival_cost']):,.0f} a year survives shutdown and is "
                f"netted off the saving")

    # successor bump (REQ 39): a platform on the critical path of a replace goes up one step
    for row in rows:
        blockers = [r["app_id"] for r in rows
                    if r.get("replacement_app_id") == row["app_id"] and r["disposition"] == "replace"]
        if blockers:
            was = row["priority"]
            row["priority"] = step_priority(was, +1)
            row["_priority_reason"] = (f"stepped up from {was}: {', '.join(blockers)} cannot complete its "
                                       f"replacement until this rollout finishes")
            row["_reasons"].append(f"named successor for {', '.join(blockers)}, so its rollout is on the "
                                   f"critical path of another initiative")

    for row in rows:
        row["suppressed_recommendation"] = row["_suppressed"]
        row["suppression_reason"] = "; ".join(row["_reasons"]) if row["_reasons"] else None
    return rows


def compute_provenance(row):
    """Completeness and confidence (REQ 12/28). 'unknown' counts as missing, not as a value."""
    missing = [f for f in COMPLETENESS_FIELDS
               if row.get(f) is None or (isinstance(row.get(f), str) and row[f] == "unknown")]
    row["completeness_score"] = round(1 - len(missing) / len(COMPLETENESS_FIELDS), 3)
    row["missing_fields"] = "; ".join(missing) if missing else None
    all_scores_present = all(row.get(n) is not None for n, _dim, w, _d in CRITERIA if w > 0)
    if row["completeness_score"] >= 0.90 and all_scores_present:
        row["confidence"] = "high"
    elif row["completeness_score"] >= 0.40:
        row["confidence"] = "medium"
    else:
        row["confidence"] = "low"
    return row


def phrase(name, score, flag):
    return f"{name} {score:.2f} {'passes' if flag == 'P' else 'FAILS'}" if score is not None \
        else f"{name} could not be scored (no populated criteria) and is treated as a fail"


def compose_rationale(row):
    head = (f"{row['vtcr_key']}: "
            + "; ".join([
                phrase("business value", row["business_value_score"], row["v_pass"]),
                phrase("technical health", row["technical_health_score"], row["t_pass"]),
                phrase("cost efficiency", row["cost_efficiency_score"], row["c_pass"]),
                phrase("risk posture", row["risk_posture_score"], row["r_pass"]),
            ]) + f" against a 3.0 gate on every dimension.")
    parts = [head, row["_evidence"]]
    if row["suppression_reason"]:
        parts.append("Guardrails: " + row["suppression_reason"] + ".")
    net = row["net_saving_annual"]
    if row["gross_saving_annual"]:
        successor = 0 if row["replacement_cost_already_in_baseline"] \
            else money(row["replacement_ongoing_tco"])
        base = ("licence reharvest, net of a seat reserve for the users being absorbed"
                if row["_gross_saving_basis"] == "licence_reharvest" else "current run-rate")
        arith = (f"Saving arithmetic ({base}): ${row['gross_saving_annual']:,.0f} gross "
                 f"- ${successor:,.0f} successor run cost "
                 f"- ${money(row['amortised_one_time_migration_cost']):,.0f} amortised one-time "
                 f"- ${money(row['residual_archival_cost']):,.0f} residual archival "
                 f"= ${net:,.0f} net a year (${row['net_saving_five_year']:,.0f} over five years, "
                 f"undiscounted), realised after {row['realization_lag_months']} months.")
        if row["replacement_cost_already_in_baseline"]:
            arith += (" The successor's own cost is already a paid line in this portfolio, so it is "
                      "deliberately NOT subtracted a second time.")
        parts.append(arith)
    elif row["disposition"] == "retain":
        parts.append("No saving is claimed for this row, and no spend either: the recommendation is "
                     "to keep it exactly as it is - "
                     + row["retain_or_invest_basis"] + ".")
    else:
        parts.append("No saving is claimed for this row; the recommendation is to keep and fund it.")
    parts.append(f"Disposition {row['disposition']}, priority {row['priority']} "
                 f"({row['_priority_reason']}), action {row['action']}, confidence {row['confidence']}"
                 + (f"; unresolved fields: {row['missing_fields']}." if row["missing_fields"] else "."))
    return " ".join(parts)


def build():
    rows = [compute(dict(r)) for r in ROWS]
    for r in rows:
        # urgency is an input roll-up, and the override priority rule reads it
        r["urgency_score"] = urgency(r)
    rows = [compute_savings(r) for r in rows]
    rows = apply_constraints_and_priority(rows)
    for r in rows:
        compute_provenance(r)
        r["rationale"] = compose_rationale(r)
    return rows


# ---------------------------------------------------------------------------------------
# 5. SANITY CHECKS - the arithmetic has to hold before anything is written
# ---------------------------------------------------------------------------------------
def sanity_checks(rows):
    problems = []
    for r in rows:
        a = r["app_id"]
        cats = sum(money(r[c]) for c in
                   ["cost_licence_subscription", "cost_upgrade_and_modules",
                    "cost_maintenance_dev_labour", "cost_infrastructure_peripherals",
                    "cost_indirect_and_training"])
        if cats != r["tco_five_category_subtotal"]:
            problems.append(f"{a}: five-category subtotal mismatch")
        if cats + money(r["consumption_based_cost"]) != r["annual_tco_recurring"]:
            problems.append(f"{a}: categories + consumption != annual_tco_recurring")
        if r["five_year_cumulative_tco"] != r["annual_tco_recurring"] * 5:
            problems.append(f"{a}: five-year TCO is not 5 x the run-rate")
        if r["licence_utilisation_rate"] is not None:
            if abs(r["licence_utilisation_rate"] - r["active_users"] / r["licences_purchased"]) > 0.0005:
                problems.append(f"{a}: utilisation != active / purchased")
            if r["unused_licence_count"] != r["licences_purchased"] - r["active_users"]:
                problems.append(f"{a}: unused licence count wrong")
        if r["net_saving_annual"] < 0:
            problems.append(f"{a}: NEGATIVE net saving {r['net_saving_annual']:,} - deliberate? flag it")
        if r["net_saving_five_year"] != r["net_saving_annual"] * 5:
            problems.append(f"{a}: five-year saving is not 5 x annual")
        if r["disposition"] not in DISPOSITIONS:
            problems.append(f"{a}: disposition '{r['disposition']}' outside the five agreed terms")
        # retain means no spend, so it cannot claim a saving or carry a funded action (v2)
        if r["disposition"] == "retain":
            if r["gross_saving_annual"]:
                problems.append(f"{a}: retain claims a ${r['gross_saving_annual']:,} saving - "
                                f"retain means leave it alone, which returns nothing")
            if r["action"] in FUNDED_ACTIONS:
                problems.append(f"{a}: retain with action '{r['action']}' - that action costs "
                                f"money, so the row is an invest")
        # invest is a deliberate injection of money or effort, so a bare 'none' contradicts it
        if r["disposition"] == "invest" and r["action"] == "none":
            problems.append(f"{a}: invest with action 'none' - nothing is being funded, so under "
                            f"the five-term vocabulary this row is a retain")
        # the basis records the GATE result, so it is keyed on that and survives an override
        # (APP-003 gates to retain and is then overridden to consolidate; both facts are kept)
        if (r["_gate_disposition"] in ("retain", "invest")) != (r["retain_or_invest_basis"] is not None):
            problems.append(f"{a}: retain_or_invest_basis is out of step with the gate result "
                            f"'{r['_gate_disposition']}'")
        if r["_split_conflict"]:
            problems.append(f"{a}: {r['_split_conflict']}")
        # the rule's whole justification: an invest can always name a failing dimension
        if r["disposition"] == "invest" and r["vtcr_key"] == "PPPP":
            problems.append(f"{a}: invest on an all-pass key - an invest must be able to name "
                            f"the dimension it is funding")
        # cost never kills an application on its own (Bina's Q2 answer)
        if r["vtcr_key"] == "PPFP" and r["disposition"] == "retire":
            problems.append(f"{a}: a cost-efficiency-only failure resolved to retire")
        if r["action"] not in ACTIONS:
            problems.append(f"{a}: action '{r['action']}' outside the vocabulary")
        if r["priority"] not in PRIORITY_LADDER:
            problems.append(f"{a}: priority '{r['priority']}' outside the vocabulary")
        for name, _dim, _w, _d in CRITERIA:
            v = r.get(name)
            if v is not None and v not in SCORE_STEPS:
                problems.append(f"{a}: {name}={v} is not on the 1-5 half-step scale")
        if r["disposition"] in ("consolidate", "replace") and not r.get("replacement_app_id"):
            problems.append(f"{a}: {r['disposition']} with no named successor (REQ 58)")
        missing_cols = [c for c in COLUMN_ORDER if c not in r]
        if missing_cols:
            problems.append(f"{a}: columns not produced: {missing_cols}")
    return problems


def comparison_table(rows):
    """v1 vs v2, row by row, with every change classified as intended or not."""
    lines, unexpected = [], []
    header = (f"{'app_id':8} {'app_name':34} {'key':5} {'v1 disp':12} {'v2 disp':12} "
              f"{'priority':10} {'changed':8} intended?")
    lines.append(header)
    lines.append("-" * len(header))
    for r in rows:
        v1_disp, v1_pri = V1_OUTPUT[r["app_id"]]
        changed = v1_disp != r["disposition"]
        want, why = INTENDED_V2_CHANGES.get(r["app_id"], (None, None))
        if changed and want != r["disposition"]:
            unexpected.append(f"{r['app_id']} {v1_disp} -> {r['disposition']}"
                              + (f" (v2 intended {want})" if want else " (no change intended)"))
            verdict = "NO - NOT AN INTENDED CHANGE"
        elif changed:
            verdict = "YES - " + why[:44]
        elif want:
            unexpected.append(f"{r['app_id']} should have changed to {want} and did not")
            verdict = f"NO - was meant to become {want}"
        else:
            verdict = "n/a - unchanged"
        pri = r["priority"] if r["priority"] == v1_pri else f"{v1_pri} -> {r['priority']}"
        lines.append(f"{r['app_id']:8} {r['app_name'][:34]:34} {r['vtcr_key']:5} "
                     f"{v1_disp:12} {r['disposition']:12} {pri:10} "
                     f"{'yes' if changed else 'no':8} {verdict}")
    return lines, unexpected


def verification_table(rows):
    """Compare every computed disposition against the intended disposition in roster-design.md."""
    lines, mismatches = [], []
    header = (f"{'app_id':8} {'app_name':34} {'key':5} {'intended':12} {'computed':12} "
              f"{'match':5} {'pri(int)':9} {'pri(comp)':9} match")
    lines.append(header)
    lines.append("-" * len(header))
    for r in rows:
        ok = r["disposition"] == r["_intended_disposition"]
        pok = r["priority"] == r["_intended_priority"]
        if not ok:
            mismatches.append(r["app_id"])
        lines.append(f"{r['app_id']:8} {r['app_name'][:34]:34} {r['vtcr_key']:5} "
                     f"{r['_intended_disposition']:12} {r['disposition']:12} "
                     f"{'YES' if ok else 'NO':5} {r['_intended_priority']:9} {r['priority']:9} "
                     f"{'YES' if pok else 'NO'}")
    return lines, mismatches

# ---------------------------------------------------------------------------------------
# 6. PORTFOLIO ROLL-UP
# ---------------------------------------------------------------------------------------
def rollup(rows):
    total = sum(r["annual_tco_recurring"] for r in rows)
    target = round(total * TARGET_REDUCTION_PCT)
    net_all = sum(r["net_saving_annual"] for r in rows)
    constrained = [r for r in rows if r["retention_override_applied"]
                   or (r["realization_lag_months"] or 0) >= 12]
    net_unconstrained = net_all - sum(r["net_saving_annual"] for r in constrained)
    return dict(
        total_annual_tco=total,
        target=target,
        net_saving_all=net_all,
        net_saving_excluding_constrained_and_deferred=net_unconstrained,
        constrained_or_deferred=[r["app_id"] for r in constrained],
        spread=Counter(r["disposition"] for r in rows),
        priority_spread=Counter(r["priority"] for r in rows),
        keys=Counter(r["vtcr_key"] for r in rows),
    )


# ---------------------------------------------------------------------------------------
# 7. SHEET COPY (plain language, for a non-engineer)
# ---------------------------------------------------------------------------------------
def readme_blocks(rows, summary):
    t = summary
    return [
        ("h1", "Application Rationalization - synthetic demo dataset"),
        ("p", f"THIS IS VERSION 2 ({DATASET_VERSION}). It replaces "
              f"App-Rationalization-Dummy-Dataset-v1.xlsx. Everything is read 'as at' "
              f"{ANALYSIS_DATE.isoformat()}, which is the date the contract windows and sign-in dates "
              f"are measured against."),

        ("h2", "What changed from v1, and who decided it"),
        ("p", "Bina Din reviewed v1 and answered the five open questions it carried. Her answers are "
              "recorded word for word on the 'Notes & assumptions' sheet, each one next to the change "
              "it produced. Two of them changed this file:"),
        ("bullet", "THERE ARE NOW FIVE RECOMMENDATION WORDS, NOT FOUR: retain, invest, consolidate, "
                   "replace, retire. This is Bina's decision, in her words: \"No, separate by invest, "
                   "retain, consolidate, replace, and retire.\" In v1, invest had to mean two different "
                   "things - a healthy application we leave alone, and an application we are putting "
                   "money into - and the only thing telling them apart was the priority column. Now "
                   "retain means healthy, leave it alone, spend nothing, and invest means we are "
                   "deliberately funding a remediation or an enhancement. Three applications that said "
                   "invest in v1 say retain here: Epic Hyperspace, TigerConnect and Power BI. Nothing "
                   "else moved."),
        ("bullet", "PATIENT-CARE CRITICALITY NOW CARRIES THE DOUBLE WEIGHT in the business value score, "
                   "and governance and compliance drops to single weight. That is Bina's answer to the "
                   "question v1 raised about which value signal a health system should weigh twice. The "
                   "column that scores it is renamed from ov_enhance_services to "
                   "ov_patient_care_criticality so it says what it measures. Every score was "
                   "recalculated; no application's pass/fail result changed, because the applications "
                   "that matter clinically already scored well on both."),
        ("p", "Everything else is deliberately identical to v1: the same 20 applications, the same real "
              "product names with everything else invented, the same overlap clusters, the same trap "
              "cases, and exactly the same cost and savings arithmetic. The portfolio run-rate and every "
              "saving figure in this file are unchanged from v1 to the dollar."),

        ("h2", "What this file is"),
        ("p", "A 20-application pretend portfolio for a pretend US health system, carrying every data "
              "point needed to actually run an application rationalization exercise: what each "
              "application is, who owns it, what it costs, how much of it is used, what it depends on, "
              "how it scores, what we recommend doing with it, and what that recommendation saves."),
        ("p", "It exists so the team can build and demonstrate the tool against realistic data before "
              "any real portfolio is loaded, and so a reviewer can check the tool's answers against a "
              "known answer key."),

        ("h2", "What is real and what is invented"),
        ("p", "The vendor and product names are real, widely known commercial products a health system "
              "would plausibly run. That was a deliberate instruction, so the demo reads as recognisable."),
        ("p", "EVERYTHING ELSE IS INVENTED. Every cost, contract identifier, contract date, licence "
              "count, user count, version, score, saving, owner name, legal entity, business unit and "
              "cost centre on this sheet was made up for this dataset. The three legal entities "
              "(Lakeshore Health Partners, Lakeshore Medical Group, Riverbend Community Hospital) do not "
              "exist. The owner names are obviously fictional. No real organisation's portfolio, "
              "contract terms or internal figures appear anywhere in this file, and none may be added."),

        ("h2", "How to use it"),
        ("bullet", "'Applications' is the dataset - one row per application. Read it left to right: "
                   "what it is, who owns it, what it does, how much it is used, what it costs, what the "
                   "contract says, what depends on it, how it scores, what we recommend, what it saves."),
        ("bullet", "'Data dictionary' explains every column in one sentence, with the requirement number "
                   "that asked for it."),
        ("bullet", "'Scoring model' shows exactly how a recommendation is produced from the scores. If "
                   "someone disagrees with an answer, this is the sheet to argue with."),
        ("bullet", "'Overlap clusters' is the consolidation story: which applications do the same job, "
                   "which one we keep, why, and what folding the others in saves."),
        ("bullet", f"'Trap cases' is the reviewer's answer key: {len(TRAP_CASES)} rows built to catch a "
                   f"specific mistake, each with the right answer and the wrong answer it is designed to "
                   f"catch. Three of them exist because of the five-term vocabulary."),
        ("bullet", "'Notes & assumptions' records every modelling choice, Bina's five answers in her own "
                   "words with the decision each one produced, and what is still open."),

        ("h2", "The headline numbers in this dataset"),
        ("bullet", f"Portfolio annual run-rate: ${t['total_annual_tco']:,.0f} across 20 applications."),
        ("bullet", f"A 15% reduction target is therefore ${t['target']:,.0f} a year."),
        ("bullet", f"Net annual saving identified: ${t['net_saving_all']:,.0f} - and "
                   f"${t['net_saving_excluding_constrained_and_deferred']:,.0f} of that is available "
                   f"without counting the rows that are blocked or deferred past twelve months, which is "
                   f"the honest version of the number."),
        ("bullet", "Recommendations: " + ", ".join(f"{k} {v}" for k, v in
                   sorted(t['spread'].items(), key=lambda x: -x[1])) + "."),

        ("h2", "The five open questions v1 raised are now closed"),
        ("p", "All five were answered by Bina Din on 2026-08-14. Her exact words and the decision each "
              "one produced are on the 'Notes & assumptions' sheet under 'Answers from Bina Din'. In "
              "short: keep risk in the end-user slot (Q1); accept that cost moves the queue but never "
              "on its own makes something a retire (Q2); move the value dimension's double weight to "
              "patient-care criticality (Q3); keep the modelled peer cost band and label it as modelled "
              "(Q4); and split invest into invest and retain (Q5)."),

        ("h2", "Still open - these need Ryo or Bina, not another modelling pass"),
        ("bullet", "O1. TigerConnect was designed as priority Low but the lookup table returns Very Low "
                   "for every all-pass row. Either the design note is a slip, or a cluster survivor that "
                   "carries absorb work should be bumped one step - in which case Power BI moves too. "
                   "Bina's call. Carried over from v1 unresolved."),
        ("bullet", "O2. Where the peer cost band comes from for the real engagement. Bina confirmed we "
                   "continue with the model for the demo, which settles this dataset but not the "
                   "engagement: a real portfolio will want a benchmark source named."),
        ("bullet", "O3. Two cluster survivors carry action 'absorb' rather than 'monitor' while sitting "
                   "on a retain disposition. Absorb is the more specific verb and describes real work, "
                   "but that work is costed on the absorbed rows, not the survivor. Confirm 'absorb' is "
                   "the verb you want to see against a retain, or say the word and they become 'monitor'."),
    ]


def notes_blocks(rows, summary):
    t = summary
    app = {r["app_id"]: r for r in rows}
    retained = [r["app_id"] for r in rows if r["disposition"] == "retain"]
    invested = [r["app_id"] for r in rows if r["disposition"] == "invest"]
    survivors = [r for r in rows if r.get("cluster_role") == "survivor"]
    surv_retain = [r["app_id"] for r in survivors if r["disposition"] == "retain"]
    blank_scores = sum(1 for n, _dim, w, _d in CRITERIA
                       if w > 0 and app["APP-010"].get(n) is None)
    return [
        ("h1", "Notes and assumptions"),
        ("p", f"Every modelling choice made in producing this dataset, so a reviewer can challenge the "
              f"rule rather than the number. Nothing below is measured; all of it is modelled. This is "
              f"{DATASET_VERSION}."),

        ("h2", "Answers from Bina Din, 2026-08-14 - the five v1 open questions, closed"),
        ("p", "Each answer is quoted exactly as given, followed by what was changed in this file as a "
              "result. Bina Din is the subject-matter expert for this engagement; these are her rulings, "
              "not the modelling team's preferences."),
        ("bullet", "Q1 asked whether it is the right trade to keep the borrowed engine's mechanics but "
                   "put RISK in the fourth lens slot, with end-user "
                   "perception still collected at weight 0. BINA: \"Yes\", keep it. DECISION: kept "
                   "exactly as v1 built it. Risk occupies position 4, r_end_user_perceived_quality is "
                   "still collected and still carries weight 0. No change to this file."),
        ("bullet", "Q2 asked whether we accept that cost moves an application's PRIORITY but can never "
                   "on its own make it a retire. BINA: \"Yes\", accept it. DECISION: accepted and now "
                   "enforced rather than merely described - the generator refuses to write if a "
                   "cost-efficiency-only failure (key PPFP) ever resolves to retire. The three PPFP rows "
                   "in this portfolio all come out invest, and the invest names cost efficiency as the "
                   "thing being funded. Trap case T14 tests it."),
        ("bullet", "Q3 asked whether patient-care criticality, rather than the engine's governance and "
                   "compliance criterion, deserves the weight-2 slot in the value dimension. BINA: "
                   "\"Yes\" - make patient-care criticality the double-weighted value criterion. "
                   "DECISION: ov_patient_care_criticality now carries weight 2 and "
                   "ov_governance_compliance moves to weight 1. The criterion was called "
                   "ov_enhance_services in v1, after the reference tool's slot; it is renamed here "
                   "because a double-weighted criterion should say what it scores. All 20 business value "
                   "scores were recalculated. No gate result changed, because the applications that are "
                   "clinically critical already scored high on both criteria - so the re-weight is the "
                   "right principle without being a disruptive one on this roster."),
        ("bullet", "Q4 asked where the peer cost band behind cost efficiency should come from, given it "
                   "is modelled rather than measured and we have no external benchmark. BINA: "
                   "\"Continue to use the model\". DECISION: the modelled peer band stands, and it is "
                   "labelled as MODELLED wherever it appears (see A5 below and the data dictionary entry "
                   "for c_cost_per_active_user_vs_peers). It is a modelled band, not a measured "
                   "benchmark, and nothing in this file claims otherwise."),
        ("bullet", "Q5 asked whether invest doing double duty - covering both 'healthy, leave alone' and "
                   "'fund a remediation', separated only by priority - reads correctly in front of a "
                   "steering committee. BINA: \"No, separate by invest, retain, consolidate, replace, "
                   "and retire.\" DECISION: five terms. See D1 below for the rule, and the 'Scoring "
                   "model' sheet for the full 16-row mapping. Three rows moved from invest to retain: "
                   + ", ".join(sorted(INTENDED_V2_CHANGES)) + "."),

        ("h2", "Decisions already taken by the team"),
        ("bullet", "D1 - RETAIN AND INVEST ARE SEPARATE WORDS, AND THE GATES DECIDE WHICH (v2, replacing "
                   "v1's D1). retain means healthy, leave it alone, spend nothing. invest means a "
                   "deliberate injection of money or effort. The rule is read off the four gates and "
                   "nothing else: if all four dimensions pass there is no failing dimension to fund, so "
                   "the answer is retain; if any dimension fails, that dimension IS what the money buys, "
                   "so the answer is invest and the recommendation can always name it. That is exactly "
                   "one key of the sixteen - the all-pass key PPPP - which is why the other fifteen rows "
                   "of the lookup table needed no change at all when the vocabulary went from four terms "
                   "to five. PRIORITY NO LONGER CARRIES ANY PART OF THIS DISTINCTION; it is only ever "
                   "about urgency now. In this dataset retain is " + ", ".join(retained)
                   + " and invest is " + ", ".join(invested) + "."),
        ("bullet", "D2 - A cluster survivor keeps its own gate disposition. Cluster membership is an "
                   "override that forces consolidate, but applied literally it stamps consolidate on the "
                   "surviving application too, which reads as 'consolidate Power BI into Power BI'. So "
                   "the override applies to non-survivor members only. The survivor keeps its gate "
                   "result, carries cluster_role = survivor and action = absorb, and its "
                   "consolidation_saving is left blank so a column sum cannot double count. UNDER FIVE "
                   f"TERMS THAT GATE RESULT IS USUALLY NOW retain - {len(surv_retain)} of the "
                   f"{len(survivors)} survivors here ({', '.join(surv_retain)}) - AND IT READS BETTER "
                   f"THAN v1 DID. In v1 the same rows came out 'invest, priority Very Low', which a "
                   f"steering committee hears as a funding request against an application nobody is "
                   f"proposing to spend money on. 'Retain' says what is actually meant: this is the one "
                   f"we are keeping, the money is being spent on moving the others onto it. The two "
                   f"survivors that are NOT retain fail a gate on their own account and are honestly "
                   f"invest: Luma Health on risk, Dragon Copilot on cost efficiency."),
        ("bullet", "D3 - The action verb 'monitor' is what a retain row carries when there is nothing to "
                   "do at all (Epic Hyperspace). 'none' is no longer used on a keep-it row: it read as "
                   "missing data rather than as a decision. The two retain rows that are cluster "
                   "survivors keep the more specific verb 'absorb' - see open item O3."),

        ("h2", "Limitations of the borrowed engine, stated rather than hidden"),
        ("bullet", "L1 - COST NEVER MAKES THE WORD A RETIRE. In the 16-row table the cost dimension "
                   "moves the PRIORITY, and on an otherwise healthy row it turns a retain into an "
                   "invest; business value and technical health decide the word everywhere else. So an "
                   "application can be wildly over-priced and still come out invest. That is a property "
                   "of the gated four-lens approach, not a bug in this dataset - and it is why every cost finding "
                   "in this portfolio surfaces as a saving line and a priority, not as a retire. BINA "
                   "ACCEPTED THIS EXPLICITLY (Q2 above), so from v2 it is enforced in code rather than "
                   "just documented: the generator refuses to write if a cost-only failure ever produces "
                   "a retire."),
        ("bullet", "L2 - The reference tool's disposition vocabulary is wider than what its own lookup "
                   "can return, and one step of its priority ladder is never reached. Neither its words "
                   "nor its values are reproduced here (see the 'Scoring model' sheet). Our own table "
                   "uses all five terms and all five ladder steps: Low is returned on the FPPP row."),
        ("bullet", "L3 - The licensed cost calculator has no line for one-time implementation, no line "
                   "for decommissioning or archival, no line for a replacement's run cost and no line "
                   "for consumption or per-request pricing. All four are OUR extensions here, marked as "
                   "such in the data dictionary, and all four are exactly the fields a savings number "
                   "needs. Without them a retire looks like it returns 100% of an application's cost."),
        ("bullet", "L4 - The licensed cost model is undiscounted and has no NPV, no inflation and no "
                   "depreciation. We kept that convention deliberately and say so: five-year figures "
                   "here are flat multiples of the annual run-rate, and the 15% target is a run-rate "
                   "figure, not a five-year one. Conflating the two is how a savings number gets lost."),

        ("h2", "Scoring assumptions"),
        ("bullet", "A1 - Scores are 1 to 5 in half steps (nine options), matching the licensed engine, "
                   "not five integers. Each dimension is a weighted arithmetic mean, and the pass "
                   "threshold is 3.0 with >= passing, so exactly 3.0 passes."),
        ("bullet", "A2 - Sparse rows are renormalised, not zero-filled. A dimension's weighted mean is "
                   "taken over the criteria that actually carry a value. If we zero-filled blanks, a "
                   "poorly documented application would drift toward retire for lack of evidence, which "
                   "is the single most damaging defect in the workshop-grid approach we started from. "
                   "APP-010 exercises this: "
                   f"{blank_scores} of its weighted score inputs are blank and it still reaches a "
                   "recommendation, at reduced confidence, with the missing field named."),
        ("bullet", "A3 - No fact is scored twice. Version currency, end-of-support proximity, unsupported "
                   "version status and vendor viability belong to technical health and appear nowhere "
                   "else. Single points of failure, missing DR, hardening, PHI exposure, residency, SOC 2 "
                   "and HITRUST posture and lock-in belong to risk. Cost per active user, seat waste and "
                   "consumption variance belong to cost efficiency. Criticality, usage breadth, process "
                   "centrality and owner-stated importance belong to business value. If an obsolescence "
                   "fact were allowed to fail both technical health and risk, one fact would fail two of "
                   "four gates and the engine would systematically over-recommend retiring old but "
                   "adequate applications. APP-017 is the row that proves the partition holds."),
        ("bullet", "A4 - Risk is scored as POSTURE, not as exposure: 5 means the risk is well controlled, "
                   "not that no risk exists. A mission-critical clinical system with tested DR and "
                   "current certifications passes the risk gate even though its inherent clinical "
                   "consequence is high. Raw exposure ratings are what bar an application from wave 1, "
                   "and they are held separately from the score."),
        ("bullet", "A5 - The peer cost band behind c_cost_per_active_user_vs_peers is MODELLED, NOT "
                   "MEASURED. This portfolio has one laboratory application and one ERP, so there is no "
                   "internal peer group for them; those scores are set against an assumed external band. "
                   "Bina's answer to Q4 was to continue to use the model, so the band stands as it is - "
                   "but it is a modelled band and this file says so wherever it appears. Sourcing a real "
                   "benchmark is still open for the live engagement (O2)."),
        ("bullet", "A11 (added in v2, which is why the number sits out of sequence - assumption IDs are "
                   "stable once issued) - The value dimension's double weight sits on patient-care "
                   "criticality from v2 "
                   "(Bina, Q3). The consequence worth knowing: a value score can now clear the 3.0 gate "
                   "on clinical criticality while governance and compliance is mediocre, where in v1 the "
                   "reverse was true. Nothing in this roster actually flips on it - every score moved by "
                   "at most 0.17 and no gate changed - so the re-weight is currently a statement of "
                   "principle rather than a change of answer. On a 600-application portfolio it will "
                   "bite, and that is the point of making it now."),

        ("h2", "Cost and savings assumptions"),
        ("bullet", "A6 - The annual run-rate is the five licensed cost categories plus our consumption "
                   "extension. One-time implementation is held separately and never netted into the "
                   "run-rate. Five-year figures are the annual figure times five, undiscounted."),
        ("bullet", "A7 - Savings net off post-disposition cost. Net saving = gross - the successor's "
                   "ongoing run cost - amortised one-time migration (spread over five years) - residual "
                   "archival cost where a retention obligation outlives the application. A retire never "
                   "returns 100% of an application's cost in this dataset."),
        ("bullet", "A8 - Where the successor is ALREADY a paid line in this portfolio, its cost is not "
                   f"subtracted a second time. {app['APP-018']['app_id']} is the case: its successor is "
                   f"live and dual-running as its own row, so what comes off the saving is amortised "
                   f"migration and the residual payroll archive, not the successor's subscription."),
        ("bullet", "A9 - The unused-licence reharvest on APP-011 is 60% of unused licence spend (only "
                   "seats that can actually be dropped at renewal), minus an explicit 25-seat reserve "
                   "for the users absorbed from APP-010. That reserve is the anti-double-counting line: "
                   "the same dollars cannot both fund the absorbed users and be handed back."),
        ("bullet", "A10 - Priority is computed, never typed. It comes from the lookup table, then three "
                   "documented adjustments: a redundancy-override row takes its priority from urgency, "
                   "lifecycle stage and saving size; a retire constrained by a live retention obligation "
                   "steps DOWN one (the archive has to be built before anything can be switched off); and "
                   "a platform on the critical path of another application's replacement steps UP one."),

        ("h2", "Known deviations from the roster design"),
        ("bullet", f"APP-012 was designed as priority Low, but the all-pass key returns Very Low from "
                   f"the lookup table on every other row. Rather than hand-patch the output we left the "
                   f"computed value ({app['APP-012']['priority']}) and are flagging it: either the design "
                   f"note is a slip, or survivors that carry absorb work should be bumped a step, in "
                   f"which case APP-014 would move too. Bina's call - still open as O1, and unaffected "
                   f"by the five-term change, which moved the word on this row but not the priority."),
        ("bullet", "The roster's indicative completeness score for APP-010 was around 0.35; computed here "
                   f"it is {app['APP-010']['completeness_score']:.2f}, because our denominator is the 44 "
                   f"fields the scoring model actually consumes rather than the full 276-attribute "
                   f"dictionary. The confidence it drives (medium) is as designed."),
        ("bullet", "The Applications sheet carries more columns than the 60-75 originally sketched. The "
                   "enumerated minimum field list needs them; nothing has been padded."),

        ("h2", "Coverage gaps this dataset does not exercise"),
        ("bullet", "All 20 rows are commercial products, so there is no custom in-house application and "
                   "the 'a custom application cannot be upgraded to the vendor's current release' "
                   "exclusion is untested. Add one in the wider generated set, where invented product "
                   "names are fine."),
        ("bullet", "Child tables in the full data dictionary - dependency edges, per-component cost "
                   "basis labels, runbook tasks, cycle snapshots, capability roll-ups - are out of scope "
                   "for this one-row-per-application deliverable. Waves, effort bands and execution "
                   "status are represented only by the urgency inputs and the realization lag."),
    ]


# ---------------------------------------------------------------------------------------
# 8. WRITERS
# ---------------------------------------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
BAND_FILL = PatternFill("solid", fgColor="DCE6F1")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_COLS = {c[1] for c in COLUMNS if c[3].startswith("money")}
RATE_COLS = {c[1] for c in COLUMNS if c[3].startswith("float")}
SCORE_COLS = {c[1] for c in COLUMNS if c[3].startswith("score")}


def cell_value(v):
    return "" if v is None else v


def write_prose(ws, blocks, width=118):
    ws.column_dimensions["A"].width = width
    r = 1
    for kind, text in blocks:
        c = ws.cell(row=r, column=1, value=("- " + text) if kind == "bullet" else text)
        if kind == "h1":
            c.font = Font(bold=True, size=15, color="1F3864")
            r += 1
        elif kind == "h2":
            c.font = Font(bold=True, size=12, color="1F3864")
        else:
            c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = max(14, 13 * (1 + len(text) // 105))
        r += 1
    ws.sheet_view.showGridLines = False


def write_table(ws, headers, rows, widths=None, wrap_cols=(), start=1, number_formats=None):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start, column=j, value=h)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BOX
        c.alignment = Alignment(wrap_text=True, vertical="bottom")
    for i, row in enumerate(rows, start=start + 1):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=cell_value(v))
            c.border = BOX
            c.font = Font(size=10)
            head = headers[j - 1]
            if head in wrap_cols:
                c.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                c.alignment = Alignment(vertical="top")
            if number_formats and head in number_formats:
                c.number_format = number_formats[head]
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=start + 1, column=1)


def sheet_applications(wb, rows):
    ws = wb.create_sheet("Applications")
    themes = {c[1]: c[0] for c in COLUMNS}
    # theme banner above the header row
    for j, name in enumerate(COLUMN_ORDER, start=1):
        c = ws.cell(row=1, column=j, value=themes[name])
        c.fill, c.font, c.border = BAND_FILL, Font(size=8, italic=True, color="1F3864"), BOX
        c.alignment = Alignment(wrap_text=True, vertical="bottom")
        c2 = ws.cell(row=2, column=j, value=name)
        c2.fill, c2.font, c2.border = HEAD_FILL, HEAD_FONT, BOX
        c2.alignment = Alignment(wrap_text=True, vertical="bottom")
    for i, r in enumerate(rows, start=3):
        for j, name in enumerate(COLUMN_ORDER, start=1):
            c = ws.cell(row=i, column=j, value=cell_value(r[name]))
            c.border = BOX
            c.font = Font(size=10)
            c.alignment = Alignment(vertical="top",
                                    wrap_text=name in ("description", "rationale", "suppression_reason",
                                                       "missing_fields", "suppressed_recommendation"))
            if name in MONEY_COLS:
                c.number_format = '#,##0'
            elif name in RATE_COLS:
                c.number_format = '0.000'
            elif name in SCORE_COLS:
                c.number_format = '0.0'
    for j, name in enumerate(COLUMN_ORDER, start=1):
        if name in ("description", "rationale"):
            w = 70
        elif name in ("suppression_reason", "missing_fields", "app_name", "vendor_name"):
            w = 34
        elif name in ("evidence",):
            w = 40
        else:
            w = min(26, max(11, len(name) * 0.95 + 2))
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 46
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMN_ORDER))}{len(rows) + 2}"
    return ws


def sheet_data_dictionary(wb):
    ws = wb.create_sheet("Data dictionary")
    headers = ["Theme", "Column", "Plain-language definition", "Data type / allowed values",
               "Motivating requirement IDs", "Input or computed"]
    body = [[t, n, defn, typ, reqs, "computed" if defn.startswith(COMPUTED_MARK) else "input"]
            for t, n, defn, typ, reqs in COLUMNS]
    write_table(ws, headers, body, widths=[22, 38, 78, 40, 20, 16],
                wrap_cols={"Plain-language definition", "Data type / allowed values"})
    ws.auto_filter.ref = f"A1:F{len(body) + 1}"
    return ws


def sheet_scoring_model(wb, rows):
    ws = wb.create_sheet("Scoring model")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    for col, w in zip("BCDEFGHIJKL", [14, 13, 13, 34, 34, 20, 22, 16, 22, 44, 24]):
        ws.column_dimensions[col].width = w
    r = 1

    def head(text, size=13):
        nonlocal r
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, size=size, color="1F3864")
        r += 2

    def para(text):
        nonlocal r
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = max(14, 13 * (1 + len(text) // 150))
        r += 2

    def table(headers, body):
        nonlocal r
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=j, value=h)
            c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BOX
            c.alignment = Alignment(wrap_text=True, vertical="bottom")
        r += 1
        for line in body:
            for j, v in enumerate(line, start=1):
                c = ws.cell(row=r, column=j, value=cell_value(v))
                c.border, c.font = BOX, Font(size=10)
                c.alignment = Alignment(wrap_text=j >= 5, vertical="top")
            r += 1
        r += 1

    head("How a recommendation is produced", 15)
    para("Read this sheet if you want to argue with an answer. Every number on the Applications sheet "
         "comes out of the six steps below, in this order. Nothing is typed in by hand.")
    para("STEP 1 score. Eighteen inputs are scored 1 to 5 in half steps (nine options, not five). "
         "STEP 2 average. Each of the four dimensions is a weighted average of its own inputs, "
         "renormalised over the inputs that actually carry a value. STEP 3 gate. Each dimension passes "
         "if it reaches 3.0; exactly 3.0 passes. STEP 4 key. The four pass/fail results are read in the "
         "fixed order value, technical health, cost, risk, giving a four-character key such as PPFP. "
         "STEP 5 look up. The key is looked up in the 16-row table below, which returns one of our five "
         "words and a priority. STEP 6 guardrails. Lifecycle, sourcing, redundancy and retention "
         "overrides are applied AFTER the lookup, and every one of them is recorded on the row with its "
         "reason - a suppressed recommendation with a stated reason is a better artifact than a silently "
         "different answer.")

    head("The four dimensions and their weights")
    para("The STRUCTURE follows the licensed Info-Tech portfolio rationalization tool - four weighted "
         "lenses, a pass threshold per lens, a four-character key and a 16-row lookup - and that tool is "
         "cited rather than reproduced here: it is licensed third-party material, this repository is "
         "public, and every dimension name, criterion name, threshold and output word below is the "
         "team's own. Two of the four lens slots are deliberately re-bound: position 3 carries our Cost "
         "Efficiency dimension (REQ 23) and position 4 our Risk dimension (REQ 24/65), rather than the "
         "lenses the reference tool ships in those slots. The reference tool supports substituting lens "
         "sets, so this is a supported customization rather than a hack. A health system's risk axes "
         "change a disposition; a satisfaction survey does not. End-user perceived quality is still "
         "collected, at weight zero.")
    table(["Position", "Our dimension", "Weight sum", "Pass threshold",
           "Column on Applications"],
          [[i + 1, dn, sum(w for _n, dim, w, _d in CRITERIA if dim == k), PASS_THRESHOLD, col]
           for i, (k, dn, col, _flag) in enumerate(DIMENSIONS)])

    head("The 18 scored inputs")
    table(["Input", "Dimension", "Weight", "Normalised weight", "What it scores"],
          [[n, dim, w,
            (round(w / sum(x[2] for x in CRITERIA if x[1] == dim), 4) if w else 0),
            desc] for n, dim, w, desc in CRITERIA])
    para("CHANGED IN v2, on Bina Din's ruling: within the value dimension the double weight moved from "
         "our governance-and-compliance criterion to PATIENT-CARE CRITICALITY, which is the "
         "signal a health system should weigh twice. Governance and compliance drops to weight 1. The "
         "criterion that scores patient-care criticality was called ov_enhance_services in v1, named "
         "after the reference tool's slot; it is renamed ov_patient_care_criticality here, "
         "because a double-weighted criterion ought to say what it measures. The value dimension's weight "
         "sum is 6 either way, so the two versions are directly comparable. Every business value score in "
         "this file was recalculated on the new weights; the largest single move was 0.17 and no "
         "application's pass/fail result changed, because the clinically critical applications already "
         "scored well on both criteria.")
    para("Weight 0 means the input is collected and stored but contributes nothing - the reference "
         "tool's own mechanism for parking a criterion, and the one piece of its mechanics we use "
         "directly. Two inputs sit at zero on purpose: the absolute-dollar cost band (because absolute "
         "cost fails every large enterprise system on sight, while the requirement defines cost "
         "EFFICIENCY relatively - cost per active user against peers) and end-user perceived quality "
         "(which a satisfaction survey should not be able to move a disposition on its own).")

    head("The five words we emit, and the one rule that separates the first two")
    para("Bina Din's ruling of 2026-08-14, in her words: \"No, separate by invest, retain, consolidate, "
         "replace, and retire.\" v1 emitted four words, with invest covering both a healthy application "
         "we leave alone and one we are funding, told apart only by the priority column. There are five "
         "words from v2. retain = healthy, leave it alone, spend nothing. invest = a deliberate "
         "injection of money or effort, a remediation or an enhancement. consolidate = fold it into "
         "another application that keeps the capability alive. replace = substitute a different product "
         "for the same capability. retire = switch it off; the capability goes away or is already "
         "covered elsewhere.")
    para("THE RETAIN / INVEST RULE, which is read off the gates and nothing else: if all four dimensions "
         "PASS, there is no failing dimension to fund, so the answer is retain. If any dimension FAILS, "
         "that dimension is what the money buys, so the answer is invest - and an invest recommendation "
         "can therefore always name what it is funding. PPPF funds risk. PPFP funds cost efficiency. "
         "PFPP funds technical health. That rule touches exactly one row of the sixteen, the all-pass "
         "key PPPP, which is why the other fifteen rows are unchanged from v1: every row that returned "
         "invest in v1 already fails at least one gate. Priority no longer carries any part of this "
         "distinction - it is only ever about urgency. The rule lives in one function, "
         "retain_or_invest(), and the generator additionally CHECKS the table against it on every row, so "
         "the table and the rule cannot drift apart. The 'retain_or_invest_basis' column on the "
         "Applications sheet carries the per-row output of that rule.")
    para("Note what the rule deliberately does NOT do: it does not special-case cost. A cost-efficiency "
         "failure with everything else passing (PPFP) is an invest, never a retire - being expensive is "
         "a reason to fund a fix or negotiate, not a reason to switch a working clinical system off. "
         "That is Bina's answer to the cost question (Q2), and the generator refuses to write the file "
         "if any PPFP row ever resolves to retire.")

    head("The 16-row lookup table - this table IS the mapping decision")
    para("The key on the left is the exhaustive enumeration of four pass/fail gates. The word and the "
         "priority on the right are OURS, from the five agreed terms and our own five-step ladder. The "
         "licensed reference tool has a lookup of the same shape carrying its own vocabulary; it is "
         "cited, not reproduced, and nothing in this table is copied from it - so read this as the "
         "team's mapping decision rather than as a rename of anyone else's. Two rows show why the "
         "decision needs making: FFPP maps to consolidate rather than retire, because value failing "
         "while technical health passes is a case where the capability should MOVE, not disappear; and "
         "FPFP and FPFF split away from FPPP/FPPF, because a low-value application that is also "
         "expensive or risky is a retire while a low-value application that is cheap and safe is only a "
         "consolidate. Every row is configuration: a client can argue with any of them directly.")
    used = Counter(r["vtcr_key"] for r in rows)
    table(["Key", "Value", "Technical health", "Cost efficiency", "Risk",
           "OUR disposition (five terms, v2)", "OUR priority",
           "v1 emitted (four terms)", "Changed in v2?", "Rows in this dataset"],
          [[k, "Pass" if k[0] == "P" else "Fail", "Pass" if k[1] == "P" else "Fail",
            "Pass" if k[2] == "P" else "Fail", "Pass" if k[3] == "P" else "Fail",
            v[0], v[1],
            V1_DISPOSITION_MAP[k],
            ("YES - the only changed row. All four gates pass, so nothing needs funding and "
             "'invest' was overstating it." if V1_DISPOSITION_MAP[k] != v[0] else "no"),
            ", ".join(r["app_id"] for r in rows if r["vtcr_key"] == k) or "-"]
           for k, v in DISPOSITION_TABLE.items()])
    para(f"Keys exercised by these 20 rows: {', '.join(f'{k} x{v}' for k, v in sorted(used.items()))}. "
         f"Four of the sixteen patterns are unexercised, which is a property of a 20-row roster, not a "
         f"gap in the table.")
    para("Read the 'Changed in v2?' column together with the rule above: exactly one of the sixteen "
         "patterns moved, and it moved because it is the only pattern where no dimension fails. Every "
         "other pattern already had a failing dimension to point at, so 'invest' was always the honest "
         "word for it and stays.")

    head("Guardrails, in the order they run after the lookup")
    table(["Order", "Guardrail", "What it does", "Rows it touches here"],
          [[1, "Lifecycle (Birth/Growth)",
            "Bars retire and replace outright and records the guard state. Does NOT bar consolidate.",
            ", ".join(r["app_id"] for r in rows if r["lifecycle_exclusion_applied"])],
           [2, "Sourcing (SaaS)",
            "A SaaS product cannot be rebuilt or re-platformed in place, so a replace becomes product "
            "substitution and the action is annotated.",
            ", ".join(r["app_id"] for r in rows if r["sourcing_exclusion_applied"])],
           [3, "Redundancy (cluster membership)",
            "Forces consolidate on non-survivor cluster members, whatever the gates said - including on "
            "an application whose four gates all pass. Survivors keep their gate disposition (D2).",
            ", ".join(r["app_id"] for r in rows if r["redundancy_override_applied"])],
           [4, "Retention (obligation not expired)",
            "A retire stays a retire but becomes CONSTRAINED: priority steps down one, the residual "
            "archival cost is netted off the saving, and the constraint's lift date is recorded.",
            ", ".join(r["app_id"] for r in rows if r["retention_override_applied"])],
           [5, "Successor on a critical path",
            "A platform another application's replacement depends on steps UP one priority.",
            ", ".join(r["app_id"] for r in rows
                      if "critical path" in (r["suppression_reason"] or ""))]])
    return ws


def sheet_clusters(wb, rows):
    ws = wb.create_sheet("Overlap clusters")
    by_id = {r["app_id"]: r for r in rows}
    body = []
    for c in CLUSTERS:
        absorbed = [m for m in c["members"] if m != c["survivor"]]
        gross = sum(by_id[m]["gross_saving_annual"] for m in absorbed)
        net = sum(by_id[m]["net_saving_annual"] for m in absorbed)
        body.append([
            c["cluster_id"], c["capability"],
            "; ".join(f"{m} {by_id[m]['app_name']}" for m in c["members"]),
            f"{c['survivor']} {by_id[c['survivor']]['app_name']}",
            c["survivor_reason"],
            "; ".join(f"{m} {by_id[m]['app_name']} ({by_id[m]['disposition']}, "
                      f"{by_id[m]['priority']})" for m in absorbed),
            c["user_overlap_count"], gross, net,
            "yes" if c["dependency_check_passed"] else "NO - see note",
            c["wave"], c["note"],
        ])
    headers = ["Cluster", "Capability", "Members", "Survivor", "Why that survivor",
               "Absorbed (and their computed disposition)", "Users holding 2+ of the members",
               "Gross annual saving", "Net annual saving", "Dependency check passed", "Wave", "Note"]
    write_table(ws, headers, body, widths=[9, 26, 46, 30, 62, 44, 14, 15, 15, 14, 10, 62],
                wrap_cols={"Why that survivor", "Members", "Note",
                           "Absorbed (and their computed disposition)"},
                number_formats={"Gross annual saving": '#,##0', "Net annual saving": '#,##0'})
    total_row = len(body) + 2
    ws.cell(row=total_row, column=7, value="TOTAL").font = Font(bold=True)
    for col, key in ((8, "gross_saving_annual"), (9, "net_saving_annual")):
        c = ws.cell(row=total_row, column=col,
                    value=sum(by_id[m][key] for cl in CLUSTERS for m in cl["members"]
                              if m != cl["survivor"]))
        c.font, c.number_format = Font(bold=True), '#,##0'
    return ws


def sheet_traps(wb, rows):
    ws = wb.create_sheet("Trap cases")
    by_id = {r["app_id"]: r for r in rows}
    body = []
    for num, apps, tests, correct, wrong in TRAP_CASES:
        ids = [a for a in apps.replace("->", "and").split() if a.startswith("APP-")]
        computed = " | ".join(
            f"{i}: {by_id[i]['vtcr_key']} -> {by_id[i]['disposition']} / {by_id[i]['priority']} "
            f"(confidence {by_id[i]['confidence']})" for i in ids)
        body.append([num, apps, tests, correct, wrong, computed])
    headers = ["#", "Row(s)", "What it tests", "The correct output",
               "The wrong answer it is designed to catch", "What this dataset actually computes"]
    write_table(ws, headers, body, widths=[6, 20, 54, 60, 60, 58],
                wrap_cols={"What it tests", "The correct output",
                           "The wrong answer it is designed to catch",
                           "What this dataset actually computes"})
    return ws


def write_xlsx(rows, summary, path):
    wb = Workbook()
    wb.remove(wb.active)
    write_prose(wb.create_sheet("Read me first"), readme_blocks(rows, summary))
    sheet_applications(wb, rows)
    sheet_data_dictionary(wb)
    sheet_scoring_model(wb, rows)
    sheet_clusters(wb, rows)
    sheet_traps(wb, rows)
    write_prose(wb.create_sheet("Notes & assumptions"), notes_blocks(rows, summary))
    wb.save(path)


def csv_value(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return v


def write_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(COLUMN_ORDER)
        for r in rows:
            w.writerow([csv_value(r[c]) for c in COLUMN_ORDER])


def write_changelog(rows, summary, path):
    """CHANGELOG-v2.md - short, plain language, no jargon."""
    changed = [r for r in rows if V1_OUTPUT[r["app_id"]][0] != r["disposition"]]
    md = [
        "# What changed in v2, and why\n",
        f"`{XLSX_NAME}` | `{CSV_NAME}` | {DATASET_VERSION}. Replaces "
        f"`App-Rationalization-Dummy-Dataset-v1.xlsx`.\n",
        "Bina Din, the subject-matter expert, reviewed v1 and answered the five open questions it "
        "carried. Two of her answers changed the file. Her exact words are on the 'Notes & assumptions' "
        "sheet, each next to the change it produced.\n",

        "## 1. There are five recommendation words now, not four\n",
        "Bina's words: *\"No, separate by invest, retain, consolidate, replace, and retire.\"*\n",
        "In v1 the word **invest** had to do two jobs. It meant \"this application is healthy, leave it "
        "alone\" AND it meant \"we are putting money into this application\". The only thing telling "
        "them apart was the priority column, which is a lot to ask of a reader in a steering committee. "
        "So v1 said *invest, priority Very Low* about Epic Hyperspace - the single largest line item in "
        "the portfolio - when what it meant was *leave it alone*.\n",
        "v2 has a separate word for each:\n",
        "| Word | What it means |\n|---|---|\n"
        "| **retain** | Healthy. Leave it alone. Spend nothing. |\n"
        "| **invest** | Deliberately put money or effort in - fund a remediation or an enhancement. |\n"
        "| **consolidate** | Fold it into another application that keeps the capability alive. |\n"
        "| **replace** | Swap in a different product for the same capability. |\n"
        "| **retire** | Switch it off. The capability goes away, or is already covered elsewhere. |\n",

        "### How the tool decides between retain and invest\n",
        "One rule, read straight off the four pass/fail gates:\n",
        "- **All four dimensions pass** -> there is no failing dimension to fund, so nothing needs "
        "money. **retain**.\n"
        "- **Any dimension fails** -> that dimension is exactly what the money would buy. **invest**, "
        "and the recommendation names it.\n",
        "That is why an invest can always answer \"invest in what?\" - risk, or cost efficiency, or "
        "technical health. It also means only one of the sixteen possible pass/fail patterns changed "
        "meaning between v1 and v2: the all-pass one. Every other pattern already had a failing "
        "dimension to point at, so invest was always the honest word for it.\n",
        "Priority no longer carries any part of this distinction. Priority is only about urgency now.\n",

        "### Which applications moved\n",
        f"Three of the twenty, all of them from invest to retain, and no priority changed:\n",
        "| Application | v1 said | v2 says | Why |\n|---|---|---|---|",
    ]
    for r in changed:
        md.append(f"| {r['app_id']} {r['app_name']} | {V1_OUTPUT[r['app_id']][0]} / "
                  f"{V1_OUTPUT[r['app_id']][1]} | **{r['disposition']}** / {r['priority']} | "
                  f"{INTENDED_V2_CHANGES[r['app_id']][1]} |")
    md.append("")
    md.append("Two of those three are cluster survivors - the application a group of overlapping "
              "products gets folded into. v1 called them *invest, priority Very Low*, which sounds like "
              "a funding request against an application nobody was proposing to spend money on. "
              "**retain** says what is actually meant: this is the one we are keeping, and the money is "
              "being spent on moving the others onto it. That reads considerably better than v1 did.\n")
    md.append("The seventeen other applications say exactly what they said in v1.\n")

    md.append("## 2. Patient-care criticality now counts double in the value score\n")
    md.append("Bina's answer to the question of which value signal a health system should weigh twice "
              "was **yes** - patient-care criticality, not the scoring engine's own 'governance and "
              "compliance' criterion. So:\n")
    md.append("- Patient-care criticality moves from weight 1 to **weight 2**.\n"
              "- Governance and compliance moves from weight 2 to **weight 1**.\n"
              "- The column that holds the score is renamed from `ov_enhance_services` to "
              "`ov_patient_care_criticality`, so it says what it measures. This is the one column name "
              "that changed between v1 and v2.\n")
    md.append("Every business value score was recalculated. The largest single change was 0.17 of a "
              "point, and **no application's pass/fail result changed** - the applications that are "
              "clinically critical already scored well on both criteria. So this is the right principle "
              "without being a disruptive change on a 20-row roster. On a 600-application portfolio it "
              "will matter more, which is the argument for making it now.\n")

    md.append("## 3. Bina's other three answers confirmed v1 as it stood\n")
    md.append("- **Risk in the end-user slot** - keep it. Risk stays as the fourth scoring dimension "
              "and end-user perception is still collected at zero weight.\n"
              "- **Cost moves priority but never on its own makes something a retire** - accepted. From "
              "v2 this is enforced in the generator rather than just described: if an application fails "
              "only on cost and the engine ever returns retire, the script refuses to write the file.\n"
              "- **The modelled peer cost band** - continue to use the model. It stands, and it is "
              "labelled as modelled rather than measured everywhere it appears.\n")

    md.append("## What did NOT change\n")
    md.append("Deliberately identical to v1: the same 20 applications, the same real product names with "
              "every other value invented, the same overlap clusters and survivors, the same trap cases, "
              "and the same cost and savings arithmetic. Savings still net off a replacement's run cost "
              "and any residual archival cost, and a cluster survivor still keeps its own gate result "
              "rather than being stamped consolidate.\n")
    md.append(f"The money is unchanged to the dollar: portfolio run-rate "
              f"${summary['total_annual_tco']:,.0f}, net annual saving identified "
              f"${summary['net_saving_all']:,.0f}, of which "
              f"${summary['net_saving_excluding_constrained_and_deferred']:,.0f} is available once the "
              f"blocked and deferred rows are set aside.\n")
    md.append("## New in the reviewer's answer key\n")
    md.append("The 'Trap cases' sheet went from twelve rows to fourteen. One was rewritten and two are "
              "new, all because of the vocabulary change:\n")
    md.append("- **T12 rewritten.** In v1 this tested that two applications with different problems both "
              "came out invest and were told apart by priority. It now tests that they come out on "
              "*different words*: Epic Hyperspace retain, Luma Health invest.\n"
              "- **T13 new.** A cluster survivor must come out retain - not consolidate (which would "
              "read as 'consolidate Power BI into Power BI') and not invest (nothing about the survivor "
              "is being funded).\n"
              "- **T14 new.** An application that fails only on cost must come out invest. Not retire, "
              "which Bina's answer rules out, and not retain, which would ignore a failing gate.\n")
    md.append("Trap case T1 also gained a new wrong answer to catch: Sunquest CoPath Plus is old but "
              "adequate, and it must still come out **invest** rather than retire *or* retain. Its "
              "technical health genuinely fails, so there is genuinely something to fund.\n")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(md))


def write_readme(rows, summary, verify_lines, path):
    t = summary
    md = []
    md.append("# Application Rationalization - synthetic demo dataset\n")
    md.append(f"`{XLSX_NAME}` | `{CSV_NAME}` | `generate_dataset.py` | {DATASET_VERSION}\n")
    md.append("## What this is\n")
    md.append("A 20-application pretend portfolio for a pretend US health system, carrying every data "
              "point needed to run an application rationalization exercise end to end: identity and "
              "lifecycle, ownership, capability, usage against entitlement, a full cost breakdown, "
              "contract terms, dependencies, the 18 scored inputs, the four dimension scores, the "
              "pass/fail gate, the recommendation with its rationale, the overlap clusters and the "
              "savings arithmetic.\n")
    md.append("**The product and vendor names are real. Everything else is invented.** Every cost, "
              "contract identifier and date, licence and user count, version, score, saving, owner name, "
              "legal entity, business unit and cost centre was made up for this dataset. No real "
              "organisation's portfolio, contract terms or internal figures appear anywhere.\n")
    md.append("## Files\n")
    md.append(f"| File | What it is |\n|---|---|\n"
              f"| `{XLSX_NAME}` | The deliverable. Seven sheets: Read me first, Applications, "
              f"Data dictionary, Scoring model, Overlap clusters, Trap cases, Notes & assumptions. |\n"
              f"| `{CSV_NAME}` | The Applications sheet as flat CSV, for loading into code. |\n"
              f"| `generate_dataset.py` | The generator. Re-run it to reproduce or extend the dataset. |\n")
    md.append("## How the generator works\n")
    md.append("Row data is explicit and readable - one dict per application, hand-written, so anyone can "
              "edit a number or add a row. Everything derived is computed by functions, never typed:\n")
    md.append("1. **Score.** 18 inputs per app, 1-5 in half steps, four dimensions: business value, "
              "technical health, cost efficiency, risk posture.\n"
              "2. **Average.** Each dimension is a weighted mean, renormalised over the inputs that "
              "actually carry a value, so a sparse row is reported as incomplete rather than being "
              "quietly scored down toward retire.\n"
              "3. **Gate.** Each dimension passes at 3.0 (>= passes).\n"
              "4. **Key.** The four flags concatenate to a four-character key, read value / technical "
              "health / cost / risk.\n"
              "5. **Look up.** A 16-row table returns our term from retain / invest / consolidate / "
              "replace / retire, plus a priority. The table is the team's own mapping decision; the "
              "licensed reference tool is cited, never reproduced. "
              "retain and invest are separated by one rule: all four gates passing "
              "means nothing needs funding (retain), any gate failing means that dimension is what the "
              "money buys (invest). The table is configuration; the 'Scoring model' sheet documents "
              "the mapping.\n"
              "6. **Guardrails, after the lookup.** Lifecycle (Birth/Growth bars retire and replace), "
              "sourcing (SaaS cannot be re-platformed in place), redundancy (cluster membership forces "
              "consolidate on non-survivors only), retention (a retire with a live obligation becomes "
              "constrained). Every one is recorded on the row with its reason.\n"
              "7. **Cost and savings.** Five licensed cost categories plus our consumption extension = "
              "annual run-rate; one-time implementation held separately; five-year figures flat and "
              "undiscounted. Net saving = gross - successor run cost - amortised one-time - residual "
              "archival.\n"
              "8. **Verify.** Every computed disposition is compared against the intended disposition in "
              "the roster design, and the script fails if any row disagrees.\n")
    md.append("## v1 vs v2 - every row\n")
    md.append("| app | name | key | v1 disposition | v2 disposition | priority | changed? |"
              "\n|---|---|---|---|---|---|---|")
    for r in rows:
        v1_disp, v1_pri = V1_OUTPUT[r["app_id"]]
        changed = v1_disp != r["disposition"]
        md.append(f"| {r['app_id']} | {r['app_name']} | {r['vtcr_key']} | {v1_disp} | "
                  f"{'**' + r['disposition'] + '**' if changed else r['disposition']} | "
                  f"{r['priority'] if r['priority'] == v1_pri else v1_pri + ' -> ' + r['priority']} | "
                  f"{'yes, intended' if changed else 'no'} |")
    md.append("")
    md.append("## Verification - computed vs intended\n")
    md.append("| app | name | key | intended | computed | match | intended priority | "
              "computed priority |\n|---|---|---|---|---|---|---|---|")
    for r in rows:
        md.append(f"| {r['app_id']} | {r['app_name']} | {r['vtcr_key']} | "
                  f"{r['_intended_disposition']} | {r['disposition']} | "
                  f"{'yes' if r['disposition'] == r['_intended_disposition'] else 'NO'} | "
                  f"{r['_intended_priority']} | {r['priority']} |")
    md.append("")
    md.append("## Headline numbers\n")
    md.append(f"- Portfolio annual run-rate: **${t['total_annual_tco']:,.0f}** across 20 applications.\n"
              f"- A 15% reduction target is **${t['target']:,.0f}** a year.\n"
              f"- Net annual saving identified: **${t['net_saving_all']:,.0f}**, of which "
              f"**${t['net_saving_excluding_constrained_and_deferred']:,.0f}** is available excluding the "
              f"constrained and deferred rows ({', '.join(t['constrained_or_deferred'])}) - the honest "
              f"version of the number.\n"
              f"- Dispositions: " + ", ".join(f"{k} {v}" for k, v in
                                              sorted(t['spread'].items(), key=lambda x: -x[1])) + ".\n")
    md.append("## Assumptions worth knowing\n")
    for kind, text in notes_blocks(rows, summary):
        if kind == "bullet":
            md.append(f"- {text}")
        elif kind == "h2":
            md.append(f"\n### {text}\n")
    # Bina's five answers are already carried above, under their own heading in notes_blocks.
    md.append("\n## Still open - for Ryo and Bina\n")
    for kind, text in readme_blocks(rows, summary):
        if kind == "bullet" and text.startswith("O"):
            md.append(f"- {text}")
    md.append("")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(md))


# ---------------------------------------------------------------------------------------
# 9. MAIN
# ---------------------------------------------------------------------------------------
def main():
    rows = build()

    problems = sanity_checks(rows)
    lines, mismatches = verification_table(rows)
    comp_lines, unexpected = comparison_table(rows)

    print("v1 -> v2 COMPARISON")
    print("\n".join(comp_lines))
    if unexpected:
        print(f"\nUNINTENDED DISPOSITION CHANGES vs v1: {unexpected}")
        problems.append(f"v2 changed dispositions v1 did not intend to change: {unexpected}")
    else:
        print(f"\nEvery v1 -> v2 disposition change is an intended one "
              f"({', '.join(sorted(INTENDED_V2_CHANGES))} - all PPPP invest -> retain). "
              f"No priority moved.")

    print("\nCOMPUTED vs INTENDED (roster design)")
    print("\n".join(lines))
    print()
    if mismatches:
        print(f"DISPOSITION MISMATCHES: {mismatches} - fix the INPUTS (scores, costs, usage), "
              f"never the output.")
    else:
        print("All 20 computed dispositions match the intended dispositions in roster-design.md.")
    pri_mismatch = [r["app_id"] for r in rows if r["priority"] != r["_intended_priority"]]
    print(f"Priority mismatches (recorded as deviations, not patched): {pri_mismatch or 'none'}")
    key_mismatch = [f"{r['app_id']} {r['_intended_key']}->{r['vtcr_key']}"
                    for r in rows if r["vtcr_key"] != r["_intended_key"]]
    print(f"Pattern-key mismatches vs the design's intended key: {key_mismatch or 'none'}")

    if problems:
        print("\nARITHMETIC / VOCABULARY PROBLEMS:")
        for p in problems:
            print("  -", p)
    else:
        print("Arithmetic and vocabulary checks: all clean.")

    summary = rollup(rows)
    print(f"\nPortfolio annual run-rate      ${summary['total_annual_tco']:,}")
    print(f"15% target                     ${summary['target']:,}")
    print(f"Net annual saving (all)        ${summary['net_saving_all']:,}")
    print(f"Net annual saving (available)  "
          f"${summary['net_saving_excluding_constrained_and_deferred']:,} "
          f"excluding {', '.join(summary['constrained_or_deferred'])}")
    print(f"Disposition spread             {dict(summary['spread'])}")
    print(f"Priority spread                {dict(summary['priority_spread'])}")
    print(f"Columns on the Applications sheet: {len(COLUMN_ORDER)}")

    if mismatches or problems:
        raise SystemExit("Refusing to write output while the dataset disagrees with its design.")

    write_xlsx(rows, summary, os.path.join(OUT_DIR, XLSX_NAME))
    write_csv(rows, os.path.join(OUT_DIR, CSV_NAME))
    write_readme(rows, summary, lines, os.path.join(OUT_DIR, README_NAME))
    write_changelog(rows, summary, os.path.join(OUT_DIR, CHANGELOG_NAME))
    print(f"\nWrote:\n  {XLSX_NAME}\n  {CSV_NAME}\n  {README_NAME}\n  {CHANGELOG_NAME}\n"
          f"in {OUT_DIR}")


if __name__ == "__main__":
    main()


